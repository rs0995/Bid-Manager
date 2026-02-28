import datetime
import json
import html
import os
import shutil
import sqlite3
import subprocess
import sys
import threading
import tempfile
import zipfile
import base64
import time
import importlib
import textwrap
from concurrent.futures import ThreadPoolExecutor
import queue as py_queue

from PySide6.QtCore import QByteArray, QDateTime, QEasingCurve, QEvent, QItemSelectionModel, QPoint, QPropertyAnimation, QRegularExpression, QTime, QTimer, Qt
from PySide6.QtGui import QColor, QFont, QImage, QKeyEvent, QPainter, QPixmap, QPolygon, QRegularExpressionValidator
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCalendarWidget,
    QCheckBox,
    QComboBox,
    QCompleter,
    QDateTimeEdit,
    QDialog,
    QFileDialog,
    QFormLayout,
    QFrame,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMessageBox,
    QInputDialog,
    QPushButton,
    QTabWidget,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextBrowser,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QGridLayout,
    QSizePolicy,
    QStyle,
    QVBoxLayout,
    QWidget,
)

import app_core as core
from frontend.api_client import BidApiClient
from templates_ui import TemplatesPage, import_templates_into_project

FRONTEND_REMOTE_ONLY = str(os.getenv("BID_FRONTEND_REMOTE_ONLY", "") or "").strip().lower() in {"1", "true", "yes", "on"}


class BackendModeScraperProxy:
    def __init__(self):
        self.local = core.ScraperBackend

    def _mode(self):
        if FRONTEND_REMOTE_ONLY:
            return "remote"
        raw = str(core.get_user_setting("backend_mode", "local") or "local").strip().lower()
        return "remote" if raw == "remote" else "local"

    def _remote_config(self):
        url = str(core.get_user_setting("backend_url", "") or "").strip().rstrip("/")
        api_key = str(core.get_user_setting("backend_api_key", "") or "").strip()
        return url, api_key

    def _remote_enabled(self):
        url, api_key = self._remote_config()
        return self._mode() == "remote" and bool(url and api_key)

    def _new_client(self):
        url, api_key = self._remote_config()
        if not url or not api_key:
            raise RuntimeError("Remote backend URL/API key is not configured.")
        return BidApiClient(base_url=url, api_key=api_key, timeout_seconds=120)

    def _encode_local_db_b64(self):
        db_path = core._resolve_path(core.DB_FILE)
        if not os.path.exists(db_path):
            core.init_db()
        with open(db_path, "rb") as f:
            return base64.b64encode(f.read()).decode("ascii")

    def _merge_folder_tree(self, src_root, dst_root):
        copied = 0
        if not os.path.isdir(src_root):
            return copied
        os.makedirs(dst_root, exist_ok=True)
        for root, _dirs, files in os.walk(src_root):
            rel = os.path.relpath(root, src_root)
            target_dir = dst_root if rel in (".", "") else os.path.join(dst_root, rel)
            os.makedirs(target_dir, exist_ok=True)
            for name in files:
                src = os.path.join(root, name)
                dst = os.path.join(target_dir, name)
                try:
                    shutil.copy2(src, dst)
                    copied += 1
                except Exception:
                    pass
        return copied

    def _apply_remote_artifact(self, zip_path):
        copied_download_files = 0
        tmp_dir = tempfile.mkdtemp(prefix="bm_remote_")
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                zf.extractall(tmp_dir)

            state_db = os.path.join(tmp_dir, "__state", "tender_manager.db")
            if os.path.isfile(state_db):
                dst_db = core._resolve_path(core.DB_FILE)
                os.makedirs(os.path.dirname(dst_db) or ".", exist_ok=True)
                shutil.copy2(state_db, dst_db)

            for item in os.listdir(tmp_dir):
                if item == "__state":
                    continue
                src = os.path.join(tmp_dir, item)
                dst = os.path.join(core.BASE_DOWNLOAD_DIRECTORY, item)
                if os.path.isdir(src):
                    copied_download_files += self._merge_folder_tree(src, dst)
                elif os.path.isfile(src):
                    os.makedirs(core.BASE_DOWNLOAD_DIRECTORY, exist_ok=True)
                    try:
                        shutil.copy2(src, os.path.join(core.BASE_DOWNLOAD_DIRECTORY, item))
                        copied_download_files += 1
                    except Exception:
                        pass
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        return copied_download_files

    def _run_remote_action(self, action, payload):
        client = self._new_client()
        body = dict(payload or {})
        job = client.create_job(action=action, payload=body, build_artifact=True)
        job_id = str(job.get("job_id") or "")
        if not job_id:
            raise RuntimeError("Remote job creation failed.")

        while True:
            job = client.get_job(job_id)
            status = str(job.get("status") or "")
            if status == "captcha_required":
                cap = job.get("captcha") or {}
                challenge_id = str(cap.get("challenge_id") or "")
                image_b64 = str(cap.get("image_base64") or "")
                if not challenge_id or not image_b64:
                    raise RuntimeError("Remote captcha payload is invalid.")
                img_data = base64.b64decode(image_b64.encode("ascii"))
                core.captcha_req_queue.put(img_data)
                answer = core.captcha_res_queue.get()
                if not answer:
                    raise RuntimeError("Captcha cancelled.")
                client.submit_captcha(job_id=job_id, challenge_id=challenge_id, value=str(answer))
            elif status == "failed":
                raise RuntimeError(str(job.get("error") or "Remote job failed."))
            elif status == "completed":
                result = job.get("result") or {}
                if bool(result.get("artifact_available")):
                    tmp_zip = os.path.join(tempfile.gettempdir(), f"bm_artifact_{job_id}.zip")
                    client.download_artifact(job_id, tmp_zip)
                    copied = self._apply_remote_artifact(tmp_zip)
                    try:
                        os.remove(tmp_zip)
                    except Exception:
                        pass
                    core.log_to_gui(f"Remote sync complete. Download files synced: {copied}")
                return True
            else:
                time.sleep(1.0)

    def _run_or_local(self, method_name, action, payload, *args, **kwargs):
        if self._mode() == "remote":
            url, api_key = self._remote_config()
            if not url or not api_key:
                raise RuntimeError("Remote backend is required. Configure Backend URL and API key in Settings.")
            return self._run_remote_action(action=action, payload=payload)
        return getattr(self.local, method_name)(*args, **kwargs)

    def get_setting(self, key, default=None):
        return self.local.get_setting(key, default)

    def set_setting(self, key, value):
        return self.local.set_setting(key, value)

    def get_websites(self):
        return self.local.get_websites()

    def add_website_logic(self, name, url, status_url):
        return self.local.add_website_logic(name, url, status_url)

    def delete_website_logic(self, website_id):
        return self.local.delete_website_logic(website_id)

    def clear_saved_scraper_details_logic(self, clear_orgs=False, clear_active=False, clear_archived=False):
        return self.local.clear_saved_scraper_details_logic(
            clear_orgs=clear_orgs,
            clear_active=clear_active,
            clear_archived=clear_archived,
        )

    def fetch_organisations_logic(self, website_id):
        return self._run_or_local(
            "fetch_organisations_logic",
            action="fetch_organisations",
            payload={"website_id": int(website_id)},
            website_id=int(website_id),
        )

    def sync_remote_state(self):
        if not self._remote_enabled():
            return True
        return self._run_remote_action(action="sync_state", payload={})

    def push_local_state(self):
        if not self._remote_enabled():
            return True
        return self._run_remote_action(
            action="sync_state",
            payload={"db_snapshot_base64": self._encode_local_db_b64()},
        )

    def fetch_tenders_logic(self, website_id):
        return self._run_or_local(
            "fetch_tenders_logic",
            action="fetch_tenders",
            payload={"website_id": int(website_id)},
            website_id=int(website_id),
        )

    def download_tenders_logic(self, website_id, target_db_ids=None, forced_mode=None):
        return self._run_or_local(
            "download_tenders_logic",
            action="download_tenders",
            payload={
                "website_id": int(website_id),
                "target_db_ids": target_db_ids,
                "forced_mode": forced_mode,
            },
            website_id=int(website_id),
            target_db_ids=target_db_ids,
            forced_mode=forced_mode,
        )

    def check_tender_status_logic(self, website_id, archived_only=False):
        return self._run_or_local(
            "check_tender_status_logic",
            action="check_status",
            payload={"website_id": int(website_id), "archived_only": bool(archived_only)},
            website_id=int(website_id),
            archived_only=bool(archived_only),
        )

    def download_tender_results_logic(self, website_id):
        return self._run_or_local(
            "download_tender_results_logic",
            action="download_results",
            payload={"website_id": int(website_id)},
            website_id=int(website_id),
        )

    def download_single_tender_logic(self, tender_db_id, mode):
        return self._run_or_local(
            "download_single_tender_logic",
            action="single_download",
            payload={"tender_db_id": int(tender_db_id), "mode": str(mode)},
            tender_db_id=int(tender_db_id),
            mode=str(mode),
        )

    def _remote_download_to_folder(self, source_tender_id, destination_folder, mode):
        self._run_remote_action(
            action="deliver_tender_docs",
            payload={"source_tender_id": str(source_tender_id or "").strip(), "mode": str(mode or "full").strip().lower()},
        )
        safe_id = core.re.sub(r'[\\/*?:"<>|]', "", str(source_tender_id or "").strip())
        conn = sqlite3.connect(core.DB_FILE)
        try:
            row = conn.execute(
                "SELECT COALESCE(folder_path,'') FROM tenders WHERE tender_id=? AND COALESCE(is_archived,0)=0 ORDER BY id DESC LIMIT 1",
                (str(source_tender_id),),
            ).fetchone()
        finally:
            conn.close()
        folder_path = str(row[0] or "") if row else ""
        src = folder_path if folder_path and os.path.isdir(folder_path) else os.path.join(core.BASE_DOWNLOAD_DIRECTORY, safe_id)
        copied = self._merge_folder_tree(src, str(destination_folder))
        return bool(copied or os.path.isdir(src))

    def download_docs_for_tender_to_folder(self, source_tender_id, destination_folder):
        if not self._remote_enabled():
            return self.local.download_docs_for_tender_to_folder(source_tender_id, destination_folder)
        return self._remote_download_to_folder(source_tender_id, destination_folder, "full")

    def download_updates_for_tender_to_folder(self, source_tender_id, destination_folder):
        if not self._remote_enabled():
            return self.local.download_updates_for_tender_to_folder(source_tender_id, destination_folder)
        return self._remote_download_to_folder(source_tender_id, destination_folder, "update")

def auto_fit_table_rows(table, min_height=24, max_height=None):
    table.resizeRowsToContents()
    for row in range(table.rowCount()):
        h = table.rowHeight(row)
        if h < min_height:
            table.setRowHeight(row, min_height)
        elif max_height is not None and h > max_height:
            table.setRowHeight(row, max_height)


def _project_tender_info_key(project_id):
    pid = int(project_id or 0)
    return f"project_tender_extra_info_{pid}"


def _load_project_tender_info(project_id):
    raw = core.get_user_setting(_project_tender_info_key(project_id), [])
    if not isinstance(raw, list):
        return []
    out = []
    for row in raw:
        if not isinstance(row, dict):
            continue
        k = str(row.get("key", "") or "").strip()
        v = str(row.get("value", "") or "").strip()
        if k and v:
            out.append({"key": k, "value": v})
    return out


def _save_project_tender_info(project_id, items):
    payload = []
    for row in (items or []):
        if not isinstance(row, dict):
            continue
        k = str(row.get("key", "") or "").strip()
        v = str(row.get("value", "") or "").strip()
        if k and v:
            payload.append({"key": k, "value": v})
    core.set_user_setting(_project_tender_info_key(project_id), payload)


def _time_remaining_from_text(deadline_text):
    txt = str(deadline_text or "").strip()
    if not txt:
        return "-"
    fmts = [
        "%d-%m-%Y %I:%M %p",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ]
    dt = None
    for fmt in fmts:
        try:
            dt = datetime.datetime.strptime(txt, fmt)
            break
        except Exception:
            continue
    if dt is None:
        return "-"
    now = datetime.datetime.now()
    secs = int((dt - now).total_seconds())
    if secs <= 0:
        return "Expired"
    days = secs // 86400
    hours = (secs % 86400) // 3600
    mins = (secs % 3600) // 60
    if days > 0:
        return f"{days}d {hours}h {mins}m"
    if hours > 0:
        return f"{hours}h {mins}m"
    return f"{mins}m"


def _collect_project_tender_base_info(project_id):
    pid = int(project_id or 0)
    if pid <= 0:
        return []
    conn = sqlite3.connect(core.DB_FILE)
    try:
        p = conn.execute(
            """SELECT COALESCE(title,''), COALESCE(description,''), COALESCE(client_name,''), COALESCE(project_value,''),
                      COALESCE(prebid,''), COALESCE(deadline,''), COALESCE(source_tender_id,'')
               FROM projects WHERE id=?""",
            (pid,),
        ).fetchone()
        if not p:
            return []
        proj_tender_id = str(p[6] or "").strip() or str(p[0] or "").strip()

        t_row = None
        if proj_tender_id:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(tenders)").fetchall()]
            has_closing_time = "closing_time" in set(cols)
            closing_time_sql = "COALESCE(closing_time,'')" if has_closing_time else "''"
            t_row = conn.execute(
                """SELECT COALESCE(tender_id,''), COALESCE(work_description,''), COALESCE(title,''),
                          COALESCE(tender_value,''), COALESCE(pre_bid_meeting_date,''), COALESCE(closing_date,''), """
                + closing_time_sql +
                """
                   FROM tenders WHERE tender_id=? ORDER BY id DESC LIMIT 1""",
                (proj_tender_id,),
            ).fetchone()
    finally:
        conn.close()

    work = str(p[1] or "").strip()
    client = str(p[2] or "").strip()
    value = str(p[3] or "").strip()
    prebid = str(p[4] or "").strip()
    deadline = str(p[5] or "").strip()
    tender_id = proj_tender_id

    if t_row:
        tender_id = str(t_row[0] or "").strip() or tender_id
        t_work = str(t_row[1] or "").strip() or str(t_row[2] or "").strip()
        t_work = " ".join(t_work.split())
        t_value = str(t_row[3] or "").strip()
        t_prebid = str(t_row[4] or "").strip()
        t_deadline = f"{str(t_row[5] or '').strip()} {str(t_row[6] or '').strip()}".strip()
        if not work:
            work = t_work
        if not value:
            value = t_value
        if not prebid:
            prebid = t_prebid
        if not deadline:
            deadline = t_deadline

    work = " ".join(str(work or "").split())
    return [
        {"key": "Tender ID", "value": tender_id or "-"},
        {"key": "Name of Work", "value": work or "-"},
        {"key": "Client", "value": client or "-"},
        {"key": "Value", "value": value or "-"},
        {"key": "Prebid", "value": prebid or "-"},
        {"key": "Deadline", "value": deadline or "-"},
        {"key": "Time Remaining", "value": _time_remaining_from_text(deadline)},
    ]


def _merge_tender_info_rows(base_rows, stored_rows):
    merged = []
    idx_by_key = {}

    def _norm_rows(rows):
        out = []
        for row in (rows or []):
            if not isinstance(row, dict):
                continue
            key = str(row.get("key", "") or "").strip()
            value = str(row.get("value", "") or "").strip()
            if key:
                out.append({"key": key, "value": value})
        return out

    for row in _norm_rows(base_rows):
        idx_by_key[row["key"].lower()] = len(merged)
        merged.append(row)

    for row in _norm_rows(stored_rows):
        lk = row["key"].lower()
        if lk in idx_by_key:
            merged[idx_by_key[lk]] = row
        else:
            idx_by_key[lk] = len(merged)
            merged.append(row)

    deadline_val = ""
    time_idx = None
    for i, row in enumerate(merged):
        lk = str(row.get("key", "")).strip().lower()
        if lk == "deadline":
            deadline_val = str(row.get("value", "") or "").strip()
        if lk == "time remaining":
            time_idx = i
    if deadline_val:
        tr = _time_remaining_from_text(deadline_val)
        if time_idx is None:
            merged.append({"key": "Time Remaining", "value": tr})
        else:
            merged[time_idx] = {"key": "Time Remaining", "value": tr}
    return merged


def _value_from_tender_rows(rows, key_name):
    target = str(key_name or "").strip().lower()
    for row in (rows or []):
        if not isinstance(row, dict):
            continue
        k = str(row.get("key", "") or "").strip().lower()
        if k == target:
            return str(row.get("value", "") or "").strip()
    return ""


def _apply_tender_info_to_project(project_id, rows):
    pid = int(project_id or 0)
    if pid <= 0:
        return
    tender_id = _value_from_tender_rows(rows, "Tender ID")
    name_of_work = _value_from_tender_rows(rows, "Name of Work")
    client = _value_from_tender_rows(rows, "Client")
    value = _value_from_tender_rows(rows, "Value")
    prebid = _value_from_tender_rows(rows, "Prebid")
    deadline = _value_from_tender_rows(rows, "Deadline")
    name_of_work = " ".join(str(name_of_work or "").split())
    conn = sqlite3.connect(core.DB_FILE)
    try:
        conn.execute(
            "UPDATE projects SET source_tender_id=?, description=?, client_name=?, project_value=?, prebid=?, deadline=? WHERE id=?",
            (
                tender_id or None,
                name_of_work,
                client,
                value,
                prebid,
                deadline,
                pid,
            ),
        )
        conn.commit()
    finally:
        conn.close()


def _open_tender_info_editor(parent, rows, title="Edit Tender Info"):
    dlg = QDialog(parent)
    dlg.setWindowTitle(title)
    dlg.resize(760, 460)
    root = QVBoxLayout(dlg)
    root.setContentsMargins(12, 10, 12, 10)
    root.setSpacing(8)
    root.addWidget(QLabel("Edit tender fields/values. Double-click a cell to edit."))

    table = QTableWidget(0, 2, dlg)
    table.setHorizontalHeaderLabels(["Field", "Value"])
    table.setSelectionBehavior(QAbstractItemView.SelectRows)
    table.setSelectionMode(QAbstractItemView.SingleSelection)
    table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed)
    table.horizontalHeader().setStretchLastSection(True)
    table.setColumnWidth(0, 220)
    table.verticalHeader().setVisible(False)
    root.addWidget(table, 1)

    def add_row(field_txt="", value_txt=""):
        r = table.rowCount()
        table.insertRow(r)
        table.setItem(r, 0, QTableWidgetItem(str(field_txt or "")))
        table.setItem(r, 1, QTableWidgetItem(str(value_txt or "")))

    for row in (rows or []):
        add_row(str(row.get("key", "")), str(row.get("value", "")))

    tools = QHBoxLayout()
    add_btn = QPushButton("Add")
    del_btn = QPushButton("Delete")
    tools.addWidget(add_btn)
    tools.addWidget(del_btn)
    tools.addStretch(1)
    root.addLayout(tools)

    def on_add():
        add_row("", "")
        table.setCurrentCell(table.rowCount() - 1, 0)
        table.editItem(table.item(table.rowCount() - 1, 0))

    def on_delete():
        row = table.currentRow()
        if row < 0:
            return
        table.removeRow(row)

    add_btn.clicked.connect(on_add)
    del_btn.clicked.connect(on_delete)

    btns = QHBoxLayout()
    ok_btn = QPushButton("Save")
    ok_btn.setObjectName("PrimaryButton")
    cancel_btn = QPushButton("Cancel")
    btns.addStretch(1)
    btns.addWidget(ok_btn)
    btns.addWidget(cancel_btn)
    root.addLayout(btns)

    out = {"rows": None}

    def on_save():
        # Ensure active in-cell editor commits text before reading table items.
        fw = QApplication.focusWidget()
        if fw is not None:
            fw.clearFocus()
        table.setFocus()
        QApplication.processEvents()
        merged = []
        for r in range(table.rowCount()):
            key_item = table.item(r, 0)
            val_item = table.item(r, 1)
            key_txt = str(key_item.text() if key_item else "").strip()
            val_txt = str(val_item.text() if val_item else "").strip()
            if key_txt and val_txt:
                merged.append({"key": key_txt, "value": val_txt})
        out["rows"] = merged
        dlg.accept()

    ok_btn.clicked.connect(on_save)
    cancel_btn.clicked.connect(dlg.reject)
    if dlg.exec() != QDialog.Accepted:
        return None
    return out["rows"] if isinstance(out["rows"], list) else []


class TrapezoidToggleButton(QPushButton):
    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.setFixedSize(24, 58)
        self.setCursor(Qt.PointingHandCursor)
        self.setFlat(True)
        self._hover = False

    def enterEvent(self, event):
        self._hover = True
        self.update()
        super().enterEvent(event)

    def leaveEvent(self, event):
        self._hover = False
        self.update()
        super().leaveEvent(event)

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        w = self.width()
        h = self.height()
        poly = QPolygon([QPoint(4, 0), QPoint(w - 3, 6), QPoint(w - 3, h - 6), QPoint(4, h)])
        bg = QColor("#d8dee7" if not self._hover else "#c7d3e4")
        p.setPen(QColor("#b7c6d8"))
        p.setBrush(bg)
        p.drawPolygon(poly)
        p.setPen(QColor("#22415f"))
        p.drawText(self.rect(), Qt.AlignCenter, self.text())
        p.end()


class AutoPopupDateTimeEdit(QDateTimeEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCalendarPopup(True)
        self.setDisplayFormat("dd-MM-yyyy hh:mm AP")
        self.setStyleSheet(
            "QDateTimeEdit::drop-down { width: 0px; border: none; }"
            "QDateTimeEdit::down-arrow { image: none; width: 0px; height: 0px; }"
        )

    def mousePressEvent(self, event):
        super().mousePressEvent(event)
        if event.button() == Qt.LeftButton:
            QTimer.singleShot(0, self.open_popup)

    def open_popup(self):
        if getattr(self, "_popup_open", False):
            return
        self._popup_open = True
        try:
            popup = QDialog(self, Qt.Popup | Qt.FramelessWindowHint)
            popup.setObjectName("DateTimePopup")
            target_w = max(300, int(self.width()))
            popup.setFixedWidth(target_w)
            root = QVBoxLayout(popup)
            root.setContentsMargins(6, 6, 6, 6)
            root.setSpacing(6)

            cal = QCalendarWidget(popup)
            cal.setGridVisible(False)
            cal.setSelectedDate(self.date())
            cal.setFixedWidth(target_w - 12)
            cal.setMaximumHeight(190)
            root.addWidget(cal)

            time_row = QHBoxLayout()
            time_row.setContentsMargins(0, 0, 0, 0)
            time_row.setSpacing(6)
            cur = self.time()
            h24 = int(cur.hour())
            minute = int(cur.minute())
            ap = "AM"
            h12 = h24
            if h24 == 0:
                h12 = 12
            elif h24 == 12:
                ap = "PM"
                h12 = 12
            elif h24 > 12:
                ap = "PM"
                h12 = h24 - 12
            hour_edit = QLineEdit(f"{h12:02d}", popup)
            hour_edit.setFixedHeight(34)
            hour_edit.setMaxLength(2)
            hour_edit.setAlignment(Qt.AlignCenter)
            hour_edit.setValidator(QRegularExpressionValidator(QRegularExpression(r"^(0?[1-9]|1[0-2])?$"), popup))
            hour_edit.setText("")
            hour_edit.setPlaceholderText("HH")
            minute_edit = QLineEdit(f"{minute:02d}", popup)
            minute_edit.setFixedHeight(34)
            minute_edit.setMaxLength(2)
            minute_edit.setAlignment(Qt.AlignCenter)
            minute_edit.setValidator(QRegularExpressionValidator(QRegularExpression(r"^[0-5]?[0-9]?$"), popup))
            minute_edit.setText("")
            minute_edit.setPlaceholderText("MM")
            am_btn = QPushButton("AM", popup)
            pm_btn = QPushButton("PM", popup)
            am_btn.setCheckable(True)
            pm_btn.setCheckable(True)
            am_btn.setFixedSize(56, 30)
            pm_btn.setFixedSize(56, 30)
            am_btn.setStyleSheet(
                "QPushButton { border: 1px solid #b7c6d8; background: #eef3f8; }"
                "QPushButton:checked { background: #2f80c1; color: white; border: 1px solid #2f80c1; }"
            )
            pm_btn.setStyleSheet(
                "QPushButton { border: 1px solid #b7c6d8; background: #eef3f8; }"
                "QPushButton:checked { background: #2f80c1; color: white; border: 1px solid #2f80c1; }"
            )
            am_btn.setChecked(ap == "AM")
            pm_btn.setChecked(ap == "PM")

            def _set_meridiem(which):
                am_btn.setChecked(which == "AM")
                pm_btn.setChecked(which == "PM")

            am_btn.clicked.connect(lambda: _set_meridiem("AM"))
            pm_btn.clicked.connect(lambda: _set_meridiem("PM"))

            time_row.addWidget(hour_edit)
            time_row.addWidget(QLabel(":"))
            time_row.addWidget(minute_edit)
            meridiem_col = QVBoxLayout()
            meridiem_col.setContentsMargins(0, 0, 0, 0)
            meridiem_col.setSpacing(2)
            meridiem_col.addWidget(am_btn)
            meridiem_col.addWidget(pm_btn)
            time_row.addLayout(meridiem_col)
            root.addLayout(time_row)

            btn_row = QHBoxLayout()
            btn_row.setContentsMargins(0, 0, 0, 0)
            btn_row.addStretch(1)
            ok_btn = QPushButton("OK", popup)
            cancel_btn = QPushButton("Cancel", popup)
            ok_btn.setFixedHeight(28)
            cancel_btn.setFixedHeight(28)
            btn_row.addWidget(ok_btn)
            btn_row.addWidget(cancel_btn)
            root.addLayout(btn_row)

            ok_btn.clicked.connect(popup.accept)
            cancel_btn.clicked.connect(popup.reject)
            cal.activated.connect(lambda _d: popup.accept())

            popup.adjustSize()
            pos = self.mapToGlobal(QPoint(0, self.height() + 2))
            popup.move(pos)
            if popup.exec() == QDialog.Accepted:
                try:
                    hh = int((hour_edit.text() or "").strip())
                    mm = int((minute_edit.text() or "").strip())
                except Exception:
                    hh = h12
                    mm = minute
                hh = max(1, min(12, hh))
                mm = max(0, min(59, mm))
                is_pm = pm_btn.isChecked()
                if hh == 12:
                    h24_new = 12 if is_pm else 0
                else:
                    h24_new = hh + (12 if is_pm else 0)
                dt = QDateTime(cal.selectedDate(), QTime(h24_new, mm, 0))
                if dt.isValid():
                    self.setDateTime(dt)
        finally:
            self._popup_open = False


class CreateProjectDialog(QDialog):
    def __init__(self, parent=None, prefill=None, batch_prefills=None, start_index=0):
        super().__init__(parent)
        self.prefill = prefill or {}
        self._batch_prefills = list(batch_prefills or [])
        self._batch_mode = len(self._batch_prefills) > 1
        self._batch_index = max(0, min(int(start_index or 0), len(self._batch_prefills) - 1)) if self._batch_prefills else 0
        if self._batch_prefills:
            self.prefill = dict(self._batch_prefills[self._batch_index] or {})
        self._selected_tender_id = str(self.prefill.get("tender_id", "") or "").strip()
        self.setWindowTitle("Create New Tender Project")
        self.resize(680, 500)

        root = QVBoxLayout(self)
        root.setContentsMargins(6, 8, 6, 8)
        root.setSpacing(6)
        form = QFormLayout()
        form.setHorizontalSpacing(8)
        form.setVerticalSpacing(4)
        form.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        self.title_edit = QLineEdit(str(self.prefill.get("title", "")))
        self.title_edit.setPlaceholderText("Tender Id")
        self.title_edit.setFixedHeight(34)
        self.title_edit.setMinimumWidth(180)
        self.title_edit.setMaximumWidth(240)
        self.fetch_tender_btn = QPushButton("Fetch")
        self.fetch_tender_btn.setFixedHeight(34)
        self.fetch_tender_btn.setMinimumWidth(84)
        self.fetch_tender_btn.clicked.connect(self._fetch_tender_by_id_clicked)
        tender_row = QWidget()
        tender_row_lay = QHBoxLayout(tender_row)
        tender_row_lay.setContentsMargins(0, 0, 0, 0)
        tender_row_lay.setSpacing(6)
        tender_row_lay.setAlignment(Qt.AlignVCenter)
        tender_row_lay.addWidget(self.title_edit, 0)
        tender_row_lay.addWidget(self.fetch_tender_btn, 0)
        tender_row_lay.addStretch(1)
        tender_row.setFixedHeight(34)
        tender_row.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        form.addRow("Tender Id", tender_row)

        self.client_combo = QComboBox()
        self.client_combo.setEditable(True)
        self.client_combo.setFixedHeight(34)
        self.client_combo.setInsertPolicy(QComboBox.NoInsert)
        self.client_combo.installEventFilter(self)
        self.client_combo.addItems(self._load_client_options())
        self._org_nav_prefix = ""
        self._org_nav_last_prefix = ""
        self._org_nav_last_index = -1
        if self.client_combo.lineEdit() is not None:
            self.client_combo.lineEdit().setPlaceholderText("Organisation")
            self.client_combo.lineEdit().installEventFilter(self)
            self.client_combo.lineEdit().textEdited.connect(self._on_org_text_edited)
        self.add_custom_client_btn = QPushButton("Add")
        self.add_custom_client_btn.setMinimumWidth(84)
        self.add_custom_client_btn.setFixedHeight(34)
        self.add_custom_client_btn.clicked.connect(self._add_custom_client_option)
        prefill_client = str(self.prefill.get("client_name", "")).strip()
        if prefill_client:
            idx = self.client_combo.findText(prefill_client)
            if idx >= 0:
                self.client_combo.setCurrentIndex(idx)
            else:
                self.client_combo.setCurrentText(prefill_client)
        org_row = QWidget()
        org_row_lay = QHBoxLayout(org_row)
        org_row_lay.setContentsMargins(0, 0, 0, 0)
        org_row_lay.setSpacing(8)
        org_row_lay.setAlignment(Qt.AlignVCenter)
        self.client_combo.setMinimumWidth(320)
        self.client_combo.setMaximumWidth(16777215)
        self.client_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        org_row.setMinimumHeight(34)
        org_row.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        org_row_lay.addWidget(self.client_combo, 1)
        org_row_lay.addWidget(self.add_custom_client_btn, 0)
        form.addRow("Organisation", org_row)

        self.value_edit = QLineEdit(str(self.prefill.get("project_value", "") or ""))
        self.value_edit.setFixedHeight(34)
        self.value_edit.setPlaceholderText("Tender Value")
        self.value_edit.setValidator(QRegularExpressionValidator(QRegularExpression(r"^[0-9,]*$"), self))
        self.value_edit.editingFinished.connect(self._format_value_field)
        form.addRow("Value", self.value_edit)

        now = datetime.datetime.now()
        self.prebid_edit = AutoPopupDateTimeEdit()
        self.prebid_edit.setDateTime(QDateTime(now))
        self.prebid_edit.setFixedHeight(34)
        self.deadline_edit = AutoPopupDateTimeEdit()
        self.deadline_edit.setDateTime(QDateTime(now))
        self.deadline_edit.setFixedHeight(34)
        self.prebid_edit.installEventFilter(self)
        self.deadline_edit.installEventFilter(self)
        if self.prebid_edit.lineEdit() is not None:
            self.prebid_edit.lineEdit().installEventFilter(self)
        if self.deadline_edit.lineEdit() is not None:
            self.deadline_edit.lineEdit().installEventFilter(self)
        dt_row = QWidget()
        dt_row_lay = QHBoxLayout(dt_row)
        dt_row_lay.setContentsMargins(0, 0, 0, 0)
        dt_row_lay.setSpacing(6)
        dt_row_lay.setAlignment(Qt.AlignVCenter)
        dt_row.setFixedHeight(34)
        dt_row.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.prebid_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.deadline_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.prebid_edit.setMinimumWidth(230)
        self.deadline_edit.setMinimumWidth(230)
        dt_row_lay.addWidget(self.prebid_edit, 1)
        sub_deadline_lbl = QLabel("Submission Deadline")
        sub_deadline_lbl.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        dt_row_lay.addWidget(sub_deadline_lbl, 0)
        dt_row_lay.addWidget(self.deadline_edit, 1)
        form.addRow("Prebid", dt_row)

        self.status_combo = QComboBox()
        self.status_combo.addItems(["Active", "Archived"])
        self.status_combo.setFixedHeight(34)
        prefill_status = str(self.prefill.get("status", "Active") or "Active").strip()
        sidx = self.status_combo.findText(prefill_status, Qt.MatchFixedString)
        if sidx >= 0:
            self.status_combo.setCurrentIndex(sidx)
        form.addRow("Status", self.status_combo)

        self.desc_edit = QTextEdit()
        self.desc_edit.setPlainText(str(self.prefill.get("description", "")))
        self.desc_edit.setFixedHeight(120)
        form.addRow("Name of Work", self.desc_edit)

        root.addLayout(form)

        btn_row = QHBoxLayout()
        clear_btn = QPushButton("Clear")
        self.prev_btn = QPushButton("Previous")
        self.next_btn = QPushButton("Next")
        self.batch_info_lbl = QLabel("")
        cancel_btn = QPushButton("Cancel")
        save_btn = QPushButton("Create Project")
        clear_btn.setObjectName("ClearActionButton")
        save_btn.setObjectName("PrimaryButton")
        for b in (clear_btn, self.prev_btn, self.next_btn, cancel_btn, save_btn):
            b.setFixedHeight(40)
            b.setMinimumWidth(110)
        clear_btn.clicked.connect(self._clear_fields)
        self.prev_btn.clicked.connect(self._go_prev_batch)
        self.next_btn.clicked.connect(self._go_next_batch)
        cancel_btn.clicked.connect(self.reject)
        save_btn.clicked.connect(self._save)
        btn_row.addWidget(clear_btn)
        btn_row.addWidget(self.prev_btn)
        btn_row.addWidget(self.next_btn)
        btn_row.addWidget(self.batch_info_lbl)
        btn_row.addStretch(1)
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(save_btn)
        root.addLayout(btn_row)
        self._init_tender_autofill()
        self._update_batch_nav()

    def _init_tender_autofill(self):
        self._tender_records = self._load_tender_records()
        self._tender_by_id = {}
        self._tender_by_title = {}
        for rec in self._tender_records:
            tid = str(rec.get("tender_id", "")).strip()
            title = str(rec.get("title", "")).strip()
            if tid:
                self._tender_by_id.setdefault(tid.lower(), rec)
            if title:
                self._tender_by_title.setdefault(title.lower(), rec)
        self.title_edit.installEventFilter(self)
        self.desc_edit.installEventFilter(self)
        # Popup completer intentionally disabled for Tender Id for now.
        self.title_edit.setCompleter(None)
        self._format_value_field()
        self._set_prebid_from_text(self.prefill.get("prebid", ""))

    def _load_tender_records(self):
        conn = sqlite3.connect(core.DB_FILE)
        try:
            rows = conn.execute(
                "SELECT COALESCE(tender_id,''), COALESCE(title,''), COALESCE(org_chain,''), "
                "COALESCE(tender_value,''), COALESCE(pre_bid_meeting_date,''), "
                "COALESCE(closing_date,''), COALESCE(closing_time,''), COALESCE(is_archived,0) "
                "FROM tenders ORDER BY id DESC"
            ).fetchall()
        except Exception:
            rows = []
        finally:
            conn.close()
        out = []
        for tid, title, org, val, prebid, cdate, ctime, is_archived in rows:
            out.append(
                {
                    "tender_id": str(tid or "").strip(),
                    "title": str(title or "").strip(),
                    "org_chain": str(org or "").strip(),
                    "value": str(val or "").strip(),
                    "prebid": str(prebid or "").strip(),
                    "closing_date": str(cdate or "").strip(),
                    "closing_time": str(ctime or "").strip(),
                    "is_archived": int(is_archived or 0),
                }
            )
        return out

    def _set_deadline_from_text(self, txt):
        self._set_datetime_edit_from_text(self.deadline_edit, txt)

    def _set_prebid_from_text(self, txt):
        self._set_datetime_edit_from_text(self.prebid_edit, txt)

    def _set_datetime_edit_from_text(self, edit, txt):
        text = str(txt or "").strip()
        if not text:
            return
        formats = [
            "dd-MM-yyyy hh:mm AP",
            "dd-MMM-yyyy hh:mm AP",
            "dd-MM-yyyy hh:mm",
            "dd-MMM-yyyy hh:mm",
            "dd/MM/yyyy hh:mm AP",
            "dd/MM/yyyy hh:mm",
            "yyyy-MM-dd hh:mm AP",
            "yyyy-MM-dd hh:mm",
        ]
        for fmt in formats:
            dt = QDateTime.fromString(text, fmt)
            if dt.isValid():
                edit.setDateTime(dt)
                return

    def _format_value_field(self):
        raw = str(self.value_edit.text() or "").strip()
        if not raw:
            return
        digits = raw.replace(",", "")
        if digits.isdigit():
            self.value_edit.setText(f"{int(digits):,}")

    def _autofill_from_text(self, raw_text, mode="any"):
        raw = str(raw_text or "").strip()
        if not raw:
            return False
        raw_l = raw.lower()
        rec = None
        if mode in ("tender_id", "any"):
            rec = self._tender_by_id.get(raw_l)
        if rec is None and mode in ("title", "any"):
            rec = self._tender_by_title.get(raw_l)
        if rec is None:
            for row in self._tender_records:
                tid_txt = str(row.get("tender_id", "")).lower()
                title_txt = str(row.get("title", "")).lower()
                ok = False
                if mode == "tender_id":
                    ok = raw_l in tid_txt
                elif mode == "title":
                    ok = raw_l in title_txt
                else:
                    ok = (raw_l in tid_txt) or (raw_l in title_txt)
                if ok:
                    rec = row
                    break
        if rec is None:
            return False
        tid = str(rec.get("tender_id", "")).strip()
        title = str(rec.get("title", "")).strip()
        org = str(rec.get("org_chain", "")).strip()
        if tid:
            self.title_edit.setText(tid)
            self._selected_tender_id = tid
        if title:
            self.desc_edit.setPlainText(title)
        if org:
            existing = [str(self.client_combo.itemText(i)).strip() for i in range(self.client_combo.count())]
            if org.lower() not in {x.lower() for x in existing if x}:
                self.client_combo.addItem(org)
                existing.append(org)
                self._save_client_options(existing)
            self.client_combo.setCurrentText(org)
        self.value_edit.setText(str(rec.get("value", "") or ""))
        self._format_value_field()
        self._set_prebid_from_text(str(rec.get("prebid", "") or ""))
        self.status_combo.setCurrentText("Archived" if int(rec.get("is_archived", 0) or 0) else "Active")
        cdate = str(rec.get("closing_date", "") or "").strip()
        ctime = str(rec.get("closing_time", "") or "").strip()
        self._set_deadline_from_text(f"{cdate} {ctime}".strip())
        return True

    def _cycle_org_by_prefix(self, key):
        le = self.client_combo.lineEdit()
        if le is None:
            return False
        prefix = str(self._org_nav_prefix or "").strip().lower()
        if not prefix:
            current_idx = int(self.client_combo.currentIndex())
            current_txt = ""
            if current_idx >= 0:
                current_txt = str(self.client_combo.itemText(current_idx) or "").strip().lower()
            typed_now = str(le.text() or "").strip().lower()
            if typed_now and typed_now != current_txt:
                prefix = typed_now
            else:
                prefix = str(self._org_nav_last_prefix or "").strip().lower()
                if not prefix:
                    prefix = typed_now
        count = int(self.client_combo.count())
        if count <= 0:
            return False
        matches = []
        for i in range(count):
            txt = str(self.client_combo.itemText(i) or "")
            if not prefix or txt.lower().startswith(prefix):
                matches.append(i)
        if not matches:
            return False
        base_index = -1
        if (
            self._org_nav_last_prefix
            and self._org_nav_last_prefix.lower() == prefix
            and self._org_nav_last_index in matches
        ):
            base_index = self._org_nav_last_index
        else:
            current = int(self.client_combo.currentIndex())
            if current in matches:
                base_index = current
        if base_index in matches:
            pos = matches.index(base_index)
            if key == Qt.Key_Down:
                nxt = matches[(pos + 1) % len(matches)]
            else:
                nxt = matches[(pos - 1) % len(matches)]
        else:
            nxt = matches[0] if key == Qt.Key_Down else matches[-1]
        self.client_combo.setCurrentIndex(nxt)
        self._org_nav_last_prefix = prefix
        self._org_nav_last_index = nxt
        return True

    def _on_org_text_edited(self, text):
        self._org_nav_prefix = str(text or "").strip()
        self._org_nav_last_prefix = self._org_nav_prefix
        self._org_nav_last_index = -1

    def _fetch_tender_by_id_clicked(self):
        tid = str(self.title_edit.text() or "").strip()
        if not tid:
            QMessageBox.information(self, "Fetch Tender", "Enter a Tender Id first.")
            return
        rec = self._tender_by_id.get(tid.lower())
        if rec is None:
            QMessageBox.information(self, "Fetch Tender", "No tender found in Database.")
            return
        self._autofill_from_text(tid, mode="tender_id")

    def _apply_prefill_data(self, data):
        d = dict(data or {})
        self.prefill = d
        self._selected_tender_id = str(d.get("tender_id", "") or "").strip()
        self.title_edit.setText(str(d.get("title", "") or "").strip())
        self.client_combo.setCurrentText(str(d.get("client_name", "") or "").strip())
        self.value_edit.setText(str(d.get("project_value", "") or "").strip())
        self._format_value_field()
        self._set_prebid_from_text(str(d.get("prebid", "") or "").strip())
        self._set_deadline_from_text(str(d.get("deadline", "") or "").strip())
        self.desc_edit.setPlainText(str(d.get("description", "") or "").strip())
        self.status_combo.setCurrentText(str(d.get("status", "Active") or "Active").strip() or "Active")

    def _update_batch_nav(self):
        if self._batch_mode:
            self.prev_btn.setVisible(True)
            self.next_btn.setVisible(True)
            self.batch_info_lbl.setVisible(True)
            self.prev_btn.setEnabled(self._batch_index > 0)
            self.next_btn.setEnabled(self._batch_index < len(self._batch_prefills) - 1)
            self.batch_info_lbl.setText(f"Tender {self._batch_index + 1} / {len(self._batch_prefills)}")
        else:
            self.prev_btn.setVisible(False)
            self.next_btn.setVisible(False)
            self.batch_info_lbl.setVisible(False)

    def _go_prev_batch(self):
        if not self._batch_mode or self._batch_index <= 0:
            return
        self._batch_index -= 1
        self._apply_prefill_data(self._batch_prefills[self._batch_index])
        self._update_batch_nav()

    def _go_next_batch(self):
        if not self._batch_mode or self._batch_index >= len(self._batch_prefills) - 1:
            return
        self._batch_index += 1
        self._apply_prefill_data(self._batch_prefills[self._batch_index])
        self._update_batch_nav()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Tab:
            if obj is self.title_edit:
                raw = str(self.title_edit.text() or "").strip().lower()
                if raw and raw not in self._tender_by_id:
                    for row in self._tender_records:
                        tid_txt = str(row.get("tender_id", "")).strip()
                        if tid_txt.lower().startswith(raw):
                            self.title_edit.setText(tid_txt)
                            break
                self._autofill_from_text(self.title_edit.text(), mode="tender_id")
            elif obj is self.desc_edit:
                self._autofill_from_text(self.desc_edit.toPlainText(), mode="title")
        if event.type() == QEvent.MouseButtonPress:
            prebid_le = self.prebid_edit.lineEdit()
            deadline_le = self.deadline_edit.lineEdit()
            if obj is self.prebid_edit or (prebid_le is not None and obj is prebid_le):
                QTimer.singleShot(0, self.prebid_edit.open_popup)
                event.accept()
                return True
            elif obj is self.deadline_edit or (deadline_le is not None and obj is deadline_le):
                QTimer.singleShot(0, self.deadline_edit.open_popup)
                event.accept()
                return True
        if event.type() == QEvent.KeyPress and obj is self.client_combo.lineEdit():
            if event.key() in (Qt.Key_Down, Qt.Key_Up):
                if self._cycle_org_by_prefix(event.key()):
                    event.accept()
                    return True
        if event.type() == QEvent.KeyPress and obj is self.client_combo:
            if event.key() in (Qt.Key_Down, Qt.Key_Up):
                if self._cycle_org_by_prefix(event.key()):
                    event.accept()
                    return True
        if event.type() == QEvent.FocusOut and obj is self.client_combo.lineEdit():
            self._org_nav_prefix = ""
            self._org_nav_last_prefix = ""
            self._org_nav_last_index = -1
        return super().eventFilter(obj, event)

    def _load_client_options(self):
        opts = []
        raw = core.ScraperBackend.get_setting("project_client_options", "[]")
        try:
            opts.extend([str(x).strip() for x in json.loads(raw) if str(x).strip()])
        except Exception:
            pass

        conn = sqlite3.connect(core.DB_FILE)
        try:
            rows = conn.execute(
                "SELECT DISTINCT TRIM(client_name) FROM projects WHERE TRIM(COALESCE(client_name,''))!='' ORDER BY client_name"
            ).fetchall()
            opts.extend([r[0] for r in rows if r and r[0]])
            org_rows = conn.execute(
                "SELECT DISTINCT TRIM(name) FROM organizations WHERE TRIM(COALESCE(name,''))!='' ORDER BY name"
            ).fetchall()
            opts.extend([r[0] for r in org_rows if r and r[0]])
        finally:
            conn.close()

        opts = sorted(set(opts), key=lambda x: x.lower())
        if opts:
            return opts
        return [""]

    def _save_client_options(self, options):
        clean = sorted(set([str(x).strip() for x in options if str(x).strip()]), key=lambda x: x.lower())
        core.ScraperBackend.set_setting("project_client_options", json.dumps(clean))

    def _add_custom_client_option(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Add Organisation")
        dlg.resize(520, 170)
        root = QVBoxLayout(dlg)
        root.addWidget(QLabel("Organisation"))
        name_edit = QLineEdit()
        name_edit.setPlaceholderText("Enter organisation name")
        root.addWidget(name_edit)
        btns = QHBoxLayout()
        btns.addStretch(1)
        ok_btn = QPushButton("OK")
        cancel_btn = QPushButton("Cancel")
        ok_btn.setObjectName("PrimaryButton")
        btns.addWidget(ok_btn)
        btns.addWidget(cancel_btn)
        root.addLayout(btns)
        ok_btn.clicked.connect(dlg.accept)
        cancel_btn.clicked.connect(dlg.reject)
        if dlg.exec() != QDialog.Accepted:
            return
        name = str(name_edit.text() or "").strip()
        if not name:
            return
        existing = [str(self.client_combo.itemText(i)).strip() for i in range(self.client_combo.count())]
        if name.lower() not in {x.lower() for x in existing if x}:
            self.client_combo.addItem(name)
            existing.append(name)
            self._save_client_options(existing)
        self.client_combo.setCurrentText(name)

    def _clear_fields(self):
        self.title_edit.clear()
        if self.client_combo.count():
            self.client_combo.setCurrentIndex(0)
        self.client_combo.setCurrentText("")
        self.value_edit.clear()
        self.prebid_edit.setDateTime(QDateTime(datetime.datetime.now()))
        self.status_combo.setCurrentText("Active")
        self.desc_edit.clear()
        self.deadline_edit.setDateTime(QDateTime(datetime.datetime.now()))
        self._selected_tender_id = ""

    def _save(self):
        title = self.title_edit.text().strip()
        if not title:
            QMessageBox.critical(self, "Error", "Title required")
            return

        client_name = self.client_combo.currentText().strip()
        if not client_name:
            QMessageBox.critical(self, "Error", "Organisation is required")
            return

        deadline = self.deadline_edit.dateTime().toString("dd-MM-yyyy hh:mm AP")
        self._format_value_field()
        project_value = self.value_edit.text().strip()
        prebid = self.prebid_edit.dateTime().toString("dd-MM-yyyy hh:mm AP")
        status = str(self.status_combo.currentText() or "Active").strip() or "Active"
        safe_title = "".join(c for c in title if c.isalnum() or c in (" ", "_")).strip()
        folder_path = os.path.join(core.ROOT_FOLDER, safe_title)
        std_folders = core.ensure_project_standard_folders(folder_path)

        source_tender_id = str(self._selected_tender_id or self.prefill.get("tender_id", "") or "").strip()
        tender_folder_path = str(self.prefill.get("tender_folder_path", "") or "").strip()

        conn = sqlite3.connect(core.DB_FILE)
        c = conn.cursor()
        try:
            if source_tender_id:
                exists = c.execute(
                    "SELECT id FROM projects WHERE source_tender_id=? LIMIT 1",
                    (source_tender_id,),
                ).fetchone()
                if exists:
                    QMessageBox.information(
                        self,
                        "Duplicate Tender",
                        f"A project for Tender ID '{source_tender_id}' already exists.",
                    )
                    return

            c.execute(
                "INSERT INTO projects (title, client_name, deadline, description, folder_path, source_tender_id, project_value, prebid, status) VALUES (?,?,?,?,?,?,?,?,?)",
                (
                    title,
                    client_name,
                    deadline,
                    self.desc_edit.toPlainText().strip(),
                    folder_path,
                    source_tender_id or None,
                    project_value,
                    prebid,
                    status,
                ),
            )
            conn.commit()
        finally:
            conn.close()

        if tender_folder_path and os.path.isdir(tender_folder_path):
            copied = core.copy_tree_contents(tender_folder_path, std_folders["tender_docs"])
            if copied:
                core.log_to_gui(f"Copied {copied} tender item(s) to project Tender Docs: {title}")

        current = [self.client_combo.itemText(i) for i in range(self.client_combo.count())]
        if client_name and client_name not in current:
            current.append(client_name)
            self._save_client_options(current)
        if self._batch_mode and self._batch_prefills:
            self._batch_prefills.pop(self._batch_index)
            if not self._batch_prefills:
                self.accept()
                return
            if self._batch_index >= len(self._batch_prefills):
                self._batch_index = len(self._batch_prefills) - 1
            self._apply_prefill_data(self._batch_prefills[self._batch_index])
            self._update_batch_nav()
            QMessageBox.information(self, "Project Created", "Project created. You can continue with next/previous tender.")
            return
        self.accept()


class ProjectsPage(QWidget):
    headers = ["Sr", "ID", "Tender Id", "Name of Work", "Client", "Value", "Prebid", "Deadline", "Time Remaining"]
    select_sql = "SELECT id, title, description, client_name, project_value, prebid, deadline, status, folder_path, COALESCE(source_tender_id,'') FROM projects"

    def __init__(self, controller):
        super().__init__()
        self.controller = controller
        self._rows = []
        self._rowfit_timer = QTimer(self)
        self._rowfit_timer.setSingleShot(True)
        self._rowfit_timer.timeout.connect(self._fit_rows)
        self._remaining_timer = QTimer(self)
        self._remaining_timer.setInterval(60 * 1000)
        self._remaining_timer.timeout.connect(self._refresh_time_remaining_cells)
        self._layout_persist_timer = QTimer(self)
        self._layout_persist_timer.setSingleShot(True)
        self._layout_persist_timer.timeout.connect(self._persist_projects_column_layout)
        self._col_visibility = self._load_projects_json_setting("projects_cols_visibility", {})
        self._col_order = self._load_projects_json_setting("projects_cols_order", [])
        self._col_widths = self._load_projects_json_setting("projects_cols_widths", {})
        self._edit_project_id = None
        self._project_entry_mode = str(core.get_user_setting("projects_entry_mode", "inline") or "inline").strip().lower()
        if self._project_entry_mode not in ("inline", "popup"):
            self._project_entry_mode = "inline"

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        header = QFrame()
        header.setObjectName("ProjectsHeaderBar")
        hlay = QHBoxLayout(header)
        hlay.setContentsMargins(18, 8, 18, 8)
        title = QLabel("Projects")
        title.setObjectName("ProjectsHeaderTitle")
        self.open_folder_btn = QPushButton("Open Folder")
        self.open_folder_btn.setObjectName("ProjectsOpenFolderButton")
        self.open_folder_btn.setProperty("compact", True)
        self.open_folder_btn.setFixedHeight(34)
        self.open_folder_btn.setMinimumWidth(112)
        self.open_folder_btn.clicked.connect(self.open_projects_root_folder)
        hlay.addWidget(title)
        hlay.addStretch(1)
        hlay.addWidget(self.open_folder_btn)
        root.addWidget(header)

        panel = QFrame()
        panel.setObjectName("ProjectsToolbarPanel")
        play = QVBoxLayout(panel)
        play.setContentsMargins(18, 10, 18, 8)
        play.setSpacing(8)

        bar = QHBoxLayout()
        bar.setSpacing(8)
        bar.addWidget(QLabel("Search Projects:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("")
        self.search_edit.setProperty("isSearchBar", True)
        self.search_edit.setFixedHeight(34)
        self.search_edit.setMaximumWidth(260)
        self.search_edit.textChanged.connect(self.filter_projects)

        self.new_btn = QPushButton("+ New Project")
        self.new_btn.setObjectName("PrimaryButton")
        self.open_btn = QPushButton("Open Project")
        self.open_btn.setObjectName("AccentBlueButton")
        self.edit_tender_info_btn = QPushButton("Edit")
        self.edit_tender_info_btn.setObjectName("SecondaryButton")
        self.delete_btn = QPushButton("Delete Project")
        self.delete_btn.setObjectName("DangerButton")

        self.new_btn.clicked.connect(self.open_create_project)
        self.open_btn.clicked.connect(self.open_selected)
        self.edit_tender_info_btn.clicked.connect(self.edit_selected_project_tender_info)
        self.delete_btn.clicked.connect(self.delete_selected)
        for b in (self.new_btn, self.open_btn, self.edit_tender_info_btn, self.delete_btn):
            b.setProperty("compact", True)
            b.setFixedHeight(34)
        self.new_btn.setMinimumWidth(126)
        self.open_btn.setMinimumWidth(118)
        self.edit_tender_info_btn.setMinimumWidth(132)
        self.delete_btn.setMinimumWidth(122)

        bar.addWidget(self.search_edit, 1)
        bar.addWidget(self.new_btn)
        bar.addStretch(1)
        bar.addWidget(self.open_btn)
        bar.addWidget(self.edit_tender_info_btn)
        bar.addWidget(self.delete_btn)
        play.addLayout(bar)

        form_wrap = QFrame()
        form_wrap.setObjectName("ProjectsInlineForm")
        form = QFormLayout(form_wrap)
        form.setContentsMargins(0, 0, 0, 0)
        form.setHorizontalSpacing(8)
        form.setVerticalSpacing(6)

        self.form_tender_id = QLineEdit()
        self.form_tender_id.setPlaceholderText("Tender Id / Project Title")
        self.form_tender_id.setFixedHeight(34)
        self.form_fetch_btn = QPushButton("Fetch")
        self.form_fetch_btn.setFixedHeight(34)
        self.form_fetch_btn.setMinimumWidth(84)
        self.form_fetch_btn.clicked.connect(self._fetch_inline_tender)
        self.form_client = QComboBox()
        self.form_client.setEditable(True)
        self.form_client.setFixedHeight(34)
        self.form_client.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
        self.form_client.setMinimumContentsLength(16)
        self.form_client.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.form_client.setMinimumWidth(360)
        self.form_client.addItems(self._load_client_options())
        self.form_add_org = QPushButton("Add")
        self.form_add_org.setFixedHeight(34)
        self.form_add_org.setMinimumWidth(84)
        self.form_add_org.clicked.connect(self._add_inline_org)
        top_row = QWidget()
        top_row_l = QHBoxLayout(top_row)
        top_row_l.setContentsMargins(0, 0, 0, 0)
        top_row_l.setSpacing(8)
        self.form_tender_id.setMinimumWidth(130)
        self.form_tender_id.setMaximumWidth(220)
        top_row_l.addWidget(self.form_tender_id, 3)
        top_row_l.addWidget(self.form_fetch_btn, 0)
        top_row_l.addWidget(QLabel("Organisation"), 0)
        top_row_l.addWidget(self.form_client, 8)
        top_row_l.addWidget(self.form_add_org, 0)
        form.addRow("Tender Id", top_row)

        self.form_value = QLineEdit()
        self.form_value.setPlaceholderText("Tender Value")
        self.form_value.setFixedHeight(34)
        self.form_value.setValidator(QRegularExpressionValidator(QRegularExpression(r"^[0-9,]*$"), self))
        self.form_value.editingFinished.connect(self._format_inline_value)

        row_vpd = QWidget()
        row_vpd_l = QHBoxLayout(row_vpd)
        row_vpd_l.setContentsMargins(0, 0, 0, 0)
        row_vpd_l.setSpacing(8)
        self.form_prebid = AutoPopupDateTimeEdit()
        self.form_prebid.setFixedHeight(34)
        self.form_deadline = AutoPopupDateTimeEdit()
        self.form_deadline.setFixedHeight(34)
        self.form_prebid.setDateTime(QDateTime(datetime.datetime.now()))
        self.form_deadline.setDateTime(QDateTime(datetime.datetime.now()))
        self.form_prebid.installEventFilter(self)
        self.form_deadline.installEventFilter(self)
        if self.form_prebid.lineEdit() is not None:
            self.form_prebid.lineEdit().installEventFilter(self)
        if self.form_deadline.lineEdit() is not None:
            self.form_deadline.lineEdit().installEventFilter(self)
        row_vpd_l.addWidget(self.form_value, 2)
        row_vpd_l.addWidget(QLabel("Prebid"), 0)
        row_vpd_l.addWidget(self.form_prebid, 2)
        row_vpd_l.addWidget(QLabel("Submission Deadline"), 0)
        row_vpd_l.addWidget(self.form_deadline, 2)
        form.addRow("Value", row_vpd)

        self.form_work = QTextEdit()
        self.form_work.setFixedHeight(88)
        form.addRow("Name of Work", self.form_work)

        action_row = QHBoxLayout()
        self.form_add_btn = QPushButton("Add")
        self.form_update_btn = QPushButton("Update")
        self.form_clear_btn = QPushButton("Clear")
        self.form_add_btn.setObjectName("PrimaryButton")
        self.form_clear_btn.setObjectName("ClearActionButton")
        for b in (self.form_add_btn, self.form_update_btn, self.form_clear_btn):
            b.setFixedHeight(34)
            b.setMinimumWidth(110)
        self.form_add_btn.clicked.connect(self.add_from_inline_form)
        self.form_update_btn.clicked.connect(self.update_from_inline_form)
        self.form_clear_btn.clicked.connect(self.clear_inline_form)
        action_row.setSpacing(12)
        action_row.addStretch(1)
        action_row.addWidget(self.form_add_btn, 0, Qt.AlignCenter)
        action_row.addWidget(self.form_update_btn, 0, Qt.AlignCenter)
        action_row.addWidget(self.form_clear_btn, 0, Qt.AlignCenter)
        action_row.addStretch(1)
        form.addRow("", action_row)
        play.addWidget(form_wrap)
        root.addWidget(panel)

        table_wrap = QFrame()
        table_wrap.setObjectName("ProjectsTablePanel")
        twl = QVBoxLayout(table_wrap)
        twl.setContentsMargins(18, 10, 18, 12)

        self.table = QTableWidget(0, len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(False)
        self.table.doubleClicked.connect(self.open_from_index)
        self.table.setWordWrap(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.horizontalHeader().customContextMenuRequested.connect(self._show_projects_header_menu)
        self.table.setColumnHidden(1, True)
        self.table.setColumnWidth(0, 52)
        self.table.setColumnWidth(2, 180)
        self.table.setColumnWidth(3, 360)
        self.table.setColumnWidth(4, 220)
        self.table.setColumnWidth(5, 140)
        self.table.setColumnWidth(6, 140)
        self.table.setColumnWidth(7, 160)
        self.table.setColumnWidth(8, 150)
        self.table.horizontalHeader().sectionResized.connect(lambda *_args: self._on_project_table_layout_changed())
        self.table.horizontalHeader().sectionMoved.connect(lambda *_args: self._on_project_table_layout_changed())
        self.table.installEventFilter(self)
        self.table.itemSelectionChanged.connect(self._load_selected_into_form)
        twl.addWidget(self.table, 1)
        root.addWidget(table_wrap, 1)

        self._init_inline_tender_autofill()
        self._apply_entry_mode()
        self._apply_projects_column_layout()
        self.load_projects()
        self._remaining_timer.start()

    def _fetch_rows(self):
        conn = sqlite3.connect(core.DB_FILE)
        try:
            return conn.execute(self.select_sql).fetchall()
        finally:
            conn.close()

    def _format_row(self, sr, r):
        deadline_txt = str(r[6] or "")
        return [
            str(sr),
            str(r[0] or ""),
            str(r[1] or ""),
            str(r[2] or ""),
            str(r[3] or ""),
            str(r[4] or ""),
            str(r[5] or ""),
            deadline_txt,
            self._time_remaining_text(deadline_txt),
        ]

    def load_projects(self):
        self._rows = self._fetch_rows()
        self._fill_table(self._rows)

    def _fill_table(self, rows):
        self.table.setRowCount(0)
        for i, r in enumerate(rows, 1):
            vals = self._format_row(i, r)
            row_idx = self.table.rowCount()
            self.table.insertRow(row_idx)
            for c, v in enumerate(vals):
                item = QTableWidgetItem(v)
                if c in (0, 1) or c >= 5:
                    item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, c, item)
        self._fit_rows()
        self._refresh_time_remaining_cells()

    def _fit_rows(self):
        auto_fit_table_rows(self.table, min_height=24, max_height=None)

    def _schedule_row_fit(self):
        self._rowfit_timer.start(80)

    def _on_project_table_layout_changed(self):
        self._schedule_row_fit()
        self._layout_persist_timer.start(160)

    def _time_remaining_text(self, deadline_text):
        raw = str(deadline_text or "").strip()
        if not raw:
            return ""
        formats = [
            "dd-MM-yyyy hh:mm AP",
            "dd-MMM-yyyy hh:mm AP",
            "dd-MM-yyyy hh:mm",
            "dd-MMM-yyyy hh:mm",
            "dd/MM/yyyy hh:mm AP",
            "dd/MM/yyyy hh:mm",
            "yyyy-MM-dd hh:mm AP",
            "yyyy-MM-dd hh:mm",
        ]
        dt = QDateTime()
        for fmt in formats:
            parsed = QDateTime.fromString(raw, fmt)
            if parsed.isValid():
                dt = parsed
                break
        if not dt.isValid():
            return ""
        now = QDateTime.currentDateTime()
        secs = now.secsTo(dt)
        if secs <= 0:
            return "Expired"
        total_minutes = secs // 60
        total_hours = total_minutes // 60
        minutes = int(total_minutes % 60)
        days = int(total_hours // 24)
        hours = int(total_hours % 24)
        if days >= 7:
            weeks = days // 7
            rem_days = days % 7
            return f"{weeks}w {rem_days}d, {hours:02d}:{minutes:02d}"
        return f"{days}d, {hours:02d}:{minutes:02d}"

    def _refresh_time_remaining_cells(self):
        if self.table.columnCount() < 9:
            return
        for row in range(self.table.rowCount()):
            d_item = self.table.item(row, 7)
            t_item = self.table.item(row, 8)
            if d_item is None:
                continue
            txt = self._time_remaining_text(d_item.text())
            if t_item is None:
                t_item = QTableWidgetItem(txt)
                t_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, 8, t_item)
            else:
                t_item.setText(txt)
                t_item.setTextAlignment(Qt.AlignCenter)

    def _load_projects_json_setting(self, key, default):
        raw = core.get_user_setting(key, default)
        if isinstance(raw, (dict, list)):
            return raw
        try:
            return json.loads(raw)
        except Exception:
            return default

    def _save_projects_json_setting(self, key, value):
        try:
            core.set_user_setting(key, json.dumps(value))
        except Exception:
            core.set_user_setting(key, value)

    def _show_projects_header_menu(self, pos):
        menu = QMenu(self.table.horizontalHeader())
        a_cols = menu.addAction("Manage Columns")
        chosen = menu.exec(self.table.horizontalHeader().mapToGlobal(pos))
        if chosen is a_cols:
            self.open_manage_columns_dialog()

    def _init_inline_tender_autofill(self):
        self._inline_tender_records = []
        self._inline_tender_by_id = {}
        conn = sqlite3.connect(core.DB_FILE)
        try:
            rows = conn.execute(
                "SELECT COALESCE(tender_id,''), COALESCE(title,''), COALESCE(org_chain,''), "
                "COALESCE(tender_value,''), COALESCE(pre_bid_meeting_date,''), COALESCE(closing_date,''), COALESCE(closing_time,''), COALESCE(is_archived,0) "
                "FROM tenders ORDER BY id DESC"
            ).fetchall()
        except Exception:
            rows = []
        finally:
            conn.close()
        for tid, title, org, val, prebid, cdate, ctime, is_archived in rows:
            rec = {
                "tender_id": str(tid or "").strip(),
                "title": str(title or "").strip(),
                "org_chain": str(org or "").strip(),
                "value": str(val or "").strip(),
                "prebid": str(prebid or "").strip(),
                "closing_date": str(cdate or "").strip(),
                "closing_time": str(ctime or "").strip(),
                "is_archived": int(is_archived or 0),
            }
            self._inline_tender_records.append(rec)
            if rec["tender_id"]:
                self._inline_tender_by_id.setdefault(rec["tender_id"].lower(), rec)

    def _fetch_inline_tender(self):
        tid = str(self.form_tender_id.text() or "").strip()
        if not tid:
            QMessageBox.information(self, "Fetch Tender", "Enter a Tender Id first.")
            return
        rec = self._inline_tender_by_id.get(tid.lower())
        if rec is None:
            low = tid.lower()
            for row in self._inline_tender_records:
                t = str(row.get("tender_id", "")).lower()
                if t.startswith(low):
                    rec = row
                    self.form_tender_id.setText(str(row.get("tender_id", "")))
                    break
        if rec is None:
            QMessageBox.information(self, "Fetch Tender", "No tender found in Database.")
            return
        self._autofill_inline_from_tender(rec)

    def _autofill_inline_from_tender(self, rec):
        if not isinstance(rec, dict):
            return
        title = str(rec.get("title", "") or "").strip()
        org = str(rec.get("org_chain", "") or "").strip()
        val = str(rec.get("value", "") or "").strip()
        prebid = str(rec.get("prebid", "") or "").strip()
        cdate = str(rec.get("closing_date", "") or "").strip()
        ctime = str(rec.get("closing_time", "") or "").strip()
        if title:
            self.form_work.setPlainText(title)
        if org:
            self.form_client.setCurrentText(org)
            existing = [self.form_client.itemText(i).strip() for i in range(self.form_client.count())]
            if org.lower() not in {x.lower() for x in existing if x}:
                self.form_client.addItem(org)
                existing.append(org)
                self._save_client_options(existing)
        if val:
            self.form_value.setText(val)
            self._format_inline_value()
        pdt = self._parse_datetime_text(prebid)
        if pdt.isValid():
            self.form_prebid.setDateTime(pdt)
        ddt = self._parse_datetime_text(f"{cdate} {ctime}".strip())
        if ddt.isValid():
            self.form_deadline.setDateTime(ddt)

    def _load_client_options(self):
        opts = []
        raw = core.ScraperBackend.get_setting("project_client_options", "[]")
        try:
            opts.extend([str(x).strip() for x in json.loads(raw) if str(x).strip()])
        except Exception:
            pass
        conn = sqlite3.connect(core.DB_FILE)
        try:
            rows = conn.execute(
                "SELECT DISTINCT TRIM(client_name) FROM projects WHERE TRIM(COALESCE(client_name,''))!='' ORDER BY client_name"
            ).fetchall()
            opts.extend([r[0] for r in rows if r and r[0]])
            org_rows = conn.execute(
                "SELECT DISTINCT TRIM(name) FROM organizations WHERE TRIM(COALESCE(name,''))!='' ORDER BY name"
            ).fetchall()
            opts.extend([r[0] for r in org_rows if r and r[0]])
        finally:
            conn.close()
        opts = sorted(set(opts), key=lambda x: x.lower())
        return opts or [""]

    def _save_client_options(self, options):
        clean = sorted(set([str(x).strip() for x in options if str(x).strip()]), key=lambda x: x.lower())
        core.ScraperBackend.set_setting("project_client_options", json.dumps(clean))

    def _add_inline_org(self):
        name, ok = QInputDialog.getText(self, "Add Organisation", "Organisation")
        if not ok:
            return
        name = str(name or "").strip()
        if not name:
            return
        existing = [self.form_client.itemText(i).strip() for i in range(self.form_client.count())]
        if name.lower() not in {x.lower() for x in existing if x}:
            self.form_client.addItem(name)
            existing.append(name)
            self._save_client_options(existing)
        self.form_client.setCurrentText(name)

    def _format_inline_value(self):
        raw = str(self.form_value.text() or "").strip()
        if not raw:
            return
        digits = raw.replace(",", "")
        if digits.isdigit():
            self.form_value.setText(f"{int(digits):,}")

    def _parse_datetime_text(self, txt):
        text = str(txt or "").strip()
        if not text:
            return QDateTime()
        formats = [
            "dd-MM-yyyy hh:mm AP",
            "dd-MMM-yyyy hh:mm AP",
            "dd-MM-yyyy hh:mm",
            "dd-MMM-yyyy hh:mm",
            "dd/MM/yyyy hh:mm AP",
            "dd/MM/yyyy hh:mm",
            "yyyy-MM-dd hh:mm AP",
            "yyyy-MM-dd hh:mm",
        ]
        for fmt in formats:
            dt = QDateTime.fromString(text, fmt)
            if dt.isValid():
                return dt
        return QDateTime()

    def clear_inline_form(self):
        self._edit_project_id = None
        self.form_tender_id.clear()
        self.form_client.setCurrentText("")
        self.form_value.clear()
        now = QDateTime(datetime.datetime.now())
        self.form_prebid.setDateTime(now)
        self.form_deadline.setDateTime(now)
        self.form_work.clear()
        self.table.clearSelection()

    def _selected_single_project_id(self):
        pids = self._selected_project_ids()
        return pids[0] if pids else None

    def edit_selected_project_tender_info(self):
        pid = self._selected_single_project_id()
        if not pid:
            QMessageBox.information(self, "Edit", "Select one project first.")
            return
        base_rows = _collect_project_tender_base_info(pid)
        stored_rows = _load_project_tender_info(pid)
        rows = _merge_tender_info_rows(base_rows, stored_rows)
        updated = _open_tender_info_editor(self, rows, title="Edit")
        if updated is None:
            return
        _save_project_tender_info(pid, updated)
        _apply_tender_info_to_project(pid, updated)
        self.load_projects()
        try:
            details_page = getattr(self.controller, "project_details_page", None)
            if details_page is not None and int(getattr(details_page, "project_id", 0) or 0) == int(pid):
                details_page.refresh_tender_meta()
        except Exception:
            pass

    def _load_selected_into_form(self):
        pid = self._selected_single_project_id()
        if not pid:
            return
        row = None
        for r in self._rows:
            if str(r[0] or "").strip() == str(pid):
                row = r
                break
        if row is None:
            return
        self._edit_project_id = str(row[0] or "").strip()
        self.form_tender_id.setText(str(row[1] or ""))
        self.form_work.setPlainText(str(row[2] or ""))
        self.form_client.setCurrentText(str(row[3] or ""))
        self.form_value.setText(str(row[4] or ""))
        self._format_inline_value()
        prebid_dt = self._parse_datetime_text(str(row[5] or ""))
        if prebid_dt.isValid():
            self.form_prebid.setDateTime(prebid_dt)
        deadline_dt = self._parse_datetime_text(str(row[6] or ""))
        if deadline_dt.isValid():
            self.form_deadline.setDateTime(deadline_dt)

    def _ensure_unique_folder(self, base):
        base_clean = "".join(c for c in str(base or "") if c.isalnum() or c in (" ", "_", "-")).strip() or "Tender Project"
        path = os.path.join(core.ROOT_FOLDER, base_clean)
        if not os.path.exists(path):
            return path
        i = 2
        while True:
            p = os.path.join(core.ROOT_FOLDER, f"{base_clean}_{i}")
            if not os.path.exists(p):
                return p
            i += 1

    def add_from_inline_form(self):
        title = self.form_tender_id.text().strip()
        if not title:
            QMessageBox.warning(self, "Add Project", "Tender Id is required.")
            return
        client = self.form_client.currentText().strip()
        if not client:
            QMessageBox.warning(self, "Add Project", "Organisation is required.")
            return
        self._format_inline_value()
        source_tender_id = title
        conn = sqlite3.connect(core.DB_FILE)
        try:
            exists = conn.execute(
                "SELECT id FROM projects WHERE source_tender_id=? LIMIT 1",
                (source_tender_id,),
            ).fetchone()
            if exists:
                QMessageBox.information(self, "Add Project", f"A project for Tender ID '{source_tender_id}' already exists.")
                return
            folder_path = self._ensure_unique_folder(title)
            core.ensure_project_standard_folders(folder_path)
            conn.execute(
                "INSERT INTO projects (title, client_name, deadline, description, folder_path, source_tender_id, project_value, prebid, status) VALUES (?,?,?,?,?,?,?,?,?)",
                (
                    title,
                    client,
                    self.form_deadline.dateTime().toString("dd-MM-yyyy hh:mm AP"),
                    self.form_work.toPlainText().strip(),
                    folder_path,
                    source_tender_id,
                    self.form_value.text().strip(),
                    self.form_prebid.dateTime().toString("dd-MM-yyyy hh:mm AP"),
                    "Active",
                ),
            )
            conn.commit()
        finally:
            conn.close()
        current = [self.form_client.itemText(i).strip() for i in range(self.form_client.count())]
        if client and client not in current:
            current.append(client)
            self._save_client_options(current)
        self.load_projects()
        self.clear_inline_form()

    def update_from_inline_form(self):
        pid = self._edit_project_id or self._selected_single_project_id()
        if not pid:
            QMessageBox.warning(self, "Update Project", "Select a project row to update.")
            return
        title = self.form_tender_id.text().strip()
        client = self.form_client.currentText().strip()
        if not title or not client:
            QMessageBox.warning(self, "Update Project", "Tender Id and Organisation are required.")
            return
        self._format_inline_value()
        conn = sqlite3.connect(core.DB_FILE)
        try:
            row = conn.execute("SELECT folder_path FROM projects WHERE id=?", (pid,)).fetchone()
            folder_path = str((row[0] if row else "") or "").strip()
            if not folder_path:
                folder_path = self._ensure_unique_folder(title)
                core.ensure_project_standard_folders(folder_path)
            conn.execute(
                "UPDATE projects SET title=?, client_name=?, deadline=?, description=?, folder_path=?, source_tender_id=?, project_value=?, prebid=?, status=? WHERE id=?",
                (
                    title,
                    client,
                    self.form_deadline.dateTime().toString("dd-MM-yyyy hh:mm AP"),
                    self.form_work.toPlainText().strip(),
                    folder_path,
                    title,
                    self.form_value.text().strip(),
                    self.form_prebid.dateTime().toString("dd-MM-yyyy hh:mm AP"),
                    str(conn.execute("SELECT COALESCE(status,'Active') FROM projects WHERE id=?", (pid,)).fetchone()[0] or "Active"),
                    pid,
                ),
            )
            conn.commit()
        finally:
            conn.close()
        self.load_projects()

    def _apply_entry_mode(self):
        inline = (self._project_entry_mode == "inline")
        for w in (
            self.form_tender_id, self.form_fetch_btn, self.form_client, self.form_add_org,
            self.form_value, self.form_prebid, self.form_deadline, self.form_work,
            self.form_add_btn, self.form_update_btn, self.form_clear_btn,
        ):
            w.setVisible(inline)
        # include labels/rows via parent form wrapper visibility
        parent_form = self.form_tender_id.parentWidget()
        while parent_form is not None and parent_form.objectName() != "ProjectsInlineForm":
            parent_form = parent_form.parentWidget()
        if parent_form is not None:
            parent_form.setVisible(inline)
        self.new_btn.setVisible(not inline)

    def reload_entry_mode(self):
        mode = str(core.get_user_setting("projects_entry_mode", "inline") or "inline").strip().lower()
        if mode not in ("inline", "popup"):
            mode = "inline"
        self._project_entry_mode = mode
        self._apply_entry_mode()

    def _apply_projects_column_layout(self):
        header = self.table.horizontalHeader()
        cols = list(self.headers)
        vis = dict(self._col_visibility or {})
        vis["ID"] = True  # always keep ID hidden in UI
        for i, name in enumerate(cols):
            hide = (name == "ID") or bool(vis.get(name, False))
            self.table.setColumnHidden(i, hide)

        saved_order = [c for c in (self._col_order or []) if c in cols]
        final_order = saved_order + [c for c in cols if c not in saved_order]
        for target_visual, name in enumerate(final_order):
            logical = cols.index(name)
            cur_visual = header.visualIndex(logical)
            if cur_visual != target_visual:
                header.moveSection(cur_visual, target_visual)

        for name, width in dict(self._col_widths or {}).items():
            if name in cols:
                idx = cols.index(name)
                try:
                    self.table.setColumnWidth(idx, max(40, int(width)))
                except Exception:
                    pass

    def _persist_projects_column_layout(self):
        cols = list(self.headers)
        self._col_visibility = {name: bool(self.table.isColumnHidden(i)) for i, name in enumerate(cols)}
        self._col_visibility["ID"] = True
        header = self.table.horizontalHeader()
        order = []
        for visual in range(header.count()):
            logical = header.logicalIndex(visual)
            if 0 <= logical < len(cols):
                order.append(cols[logical])
        self._col_order = order
        self._col_widths = {name: int(self.table.columnWidth(i)) for i, name in enumerate(cols)}
        self._save_projects_json_setting("projects_cols_visibility", self._col_visibility)
        self._save_projects_json_setting("projects_cols_order", self._col_order)
        self._save_projects_json_setting("projects_cols_widths", self._col_widths)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.MouseButtonPress:
            prebid_le = self.form_prebid.lineEdit() if hasattr(self, "form_prebid") else None
            deadline_le = self.form_deadline.lineEdit() if hasattr(self, "form_deadline") else None
            if obj is getattr(self, "form_prebid", None) or (prebid_le is not None and obj is prebid_le):
                QTimer.singleShot(0, self.form_prebid.open_popup)
                event.accept()
                return True
            if obj is getattr(self, "form_deadline", None) or (deadline_le is not None and obj is deadline_le):
                QTimer.singleShot(0, self.form_deadline.open_popup)
                event.accept()
                return True
        if hasattr(self, "table") and obj is self.table and event.type() == QEvent.Resize:
            self._schedule_row_fit()
        return super().eventFilter(obj, event)

    def filter_projects(self):
        query = self.search_edit.text().strip().lower()
        if not query:
            self._fill_table(self._rows)
            return

        filtered = []
        for r in self._rows:
            fields = [str(r[1] or ""), str(r[2] or ""), str(r[3] or ""), str(r[4] or ""), str(r[5] or ""), str(r[6] or "")]
            blob = " ".join(fields).lower()
            if query in blob:
                filtered.append(r)
        self._fill_table(filtered)

    def open_manage_columns_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Manage Columns")
        dlg.resize(420, 470)
        root = QVBoxLayout(dlg)
        root.addWidget(QLabel("Drag to reorder columns and toggle visibility."))
        lst = QListWidget()
        lst.setDragDropMode(QAbstractItemView.InternalMove)
        lst.setDefaultDropAction(Qt.MoveAction)
        root.addWidget(lst, 1)
        cols = list(self.headers)
        cur_vis = dict(self._col_visibility or {})
        order = [c for c in (self._col_order or []) if c in cols] + [c for c in cols if c not in (self._col_order or [])]
        for name in order:
            it = QListWidgetItem(name)
            flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsDragEnabled | Qt.ItemIsUserCheckable
            if name == "ID":
                flags = Qt.ItemIsEnabled
            it.setFlags(flags)
            hidden = True if name == "ID" else bool(cur_vis.get(name, False))
            it.setCheckState(Qt.Unchecked if hidden else Qt.Checked)
            lst.addItem(it)
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_ok = QPushButton("Apply")
        btn_cancel = QPushButton("Cancel")
        btn_ok.setObjectName("PrimaryButton")
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        root.addLayout(btn_row)

        def apply_changes():
            vis = {}
            new_order = []
            for i in range(lst.count()):
                it = lst.item(i)
                name = str(it.text())
                new_order.append(name)
                vis[name] = False if name == "ID" else (it.checkState() != Qt.Checked)
            self._col_visibility = vis
            self._col_order = new_order
            self._apply_projects_column_layout()
            self._persist_projects_column_layout()
            dlg.accept()

        btn_ok.clicked.connect(apply_changes)
        btn_cancel.clicked.connect(dlg.reject)
        dlg.exec()

    def _selected_project_ids(self):
        pids = []
        for idx in self.table.selectionModel().selectedRows():
            row = idx.row()
            pid_item = self.table.item(row, 1)
            if pid_item and pid_item.text().strip():
                pids.append(pid_item.text().strip())
        return pids

    def open_create_project(self):
        if self._project_entry_mode == "popup":
            dlg = CreateProjectDialog(self)
            if dlg.exec() == QDialog.Accepted:
                self.load_projects()
            return
        self.clear_inline_form()
        self.form_tender_id.setFocus()

    def open_selected(self):
        pids = self._selected_project_ids()
        if not pids:
            return
        self.controller.open_project_details(pids[0])

    def open_from_index(self, index):
        row = int(index.row())
        pid_item = self.table.item(row, 1)
        if not pid_item:
            return
        pid = pid_item.text().strip()
        if pid:
            self.controller.open_project_details(pid)

    def delete_selected(self):
        pids = self._selected_project_ids()
        if not pids:
            return

        if len(pids) == 1:
            ok = QMessageBox.question(self, "Confirm", "Delete this project?") == QMessageBox.Yes
        else:
            ok = QMessageBox.question(self, "Confirm", f"Delete {len(pids)} selected projects?") == QMessageBox.Yes
        if not ok:
            return

        conn = sqlite3.connect(core.DB_FILE)
        try:
            for pid in pids:
                conn.execute("DELETE FROM checklist_items WHERE project_id=?", (pid,))
                conn.execute("DELETE FROM projects WHERE id=?", (pid,))
            conn.commit()
        finally:
            conn.close()

        self.load_projects()

    def open_projects_root_folder(self):
        folder = core._resolve_path(core.ROOT_FOLDER)
        try:
            if sys.platform.startswith("win"):
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception as e:
            QMessageBox.critical(self, "Open Folder", f"Could not open folder:\n{e}")


class ProjectDetailsPage(QWidget):
    headers = ["Sr", "Document Name", "Description", "Status", "Attachment", "ID", "Subfolder"]

    def __init__(self, controller):
        super().__init__()
        self.controller = controller
        self.project_id = None
        self.folder_path = ""
        self.source_tender_id = ""
        self.current_pdf = None
        self.current_page = 0
        self._download_executor = ThreadPoolExecutor(max_workers=1)
        self._download_future = None
        self._download_context = None
        self._rowfit_timer = QTimer(self)
        self._rowfit_timer.setSingleShot(True)
        self._rowfit_timer.timeout.connect(self._fit_rows)
        self._preview_anim = None
        self._preview_visible = False
        self._preview_target_width = 520
        self._preview_toggle_manual_y = 440
        self._preview_toggle_y_offset = 24
        self._tender_meta_min_top_height = 76
        self._tender_meta_max_rows = 6
        self._tender_meta_row_height = 24
        self._tender_meta_controls_min_height = 88
        self._tender_meta_controls_max_height = 140
        self._tender_splitter_updating = False
        self._tender_meta_auto_top_height = self._tender_meta_min_top_height
        self._show_tender_info = True
        self._project_table_user_layout = False
        self._project_table_restoring_layout = False
        self._left_box = None
        self.body_splitter = None
        self._current_deadline_text = ""
        self._tender_countdown_timer = QTimer(self)
        self._tender_countdown_timer.setInterval(1000)
        self._tender_countdown_timer.timeout.connect(self._update_tender_countdown_label)
        self._tender_countdown_timer.start()

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        top_bar = QFrame()
        top_bar.setObjectName("ProjectTopBar")
        top = QHBoxLayout(top_bar)
        top.setContentsMargins(10, 8, 10, 8)
        top.setSpacing(10)
        self.back_btn = QPushButton("< Back")
        self.back_btn.setObjectName("ProjectBackButton")
        self.back_btn.setProperty("compact", True)
        self.title_lbl = QLabel("Project Details")
        self.title_lbl.setObjectName("ProjectTitleBarText")
        self.open_explorer_btn = QPushButton("Open in Explorer")
        self.open_explorer_btn.setObjectName("ProjectOpenExplorerButton")
        self.open_explorer_btn.setProperty("compact", True)
        self.sync_btn = QPushButton("Update")
        self.sync_btn.setObjectName("ProjectUpdateButton")
        self.sync_btn.setProperty("compact", True)
        top.addWidget(self.back_btn)
        top.addWidget(self.title_lbl, 1)
        top.addWidget(self.sync_btn)
        top.addWidget(self.open_explorer_btn)
        root.addWidget(top_bar)

        checklist_panel = QFrame()
        self.checklist_panel = checklist_panel
        checklist_panel.setObjectName("ChecklistPanel")
        checklist_layout = QVBoxLayout(checklist_panel)
        self.checklist_layout = checklist_layout
        checklist_layout.setContentsMargins(5, 0, 5, 8)
        checklist_layout.setSpacing(0)

        tender_meta_top = QHBoxLayout()
        tender_meta_top.setContentsMargins(0, 0, 0, 1)
        tender_meta_top.setSpacing(4)
        tender_meta_top.addStretch(1)
        self.tender_countdown_lbl = QLabel("Time Remaining: -")
        self.tender_countdown_lbl.setObjectName("TenderCountdownLabel")
        self.edit_tender_info_btn = QPushButton("Add/Edit Info")
        self.edit_tender_info_btn.setProperty("compact", True)
        self.edit_tender_info_btn.setProperty("legacyProjectButton", True)
        tender_meta_top.addWidget(self.tender_countdown_lbl)
        tender_meta_top.addWidget(self.edit_tender_info_btn)
        checklist_layout.addLayout(tender_meta_top)

        self.tender_meta_wrap = QWidget()
        self.tender_meta_wrap_layout = QVBoxLayout(self.tender_meta_wrap)
        self.tender_meta_wrap_layout.setContentsMargins(0, 0, 0, 0)
        self.tender_meta_wrap_layout.setSpacing(2)
        self.tender_meta_wrap.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        checklist_layout.addWidget(self.tender_meta_wrap)

        self.tender_meta_view = QTextBrowser()
        self.tender_meta_view.setOpenLinks(False)
        self.tender_meta_view.setOpenExternalLinks(False)
        self.tender_meta_view.setMinimumHeight(76)
        self.tender_meta_view.setMaximumHeight(16777215)
        self.tender_meta_view.setPlaceholderText("Tender details will appear here.")
        self.tender_meta_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.tender_meta_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.tender_meta_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tender_meta_view.installEventFilter(self)
        self.tender_meta_wrap_layout.addWidget(self.tender_meta_view)

        checklist_controls = QWidget()
        self.checklist_controls_widget = checklist_controls
        checklist_controls_layout = QVBoxLayout(checklist_controls)
        self.checklist_controls_layout = checklist_controls_layout
        checklist_controls_layout.setContentsMargins(0, 2, 0, 0)
        checklist_controls_layout.setSpacing(2)
        checklist_controls_layout.setAlignment(Qt.AlignTop)
        self.tender_meta_wrap_layout.addWidget(checklist_controls)

        form = QHBoxLayout()
        form.setContentsMargins(0, 0, 0, 0)
        form.setSpacing(4)
        self.doc_name_edit = QLineEdit()
        self.doc_name_edit.setPlaceholderText("Document Name")
        self.doc_name_edit.setMaximumWidth(280)
        self.doc_name_edit.setMinimumWidth(170)
        self.desc_edit = QLineEdit()
        self.desc_edit.setPlaceholderText("Description")
        self.desc_edit.setMaximumWidth(300)
        self.desc_edit.setMinimumWidth(170)
        self.folder_combo = QComboBox()
        self.folder_combo.setMaximumWidth(170)
        form.addWidget(QLabel("Document Name:"))
        form.addWidget(self.doc_name_edit, 1)
        form.addWidget(QLabel("Description:"))
        form.addWidget(self.desc_edit, 1)
        form.addWidget(QLabel("Location (Folder):"))
        form.addWidget(self.folder_combo, 0)
        actions = QHBoxLayout()
        actions.setSpacing(5)
        actions.setContentsMargins(0, 0, 0, 0)
        self.add_btn = QPushButton("Add Item")
        self.add_btn.setObjectName("ProjectAddButton")
        self.update_btn = QPushButton("Update")
        self.update_btn.setObjectName("ProjectUpdateSelectedButton")
        self.new_folder_btn = QPushButton("+ New Folder")
        self.new_folder_btn.setObjectName("ProjectNewFolderButton")
        self.manage_folders_btn = QPushButton("Manage Folders")
        self.manage_folders_btn.setObjectName("ProjectManageFoldersButton")
        self.import_btn = QPushButton("Import Docs")
        self.import_btn.setObjectName("ProjectImportButton")
        self.import_templates_btn = QPushButton("Import Template")
        self.import_templates_btn.setObjectName("ProjectImportTemplateButton")
        self.save_template_btn = QPushButton("Save Template")
        self.save_template_btn.setObjectName("ProjectSaveTemplateButton")
        self.download_btn = QPushButton("Check for Corrigendum")
        self.download_btn.setObjectName("ProjectDownloadButton")
        self.attach_btn = QPushButton("Attach File")
        self.attach_btn.setObjectName("ProjectAttachButton")
        self.open_file_btn = QPushButton("Open File")
        self.open_file_btn.setObjectName("ProjectOpenFileButton")
        self.delete_btn = QPushButton("Delete File")
        self.delete_btn.setObjectName("ProjectDeleteButton")

        left_buttons = (
            self.new_folder_btn,
            self.manage_folders_btn,
            self.import_btn,
            self.import_templates_btn,
            self.download_btn,
            self.save_template_btn,
        )
        right_buttons = (
            self.attach_btn,
            self.open_file_btn,
            self.delete_btn,
        )
        for btn in left_buttons:
            btn.setProperty("compact", True)
            btn.setProperty("legacyProjectButton", True)
            actions.addWidget(btn)
        actions.addStretch(1)
        for btn in right_buttons:
            btn.setProperty("compact", True)
            btn.setProperty("legacyProjectButton", True)
            actions.addWidget(btn)

        for btn in (self.add_btn, self.update_btn):
            btn.setProperty("compact", True)
            btn.setProperty("legacyProjectButton", True)
            btn.setMinimumWidth(84)
            form.addWidget(btn)
        form.addStretch(1)
        self.back_btn.setProperty("legacyProjectTopButton", True)
        self.sync_btn.setProperty("legacyProjectTopButton", True)
        self.open_explorer_btn.setProperty("legacyProjectTopButton", True)
        checklist_controls_layout.addLayout(form)
        checklist_controls_layout.addLayout(actions)
        root.addWidget(checklist_panel)

        self.table = QTableWidget(0, len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(True)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setColumnHidden(5, True)
        self.table.setColumnHidden(6, True)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 52)
        self.table.setColumnWidth(1, 260)
        self.table.setColumnWidth(2, 340)
        self.table.setColumnWidth(3, 120)
        self.table.setColumnWidth(4, 220)
        self.table.horizontalHeader().setSectionsMovable(True)
        self.table.horizontalHeader().sectionResized.connect(self._on_project_table_section_resized)
        self.table.horizontalHeader().sectionMoved.connect(self._on_project_table_section_moved)
        self.table.installEventFilter(self)

        left_box = QFrame()
        self._left_box = left_box
        left_layout = QVBoxLayout(left_box)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.addWidget(self.table, 1)

        self.preview_box = QFrame()
        self.preview_box.setObjectName("PreviewBox")
        preview_layout = QVBoxLayout(self.preview_box)
        preview_layout.setContentsMargins(8, 8, 8, 8)
        preview_layout.setSpacing(6)
        preview_layout.addWidget(QLabel("Preview Area"))
        self.preview_label = QLabel("Preview Area\n(Select a file)")
        self.preview_label.setAlignment(Qt.AlignCenter)
        self.preview_label.setMinimumHeight(260)
        self.preview_label.setStyleSheet("background:#ffffff;border:1px solid #d2dce8;border-radius:8px;")
        self.preview_label.installEventFilter(self)
        preview_layout.addWidget(self.preview_label, 1)
        nav = QHBoxLayout()
        self.prev_btn = QPushButton("<")
        self.next_btn = QPushButton(">")
        self.page_edit = QLineEdit("1")
        self.page_edit.setFixedWidth(46)
        self.page_total = QLabel("1")
        nav.addWidget(self.prev_btn)
        nav.addWidget(self.page_edit)
        nav.addWidget(QLabel("/"))
        nav.addWidget(self.page_total)
        nav.addWidget(self.next_btn)
        nav.addStretch(1)
        preview_layout.addLayout(nav)
        self.preview_box.setMinimumWidth(0)
        self.preview_side_btn = TrapezoidToggleButton("<", self)
        self.preview_side_btn.show()
        self.preview_side_btn.raise_()

        body_wrap = QFrame()
        body_layout = QVBoxLayout(body_wrap)
        body_layout.setContentsMargins(10, 8, 10, 10)
        body_layout.setSpacing(0)
        self.body_splitter = QSplitter(Qt.Horizontal)
        self.body_splitter.addWidget(left_box)
        self.body_splitter.addWidget(self.preview_box)
        self.body_splitter.setChildrenCollapsible(True)
        self.body_splitter.setCollapsible(1, True)
        self.body_splitter.setStretchFactor(0, 3)
        self.body_splitter.setStretchFactor(1, 2)
        body_layout.addWidget(self.body_splitter, 1)
        root.addWidget(body_wrap, 1)

        self.back_btn.clicked.connect(self.controller.show_projects)
        self.open_explorer_btn.clicked.connect(self.open_explorer)
        self.sync_btn.clicked.connect(self.update_project_table_from_folder)
        self.preview_side_btn.clicked.connect(self.toggle_preview_panel)
        self.add_btn.clicked.connect(self.add_item)
        self.update_btn.clicked.connect(self.update_item)
        self.new_folder_btn.clicked.connect(self.create_new_folder)
        self.manage_folders_btn.clicked.connect(self.open_manage_folders_dialog)
        self.import_btn.clicked.connect(self.import_tender_docs)
        self.import_templates_btn.clicked.connect(self.import_templates_to_project)
        self.save_template_btn.clicked.connect(self.save_checklist_as_template)
        self.download_btn.clicked.connect(self.check_for_corrigendum)
        self.edit_tender_info_btn.clicked.connect(self.edit_tender_info)
        self.attach_btn.clicked.connect(self.upload_file)
        self.open_file_btn.clicked.connect(self.open_file)
        self.delete_btn.clicked.connect(self.delete_item)
        self.table.itemSelectionChanged.connect(self.on_select)
        self.prev_btn.clicked.connect(self.prev_page)
        self.next_btn.clicked.connect(self.next_page)
        self.body_splitter.splitterMoved.connect(self._on_preview_splitter_moved)
        self.body_splitter.installEventFilter(self)
        self.preview_box.installEventFilter(self)
        self._load_preview_splitter_sizes()
        self.apply_tender_info_visibility_setting()
        self._update_tender_meta_auto_height()
        self._apply_tender_meta_compact_layout()
        self._set_preview_visible(False, animate=False)
        self._adjust_checklist_column_widths()
        QTimer.singleShot(0, self._position_preview_toggle)

    def load_project(self, pid):
        self.project_id = int(pid)
        self._project_table_user_layout = False
        self._set_preview_visible(False, animate=False)
        conn = sqlite3.connect(core.DB_FILE)
        try:
            p = conn.execute(
                "SELECT title, folder_path, COALESCE(source_tender_id,'') FROM projects WHERE id=?",
                (self.project_id,),
            ).fetchone()
            if not p:
                QMessageBox.warning(self, "Project", "Project not found.")
                self.controller.show_projects()
                return

            self.title_lbl.setText(str(p[0] or "Project Details"))
            repaired_folder = core.resolve_project_folder_path(p[1], p[0])
            self.folder_path = repaired_folder
            try:
                old_folder = str(p[1] or "").strip()
                if os.path.normcase(old_folder) != os.path.normcase(repaired_folder):
                    conn.execute("UPDATE projects SET folder_path=? WHERE id=?", (repaired_folder, self.project_id))
                    conn.commit()
                    core.log_to_gui(f"Project path repaired to: {repaired_folder}")
            except Exception:
                pass
            self.source_tender_id = str(p[2] or "").strip()
            std_folders = core.ensure_project_standard_folders(self.folder_path)
            self.folder_path = std_folders.get("project_root", self.folder_path)
            if os.path.normcase(self.folder_path) != os.path.normcase(repaired_folder):
                try:
                    conn.execute("UPDATE projects SET folder_path=? WHERE id=?", (self.folder_path, self.project_id))
                    conn.commit()
                    core.log_to_gui(f"Project path fallback applied: {self.folder_path}")
                except Exception:
                    pass

            items = conn.execute(
                "SELECT id, sr_no, req_file_name, description, status, linked_file_path, subfolder FROM checklist_items WHERE project_id=? ORDER BY subfolder, sr_no",
                (self.project_id,),
            ).fetchall()
        finally:
            conn.close()

        self.refresh_folder_list()
        self.refresh_tender_meta()
        self.apply_tender_info_visibility_setting()
        self._apply_tender_meta_compact_layout()
        self._fill_items(items)
        if self._restore_project_table_layout():
            self._project_table_user_layout = True

    def _fill_items(self, items):
        self.table.setRowCount(0)
        grouped = {}
        for item in items:
            folder = self._normalize_subfolder(item[6] if len(item) > 6 else "Main")
            grouped.setdefault(folder, []).append(item)

        preferred = ["Ready Docs", "Tender Docs", "Working Docs"]
        all_folders = set(grouped.keys())
        all_folders.update(preferred)
        all_folders.add("Main")
        if os.path.isdir(self.folder_path):
            for root, dirs, _files in os.walk(self.folder_path):
                rel_root = self._normalize_subfolder(os.path.relpath(root, self.folder_path))
                all_folders.add(rel_root)
                for d in dirs:
                    rel_path = self._normalize_subfolder(os.path.relpath(os.path.join(root, d), self.folder_path))
                    all_folders.add(rel_path)
        all_folders = {self._normalize_subfolder(x) for x in all_folders}
        folders = [f for f in preferred if f in all_folders]
        extras = [f for f in sorted(all_folders, key=lambda x: x.lower()) if f not in folders and f != "Main"]
        folders.extend(extras)

        section_no = 1
        for folder in folders:
            header_row = self.table.rowCount()
            self.table.insertRow(header_row)
            sr_it = QTableWidgetItem(str(section_no))
            sr_it.setTextAlignment(Qt.AlignCenter)
            sr_font = sr_it.font()
            sr_font.setBold(True)
            sr_it.setFont(sr_font)
            sr_it.setBackground(QColor("#eceff1"))
            self.table.setItem(header_row, 0, sr_it)
            folder_it = QTableWidgetItem(str(folder))
            folder_it.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            folder_font = folder_it.font()
            folder_font.setBold(True)
            folder_it.setFont(folder_font)
            folder_it.setBackground(QColor("#eceff1"))
            self.table.setItem(header_row, 1, folder_it)
            self.table.setSpan(header_row, 1, 1, 4)
            hidden_id_it = QTableWidgetItem("")
            hidden_sub_it = QTableWidgetItem(str(folder))
            hidden_id_it.setBackground(QColor("#eceff1"))
            hidden_sub_it.setBackground(QColor("#eceff1"))
            self.table.setItem(header_row, 5, hidden_id_it)
            self.table.setItem(header_row, 6, hidden_sub_it)
            section_no += 1

            for idx, item in enumerate(grouped.get(folder, []), 1):
                att_name = os.path.basename(item[5]) if item[5] else "-"
                vals = [
                    f"{idx}.",
                    f"   {str(item[2] or '')}",
                    str(item[3] or ""),
                    str(item[4] or "Pending"),
                    str(att_name),
                    str(item[0]),
                    str(item[6] or folder),
                ]
                row = self.table.rowCount()
                self.table.insertRow(row)
                for c, v in enumerate(vals):
                    t = QTableWidgetItem(v)
                    if c == 0:
                        t.setTextAlignment(Qt.AlignCenter)
                    elif c in (2, 3):
                        t.setTextAlignment(Qt.AlignCenter)
                    if c == 3:
                        if v == "Completed":
                            t.setForeground(QColor("#0a7a5b"))
                        else:
                            t.setForeground(QColor("#d32f2f"))
                    self.table.setItem(row, c, t)
        self._fit_rows()
        self._adjust_checklist_column_widths()

    def _fit_rows(self):
        auto_fit_table_rows(self.table, min_height=24, max_height=None)

    def _schedule_row_fit(self):
        self._rowfit_timer.start(80)

    def _settings_key_project_table_layout(self):
        pid = int(getattr(self, "project_id", 0) or 0)
        return f"project_details_table_layout_{pid if pid > 0 else 'default'}"

    def _save_project_table_layout(self):
        if not getattr(self, "table", None):
            return
        if self._project_table_restoring_layout:
            return
        header = self.table.horizontalHeader()
        if header is None:
            return
        try:
            count = int(self.table.columnCount())
            widths = [int(self.table.columnWidth(i)) for i in range(count)]
            order = [int(header.logicalIndex(v)) for v in range(count)]
            core.set_user_setting(self._settings_key_project_table_layout(), {"widths": widths, "order": order})
        except Exception:
            pass

    def _restore_project_table_layout(self):
        if not getattr(self, "table", None):
            return False
        raw = core.get_user_setting(self._settings_key_project_table_layout(), None)
        if not isinstance(raw, dict):
            return False
        header = self.table.horizontalHeader()
        if header is None:
            return False
        widths = raw.get("widths")
        order = raw.get("order")
        applied = False
        self._project_table_restoring_layout = True
        try:
            if isinstance(widths, list):
                for c in range(min(self.table.columnCount(), len(widths))):
                    try:
                        w = int(widths[c])
                    except Exception:
                        continue
                    if w >= 40:
                        self.table.setColumnWidth(c, w)
                        applied = True
            if isinstance(order, list):
                for target_visual, logical in enumerate(order[: self.table.columnCount()]):
                    try:
                        logical_idx = int(logical)
                    except Exception:
                        continue
                    if logical_idx < 0 or logical_idx >= self.table.columnCount():
                        continue
                    current_visual = int(header.visualIndex(logical_idx))
                    if current_visual != target_visual and current_visual >= 0:
                        header.moveSection(current_visual, target_visual)
                        applied = True
        except Exception:
            return False
        finally:
            self._project_table_restoring_layout = False
        return applied

    def _on_project_table_section_resized(self, *_args):
        self._schedule_row_fit()
        if self._project_table_restoring_layout:
            return
        self._project_table_user_layout = True
        self._save_project_table_layout()

    def _on_project_table_section_moved(self, *_args):
        if self._project_table_restoring_layout:
            return
        self._project_table_user_layout = True
        self._save_project_table_layout()

    def _adjust_checklist_column_widths(self):
        table = getattr(self, "table", None)
        if table is None or table.columnCount() < 5:
            return
        if self._project_table_user_layout:
            return
        viewport_w = max(560, int(table.viewport().width()))
        sr_w = 52
        status_w = 120
        doc_min = 210
        desc_min = 220
        attach_min = 170
        free_w = max(300, viewport_w - sr_w - status_w)
        min_total = doc_min + desc_min + attach_min
        if free_w <= min_total:
            doc_w = doc_min
            desc_w = desc_min
            attach_w = max(140, free_w - doc_w - desc_w)
        else:
            extra = free_w - min_total
            doc_w = doc_min + int(extra * 0.36)
            desc_w = desc_min + int(extra * 0.42)
            attach_w = free_w - doc_w - desc_w
        table.setColumnWidth(0, sr_w)
        table.setColumnWidth(1, max(180, doc_w))
        table.setColumnWidth(2, max(200, desc_w))
        table.setColumnWidth(3, status_w)
        table.setColumnWidth(4, max(140, attach_w))

    def _settings_key_preview_splitter(self):
        return "project_details_preview_splitter_sizes"

    def _settings_key_show_tender_info(self):
        return "project_details_show_tender_info"

    def _tender_meta_max_top_height(self):
        return max(self._tender_meta_min_top_height, int(self._tender_meta_max_rows * self._tender_meta_row_height))

    def _compute_tender_meta_auto_height(self):
        view = getattr(self, "tender_meta_view", None)
        if view is None:
            return self._tender_meta_min_top_height
        try:
            doc = view.document()
            if doc is None:
                return self._tender_meta_min_top_height
            vp = view.viewport()
            text_w = max(220, int(vp.width()) - 8) if vp is not None else 420
            doc.setTextWidth(float(text_w))
            content_h = int(doc.size().height()) + 12
        except Exception:
            content_h = self._tender_meta_min_top_height
        content_h = max(self._tender_meta_min_top_height, content_h)
        content_h = min(content_h, self._tender_meta_max_top_height())
        return int(content_h)

    def _update_tender_meta_auto_height(self):
        self._tender_meta_auto_top_height = self._compute_tender_meta_auto_height()
        self._apply_tender_meta_compact_layout()

    def _save_preview_splitter_sizes(self):
        if not self.body_splitter:
            return
        try:
            sizes = [int(x) for x in self.body_splitter.sizes()]
            if len(sizes) >= 2 and sizes[1] <= 32:
                return
            core.set_user_setting(self._settings_key_preview_splitter(), sizes)
        except Exception:
            pass

    def _load_preview_splitter_sizes(self):
        if not self.body_splitter:
            return
        self.preview_box.setVisible(True)
        self.body_splitter.setSizes([980, 0])
        self.preview_side_btn.setText("<")

    def _apply_tender_meta_compact_layout(self):
        if not getattr(self, "tender_meta_view", None):
            return
        if not getattr(self, "checklist_controls_widget", None):
            return
        try:
            top_h = int(self._tender_meta_auto_top_height or self._tender_meta_min_top_height)
            top_h = max(self._tender_meta_min_top_height, min(top_h, self._tender_meta_max_top_height()))
            controls_hint = int(self.checklist_controls_widget.sizeHint().height())
            controls_h = max(self._tender_meta_controls_min_height, controls_hint + 10)
            if not bool(getattr(self, "_show_tender_info", True)):
                top_h = 0
            self.tender_meta_view.setMinimumHeight(top_h)
            self.tender_meta_view.setMaximumHeight(top_h)
            self.checklist_controls_widget.setMinimumHeight(controls_h)
            self.checklist_controls_widget.setMaximumHeight(controls_h)
            total_h = top_h + controls_h + 2
            if getattr(self, "tender_meta_wrap", None) is not None:
                self.tender_meta_wrap.setMinimumHeight(total_h)
                self.tender_meta_wrap.setMaximumHeight(total_h)
                self.tender_meta_wrap.resize(self.tender_meta_wrap.width(), total_h)
        except Exception:
            pass

    def _on_preview_splitter_moved(self, *_args):
        if not self.body_splitter:
            return
        sizes = self.body_splitter.sizes()
        is_collapsed = len(sizes) >= 2 and int(sizes[1]) <= 32
        if not is_collapsed:
            self._save_preview_splitter_sizes()
        self._position_preview_toggle()
        self.preview_side_btn.setText("<" if is_collapsed else ">")
        self._preview_visible = not is_collapsed
        self._adjust_checklist_column_widths()

    def apply_tender_info_visibility_setting(self):
        show = bool(core.get_user_setting(self._settings_key_show_tender_info(), True))
        self._show_tender_info = show
        self.tender_countdown_lbl.setVisible(show)
        self.edit_tender_info_btn.setVisible(show)
        self.tender_meta_view.setVisible(show)
        if show:
            # Ensure the view never stays clamped at 0 after being shown again.
            self.tender_meta_view.setMinimumHeight(self._tender_meta_min_top_height)
            self.tender_meta_view.setMaximumHeight(16777215)
            self._tender_meta_auto_top_height = max(self._tender_meta_min_top_height, self._compute_tender_meta_auto_height())
        if getattr(self, "checklist_controls_layout", None) is not None:
            # More gap between input row and action row when info box is hidden.
            self.checklist_controls_layout.setSpacing(2 if show else 6)
            self.checklist_controls_layout.setContentsMargins(0, 2 if show else 0, 0, 0)
        if getattr(self, "checklist_layout", None) is not None:
            # Keep a bit more room below controls for full action-row visibility.
            self.checklist_layout.setContentsMargins(5, 0, 5, 8 if show else 10)
        self._apply_tender_meta_compact_layout()
        if show:
            QTimer.singleShot(0, self._update_tender_meta_auto_height)

    def toggle_preview_panel(self):
        if not self.body_splitter:
            return
        sizes = self.body_splitter.sizes()
        total = max(400, int(sum(sizes)) if len(sizes) >= 2 else int(self.body_splitter.width()))
        is_collapsed = (not self.preview_box.isVisible()) or (len(sizes) >= 2 and int(sizes[1]) <= 32)
        if is_collapsed:
            saved = core.get_user_setting(self._settings_key_preview_splitter(), [980, self._preview_target_width])
            if isinstance(saved, list) and len(saved) >= 2:
                try:
                    right = max(360, int(saved[1]))
                except Exception:
                    right = self._preview_target_width
                left = max(200, total - right)
                self.preview_box.setVisible(True)
                self.body_splitter.setSizes([left, right])
            else:
                self.preview_box.setVisible(True)
                self.body_splitter.setSizes([max(200, total - self._preview_target_width), self._preview_target_width])
            self.preview_side_btn.setText(">")
            self._preview_visible = True
            self._load_preview_from_selection()
            if self.current_pdf:
                self.show_pdf_page()
        else:
            self._save_preview_splitter_sizes()
            self.body_splitter.setSizes([max(200, total - 2), 0])
            self.preview_side_btn.setText("<")
            self._preview_visible = False
        self._adjust_checklist_column_widths()
        self._position_preview_toggle()

    def _set_preview_visible(self, visible, animate=True):
        if not self.body_splitter:
            return
        sizes = self.body_splitter.sizes()
        is_collapsed = (not self.preview_box.isVisible()) or (len(sizes) >= 2 and int(sizes[1]) <= 32)
        if bool(visible) != (not is_collapsed):
            self.toggle_preview_panel()
        else:
            self._position_preview_toggle()
            self._adjust_checklist_column_widths()

    def _position_preview_toggle(self):
        if not getattr(self, "body_splitter", None):
            return
        if not getattr(self, "preview_side_btn", None):
            return
        sizes = self.body_splitter.sizes()
        if len(sizes) < 2:
            return
        bw = self.preview_side_btn.width()
        bh = self.preview_side_btn.height()
        bx = self.body_splitter.x() + int(sizes[0]) - (bw // 2)
        by = self.body_splitter.y() + max(0, (self.body_splitter.height() - bh) // 2)
        table = getattr(self, "table", None)
        if table is not None and table.viewport() is not None:
            vp = table.viewport()
            vp_top_left = vp.mapTo(self, QPoint(0, 0))
            by = int(vp_top_left.y() + max(0, (vp.height() - bh) // 2) + int(self._preview_toggle_y_offset))
        max_x = max(0, self.width() - bw)
        max_y = max(0, self.height() - bh)
        bx = min(max(0, int(bx)), max_x)
        by = min(max(0, int(by)), max_y)
        self.preview_side_btn.move(bx, by)
        if not self.preview_side_btn.isVisible():
            self.preview_side_btn.show()
        self.preview_side_btn.raise_()

    def eventFilter(self, obj, event):
        table = getattr(self, "table", None)
        preview_label = getattr(self, "preview_label", None)
        body_splitter = getattr(self, "body_splitter", None)
        preview_box = getattr(self, "preview_box", None)
        tender_splitter = getattr(self, "tender_meta_splitter", None)
        tender_handle = getattr(self, "_tender_splitter_handle", None)
        if obj is tender_handle and event.type() in (
            QEvent.Enter,
            QEvent.HoverMove,
            QEvent.MouseMove,
            QEvent.MouseButtonPress,
            QEvent.MouseButtonRelease,
            QEvent.CursorChange,
        ):
            try:
                tender_handle.setCursor(Qt.ArrowCursor)
            except Exception:
                pass
            return True
        if obj is tender_splitter and event.type() in (QEvent.CursorChange, QEvent.HoverMove):
            try:
                tender_splitter.setCursor(Qt.ArrowCursor)
            except Exception:
                pass
        if obj is table and event.type() == QEvent.Resize:
            self._schedule_row_fit()
            self._adjust_checklist_column_widths()
            self._position_preview_toggle()
        if obj in (body_splitter, preview_box) and event.type() in (QEvent.Resize, QEvent.Move, QEvent.LayoutRequest, QEvent.Show):
            self._position_preview_toggle()
        if obj is getattr(self, "tender_meta_view", None) and event.type() == QEvent.Resize:
            self._update_tender_meta_auto_height()
        if obj is preview_label and event.type() == QEvent.Resize and self.current_pdf:
            self.show_pdf_page()
        return super().eventFilter(obj, event)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._adjust_checklist_column_widths()
        self._apply_tender_meta_compact_layout()
        self._position_preview_toggle()

    def refresh_folder_list(self):
        folder_list = []
        try:
            for root, dirs, _ in os.walk(self.folder_path):
                for d in dirs:
                    full_path = os.path.join(root, d)
                    rel_path = os.path.relpath(full_path, self.folder_path)
                    folder_list.append(rel_path)
        except Exception:
            folder_list = []

        current = self.folder_combo.currentText().strip() or "Main"
        self.folder_combo.clear()
        values = [v for v in sorted(set(folder_list), key=lambda x: x.lower()) if self._normalize_subfolder(v) != "Main"]
        self.folder_combo.addItems(values)
        idx = self.folder_combo.findText(current)
        if idx >= 0:
            self.folder_combo.setCurrentIndex(idx)
        elif self.folder_combo.count():
            self.folder_combo.setCurrentIndex(0)

    def _load_extra_info_pairs(self):
        return _load_project_tender_info(getattr(self, "project_id", 0) or 0)

    def _save_extra_info_pairs(self, items):
        _save_project_tender_info(getattr(self, "project_id", 0) or 0, items)

    def _parse_deadline_datetime(self, raw):
        txt = str(raw or "").strip()
        if not txt:
            return None
        fmts = [
            "%d-%m-%Y %I:%M %p",
            "%d-%m-%Y %H:%M",
            "%d-%m-%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
        ]
        for fmt in fmts:
            try:
                return datetime.datetime.strptime(txt, fmt)
            except Exception:
                continue
        return None

    def _time_remaining_text(self, deadline_text):
        dt = self._parse_deadline_datetime(deadline_text)
        if dt is None:
            return "-"
        now = datetime.datetime.now()
        delta = dt - now
        secs = int(delta.total_seconds())
        if secs <= 0:
            return "Expired"
        days = secs // 86400
        hours = (secs % 86400) // 3600
        mins = (secs % 3600) // 60
        if days > 0:
            return f"{days}d {hours}h {mins}m"
        if hours > 0:
            return f"{hours}h {mins}m"
        return f"{mins}m"

    def _live_tender_countdown_text(self, deadline_text):
        dt = self._parse_deadline_datetime(deadline_text)
        if dt is None:
            return "-"
        now = datetime.datetime.now()
        total = int((dt - now).total_seconds())
        if total <= 0:
            return "Expired"
        days = total // 86400
        rem = total % 86400
        hours = rem // 3600
        rem %= 3600
        mins = rem // 60
        secs = rem % 60
        day_label = "Day" if days == 1 else "Days"
        return f"{days} {day_label}, {hours:02d}:{mins:02d}:{secs:02d}"

    def _update_tender_countdown_label(self):
        text = self._live_tender_countdown_text(getattr(self, "_current_deadline_text", ""))
        self.tender_countdown_lbl.setText(f"Time Remaining: {text}")

    def _render_tender_meta_rows(self, rows):
        tr = []
        for key, val in rows:
            k = html.escape(str(key or ""))
            v = html.escape(str(val or "-"))
            tr.append(
                "<tr>"
                f"<td style='width:180px; vertical-align:top; color:#234; font-weight:600; padding:1px 8px 1px 0;'>{k}</td>"
                f"<td style='vertical-align:top; color:#111; padding:1px 0; white-space:normal; word-break:break-word;'>{v}</td>"
                "</tr>"
            )
        doc = (
            "<html><body style='margin:0; font-family:Segoe UI, Arial, sans-serif; font-size:9.5pt;'>"
            "<table style='width:100%; border-collapse:collapse;'>"
            + "".join(tr)
            + "</table></body></html>"
        )
        self.tender_meta_view.setHtml(doc)
        QTimer.singleShot(0, self._update_tender_meta_auto_height)

    def refresh_tender_meta(self):
        base_rows = _collect_project_tender_base_info(getattr(self, "project_id", 0) or 0)
        stored_rows = self._load_extra_info_pairs()
        merged_rows = _merge_tender_info_rows(base_rows, stored_rows)
        self._current_deadline_text = _value_from_tender_rows(merged_rows, "Deadline")
        self._update_tender_countdown_label()
        view_rows = []
        for row in merged_rows:
            key = str(row.get("key", "") or "").strip()
            val = str(row.get("value", "") or "").strip()
            if key:
                if key.lower() == "name of work":
                    val = " ".join(val.split())
                if key.lower() != "time remaining":
                    view_rows.append((key, val or "-"))
        self._tender_meta_max_rows = 6
        self._render_tender_meta_rows(view_rows)

    def add_tender_info(self):
        if not self.project_id:
            return
        key, ok = QInputDialog.getText(self, "Add Info", "Field name:")
        if not ok:
            return
        key = str(key or "").strip()
        if not key:
            return
        val, ok = QInputDialog.getText(self, "Add Info", "Field value:")
        if not ok:
            return
        val = str(val or "").strip()
        if not val:
            return
        rows = self._load_extra_info_pairs()
        rows.append({"key": key, "value": val})
        self._save_extra_info_pairs(rows)
        self.refresh_tender_meta()

    def edit_tender_info(self):
        if not self.project_id:
            return
        base_rows = _collect_project_tender_base_info(self.project_id)
        stored_rows = self._load_extra_info_pairs()
        rows = _merge_tender_info_rows(base_rows, stored_rows)
        updated = _open_tender_info_editor(self, rows, title="Edit Info")
        if updated is None:
            return
        self._save_extra_info_pairs(updated)
        _apply_tender_info_to_project(self.project_id, updated)
        self.source_tender_id = _value_from_tender_rows(updated, "Tender ID")
        self.refresh_tender_meta()
        try:
            ppage = getattr(self.controller, "projects_page", None)
            if ppage is not None:
                ppage.load_projects()
        except Exception:
            pass

    def _selected_row(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows:
            return None
        row = rows[0].row()
        id_item = self.table.item(row, 5)
        txt = str(id_item.text()).strip() if id_item else ""
        if not txt.isdigit():
            return None
        return row

    def on_select(self):
        row = self._selected_row()
        if row is None:
            return
        self.doc_name_edit.setText(self.table.item(row, 1).text() if self.table.item(row, 1) else "")
        self.desc_edit.setText(self.table.item(row, 2).text() if self.table.item(row, 2) else "")
        sf = self.table.item(row, 6).text() if self.table.item(row, 6) else "Main"
        idx = self.folder_combo.findText(sf)
        if idx >= 0:
            self.folder_combo.setCurrentIndex(idx)
        sizes = self.body_splitter.sizes() if self.body_splitter else []
        is_preview_open = self.preview_box.isVisible() and len(sizes) >= 2 and int(sizes[1]) > 32
        if is_preview_open:
            self._load_preview_from_selection()

    def _load_preview_from_selection(self):
        row = self._selected_row()
        if row is None:
            self._close_preview_pdf()
            self.preview_label.setPixmap(QPixmap())
            self.preview_label.setText("Preview Area\n(Select a file)")
            self.page_edit.setText("1")
            self.page_total.setText("1")
            return
        item_id = str(self.table.item(row, 5).text() if self.table.item(row, 5) else "").strip()
        if not item_id.isdigit():
            self.preview_label.setPixmap(QPixmap())
            self.preview_label.setText("Preview Area\n(Select a file)")
            self.page_edit.setText("1")
            self.page_total.setText("1")
            return
        self.load_preview(int(item_id))

    def add_item(self):
        if not self.project_id:
            return
        name = self.doc_name_edit.text().strip()
        desc = self.desc_edit.text().strip()
        folder = self.folder_combo.currentText().strip() or "Main"
        if not name:
            QMessageBox.critical(self, "Error", "Document Name is required")
            return

        folder_abs = self.folder_path if folder == "Main" else os.path.join(self.folder_path, folder)
        os.makedirs(folder_abs, exist_ok=True)

        conn = sqlite3.connect(core.DB_FILE)
        try:
            count = conn.execute("SELECT count(*) FROM checklist_items WHERE project_id=?", (self.project_id,)).fetchone()[0]
            conn.execute(
                "INSERT INTO checklist_items (project_id, sr_no, req_file_name, description, subfolder) VALUES (?,?,?,?,?)",
                (self.project_id, int(count) + 1, name, desc, folder),
            )
            conn.commit()
        finally:
            conn.close()

        self.doc_name_edit.clear()
        self.desc_edit.clear()
        self.load_project(self.project_id)

    def update_item(self):
        row = self._selected_row()
        if row is None:
            return
        item_id = self.table.item(row, 5).text() if self.table.item(row, 5) else ""
        if not item_id:
            return
        conn = sqlite3.connect(core.DB_FILE)
        try:
            conn.execute(
                "UPDATE checklist_items SET req_file_name=?, description=?, subfolder=? WHERE id=?",
                (
                    self.doc_name_edit.text().strip(),
                    self.desc_edit.text().strip(),
                    self.folder_combo.currentText().strip() or "Main",
                    item_id,
                ),
            )
            conn.commit()
        finally:
            conn.close()
        self.load_project(self.project_id)

    def upload_file(self):
        row = self._selected_row()
        if row is None:
            return
        item_id = self.table.item(row, 5).text() if self.table.item(row, 5) else ""
        if not item_id:
            return
        src_path, _ = QFileDialog.getOpenFileName(self, "Select File")
        if not src_path:
            return

        subfolder = self.folder_combo.currentText().strip() or "Main"
        dest_folder = self.folder_path if subfolder == "Main" else os.path.join(self.folder_path, subfolder)
        os.makedirs(dest_folder, exist_ok=True)
        dest_path = os.path.join(dest_folder, os.path.basename(src_path))
        try:
            shutil.copy(src_path, dest_path)
            conn = sqlite3.connect(core.DB_FILE)
            try:
                conn.execute(
                    "UPDATE checklist_items SET linked_file_path=?, status='Completed' WHERE id=?",
                    (dest_path, item_id),
                )
                conn.commit()
            finally:
                conn.close()
            self.load_project(self.project_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def open_file(self):
        row = self._selected_row()
        if row is None:
            return
        item_id = self.table.item(row, 5).text() if self.table.item(row, 5) else ""
        if not item_id:
            return
        conn = sqlite3.connect(core.DB_FILE)
        try:
            rec = conn.execute("SELECT linked_file_path FROM checklist_items WHERE id=?", (item_id,)).fetchone()
        finally:
            conn.close()
        if rec and rec[0] and os.path.exists(rec[0]):
            path = rec[0]
            if sys.platform.startswith("win"):
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        else:
            QMessageBox.information(self, "Info", "No file attached.")

    def delete_item(self):
        row = self._selected_row()
        if row is None:
            return
        self._close_preview_pdf()
        item_id = self.table.item(row, 5).text() if self.table.item(row, 5) else ""
        if not item_id:
            return
        ok = QMessageBox.question(self, "Confirm", "Delete selected file/item?") == QMessageBox.Yes
        if not ok:
            return

        target_file = ""
        conn = sqlite3.connect(core.DB_FILE)
        try:
            rec = conn.execute("SELECT linked_file_path FROM checklist_items WHERE id=?", (item_id,)).fetchone()
            if rec and rec[0]:
                target_file = str(rec[0])
            conn.execute("DELETE FROM checklist_items WHERE id=?", (item_id,))
            conn.commit()
        finally:
            conn.close()
        if target_file and os.path.isfile(target_file):
            try:
                os.remove(target_file)
            except Exception as e:
                QMessageBox.warning(self, "Warning", f"Item removed from table, but file delete failed:\n{e}")
        self.load_project(self.project_id)

    def open_explorer(self):
        if not self.folder_path:
            return
        if sys.platform.startswith("win"):
            os.startfile(self.folder_path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", self.folder_path])
        else:
            subprocess.Popen(["xdg-open", self.folder_path])

    def create_new_folder(self):
        if not self.folder_path:
            return
        name, ok = QInputDialog.getText(self, "Create New Folder", "Folder Name:")
        if not ok:
            return
        name = str(name or "").strip()
        if not name:
            return
        parent = self.folder_combo.currentText().strip() or "Main"
        new_path = os.path.join(self.folder_path, name) if parent == "Main" else os.path.join(self.folder_path, parent, name)
        try:
            os.makedirs(new_path, exist_ok=False)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            return
        self.refresh_folder_list()
        self.load_project(self.project_id)

    def _normalize_subfolder(self, rel):
        txt = str(rel or "").strip()
        if not txt or txt == "." or txt.lower() == "main":
            return "Main"
        txt = txt.replace("/", os.sep).replace("\\", os.sep)
        txt = os.path.normpath(txt).strip()
        if txt in ("", ".", os.sep):
            return "Main"
        if txt.startswith(os.sep):
            txt = txt.lstrip(os.sep)
        return txt

    def _list_project_subfolders(self):
        out = []
        if not os.path.isdir(self.folder_path):
            return out
        for root, dirs, _files in os.walk(self.folder_path):
            for d in dirs:
                rel = os.path.relpath(os.path.join(root, d), self.folder_path)
                rel = self._normalize_subfolder(rel)
                if rel != "Main":
                    out.append(rel)
        return sorted(set(out), key=lambda x: x.lower())

    def _is_same_or_child_folder(self, parent_rel, candidate_rel):
        p = self._normalize_subfolder(parent_rel)
        c = self._normalize_subfolder(candidate_rel)
        if p == "Main":
            return True
        return c == p or c.startswith(p + os.sep)

    def _apply_folder_mapping_for_checklist(self, old_rel, new_rel=None, delete_mode=False):
        old_rel = self._normalize_subfolder(old_rel)
        if old_rel == "Main":
            return
        new_rel = self._normalize_subfolder(new_rel) if new_rel is not None else "Main"
        old_abs = os.path.normcase(os.path.abspath(os.path.join(self.folder_path, old_rel)))
        new_abs = os.path.normcase(os.path.abspath(os.path.join(self.folder_path, new_rel)))

        conn = sqlite3.connect(core.DB_FILE)
        rows = conn.execute(
            """SELECT id, COALESCE(subfolder,'Main'), COALESCE(linked_file_path,''), COALESCE(status,'Pending')
               FROM checklist_items
               WHERE project_id=?""",
            (self.project_id,),
        ).fetchall()
        updates = []
        for rid, subfolder, linked_file_path, status in rows:
            sf = self._normalize_subfolder(subfolder)
            if not self._is_same_or_child_folder(old_rel, sf):
                continue
            if delete_mode:
                updates.append(("Main", "", "Pending", rid))
                continue

            rest = ""
            if sf != old_rel:
                rest = sf[len(old_rel):].lstrip("\\/")
            mapped_sf = new_rel if new_rel != "Main" else "Main"
            if rest:
                mapped_sf = os.path.join(mapped_sf, rest) if mapped_sf != "Main" else rest
            mapped_sf = self._normalize_subfolder(mapped_sf)

            mapped_link = str(linked_file_path or "").strip()
            if mapped_link:
                lp_abs = os.path.normcase(os.path.abspath(mapped_link))
                if lp_abs == old_abs or lp_abs.startswith(old_abs + os.sep):
                    suffix = lp_abs[len(old_abs):].lstrip("\\/")
                    mapped_link = os.path.join(new_abs, suffix) if suffix else new_abs
            updates.append((mapped_sf, mapped_link, status, rid))

        if updates:
            conn.executemany(
                "UPDATE checklist_items SET subfolder=?, linked_file_path=?, status=? WHERE id=?",
                updates,
            )
            conn.commit()
        conn.close()

    def open_manage_folders_dialog(self):
        if not self.project_id:
            QMessageBox.warning(self, "Manage Folders", "Open a project first.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Manage Folders")
        dlg.resize(760, 460)
        root = QHBoxLayout(dlg)

        left = QVBoxLayout()
        right = QVBoxLayout()
        root.addLayout(left, 2)
        root.addLayout(right, 1)

        left.addWidget(QLabel("Project Folders"))
        tree = QTreeWidget()
        tree.setHeaderHidden(True)
        tree.setRootIsDecorated(True)
        tree.setItemsExpandable(True)
        tree.setExpandsOnDoubleClick(True)
        left.addWidget(tree, 1)

        add_name_edit = QLineEdit()
        add_name_edit.setPlaceholderText("Folder name")
        add_parent_combo = QComboBox()
        add_btn = QPushButton("Add")
        add_btn.setObjectName("PrimaryButton")

        rename_edit = QLineEdit()
        rename_edit.setPlaceholderText("New name")
        rename_btn = QPushButton("Rename")

        move_parent_combo = QComboBox()
        move_btn = QPushButton("Move")

        delete_btn = QPushButton("Delete Selected")
        delete_btn.setObjectName("DangerButton")
        refresh_btn = QPushButton("Refresh")
        close_btn = QPushButton("Close")

        right.addWidget(QLabel("Add Folder"))
        right.addWidget(add_name_edit)
        right.addWidget(add_parent_combo)
        right.addWidget(add_btn)
        right.addSpacing(6)
        right.addWidget(QLabel("Rename Selected"))
        right.addWidget(rename_edit)
        right.addWidget(rename_btn)
        right.addSpacing(6)
        right.addWidget(QLabel("Move Selected"))
        right.addWidget(move_parent_combo)
        right.addWidget(move_btn)
        right.addSpacing(8)
        right.addWidget(delete_btn)
        right.addWidget(refresh_btn)
        right.addStretch(1)
        right.addWidget(close_btn)

        def refresh_lists():
            folders = self._list_project_subfolders()
            selected_prev = selected_folder()
            tree.clear()
            node_map = {}

            root_node = QTreeWidgetItem(["Main"])
            root_node.setData(0, Qt.UserRole, "Main")
            tree.addTopLevelItem(root_node)
            node_map["Main"] = root_node

            def ensure_node(rel_path):
                rel = self._normalize_subfolder(str(rel_path or "Main").replace("\\", "/"))
                if rel in (".", ""):
                    rel = "Main"
                if rel in node_map:
                    return node_map[rel]
                parent_rel = self._normalize_subfolder(os.path.dirname(rel))
                if parent_rel in (".", ""):
                    parent_rel = "Main"
                parent_node = ensure_node(parent_rel)
                label = os.path.basename(rel) if rel != "Main" else "Main"
                node = QTreeWidgetItem([label])
                node.setData(0, Qt.UserRole, rel)
                parent_node.addChild(node)
                node_map[rel] = node
                return node

            for rel in folders:
                ensure_node(rel)

            tree.expandToDepth(1)
            if selected_prev and selected_prev in node_map:
                tree.setCurrentItem(node_map[selected_prev])
            else:
                tree.setCurrentItem(root_node)
            choices = ["Main"] + folders

            add_current = add_parent_combo.currentText().strip() or "Main"
            add_parent_combo.clear()
            add_parent_combo.addItems(choices)
            add_parent_combo.setCurrentText(add_current if add_current in choices else "Main")

            move_current = move_parent_combo.currentText().strip() or "Main"
            move_parent_combo.clear()
            move_parent_combo.addItems(choices)
            move_parent_combo.setCurrentText(move_current if move_current in choices else "Main")

        def selected_folder():
            item = tree.currentItem()
            if not item:
                return ""
            return str(item.data(0, Qt.UserRole) or "").strip()

        def add_folder():
            name = str(add_name_edit.text() or "").strip()
            parent = self._normalize_subfolder(add_parent_combo.currentText())
            if not name:
                QMessageBox.warning(dlg, "Manage Folders", "Folder name is required.")
                return
            target_rel = self._normalize_subfolder(name if parent == "Main" else os.path.join(parent, name))
            target_abs = os.path.join(self.folder_path, target_rel)
            if os.path.exists(target_abs):
                QMessageBox.warning(dlg, "Manage Folders", "Folder already exists.")
                return
            os.makedirs(target_abs, exist_ok=True)
            add_name_edit.clear()
            refresh_lists()
            self.load_project(self.project_id)

        def rename_folder():
            old_rel = selected_folder()
            if not old_rel:
                QMessageBox.warning(dlg, "Manage Folders", "Select a folder to rename.")
                return
            new_name = str(rename_edit.text() or "").strip()
            if not new_name:
                QMessageBox.warning(dlg, "Manage Folders", "New name is required.")
                return
            parent_rel = self._normalize_subfolder(os.path.dirname(old_rel))
            if parent_rel == ".":
                parent_rel = "Main"
            new_rel = self._normalize_subfolder(new_name if parent_rel == "Main" else os.path.join(parent_rel, new_name))
            if new_rel == old_rel:
                return
            old_abs = os.path.join(self.folder_path, old_rel)
            new_abs = os.path.join(self.folder_path, new_rel)
            if os.path.exists(new_abs):
                QMessageBox.warning(dlg, "Manage Folders", "Target folder name already exists.")
                return
            os.makedirs(os.path.dirname(new_abs), exist_ok=True)
            shutil.move(old_abs, new_abs)
            self._apply_folder_mapping_for_checklist(old_rel, new_rel=new_rel, delete_mode=False)
            rename_edit.clear()
            refresh_lists()
            self.load_project(self.project_id)

        def move_folder():
            old_rel = selected_folder()
            if not old_rel:
                QMessageBox.warning(dlg, "Manage Folders", "Select a folder to move.")
                return
            dest_parent = self._normalize_subfolder(move_parent_combo.currentText())
            if self._is_same_or_child_folder(old_rel, dest_parent):
                QMessageBox.warning(dlg, "Manage Folders", "Cannot move folder inside itself.")
                return
            base = os.path.basename(old_rel)
            new_rel = self._normalize_subfolder(base if dest_parent == "Main" else os.path.join(dest_parent, base))
            if new_rel == old_rel:
                return
            old_abs = os.path.join(self.folder_path, old_rel)
            new_abs = os.path.join(self.folder_path, new_rel)
            if os.path.exists(new_abs):
                QMessageBox.warning(dlg, "Manage Folders", "Target location already has this folder.")
                return
            os.makedirs(os.path.dirname(new_abs), exist_ok=True)
            shutil.move(old_abs, new_abs)
            self._apply_folder_mapping_for_checklist(old_rel, new_rel=new_rel, delete_mode=False)
            refresh_lists()
            self.load_project(self.project_id)

        def delete_folder():
            old_rel = selected_folder()
            if not old_rel:
                QMessageBox.warning(dlg, "Manage Folders", "Select a folder to delete.")
                return
            ok = QMessageBox.question(
                dlg,
                "Delete Folder",
                f"Delete folder '{old_rel}' and all its contents?\nChecklist items will be moved to Main as Pending.",
            ) == QMessageBox.Yes
            if not ok:
                return
            old_abs = os.path.join(self.folder_path, old_rel)
            if os.path.isdir(old_abs):
                shutil.rmtree(old_abs, ignore_errors=False)
            self._apply_folder_mapping_for_checklist(old_rel, delete_mode=True)
            refresh_lists()
            self.load_project(self.project_id)

        add_btn.clicked.connect(add_folder)
        rename_btn.clicked.connect(rename_folder)
        move_btn.clicked.connect(move_folder)
        delete_btn.clicked.connect(delete_folder)
        refresh_btn.clicked.connect(refresh_lists)
        close_btn.clicked.connect(dlg.close)

        refresh_lists()
        dlg.exec()

    def update_project_table_from_folder(self, notify=True):
        if not self.project_id:
            QMessageBox.warning(self, "Update", "Open a project first.")
            return
        core.ensure_project_standard_folders(self.folder_path)
        conn = sqlite3.connect(core.DB_FILE)
        c = conn.cursor()
        rows = c.execute(
            """SELECT id, COALESCE(req_file_name,''), COALESCE(description,''), COALESCE(status,'Pending'),
                      COALESCE(linked_file_path,''), COALESCE(subfolder,'Main')
               FROM checklist_items
               WHERE project_id=?""",
            (self.project_id,),
        ).fetchall()

        fs_index = {}
        for root, _dirs, files in os.walk(self.folder_path):
            rel_root = self._normalize_subfolder(os.path.relpath(root, self.folder_path))
            for fn in files:
                full = os.path.join(root, fn)
                try:
                    f_abs = os.path.normcase(os.path.abspath(full))
                except Exception:
                    continue
                fs_index[f_abs] = (rel_root, fn, full)

        to_delete_ids = set()
        linked_by_path = {}
        represented_paths = set()

        for rid, req_name, _desc, status, linked_file_path, subfolder in rows:
            lp = str(linked_file_path or "").strip()
            if not lp:
                continue
            try:
                lp_abs = os.path.normcase(os.path.abspath(lp))
            except Exception:
                to_delete_ids.add(rid)
                continue
            if lp_abs in linked_by_path:
                to_delete_ids.add(rid)
                continue
            if lp_abs not in fs_index:
                to_delete_ids.add(rid)
                continue
            linked_by_path[lp_abs] = rid
            represented_paths.add(lp_abs)
            rel_root, fn, full = fs_index[lp_abs]
            needs_update = (
                str(req_name or "") != fn
                or self._normalize_subfolder(subfolder) != rel_root
                or str(lp) != str(full)
                or str(status or "") != "Completed"
            )
            if needs_update:
                c.execute(
                    """UPDATE checklist_items
                       SET req_file_name=?, subfolder=?, linked_file_path=?, status='Completed'
                       WHERE id=?""",
                    (fn, rel_root, full, rid),
                )

        removed = 0
        if to_delete_ids:
            q = ",".join(["?"] * len(to_delete_ids))
            c.execute(f"DELETE FROM checklist_items WHERE id IN ({q})", tuple(to_delete_ids))
            removed = len(to_delete_ids)

        count_row = c.execute("SELECT COUNT(*) FROM checklist_items WHERE project_id=?", (self.project_id,)).fetchone()
        next_sr = int((count_row[0] if count_row else 0) or 0) + 1
        inserted = 0
        for f_abs, (rel_root, fn, full) in sorted(fs_index.items(), key=lambda kv: kv[1][1].lower()):
            if f_abs in represented_paths:
                continue
            c.execute(
                """INSERT INTO checklist_items
                   (project_id, sr_no, req_file_name, description, subfolder, linked_file_path, status)
                    VALUES (?, ?, ?, ?, ?, ?, 'Completed')""",
                (self.project_id, next_sr, fn, "", rel_root, full),
            )
            next_sr += 1
            inserted += 1
        conn.commit()
        conn.close()

        self.load_project(self.project_id)
        if notify:
            if inserted or removed:
                QMessageBox.information(self, "Update", f"Synced project table. Added {inserted}, removed {removed}.")
            else:
                QMessageBox.information(self, "Update", "Project table already mirrors the project folder.")

    def _close_preview_pdf(self):
        try:
            if self.current_pdf:
                self.current_pdf.close()
        except Exception:
            pass
        self.current_pdf = None

    def load_preview(self, item_id):
        self._close_preview_pdf()
        self.preview_label.setPixmap(QPixmap())
        self.preview_label.setText("Preview Area\n(Select a file)")
        self.page_edit.setText("1")
        self.page_total.setText("1")
        if not item_id:
            return
        conn = sqlite3.connect(core.DB_FILE)
        try:
            row = conn.execute(
                "SELECT linked_file_path, req_file_name, subfolder FROM checklist_items WHERE id=?",
                (item_id,),
            ).fetchone()
        finally:
            conn.close()
        if not row:
            return
        linked_path = str(row[0] or "").strip()
        req_name = str(row[1] or "").strip()
        subfolder = self._normalize_subfolder(str(row[2] or "Main"))
        candidates = []
        if linked_path:
            candidates.append(linked_path)
            if not os.path.isabs(linked_path):
                candidates.append(os.path.join(self.folder_path, linked_path))
        if req_name:
            base_folder = self.folder_path if subfolder == "Main" else os.path.join(self.folder_path, subfolder)
            candidates.append(os.path.join(base_folder, req_name))
            candidates.append(os.path.join(self.folder_path, req_name))
        fpath = ""
        for c in candidates:
            try:
                if c and os.path.exists(c):
                    fpath = c
                    break
            except Exception:
                continue
        if not fpath:
            self.preview_label.setText("Attached file not found on disk.")
            return
        ext = os.path.splitext(fpath)[1].lower()
        if ext != ".pdf":
            self.preview_label.setText(f"Preview not available for {ext} files.")
            return
        if not getattr(core, "PDF_SUPPORT", False):
            self.preview_label.setText("Install 'pymupdf' to view PDFs.")
            return
        try:
            self.current_pdf = core.fitz.open(fpath)
            self.current_page = 0
            self.page_total.setText(str(len(self.current_pdf)))
            self.show_pdf_page()
        except Exception as e:
            self.preview_label.setText(f"Error reading PDF:\n{e}")

    def show_pdf_page(self):
        if not self.current_pdf:
            return
        try:
            page = self.current_pdf.load_page(self.current_page)
            # Render to exact display size with DPI awareness to avoid blur.
            dpr = max(1.0, float(self.devicePixelRatioF()))
            target_w = max(420.0, float(self.preview_label.width() - 20) * dpr)
            target_h = max(260.0, float(self.preview_label.height() - 20) * dpr)
            zoom_x = float(target_w) / max(1.0, float(page.rect.width))
            zoom_y = float(target_h) / max(1.0, float(page.rect.height))
            zoom = max(0.5, min(zoom_x, zoom_y))
            pix = page.get_pixmap(matrix=core.fitz.Matrix(zoom, zoom), alpha=False)
            qimg = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format_RGB888).copy()
            pm = QPixmap.fromImage(qimg)
            pm.setDevicePixelRatio(dpr)
            self.preview_label.setText("")
            self.preview_label.setPixmap(pm)
            self.page_edit.setText(str(self.current_page + 1))
        except Exception as e:
            self.preview_label.setText(f"Error rendering page: {e}")

    def prev_page(self):
        if self.current_pdf and self.current_page > 0:
            self.current_page -= 1
            self.show_pdf_page()

    def next_page(self):
        if self.current_pdf and self.current_page < len(self.current_pdf) - 1:
            self.current_page += 1
            self.show_pdf_page()

    def _project_tender_docs_path(self):
        return core.ensure_project_standard_folders(self.folder_path)["tender_docs"]

    def _get_active_tender_record(self, tender_id):
        conn = sqlite3.connect(core.DB_FILE)
        try:
            return conn.execute(
                """SELECT id, COALESCE(tender_id,''), COALESCE(folder_path,''), COALESCE(is_downloaded,0), COALESCE(is_archived,0)
                   FROM tenders
                   WHERE TRIM(COALESCE(tender_id,''))=TRIM(?)
                   ORDER BY created_at DESC
                   LIMIT 1""",
                (tender_id,),
            ).fetchone()
        finally:
            conn.close()

    def _resolve_tender_download_folder(self, tender_id, db_folder_path=""):
        safe_id = core.re.sub(r'[\\/*?:"<>|]', "", str(tender_id or "").strip())
        preferred = os.path.join(core.BASE_DOWNLOAD_DIRECTORY, safe_id) if safe_id else ""
        if preferred and os.path.isdir(preferred):
            return preferred
        alt = str(db_folder_path or "").strip()
        if alt and os.path.isdir(alt):
            return alt
        return preferred

    def _folder_has_files(self, folder):
        if not folder or not os.path.isdir(folder):
            return False
        try:
            for _, _, files in os.walk(folder):
                if files:
                    return True
        except Exception:
            return False
        return False

    def _copy_special_tender_docs(self, src_root, dest_root):
        if not src_root or not os.path.isdir(src_root):
            return 0
        os.makedirs(dest_root, exist_ok=True)
        copied = 0
        keywords = ("corrigendum", "corrigenda", "addendum", "addenda", "prebid", "pre-bid", "pre bid")
        excluded_parts = {"ready docs", "tender docs", "working docs"}
        for root_dir, _, files in os.walk(src_root):
            rel_dir = os.path.relpath(root_dir, src_root)
            rel_dir_low = str(rel_dir or "").lower()
            if any(part in rel_dir_low for part in excluded_parts):
                continue
            for name in files:
                src = os.path.join(root_dir, name)
                rel_path = os.path.relpath(src, src_root)
                rel_low = str(rel_path).lower()
                if not any(k in rel_low for k in keywords):
                    continue
                dst = os.path.join(dest_root, rel_path)
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                try:
                    if os.path.exists(dst):
                        if (
                            os.path.getsize(dst) == os.path.getsize(src)
                            and int(os.path.getmtime(dst)) == int(os.path.getmtime(src))
                        ):
                            continue
                    shutil.copy2(src, dst)
                    copied += 1
                except Exception:
                    pass
        return copied

    def import_tender_docs(self):
        if not self.project_id:
            QMessageBox.warning(self, "Import Docs", "Open a project first.")
            return
        tender_id = str(self.source_tender_id or "").strip()
        if not tender_id:
            QMessageBox.warning(self, "Import Docs", "Open project is not linked to a Tender ID.")
            return

        row = self._get_active_tender_record(tender_id)
        if not row:
            QMessageBox.information(self, "Import Docs", "Tender ID of this project was not found in Active Tenders.")
            return
        _, _, db_folder_path, _, is_archived = row
        if int(is_archived or 0) != 0:
            QMessageBox.information(self, "Import Docs", "Tender ID of this project is not in Active Tenders.")
            return

        src = self._resolve_tender_download_folder(tender_id, db_folder_path)
        if not self._folder_has_files(src):
            ask = QMessageBox.question(
                self,
                "Import Docs",
                "No downloaded docs were found for this tender.\nDo you want to download now?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if ask != QMessageBox.Yes:
                return
            self._start_download_job(
                tender_id=tender_id,
                mode="full",
                destination_folder=src,
                action="import_docs",
            )
            return

        dest = self._project_tender_docs_path()
        copied = core.copy_tree_contents(
            src,
            dest,
            exclude_names={"Ready Docs", "Tender Docs", "Working Docs"},
        )
        if copied:
            core.log_to_gui(f"Imported {copied} item(s) into project Tender Docs for tender {tender_id}.")
            QMessageBox.information(self, "Import Docs", f"Imported {copied} item(s) into Tender Docs.")
            self.update_project_table_from_folder(notify=False)
        else:
            QMessageBox.information(self, "Import Docs", "No documents were downloaded for this tender.")
            self.load_project(self.project_id)

    def import_templates_to_project(self):
        if not self.project_id:
            QMessageBox.warning(self, "Import Templates", "Open a project first.")
            return
        if not self.folder_path:
            QMessageBox.warning(self, "Import Templates", "Project folder is not available.")
            return
        core.ensure_project_standard_folders(self.folder_path)
        import_templates_into_project(self, self.project_id, self.folder_path)
        self.load_project(self.project_id)

    def _ask_template_meta(self):
        conn = sqlite3.connect(core.DB_FILE)
        row = conn.execute(
            "SELECT COALESCE(title,''), COALESCE(client_name,'') FROM projects WHERE id=?",
            (self.project_id,),
        ).fetchone()
        next_no_row = conn.execute("SELECT COALESCE(MAX(template_no),0) FROM checklist_templates").fetchone()
        conn.close()
        default_name = str((row[0] if row else "") or "").strip() or "Template"
        default_org = str((row[1] if row else "") or "").strip()
        next_no = int((next_no_row[0] if next_no_row else 0) or 0) + 1

        dlg = QDialog(self)
        dlg.setWindowTitle("Save Checklist as Template")
        dlg.resize(560, 220)
        root = QVBoxLayout(dlg)
        form = QFormLayout()
        no_edit = QLineEdit(str(next_no))
        org_edit = QLineEdit(default_org)
        name_edit = QLineEdit(default_name)
        desc_edit = QLineEdit("Imported from project checklist")
        form.addRow("Template Number", no_edit)
        form.addRow("Organization", org_edit)
        form.addRow("Template Name", name_edit)
        form.addRow("Description", desc_edit)
        root.addLayout(form)
        row_btn = QHBoxLayout()
        row_btn.addStretch(1)
        cancel = QPushButton("Cancel")
        save = QPushButton("Save")
        save.setObjectName("PrimaryButton")
        row_btn.addWidget(cancel)
        row_btn.addWidget(save)
        root.addLayout(row_btn)
        cancel.clicked.connect(dlg.reject)
        save.clicked.connect(dlg.accept)
        if dlg.exec() != QDialog.Accepted:
            return None
        no_txt = str(no_edit.text() or "").strip()
        t_no = int(no_txt) if no_txt.isdigit() else next_no
        org = str(org_edit.text() or "").strip() or "General"
        tname = str(name_edit.text() or "").strip()
        tdesc = str(desc_edit.text() or "").strip()
        if not tname:
            QMessageBox.warning(self, "Template", "Template name is required.")
            return None
        return {
            "template_no": t_no,
            "organization": org,
            "template_name": tname,
            "description": tdesc,
        }

    def save_checklist_as_template(self):
        if not self.project_id:
            QMessageBox.warning(self, "Template", "Open a project first.")
            return
        meta = self._ask_template_meta()
        if not meta:
            return
        org = meta["organization"]
        tname = meta["template_name"]
        tdesc = meta["description"]
        tno = int(meta["template_no"])

        conn = sqlite3.connect(core.DB_FILE)
        try:
            items = conn.execute(
                """SELECT COALESCE(sr_no,0), COALESCE(req_file_name,''), COALESCE(description,''),
                          COALESCE(subfolder,'Main'), COALESCE(linked_file_path,'')
                   FROM checklist_items
                   WHERE project_id=?
                   ORDER BY subfolder, sr_no, id""",
                (self.project_id,),
            ).fetchall()
            if not items:
                QMessageBox.information(self, "Template", "No checklist items found to save.")
                return

            c = conn.cursor()
            c.execute(
                "INSERT INTO checklist_templates (template_no, organization, template_name, description) VALUES (?, ?, ?, ?)",
                (tno, org, tname, tdesc),
            )
            template_id = int(c.lastrowid or 0)
            folder_path = core.ensure_template_storage_folder(org, tname, template_id)
            c.execute(
                "UPDATE checklist_templates SET folder_path=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (folder_path, template_id),
            )

            copied_files = 0
            for sr_no, req_file_name, description, subfolder, linked_file_path in items:
                c.execute(
                    """INSERT INTO checklist_template_items
                       (template_id, sr_no, req_file_name, description, subfolder)
                       VALUES (?, ?, ?, ?, ?)""",
                    (
                        template_id,
                        int(sr_no or 0),
                        str(req_file_name or ""),
                        str(description or ""),
                        str(subfolder or "Main"),
                    ),
                )
                template_item_id = int(c.lastrowid or 0)
                src = str(linked_file_path or "").strip()
                if src and os.path.isfile(src):
                    item_dir = os.path.join(folder_path, f"item_{template_item_id}")
                    os.makedirs(item_dir, exist_ok=True)
                    dst_name = core.sanitize_name(os.path.basename(src), "attachment.bin")
                    dst = os.path.join(item_dir, dst_name)
                    try:
                        shutil.copy2(src, dst)
                        c.execute(
                            """INSERT INTO checklist_template_item_files
                               (template_item_id, file_name, source_name, stored_path)
                               VALUES (?, ?, ?, ?)""",
                            (
                                template_item_id,
                                dst_name,
                                os.path.basename(src),
                                dst,
                            ),
                        )
                        copied_files += 1
                    except Exception:
                        pass
            conn.commit()
        finally:
            conn.close()
        QMessageBox.information(
            self,
            "Template",
            f"Template saved successfully.\nOrganization: {org}\nTemplate: {tname}\nAttached files copied: {copied_files}",
        )

    def _start_download_job(self, tender_id, mode, destination_folder, action):
        if self._download_future and not self._download_future.done():
            QMessageBox.information(self, "Busy", "Another docs task is already in progress.")
            return
        mode_txt = str(mode or "").strip().lower()
        if mode_txt not in {"full", "update"}:
            mode_txt = "full"
        dest = str(destination_folder or "").strip()
        if not dest:
            safe_id = core.re.sub(r'[\\/*?:"<>|]', "", str(tender_id or "").strip())
            dest = os.path.join(core.BASE_DOWNLOAD_DIRECTORY, safe_id)
        os.makedirs(dest, exist_ok=True)
        self.download_btn.setEnabled(False)
        self.download_btn.setText("Checking..." if mode_txt == "update" else "Downloading...")
        self._download_context = {
            "pid": self.project_id,
            "tender_id": str(tender_id or "").strip(),
            "mode": mode_txt,
            "action": str(action or "").strip(),
            "destination_folder": dest,
            "project_tender_docs": self._project_tender_docs_path(),
            "local_sync_count": 0,
        }
        self._download_future = self._download_executor.submit(self._download_tender_docs_worker, tender_id, dest, mode_txt)
        QTimer.singleShot(300, self._check_download_worker)

    def _download_tender_docs_worker(self, tender_id, dest, mode):
        try:
            backend = getattr(self.controller, "scraper_backend", None) or core.ScraperBackend
            mode_txt = str(mode or "").strip().lower()
            if mode_txt == "update":
                ok = backend.download_updates_for_tender_to_folder(tender_id, dest)
            else:
                ok = backend.download_docs_for_tender_to_folder(tender_id, dest)
            return {"ok": bool(ok), "error": ""}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def check_for_corrigendum(self):
        if not self.project_id:
            QMessageBox.warning(self, "Check for Corrigendum", "Open a project first.")
            return
        tender_id = str(self.source_tender_id or "").strip()
        if not tender_id:
            QMessageBox.warning(self, "Check for Corrigendum", "This project is not linked to a Tender ID.")
            return
        row = self._get_active_tender_record(tender_id)
        if not row:
            QMessageBox.information(self, "Check for Corrigendum", "Tender ID of this project was not found in Active Tenders.")
            return
        _, _, db_folder_path, _, is_archived = row
        if int(is_archived or 0) != 0:
            QMessageBox.information(self, "Check for Corrigendum", "Tender ID of this project is not in Active Tenders.")
            return

        source_folder = self._resolve_tender_download_folder(tender_id, db_folder_path)
        project_tender_docs = self._project_tender_docs_path()
        local_sync = self._copy_special_tender_docs(source_folder, project_tender_docs)
        self._start_download_job(
            tender_id=tender_id,
            mode="update",
            destination_folder=source_folder,
            action="corrigendum",
        )
        if isinstance(self._download_context, dict):
            self._download_context["local_sync_count"] = int(local_sync or 0)
            self._download_context["project_tender_docs"] = project_tender_docs

    def _check_download_worker(self):
        fut = self._download_future
        if not fut:
            return
        if not fut.done():
            QTimer.singleShot(300, self._check_download_worker)
            return

        result = {"ok": False, "error": "Unknown error"}
        try:
            result = fut.result()
        except Exception as e:
            result = {"ok": False, "error": str(e)}

        ctx = self._download_context or {}
        pid = int((ctx.get("pid") if isinstance(ctx, dict) else self.project_id) or 0)
        tender_id = str((ctx.get("tender_id") if isinstance(ctx, dict) else "") or "").strip()
        action = str((ctx.get("action") if isinstance(ctx, dict) else "") or "").strip()
        mode_txt = str((ctx.get("mode") if isinstance(ctx, dict) else "") or "").strip().lower()
        destination_folder = str((ctx.get("destination_folder") if isinstance(ctx, dict) else "") or "").strip()
        local_sync_count = int((ctx.get("local_sync_count") if isinstance(ctx, dict) else 0) or 0)
        project_tender_docs = str((ctx.get("project_tender_docs") if isinstance(ctx, dict) else "") or "").strip()

        do_import = False
        if result.get("ok"):
            if mode_txt == "update" and action == "corrigendum":
                refreshed_source = self._resolve_tender_download_folder(tender_id, destination_folder)
                online_sync = self._copy_special_tender_docs(refreshed_source, project_tender_docs)
                total_sync = local_sync_count + online_sync
                core.log_to_gui(f"Corrigendum check complete for tender {tender_id}. Synced files: {total_sync}.")
                QMessageBox.information(
                    self,
                    "Check for Corrigendum",
                    f"Corrigendum/Addendum/Pre-bid check completed.\nFiles synced: {total_sync}",
                )
            elif action == "import_docs":
                core.log_to_gui(f"Download complete for import, tender {tender_id}.")
                do_import = True
            else:
                core.log_to_gui(f"Download complete for tender {tender_id}.")
                QMessageBox.information(self, "Download", "Download complete.")
        else:
            err = str(result.get("error") or "").strip()
            title = "Check for Corrigendum" if mode_txt == "update" else "Download Docs"
            if err:
                core.log_to_gui(f"Docs task error for tender {tender_id}: {err}")
                QMessageBox.warning(self, title, f"Operation failed:\n{err}")
            else:
                core.log_to_gui(f"Docs task failed for tender {tender_id}.")
                QMessageBox.warning(self, title, "Operation failed.")

        self.download_btn.setEnabled(True)
        self.download_btn.setText("Check for Corrigendum")
        if pid and self.project_id == pid:
            self.load_project(self.project_id)
        self._download_future = None
        self._download_context = None
        if do_import and self.project_id == pid:
            self.import_tender_docs()

class ViewTendersPage(QWidget):
    org_cols = ["Sr", "OrgID", "Website", "Name", "Count", "Select"]
    tender_cols = [
        "Sr", "ID", "Website", "Tender ID", "Title", "Work Description", "Value", "EMD",
        "Org Chain", "Closing Date", "Closing Time", "Pre-Bid", "Location", "Category", "Select", "Download"
    ]
    archived_cols = [
        "Sr", "ID", "Website", "Tender ID", "Title", "Work Description", "Value", "EMD",
        "Org Chain", "Closing Date", "Closing Time", "Pre-Bid", "Location", "Category", "Status", "Select", "Download"
    ]

    def __init__(self, controller):
        super().__init__()
        self.controller = controller
        self.backend = getattr(controller, "scraper_backend", None) or BackendModeScraperProxy()
        self._task_executor = ThreadPoolExecutor(max_workers=1)
        self._task_future = None
        self._auto_fetch_timer = QTimer(self)
        self._auto_fetch_timer.timeout.connect(self._run_auto_fetch_cycle)
        self.quick_search_map = {
            "orgs": core.get_user_setting("viewtenders_search_orgs", self.backend.get_setting("viewtenders_search_orgs", "")) or "",
            "tenders": core.get_user_setting("viewtenders_search_tenders", self.backend.get_setting("viewtenders_search_tenders", "")) or "",
            "archived": core.get_user_setting("viewtenders_search_archived", self.backend.get_setting("viewtenders_search_archived", "")) or "",
        }
        self.filter_map = {
            "orgs": self._load_json_setting("viewtenders_filters_orgs", {}),
            "tenders": self._load_json_setting("viewtenders_filters_tenders", {}),
            "archived": self._load_json_setting("viewtenders_filters_archived", {}),
        }
        self.sort_map = {
            "orgs": self._load_json_setting("viewtenders_sort_orgs", {}),
            "tenders": self._load_json_setting("viewtenders_sort_tenders", {}),
            "archived": self._load_json_setting("viewtenders_sort_archived", {}),
        }
        self.column_visibility = {
            "orgs": self._build_column_visibility("orgs", self.org_cols, hidden_fixed={"OrgID"}),
            "tenders": self._build_column_visibility("tenders", self.tender_cols, hidden_fixed={"ID"}),
            "archived": self._build_column_visibility("archived", self.archived_cols, hidden_fixed={"ID"}),
        }
        self.column_widths = {
            "orgs": self._load_json_setting("viewtenders_widths_orgs", {}),
            "tenders": self._load_json_setting("viewtenders_widths_tenders", {}),
            "archived": self._load_json_setting("viewtenders_widths_archived", {}),
        }
        self.column_order = {
            "orgs": self._load_json_setting("viewtenders_order_orgs", self._load_json_setting("viewtenders_cols_orgs", [])),
            "tenders": self._load_json_setting("viewtenders_order_tenders", self._load_json_setting("viewtenders_cols_tenders", [])),
            "archived": self._load_json_setting("viewtenders_order_archived", self._load_json_setting("viewtenders_cols_archived", [])),
        }
        self._reflow_timers = {}
        self._width_save_timers = {}
        self._order_save_timers = {}

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        header_bar = QFrame()
        header_bar.setObjectName("ScraperHeaderBar")
        header = QHBoxLayout(header_bar)
        header.setContentsMargins(12, 6, 12, 6)
        header.setSpacing(8)
        title = QLabel("Online Tender Scraper")
        title.setObjectName("ScraperHeaderTitle")
        header.addWidget(title, 1)
        root.addWidget(header_bar)

        controls = QFrame()
        controls.setObjectName("ScraperControls")
        controls_lay = QVBoxLayout(controls)
        controls_lay.setContentsMargins(10, 5, 10, 5)
        controls_lay.setSpacing(5)

        top = QHBoxLayout()
        top.setSpacing(6)
        self.cb_sites = QComboBox()
        self.cb_sites.setMinimumWidth(160)
        self.cb_sites.setMaximumWidth(190)
        self.cb_sites.setFixedHeight(28)
        self.cb_sites.currentIndexChanged.connect(self.on_site_changed)
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("")
        self.search_edit.setProperty("isSearchBar", True)
        self.search_edit.setMinimumWidth(252)
        self.search_edit.setMaximumWidth(324)
        self.search_edit.setFixedHeight(34)
        self.search_edit.textChanged.connect(self.on_search_changed)
        self.btn_manage_websites = QPushButton("Manage Websites")
        self.btn_clear_data = QPushButton("Clear Data")
        self.btn_download_one = QPushButton("Download")
        self.btn_manage_websites.clicked.connect(self.manage_websites_dialog)
        self.btn_clear_data.clicked.connect(self.clear_saved_details_dialog)
        self.btn_download_one.clicked.connect(self.run_single_download)
        top_uniform_w = 112
        top_uniform_h = 29
        for b in (self.btn_manage_websites, self.btn_download_one, self.btn_clear_data):
            b.setProperty("compact", True)
            b.setFixedHeight(top_uniform_h)
            if b is self.btn_manage_websites:
                b.setFixedWidth(top_uniform_w + 14)
            else:
                b.setFixedWidth(top_uniform_w)
        top.addWidget(QLabel("Website:"))
        top.addWidget(self.cb_sites)
        top.addWidget(QLabel("Search:"))
        top.addWidget(self.search_edit)
        top.addWidget(self.btn_manage_websites)
        top.addWidget(self.btn_download_one)
        top.addStretch(1)
        top.addWidget(self.btn_clear_data)
        controls_lay.addLayout(top)

        self.actions_wrap = QWidget()
        self.actions_wrap.setFixedHeight(31)
        actions = QHBoxLayout(self.actions_wrap)
        actions.setContentsMargins(0, 0, 0, 0)
        actions.setSpacing(6)
        self.btn_fetch_orgs = QPushButton("Fetch Organizations")
        self.btn_get_tenders = QPushButton("Get Tenders")
        self.btn_download_selected = QPushButton("Download Selected")
        self.btn_select_all = QPushButton("Select All")
        self.btn_download_results = QPushButton("Download Results")
        self.btn_check_status = QPushButton("Check Status")
        self.btn_add_projects = QPushButton("Add Projects")
        self.btn_filters = QPushButton("Filters")
        self.btn_advanced = QPushButton("Advanced")
        self.btn_manage_websites.setObjectName("ScraperManageButton")
        self.btn_download_one.setObjectName("ScraperDownloadButton")
        self.btn_clear_data.setObjectName("ScraperClearButton")
        self.btn_fetch_orgs.setObjectName("ScraperFetchButton")
        self.btn_get_tenders.setObjectName("ScraperGetButton")
        self.btn_download_selected.setObjectName("ScraperDownloadSelectedButton")
        self.btn_download_results.setObjectName("ScraperDownloadResultsButton")
        self.btn_check_status.setObjectName("ScraperCheckStatusButton")
        self.btn_add_projects.setObjectName("ScraperAddProjectsButton")
        self.btn_filters.setObjectName("ScraperFiltersButton")
        self.btn_fetch_orgs.clicked.connect(self.run_fetch_orgs)
        self.btn_get_tenders.clicked.connect(self.run_fetch_tenders)
        self.btn_download_selected.clicked.connect(self.run_download)
        self.btn_select_all.clicked.connect(self.select_all_tenders)
        self.btn_download_results.clicked.connect(self.run_download_results)
        self.btn_check_status.clicked.connect(self.run_status_check)
        self.btn_add_projects.clicked.connect(self.add_selected_tenders_to_new_project)
        self.btn_filters.clicked.connect(self.open_filters_dialog)
        self.btn_advanced.clicked.connect(self.open_advanced_dialog)
        for b in (
            self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected, self.btn_select_all,
            self.btn_download_results, self.btn_check_status, self.btn_add_projects,
            self.btn_filters, self.btn_advanced
        ):
            b.setProperty("compact", True)
            b.setFixedHeight(top_uniform_h)
        action_btn_widths = {
            self.btn_fetch_orgs: 168,
            self.btn_get_tenders: 164,
            self.btn_download_selected: 168,
            self.btn_select_all: 138,
            self.btn_download_results: 168,
            self.btn_check_status: 150,
            self.btn_add_projects: 136,
            self.btn_filters: 118,
            self.btn_advanced: 126,
        }
        for b, w in action_btn_widths.items():
            b.setFixedWidth(w)
        for b in (
            self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected, self.btn_select_all,
            self.btn_download_results, self.btn_check_status
        ):
            actions.addWidget(b)
        for b in (self.btn_add_projects, self.btn_filters):
            actions.addWidget(b)
        actions.addStretch(1)
        controls_lay.addWidget(self.actions_wrap)
        root.addWidget(controls)
        self._all_action_buttons = [
            self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected, self.btn_select_all,
            self.btn_download_results, self.btn_check_status, self.btn_add_projects,
            self.btn_filters,
        ]

        body_wrap = QWidget()
        body_lay = QVBoxLayout(body_wrap)
        body_lay.setContentsMargins(10, 2, 10, 8)
        body_lay.setSpacing(0)
        self.btn_export_data = QPushButton("Export Data")
        self.btn_export_data.setObjectName("ExportDataButton")
        self.btn_export_data.setProperty("compact", True)
        self.btn_export_data.setFixedHeight(22)
        self.btn_export_data.setMinimumWidth(84)
        self.btn_export_data.clicked.connect(self.open_active_tender_export_dialog)
        self.tabs = QTabWidget()
        self._export_corner_wrap = QWidget()
        self._export_corner_lay = QHBoxLayout(self._export_corner_wrap)
        self._export_corner_lay.setContentsMargins(0, 0, 10, 2)
        self._export_corner_lay.setSpacing(0)
        self._export_corner_lay.addWidget(self.btn_export_data)
        self._export_corner_lay.setAlignment(self.btn_export_data, Qt.AlignTop | Qt.AlignRight)
        self.tabs.setCornerWidget(self._export_corner_wrap, Qt.TopRightCorner)
        self.tabs.tabBar().installEventFilter(self)
        QTimer.singleShot(0, self._sync_export_corner_geometry)
        body_lay.addWidget(self.tabs, 1)
        root.addWidget(body_wrap, 1)
        self.tab_orgs = QWidget()
        self.tab_tenders = QWidget()
        self.tab_archived = QWidget()
        self.tab_logs = QWidget()
        self.tabs.addTab(self.tab_orgs, "Organizations")
        self.tabs.addTab(self.tab_tenders, "Active Tenders")
        self.tabs.addTab(self.tab_archived, "Archived Tenders")
        self.tabs.addTab(self.tab_logs, "Live Logs")

        self.table_orgs = self._make_table(self.org_cols, hidden_cols={"OrgID"}, table_key="orgs")
        self.table_tenders = self._make_table(self.tender_cols, hidden_cols={"ID"}, table_key="tenders")
        self.table_archived = self._make_table(self.archived_cols, hidden_cols={"ID"}, table_key="archived")
        self.log_view = QTextBrowser()
        self.log_view.setObjectName("LogView")

        lo = QVBoxLayout(self.tab_orgs)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.addWidget(self.table_orgs)
        lt = QVBoxLayout(self.tab_tenders)
        lt.setContentsMargins(0, 0, 0, 0)
        lt.addWidget(self.table_tenders)
        la = QVBoxLayout(self.tab_archived)
        la.setContentsMargins(0, 0, 0, 0)
        la.addWidget(self.table_archived)
        ll = QVBoxLayout(self.tab_logs)
        ll.setContentsMargins(0, 0, 0, 0)
        ll.addWidget(self.log_view)

        self.table_orgs.cellClicked.connect(self._on_org_cell_action)
        self.table_tenders.cellClicked.connect(lambda r, c: self._on_tender_cell_action(self.table_tenders, self.tender_cols, r, c, archived=False))
        self.table_archived.cellClicked.connect(lambda r, c: self._on_tender_cell_action(self.table_archived, self.archived_cols, r, c, archived=True))
        self.table_orgs.cellDoubleClicked.connect(self._on_org_cell_action)
        self.table_tenders.cellDoubleClicked.connect(lambda r, c: self._on_tender_cell_action(self.table_tenders, self.tender_cols, r, c, archived=False))
        self.table_archived.cellDoubleClicked.connect(lambda r, c: self._on_tender_cell_action(self.table_archived, self.archived_cols, r, c, archived=True))
        self.table_orgs.horizontalHeader().sectionClicked.connect(lambda c: self.on_table_header_click("orgs", c))
        self.table_tenders.horizontalHeader().sectionClicked.connect(lambda c: self.on_table_header_click("tenders", c))
        self.table_archived.horizontalHeader().sectionClicked.connect(lambda c: self.on_table_header_click("archived", c))
        self.table_orgs.customContextMenuRequested.connect(lambda pos: self._show_table_context_menu("orgs", self.table_orgs, pos))
        self.table_tenders.customContextMenuRequested.connect(lambda pos: self._show_table_context_menu("tenders", self.table_tenders, pos))
        self.table_archived.customContextMenuRequested.connect(lambda pos: self._show_table_context_menu("archived", self.table_archived, pos))
        self.table_orgs.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_tenders.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_archived.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_orgs.horizontalHeader().customContextMenuRequested.connect(lambda pos: self._show_header_context_menu("orgs", self.table_orgs, pos))
        self.table_tenders.horizontalHeader().customContextMenuRequested.connect(lambda pos: self._show_header_context_menu("tenders", self.table_tenders, pos))
        self.table_archived.horizontalHeader().customContextMenuRequested.connect(lambda pos: self._show_header_context_menu("archived", self.table_archived, pos))
        self.table_orgs.installEventFilter(self)
        self.table_tenders.installEventFilter(self)
        self.table_archived.installEventFilter(self)
        self.tabs.currentChanged.connect(self.on_tab_changed)
        self._restore_last_tab()

        try:
            self.backend.sync_remote_state()
        except Exception as e:
            core.log_to_gui(f"Remote state sync skipped: {e}")
        self.refresh_sites()
        self.refresh_auto_fetch_settings()
        self.on_tab_changed(self.tabs.currentIndex())

    def _load_json_setting(self, key, default):
        raw = core.get_user_setting(key, None)
        if raw is None:
            raw = self.backend.get_setting(key, None)
        if raw in (None, ""):
            return default
        if isinstance(raw, (dict, list)):
            return raw
        try:
            return json.loads(raw)
        except Exception:
            return default

    def _save_json_setting(self, key, value):
        try:
            payload = json.dumps(value)
            core.set_user_setting(key, payload)
            try:
                self.backend.set_setting(key, payload)
            except Exception:
                pass
        except Exception:
            pass

    def _build_column_visibility(self, key, cols, hidden_fixed=None):
        hidden_fixed = hidden_fixed or set()
        saved = self._load_json_setting(f"viewtenders_cols_{key}", None)
        if isinstance(saved, list):
            saved_set = {str(x) for x in saved}
            out = {}
            for c in cols:
                if c in hidden_fixed:
                    out[c] = False
                else:
                    out[c] = c in saved_set
            if any(out.get(c, False) for c in cols if c not in hidden_fixed):
                return out
        return {c: (c not in hidden_fixed) for c in cols}

    def _persist_column_visibility(self, key):
        cols = self._columns_for_key(key)
        vis = self.column_visibility.get(key, {})
        out = [c for c in cols if c not in ("ID", "OrgID") and bool(vis.get(c, True))]
        self._save_json_setting(f"viewtenders_cols_{key}", out)

    def _persist_filters(self, key):
        self._save_json_setting(f"viewtenders_filters_{key}", self.filter_map.get(key, {}))

    def _restore_last_tab(self):
        last_tab = str(core.get_user_setting("viewtenders_last_tab", "orgs") or "orgs").strip().lower()
        idx_map = {"orgs": 0, "tenders": 1, "archived": 2, "logs": 3}
        self.tabs.setCurrentIndex(idx_map.get(last_tab, 0))

    def _make_table(self, headers, hidden_cols=None, table_key=None):
        hidden_cols = hidden_cols or set()
        tbl = QTableWidget(0, len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.verticalHeader().setVisible(False)
        tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.setSelectionMode(QTableWidget.ExtendedSelection)
        tbl.setAlternatingRowColors(True)
        tbl.setWordWrap(True)
        tbl.setContextMenuPolicy(Qt.CustomContextMenu)
        tbl.horizontalHeader().setSectionsMovable(True)
        tbl.horizontalHeader().setSectionsClickable(True)
        tbl.horizontalHeader().setStretchLastSection(True)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        tbl.horizontalHeader().setSortIndicatorShown(False)
        tbl.horizontalHeader().sectionResized.connect(lambda *_args, t=tbl: self._schedule_table_reflow(t))
        if table_key:
            tbl.horizontalHeader().sectionResized.connect(lambda *_args, k=table_key: self._schedule_column_width_persist(k))
            tbl.horizontalHeader().sectionMoved.connect(lambda *_args, k=table_key: self._schedule_column_order_persist(k))
        self._apply_default_column_widths(tbl, headers)
        tbl.setSortingEnabled(False)
        for col_name in hidden_cols:
            if col_name in headers:
                tbl.setColumnHidden(headers.index(col_name), True)
        if table_key:
            self._restore_column_widths(table_key)
            self._restore_column_order(table_key)
        return tbl

    def _apply_default_column_widths(self, table, headers):
        widths = {
            "Sr": 48,
            "ID": 1,
            "OrgID": 1,
            "Select": 88,
            "Download": 120,
            "Website": 160,
            "Name": 380,
            "Tender ID": 190,
            "Title": 360,
            "Work Description": 420,
            "Value": 130,
            "EMD": 120,
            "Org Chain": 240,
            "Closing Date": 140,
            "Closing Time": 110,
            "Pre-Bid": 150,
            "Location": 220,
            "Category": 140,
            "Status": 130,
            "Count": 100,
        }
        minimums = {
            "Title": 260,
            "Work Description": 280,
            "Org Chain": 220,
            "Location": 180,
            "Tender ID": 170,
            "Category": 130,
            "Status": 120,
            "Value": 120,
            "EMD": 120,
            "Pre-Bid": 140,
        }
        table.horizontalHeader().setMinimumSectionSize(36)
        for idx, col in enumerate(headers):
            w = int(widths.get(col, 120))
            table.setColumnWidth(idx, w)
            if col in minimums:
                table.horizontalHeader().resizeSection(idx, max(w, minimums[col]))

    def _schedule_column_width_persist(self, key):
        timer = self._width_save_timers.get(key)
        if timer is None:
            timer = QTimer(self)
            timer.setSingleShot(True)
            timer.timeout.connect(lambda k=key: self._persist_column_widths(k))
            self._width_save_timers[key] = timer
        timer.start(220)

    def _persist_column_widths(self, key):
        table = self._table_for_key(key)
        cols = self._columns_for_key(key)
        if table is None:
            return
        payload = {}
        for idx, col in enumerate(cols):
            if col in ("ID", "OrgID"):
                continue
            try:
                payload[col] = int(table.columnWidth(idx))
            except Exception:
                continue
        self.column_widths[key] = payload
        self._save_json_setting(f"viewtenders_widths_{key}", payload)

    def _restore_column_widths(self, key):
        table = self._table_for_key(key)
        cols = self._columns_for_key(key)
        if table is None:
            return
        saved = self.column_widths.get(key, {})
        if not isinstance(saved, dict):
            return
        for col, w in saved.items():
            if col not in cols:
                continue
            try:
                idx = cols.index(col)
                table.setColumnWidth(idx, max(36, int(w)))
            except Exception:
                continue

    def _schedule_column_order_persist(self, key):
        timer = self._order_save_timers.get(key)
        if timer is None:
            timer = QTimer(self)
            timer.setSingleShot(True)
            timer.timeout.connect(lambda k=key: self._persist_column_order(k))
            self._order_save_timers[key] = timer
        timer.start(220)

    def _persist_column_order(self, key):
        table = self._table_for_key(key)
        cols = self._columns_for_key(key)
        if table is None:
            return
        header = table.horizontalHeader()
        ordered = []
        for visual in range(header.count()):
            logical = header.logicalIndex(visual)
            if logical < 0 or logical >= len(cols):
                continue
            col = cols[logical]
            if col in ("ID", "OrgID"):
                continue
            ordered.append(col)
        self.column_order[key] = ordered
        self._save_json_setting(f"viewtenders_order_{key}", ordered)

    def _restore_column_order(self, key):
        table = self._table_for_key(key)
        cols = self._columns_for_key(key)
        if table is None:
            return
        saved = self.column_order.get(key, [])
        if not isinstance(saved, list):
            return
        target = [c for c in saved if c in cols and c not in ("ID", "OrgID")]
        for c in cols:
            if c not in ("ID", "OrgID") and c not in target:
                target.append(c)
        header = table.horizontalHeader()
        # Move by visual index to rebuild saved order deterministically.
        for visual_pos, col in enumerate(target):
            logical = cols.index(col)
            cur_visual = header.visualIndex(logical)
            if cur_visual != visual_pos and cur_visual >= 0:
                header.moveSection(cur_visual, visual_pos)

    def persist_all_column_layouts(self):
        for key in ("orgs", "tenders", "archived"):
            self._persist_column_visibility(key)
            self._persist_column_widths(key)
            self._persist_column_order(key)

    def _schedule_table_reflow(self, table):
        key = id(table)
        timer = self._reflow_timers.get(key)
        if timer is None:
            timer = QTimer(self)
            timer.setSingleShot(True)
            timer.timeout.connect(lambda t=table: self._run_table_reflow(t))
            self._reflow_timers[key] = timer
        timer.start(140)

    def _run_table_reflow(self, table):
        try:
            self._fit_table_rows(table)
            QTimer.singleShot(220, lambda t=table: self._fit_table_rows(t))
        except Exception:
            pass

    def on_table_header_click(self, table_key, col_idx):
        cols = self._columns_for_key(table_key)
        if col_idx < 0 or col_idx >= len(cols):
            return
        col_name = cols[col_idx]
        state = self.sort_map.get(table_key, {}) or {}
        if state.get("column") == col_name:
            ascending = not bool(state.get("ascending", True))
        else:
            ascending = True
        new_state = {"column": col_name, "ascending": ascending}
        self.sort_map[table_key] = new_state
        self._save_json_setting(f"viewtenders_sort_{table_key}", new_state)
        self._reload_table_by_key(table_key)

    def _reload_table_by_key(self, key):
        if key == "orgs":
            self.load_org_table()
        elif key == "tenders":
            self.load_tender_table()
        elif key == "archived":
            self.load_archived_table()

    def _is_select_sort_deferred(self, table_key):
        state = self.sort_map.get(table_key, {}) or {}
        return str(state.get("column") or "") == "Select"

    def _update_visible_select_values(self, table, cols, updates_by_id):
        if table is None or not updates_by_id or "Select" not in cols:
            return
        select_idx = cols.index("Select")
        for row in range(table.rowCount()):
            id_item = table.item(row, 1)
            if not id_item:
                continue
            row_id = str(id_item.text() or "").strip()
            if row_id not in updates_by_id:
                continue
            item = table.item(row, select_idx)
            if item is None:
                item = QTableWidgetItem("")
                table.setItem(row, select_idx, item)
            item.setText(str(updates_by_id[row_id]))

    def _show_table_context_menu(self, key, table, pos):
        menu = QMenu(self)
        a_cols = menu.addAction("Manage Columns")
        a_filters = menu.addAction("Filters")
        menu.addSeparator()
        a_refresh = menu.addAction("Refresh Table")
        if key in ("tenders", "archived"):
            menu.addSeparator()
            a_select_all = menu.addAction("Toggle Select All")
        else:
            a_select_all = None

        action = menu.exec(table.viewport().mapToGlobal(pos))
        if action == a_cols:
            self.open_manage_columns_dialog(key)
        elif action == a_filters:
            self.open_filters_dialog(key)
        elif action == a_refresh:
            self._reload_table_by_key(key)
        elif a_select_all is not None and action == a_select_all:
            idx = 2 if key == "archived" else 1
            self.tabs.setCurrentIndex(idx)
            self.select_all_tenders()

    def _show_header_context_menu(self, key, table, pos):
        header = table.horizontalHeader()
        menu = QMenu(self)
        a_cols = menu.addAction("Manage Columns")
        a_filters = menu.addAction("Filters")
        menu.addSeparator()
        a_refresh = menu.addAction("Refresh Table")
        action = menu.exec(header.mapToGlobal(pos))
        if action == a_cols:
            self.open_manage_columns_dialog(key)
        elif action == a_filters:
            self.open_filters_dialog(key)
        elif action == a_refresh:
            self._reload_table_by_key(key)

    def _apply_sort_indicator(self, table_key):
        table = self._table_for_key(table_key)
        cols = self._columns_for_key(table_key)
        if table is None:
            return
        state = self.sort_map.get(table_key, {}) or {}
        col_name = state.get("column")
        if not col_name or col_name not in cols:
            table.horizontalHeader().setSortIndicatorShown(False)
            return
        col_idx = cols.index(col_name)
        order = Qt.AscendingOrder if bool(state.get("ascending", True)) else Qt.DescendingOrder
        table.horizontalHeader().setSortIndicatorShown(True)
        table.horizontalHeader().setSortIndicator(col_idx, order)

    def apply_sort(self, table_key, cols, display_rows):
        state = self.sort_map.get(table_key, {}) or {}
        col = state.get("column")
        if not col or col not in cols:
            return display_rows
        idx = cols.index(col)
        asc = bool(state.get("ascending", True))

        # Match old Tk behavior for Value: ascending keeps NA first, descending keeps NA last.
        if col == "Value":
            def to_num(v):
                try:
                    vv = str(v).replace(",", "").replace("Rs.", "").replace("INR", "").strip()
                    return float(vv)
                except Exception:
                    return None

            if asc:
                def key_value_asc(row):
                    if idx >= len(row):
                        return (0, 0.0)
                    raw = row[idx]
                    s = str(raw).strip()
                    n = to_num(raw)
                    is_na = (not s) or (s.upper() == "NA") or (n is None)
                    return (0, 0.0) if is_na else (1, n)
                return sorted(display_rows, key=key_value_asc)
            else:
                def key_value_desc(row):
                    if idx >= len(row):
                        return (1, 0.0)
                    raw = row[idx]
                    s = str(raw).strip()
                    n = to_num(raw)
                    is_na = (not s) or (s.upper() == "NA") or (n is None)
                    return (1, 0.0) if is_na else (0, -n)
                return sorted(display_rows, key=key_value_desc)

        def key_fn(row):
            if idx >= len(row):
                return (3, "")
            val = row[idx]
            s = str(val).strip()
            try:
                n = float(s.replace(",", "").replace("Rs.", "").replace("INR", "").strip())
                return (0, n)
            except Exception:
                pass
            for fmt in ("%d-%b-%Y %I:%M %p", "%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return (1, datetime.datetime.strptime(s, fmt))
                except Exception:
                    continue
            return (2, s.lower())

        return sorted(display_rows, key=key_fn, reverse=not asc)

    def _fit_table_rows(self, table, max_height=None):
        auto_fit_table_rows(table, min_height=24, max_height=max_height)

    def _current_key(self):
        idx = self.tabs.currentIndex()
        if idx == 0:
            return "orgs"
        if idx == 1:
            return "tenders"
        if idx == 2:
            return "archived"
        return "logs"

    def on_tab_changed(self, _idx):
        key = self._current_key()
        core.set_user_setting("viewtenders_last_tab", key)
        if key in self.quick_search_map:
            self.search_edit.blockSignals(True)
            self.search_edit.setText(self.quick_search_map.get(key, ""))
            self.search_edit.blockSignals(False)
        self._update_action_buttons_for_tab(key)
        self.refresh_current_table_view()

    def _update_action_buttons_for_tab(self, key):
        show_map = {
            "orgs": {
                self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected,
                self.btn_add_projects, self.btn_filters,
            },
            "tenders": {
                self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected,
                self.btn_add_projects, self.btn_filters,
            },
            "archived": {
                self.btn_download_results, self.btn_check_status, self.btn_select_all,
                self.btn_add_projects, self.btn_filters,
            },
            "logs": set(),
        }
        visible = show_map.get(key, set())
        # Keep a fixed blank row for Logs so tab strip position stays stable.
        self.actions_wrap.setVisible(True)
        self.btn_export_data.setVisible(key == "tenders")
        for b in self._all_action_buttons:
            b.setVisible(b in visible)

    def _sync_export_corner_geometry(self):
        try:
            h = max(24, int(self.tabs.tabBar().sizeHint().height()))
            self._export_corner_wrap.setFixedHeight(h)
            self.btn_export_data.setFixedHeight(max(18, min(20, h - 8)))
        except Exception:
            pass

    def on_search_changed(self, text):
        key = self._current_key()
        if key not in self.quick_search_map:
            return
        self.quick_search_map[key] = text
        core.set_user_setting(f"viewtenders_search_{key}", text)
        self.refresh_current_table_view()

    def row_matches_quick_search(self, table_key, row_values):
        q = (self.quick_search_map.get(table_key, "") or "").strip().lower()
        if not q:
            return True
        hay = " | ".join(str(v) for v in row_values).lower()
        return q in hay

    def refresh_sites(self):
        sites = self.backend.get_websites()
        values = [("ALL", "ALL: All Websites")] + [(str(k), f"{k}: {v['name']}") for k, v in sites.items()]
        self.cb_sites.blockSignals(True)
        self.cb_sites.clear()
        for sid, label in values:
            self.cb_sites.addItem(label, sid)
        preferred = core.get_user_setting("viewtenders_selected_site", self.backend.get_setting("viewtenders_selected_site", "ALL"))
        idx = 0
        for i in range(self.cb_sites.count()):
            if str(self.cb_sites.itemData(i)).upper() == str(preferred).upper():
                idx = i
                break
        self.cb_sites.setCurrentIndex(idx)
        self.cb_sites.blockSignals(False)
        self.on_site_changed()

    def get_selected_site_id(self):
        sid = str(self.cb_sites.currentData() or "ALL").strip().upper()
        if sid == "ALL":
            return None
        try:
            return int(sid)
        except Exception:
            return None

    def get_target_site_ids(self):
        sid = self.get_selected_site_id()
        if sid is not None:
            return [sid]
        return sorted(self.backend.get_websites().keys())

    def on_site_changed(self, *_args):
        sid = self.get_selected_site_id()
        core.set_user_setting("viewtenders_selected_site", "ALL" if sid is None else str(sid))
        self.load_org_table()
        self.load_tender_table()
        self.load_archived_table()

    def refresh_current_table_view(self):
        if not all(hasattr(self, n) for n in ("table_orgs", "table_tenders", "table_archived")):
            return
        key = self._current_key()
        if key == "orgs":
            self.load_org_table()
        elif key == "tenders":
            self.load_tender_table()
        elif key == "archived":
            self.load_archived_table()

    def _table_for_key(self, key):
        if key == "orgs":
            return getattr(self, "table_orgs", None)
        if key == "tenders":
            return getattr(self, "table_tenders", None)
        if key == "archived":
            return getattr(self, "table_archived", None)
        return None

    def _columns_for_key(self, key):
        if key == "orgs":
            return self.org_cols
        if key == "tenders":
            return self.tender_cols
        if key == "archived":
            return self.archived_cols
        return []

    def _get_selected_row_ids(self, table, id_col=1):
        if table is None or table.selectionModel() is None:
            return []
        ids = []
        for mi in table.selectionModel().selectedRows():
            it = table.item(mi.row(), id_col)
            if it is None:
                continue
            rid = str(it.text() or "").strip()
            if rid:
                ids.append(rid)
        return ids

    def _restore_selected_row_ids(self, table, selected_ids, id_col=1):
        if table is None or table.selectionModel() is None:
            return
        id_set = {str(x).strip() for x in (selected_ids or []) if str(x).strip()}
        if not id_set:
            return
        sel = table.selectionModel()
        first_row = None
        for row in range(table.rowCount()):
            it = table.item(row, id_col)
            rid = str(it.text() or "").strip() if it is not None else ""
            if rid in id_set:
                idx = table.model().index(row, 0)
                sel.select(idx, QItemSelectionModel.Select | QItemSelectionModel.Rows)
                if first_row is None:
                    first_row = row
        if first_row is not None:
            table.setCurrentCell(first_row, 0, QItemSelectionModel.NoUpdate)

    def apply_column_visibility(self, key):
        table = self._table_for_key(key)
        cols = self._columns_for_key(key)
        if table is None:
            return
        visibility = self.column_visibility.get(key, {})
        for idx, col in enumerate(cols):
            if col in ("ID", "OrgID"):
                table.setColumnHidden(idx, True)
                continue
            table.setColumnHidden(idx, not bool(visibility.get(col, True)))

    def _apply_persisted_layout(self, key):
        self.apply_column_visibility(key)
        self._restore_column_order(key)
        self._restore_column_widths(key)

    def row_matches_filters(self, key, row_values):
        filters = self.filter_map.get(key, {}) or {}
        if not filters:
            return True
        cols = self._columns_for_key(key)
        value_map = {c: str(row_values[i]) if i < len(row_values) else "" for i, c in enumerate(cols)}

        def to_num(v):
            try:
                vv = str(v).replace(",", "").replace("Rs.", "").replace("INR", "").strip()
                return float(vv)
            except Exception:
                return None

        for col, needle in filters.items():
            hay_raw = value_map.get(col, "")
            hay = hay_raw.lower()
            if isinstance(needle, str):
                if needle.lower() not in hay:
                    return False
                continue
            mode = str(needle.get("mode", "")).strip().lower()
            if mode == "values":
                selected = [str(x) for x in needle.get("selected", [])]
                if selected and str(hay_raw) not in selected:
                    return False
            elif mode == "equals":
                if hay != str(needle.get("value", "")).lower():
                    return False
            elif mode == "contains":
                if str(needle.get("value", "")).lower() not in hay:
                    return False
            elif mode == "number":
                lhs = to_num(hay_raw)
                rhs = to_num(needle.get("value"))
                if lhs is None or rhs is None:
                    return False
                op = needle.get("op", "=")
                ok = (
                    (op == "=" and lhs == rhs)
                    or (op == "!=" and lhs != rhs)
                    or (op == ">" and lhs > rhs)
                    or (op == ">=" and lhs >= rhs)
                    or (op == "<" and lhs < rhs)
                    or (op == "<=" and lhs <= rhs)
                )
                if not ok:
                    return False
        return True

    def _set_row(self, table, row_idx, values, center_cols=None):
        center_cols = center_cols or set()
        for c, v in enumerate(values):
            item = QTableWidgetItem(str(v))
            if c in center_cols:
                item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row_idx, c, item)

    def load_org_table(self):
        self.table_orgs.setRowCount(0)
        conn = sqlite3.connect(core.DB_FILE)
        c = conn.cursor()
        sid = self.get_selected_site_id()
        if sid is None:
            c.execute(
                """SELECT o.id, w.name, o.name, o.tender_count, o.is_selected
                   FROM organizations o JOIN websites w ON w.id=o.website_id
                   ORDER BY w.name, o.id"""
            )
        else:
            c.execute(
                """SELECT o.id, w.name, o.name, o.tender_count, o.is_selected
                   FROM organizations o JOIN websites w ON w.id=o.website_id
                   WHERE o.website_id=? ORDER BY o.id""",
                (sid,),
            )
        rows = c.fetchall()
        conn.close()

        display = []
        for r in rows:
            row_vals = (0, r[0], r[1], r[2], r[3], ("Yes" if r[4] else "No"))
            if self.row_matches_filters("orgs", row_vals) and self.row_matches_quick_search("orgs", row_vals):
                display.append(row_vals)
        display = self.apply_sort("orgs", self.org_cols, display)
        for i, rv in enumerate(display, 1):
            vals = list(rv)
            vals[0] = i
            row = self.table_orgs.rowCount()
            self.table_orgs.insertRow(row)
            self._set_row(self.table_orgs, row, vals, center_cols={0, 1, 4, 5})
        self._apply_persisted_layout("orgs")
        self._fit_table_rows(self.table_orgs)
        self._apply_sort_indicator("orgs")
        self._schedule_table_reflow(self.table_orgs)

    def split_date_time_text(self, value):
        txt = str(value or "").strip()
        if not txt:
            return "", ""
        m = core.re.search(r"^(.+?)\s+(\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{0,2})$", txt)
        if m:
            return m.group(1).strip(), m.group(2).strip().upper()
        return txt, ""

    def get_archive_paths(self, folder_path, tender_id):
        safe_id = core.re.sub(r'[\\/*?:"<>|]', "", str(tender_id or ""))
        preferred = []
        if safe_id:
            preferred.append(os.path.join(folder_path, f"{safe_id}.rar"))
            preferred.append(os.path.join(folder_path, f"{safe_id}.zip"))
        existing = [p for p in preferred if os.path.exists(p)]
        if existing:
            return existing
        generic = []
        try:
            for name in os.listdir(folder_path):
                lower = name.lower()
                if lower.endswith(".rar") or lower.endswith(".zip"):
                    generic.append(os.path.join(folder_path, name))
        except Exception:
            return []
        return sorted(generic)

    def is_already_extracted(self, folder_path, archive_paths):
        archive_names = {os.path.basename(p).lower() for p in archive_paths}
        try:
            for name in os.listdir(folder_path):
                low = name.lower()
                if low in archive_names:
                    continue
                if low.endswith(".crdownload") or low.endswith(".part"):
                    continue
                return True
        except Exception:
            return False
        return False

    def get_download_action_label(self, folder_path, tender_id):
        if not folder_path or not os.path.exists(folder_path):
            return ""
        archive_paths = self.get_archive_paths(folder_path, tender_id)
        if not archive_paths:
            return "Open"
        return "Open" if self.is_already_extracted(folder_path, archive_paths) else "Unzip & Open"

    def load_tender_table(self):
        self.table_tenders.setRowCount(0)
        conn = sqlite3.connect(core.DB_FILE)
        c = conn.cursor()
        sid = self.get_selected_site_id()
        where_parts = []
        params = []
        if sid is not None:
            where_parts.append("t.website_id=?")
            params.append(sid)
        where_parts.append("COALESCE(t.is_archived,0)=0")
        where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
        c.execute(
            f"""SELECT t.id, w.name, t.tender_id, t.title, t.work_description, t.tender_value, t.emd, t.org_chain, t.closing_date,
                       t.pre_bid_meeting_date, t.location, t.tender_category, t.status, t.is_downloaded, t.folder_path
                FROM tenders t JOIN websites w ON w.id=t.website_id {where_sql}
                ORDER BY t.created_at DESC""",
            tuple(params),
        )
        rows = c.fetchall()
        conn.close()
        display = []
        for r in rows:
            closing_date_text, closing_time_text = self.split_date_time_text(r[8])
            download_action = self.get_download_action_label(r[14], r[2])
            row_vals = (
                0, r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], closing_date_text, closing_time_text, r[9], r[10], r[11],
                ("Yes" if r[13] else "No"), download_action
            )
            if self.row_matches_filters("tenders", row_vals) and self.row_matches_quick_search("tenders", row_vals):
                display.append(row_vals)
        display = self.apply_sort("tenders", self.tender_cols, display)
        for i, rv in enumerate(display, 1):
            vals = list(rv)
            vals[0] = i
            row = self.table_tenders.rowCount()
            self.table_tenders.insertRow(row)
            self._set_row(self.table_tenders, row, vals, center_cols={0, 1, 9, 10, 14, 15})
        self._apply_persisted_layout("tenders")
        self._fit_table_rows(self.table_tenders)
        self._apply_sort_indicator("tenders")
        self._schedule_table_reflow(self.table_tenders)

    def load_archived_table(self):
        self.table_archived.setRowCount(0)
        conn = sqlite3.connect(core.DB_FILE)
        c = conn.cursor()
        sid = self.get_selected_site_id()
        if sid is None:
            c.execute(
                """SELECT t.id, w.name, t.tender_id, t.title, t.work_description, t.tender_value, t.emd, t.org_chain, t.closing_date,
                          t.pre_bid_meeting_date, t.location, t.tender_category, t.status, t.is_downloaded, t.folder_path
                   FROM tenders t JOIN websites w ON w.id=t.website_id
                   WHERE COALESCE(t.is_archived,0)=1 ORDER BY t.created_at DESC"""
            )
        else:
            c.execute(
                """SELECT t.id, w.name, t.tender_id, t.title, t.work_description, t.tender_value, t.emd, t.org_chain, t.closing_date,
                          t.pre_bid_meeting_date, t.location, t.tender_category, t.status, t.is_downloaded, t.folder_path
                   FROM tenders t JOIN websites w ON w.id=t.website_id
                   WHERE t.website_id=? AND COALESCE(t.is_archived,0)=1 ORDER BY t.created_at DESC""",
                (sid,),
            )
        rows = c.fetchall()
        conn.close()
        display = []
        for r in rows:
            closing_date_text, closing_time_text = self.split_date_time_text(r[8])
            download_action = self.get_download_action_label(r[14], r[2])
            row_vals = (
                0, r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], closing_date_text, closing_time_text, r[9], r[10], r[11], r[12],
                ("Yes" if r[13] else "No"), download_action
            )
            if self.row_matches_filters("archived", row_vals) and self.row_matches_quick_search("archived", row_vals):
                display.append(row_vals)
        display = self.apply_sort("archived", self.archived_cols, display)
        for i, rv in enumerate(display, 1):
            vals = list(rv)
            vals[0] = i
            row = self.table_archived.rowCount()
            self.table_archived.insertRow(row)
            self._set_row(self.table_archived, row, vals, center_cols={0, 1, 9, 10, 14, 15, 16})
        self._apply_persisted_layout("archived")
        self._fit_table_rows(self.table_archived)
        self._apply_sort_indicator("archived")
        self._schedule_table_reflow(self.table_archived)

    def _on_org_cell_action(self, row, col):
        if row < 0 or col < 0 or col >= len(self.org_cols):
            return
        if self.org_cols[col] != "Select":
            return
        keep_ids = self._get_selected_row_ids(self.table_orgs, 1)
        org_item = self.table_orgs.item(row, 1)
        if not org_item:
            return
        try:
            org_id = int(str(org_item.text()).strip())
        except Exception:
            return
        sid_txt = str(org_id)
        if sid_txt not in keep_ids:
            keep_ids.append(sid_txt)
        select_idx = self.org_cols.index("Select")
        curr = str(self.table_orgs.item(row, select_idx).text() if self.table_orgs.item(row, select_idx) else "No")
        target = 0 if curr == "Yes" else 1
        conn = sqlite3.connect(core.DB_FILE)
        conn.execute("UPDATE organizations SET is_selected=? WHERE id=?", (target, org_id))
        conn.commit()
        conn.close()
        if self._is_select_sort_deferred("orgs"):
            select_item = self.table_orgs.item(row, select_idx)
            if select_item is not None:
                select_item.setText("Yes" if target else "No")
        else:
            self.load_org_table()
            self._restore_selected_row_ids(self.table_orgs, keep_ids, 1)

    def _selected_tender_db_id(self, archived=False):
        table = self.table_archived if archived else self.table_tenders
        rows = table.selectionModel().selectedRows()
        if not rows:
            return None
        row = rows[0].row()
        item = table.item(row, 1)
        if not item:
            return None
        try:
            return int(str(item.text()).strip())
        except Exception:
            return None

    def _on_tender_cell_action(self, table, cols, row, col, archived=False):
        if row < 0 or col < 0 or col >= len(cols):
            return
        col_name = cols[col]
        id_item = table.item(row, 1)
        if not id_item:
            return
        try:
            db_id = int(str(id_item.text()).strip())
        except Exception:
            return

        if col_name == "Select":
            keep_ids = self._get_selected_row_ids(table, 1)
            sid_txt = str(db_id)
            if sid_txt not in keep_ids:
                keep_ids.append(sid_txt)
            select_idx = cols.index("Select")
            curr = str(table.item(row, select_idx).text() if table.item(row, select_idx) else "No")
            target = 0 if curr == "Yes" else 1
            conn = sqlite3.connect(core.DB_FILE)
            conn.execute("UPDATE tenders SET is_downloaded=? WHERE id=?", (target, db_id))
            conn.commit()
            conn.close()
            table_key = "archived" if archived else "tenders"
            if self._is_select_sort_deferred(table_key):
                select_item = table.item(row, select_idx)
                if select_item is not None:
                    select_item.setText("Yes" if target else "No")
            else:
                self.refresh_current_table_view()
                self._restore_selected_row_ids(table, keep_ids, 1)
            return

        if col_name == "Download":
            action_idx = cols.index("Download")
            action = str(table.item(row, action_idx).text() if table.item(row, action_idx) else "")
            if action in ("Open", "Unzip & Open"):
                self.unzip_and_open_folder(db_id)

    def _toggle_selected_org_rows(self):
        rows = self.table_orgs.selectionModel().selectedRows()
        if not rows:
            return False
        updates = []
        select_idx = self.org_cols.index("Select")
        for mi in rows:
            row = mi.row()
            org_item = self.table_orgs.item(row, 1)
            if not org_item:
                continue
            try:
                org_id = int(str(org_item.text()).strip())
            except Exception:
                continue
            curr = str(self.table_orgs.item(row, select_idx).text() if self.table_orgs.item(row, select_idx) else "No")
            target = 0 if curr == "Yes" else 1
            updates.append((target, org_id))
        if not updates:
            return False
        conn = sqlite3.connect(core.DB_FILE)
        conn.executemany("UPDATE organizations SET is_selected=? WHERE id=?", updates)
        conn.commit()
        conn.close()
        keep_ids = [str(org_id) for _target, org_id in updates]
        if self._is_select_sort_deferred("orgs"):
            self._update_visible_select_values(
                self.table_orgs,
                self.org_cols,
                {str(org_id): ("Yes" if target else "No") for target, org_id in updates},
            )
        else:
            self.load_org_table()
            self._restore_selected_row_ids(self.table_orgs, keep_ids, 1)
        return True

    def _toggle_selected_tender_rows(self, table, cols):
        rows = table.selectionModel().selectedRows()
        if not rows:
            return False
        select_idx = cols.index("Select")
        updates = []
        for mi in rows:
            row = mi.row()
            id_item = table.item(row, 1)
            if not id_item:
                continue
            try:
                db_id = int(str(id_item.text()).strip())
            except Exception:
                continue
            curr = str(table.item(row, select_idx).text() if table.item(row, select_idx) else "No")
            target = 0 if curr == "Yes" else 1
            updates.append((target, db_id))
        if not updates:
            return False
        conn = sqlite3.connect(core.DB_FILE)
        conn.executemany("UPDATE tenders SET is_downloaded=? WHERE id=?", updates)
        conn.commit()
        conn.close()
        keep_ids = [str(db_id) for _target, db_id in updates]
        table_key = "archived" if table is self.table_archived else "tenders"
        if self._is_select_sort_deferred(table_key):
            self._update_visible_select_values(
                table,
                cols,
                {str(db_id): ("Yes" if target else "No") for target, db_id in updates},
            )
        else:
            self.refresh_current_table_view()
            self._restore_selected_row_ids(table, keep_ids, 1)
        return True

    def eventFilter(self, obj, event):
        if hasattr(self, "tabs") and obj is self.tabs.tabBar():
            if event.type() == QEvent.Resize:
                QTimer.singleShot(0, self._sync_export_corner_geometry)
        table_orgs = getattr(self, "table_orgs", None)
        table_tenders = getattr(self, "table_tenders", None)
        table_archived = getattr(self, "table_archived", None)
        if obj in (table_orgs, table_tenders, table_archived):
            if event.type() == QEvent.Resize:
                self._schedule_table_reflow(obj)
            elif event.type() == QEvent.Wheel and bool(event.modifiers() & Qt.ShiftModifier):
                try:
                    delta = event.angleDelta().y()
                    if delta:
                        step = -8 if delta > 0 else 8
                        bar = obj.horizontalScrollBar()
                        bar.setValue(bar.value() + (step * max(1, bar.singleStep())))
                        return True
                except Exception:
                    pass
        if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Space:
            if obj is table_orgs:
                if self._toggle_selected_org_rows():
                    return True
            elif obj is table_tenders:
                if self._toggle_selected_tender_rows(self.table_tenders, self.tender_cols):
                    return True
            elif obj is table_archived:
                if self._toggle_selected_tender_rows(self.table_archived, self.archived_cols):
                    return True
        return super().eventFilter(obj, event)

    def append_log(self, text):
        self.log_view.append(str(text))
        self.log_view.verticalScrollBar().setValue(self.log_view.verticalScrollBar().maximum())

    def open_download_root_folder(self):
        folder = core._resolve_path(core.BASE_DOWNLOAD_DIRECTORY)
        try:
            if sys.platform.startswith("win"):
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception as e:
            QMessageBox.critical(self, "Open Folder", f"Could not open folder:\n{e}")

    def unzip_and_open_folder(self, db_id):
        conn = sqlite3.connect(core.DB_FILE)
        row = conn.execute("SELECT folder_path, tender_id FROM tenders WHERE id=?", (db_id,)).fetchone()
        conn.close()
        if not row:
            return
        folder_path, tender_id = row
        if not (folder_path and os.path.exists(folder_path)):
            return
        archive_paths = self.get_archive_paths(folder_path, tender_id)
        if archive_paths and not self.is_already_extracted(folder_path, archive_paths):
            for ap in archive_paths:
                self.extract_archive(ap, folder_path)
        try:
            if sys.platform.startswith("win"):
                os.startfile(folder_path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder_path])
            else:
                subprocess.Popen(["xdg-open", folder_path])
            self.on_site_changed()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not open the folder:\n{e}")

    def extract_archive(self, archive_path, dest_dir):
        low = archive_path.lower()
        try:
            if low.endswith(".zip"):
                with core.zipfile.ZipFile(archive_path, "r") as zf:
                    zf.extractall(dest_dir)
                return True
            if low.endswith(".rar"):
                try:
                    import patoolib
                    patoolib.extract_archive(archive_path, outdir=dest_dir, verbosity=-1)
                    return True
                except Exception:
                    pass
                try:
                    proc = subprocess.run(["7z", "x", "-y", f"-o{dest_dir}", archive_path], capture_output=True, text=True)
                    return proc.returncode == 0
                except Exception:
                    return False
        except Exception:
            return False
        return False

    def _run_bg(self, worker_fn, done_refresh=True, switch_to_logs=True):
        if self._task_future and not self._task_future.done():
            if switch_to_logs:
                QMessageBox.information(self, "Busy", "Another scraper task is already running.")
            return
        if switch_to_logs:
            self.tabs.setCurrentIndex(3)
        self._task_future = self._task_executor.submit(worker_fn)
        if done_refresh:
            QTimer.singleShot(300, self._poll_bg_task)

    def _poll_bg_task(self):
        if not self._task_future:
            return
        if not self._task_future.done():
            QTimer.singleShot(300, self._poll_bg_task)
            return
        try:
            self._task_future.result()
        except Exception as e:
            self.append_log(f"Task failed: {e}")
        self._task_future = None
        self.on_site_changed()

    def _sync_remote_scraper_state_if_needed(self):
        try:
            if hasattr(self.backend, "push_local_state"):
                self.backend.push_local_state()
        except Exception as e:
            core.log_to_gui(f"Remote state push failed: {e}")

    def refresh_auto_fetch_settings(self):
        enabled = bool(core.get_user_setting("scraper_auto_fetch_enabled", False))
        try:
            minutes = int(str(core.get_user_setting("scraper_auto_fetch_interval_minutes", 30) or "30").strip())
        except Exception:
            minutes = 30
        minutes = max(1, min(24 * 60, minutes))
        if enabled:
            self._auto_fetch_timer.start(minutes * 60 * 1000)
        else:
            self._auto_fetch_timer.stop()

    def _mark_auto_fetch_run(self):
        stamp = datetime.datetime.now(datetime.timezone.utc).isoformat()
        core.set_user_setting("scraper_last_auto_fetch_at", stamp)
        def _refresh_label():
            if hasattr(self.controller, "settings_page") and self.controller.settings_page is not None:
                try:
                    self.controller.settings_page.scraper_last_auto_fetch_label.setText(
                        self.controller.settings_page._format_last_auto_fetch_text()
                    )
                except Exception:
                    pass
        QTimer.singleShot(0, _refresh_label)

    def _run_auto_fetch_cycle(self):
        if self._task_future and not self._task_future.done():
            return

        def site_has_selected_orgs(site_id):
            conn = sqlite3.connect(core.DB_FILE)
            try:
                row = conn.execute(
                    """SELECT COUNT(*)
                       FROM organizations
                       WHERE website_id=?
                         AND COALESCE(is_selected,0)=1""",
                    (site_id,),
                ).fetchone()
                return int((row[0] if row else 0) or 0) > 0
            finally:
                conn.close()

        def worker():
            target_sites = self.get_target_site_ids()
            if not target_sites:
                core.log_to_gui("Auto fetch skipped: no websites configured.")
                self._mark_auto_fetch_run()
                return
            self._sync_remote_scraper_state_if_needed()
            core.log_to_gui("Auto fetch started for selected organizations and tenders.")
            for sid in target_sites:
                self.backend.fetch_organisations_logic(sid)
            eligible_sites = [sid for sid in target_sites if site_has_selected_orgs(sid)]
            if not eligible_sites:
                core.log_to_gui("Auto fetch skipped tender refresh: no organizations selected.")
                self._mark_auto_fetch_run()
                return
            for sid in eligible_sites:
                self.backend.fetch_tenders_logic(sid)
            core.log_to_gui("Auto fetch completed.")
            self._mark_auto_fetch_run()

        self._run_bg(worker, done_refresh=True, switch_to_logs=False)

    def run_auto_fetch_now(self):
        self._run_auto_fetch_cycle()

    def run_fetch_orgs(self):
        def worker():
            self._sync_remote_scraper_state_if_needed()
            for sid in self.get_target_site_ids():
                self.backend.fetch_organisations_logic(sid)
        self._run_bg(worker)

    def run_fetch_tenders(self):
        def site_has_selected_orgs(site_id):
            conn = sqlite3.connect(core.DB_FILE)
            try:
                row = conn.execute(
                    """SELECT COUNT(*)
                       FROM organizations
                       WHERE website_id=?
                         AND COALESCE(is_selected,0)=1""",
                    (site_id,),
                ).fetchone()
                return int((row[0] if row else 0) or 0) > 0
            finally:
                conn.close()

        def worker():
            self._sync_remote_scraper_state_if_needed()
            eligible_sites = [sid for sid in self.get_target_site_ids() if site_has_selected_orgs(sid)]
            if not eligible_sites:
                core.log_to_gui("No organizations selected. Please select organizations first.")
                return
            for sid in eligible_sites:
                self.backend.fetch_tenders_logic(sid)
        self._run_bg(worker)

    def run_download(self):
        def site_has_marked_tenders(site_id):
            conn = sqlite3.connect(core.DB_FILE)
            try:
                row = conn.execute(
                    """SELECT COUNT(*)
                       FROM tenders
                       WHERE website_id=?
                         AND COALESCE(is_downloaded,0)=1
                         AND COALESCE(is_archived,0)=0""",
                    (site_id,),
                ).fetchone()
                return int((row[0] if row else 0) or 0) > 0
            finally:
                conn.close()

        def worker():
            self._sync_remote_scraper_state_if_needed()
            eligible_sites = [sid for sid in self.get_target_site_ids() if site_has_marked_tenders(sid)]
            if not eligible_sites:
                core.log_to_gui("No tenders marked for download.")
                return
            for sid in eligible_sites:
                self.backend.download_tenders_logic(sid)
        self._run_bg(worker)

    def run_status_check(self):
        archived_mode = (self.tabs.currentIndex() == 2)
        def worker():
            self._sync_remote_scraper_state_if_needed()
            for sid in self.get_target_site_ids():
                self.backend.check_tender_status_logic(sid, archived_only=archived_mode)
        self._run_bg(worker)

    def run_download_results(self):
        def worker():
            self._sync_remote_scraper_state_if_needed()
            for sid in self.get_target_site_ids():
                self.backend.download_tender_results_logic(sid)
        self._run_bg(worker)

    def run_single_download(self):
        if self.tabs.currentIndex() != 1:
            QMessageBox.warning(self, "Download", "Switch to Active Tenders and select one tender.")
            return
        tender_db_id = self._selected_tender_db_id(archived=False)
        if tender_db_id is None:
            QMessageBox.warning(self, "Download", "Select/highlight one active tender first.")
            return
        mode = self._ask_single_download_mode()
        if not mode:
            return

        def worker():
            self._sync_remote_scraper_state_if_needed()
            self.backend.download_single_tender_logic(tender_db_id, mode)
        self._run_bg(worker)

    def _ask_single_download_mode(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Download Mode")
        dlg.resize(440, 170)
        root = QVBoxLayout(dlg)
        note = QLabel("Choose download mode for selected tender.")
        root.addWidget(note)
        btns = QHBoxLayout()
        btns.setSpacing(12)
        full_btn = QPushButton("Full")
        update_btn = QPushButton("Update")
        cancel_btn = QPushButton("Cancel")
        full_btn.setObjectName("PrimaryButton")
        for b in (full_btn, update_btn, cancel_btn):
            b.setMinimumWidth(120)
            b.setMinimumHeight(44)
        btns.addStretch(1)
        btns.addWidget(full_btn)
        btns.addWidget(update_btn)
        btns.addWidget(cancel_btn)
        btns.addStretch(1)
        root.addLayout(btns)
        choice = {"mode": None}
        full_btn.clicked.connect(lambda: (choice.update({"mode": "full"}), dlg.accept()))
        update_btn.clicked.connect(lambda: (choice.update({"mode": "update"}), dlg.accept()))
        cancel_btn.clicked.connect(dlg.reject)
        if dlg.exec() != QDialog.Accepted:
            return None
        return choice["mode"]

    def select_all_tenders(self):
        sid = self.get_selected_site_id()
        archived_mode = (self.tabs.currentIndex() == 2)
        conn = sqlite3.connect(core.DB_FILE)
        where_parts = []
        params = []
        if sid is not None:
            where_parts.append("website_id=?")
            params.append(sid)
        if archived_mode:
            where_parts.append("COALESCE(is_archived,0)=1")
        else:
            where_parts.append("COALESCE(is_archived,0)=0")
        where_sql = " AND ".join(where_parts)
        total = conn.execute(f"SELECT COUNT(*) FROM tenders WHERE {where_sql}", tuple(params)).fetchone()[0]
        selected = conn.execute(
            f"SELECT COUNT(*) FROM tenders WHERE {where_sql} AND COALESCE(is_downloaded,0)=1",
            tuple(params),
        ).fetchone()[0]
        target = 0 if total > 0 and selected == total else 1
        conn.execute(f"UPDATE tenders SET is_downloaded=? WHERE {where_sql}", (target, *params))
        conn.commit()
        conn.close()
        table = self.table_archived if archived_mode else self.table_tenders
        cols = self.archived_cols if archived_mode else self.tender_cols
        table_key = "archived" if archived_mode else "tenders"
        if self._is_select_sort_deferred(table_key):
            visible_updates = {}
            for row in range(table.rowCount()):
                id_item = table.item(row, 1)
                if id_item:
                    visible_updates[str(id_item.text() or "").strip()] = "Yes" if target else "No"
            self._update_visible_select_values(table, cols, visible_updates)
        else:
            self.on_site_changed()

    def manage_websites_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Manage Websites")
        dlg.resize(980, 560)
        root = QVBoxLayout(dlg)

        form_wrap = QWidget()
        form = QGridLayout(form_wrap)
        form.setContentsMargins(0, 0, 0, 0)
        e_name = QLineEdit()
        e_url = QLineEdit()
        e_surl = QLineEdit()
        form.addWidget(QLabel("Name:"), 0, 0)
        form.addWidget(e_name, 0, 1)
        form.addWidget(QLabel("Tenders URL:"), 0, 2)
        form.addWidget(e_url, 0, 3)
        form.addWidget(QLabel("Status URL:"), 0, 4)
        form.addWidget(e_surl, 0, 5)
        form.setColumnStretch(1, 2)
        form.setColumnStretch(3, 3)
        form.setColumnStretch(5, 3)
        root.addWidget(form_wrap)

        cols = ["Sr", "ID", "Name", "Tenders URL", "Status URL"]
        table = QTableWidget(0, len(cols))
        table.setHorizontalHeaderLabels(cols)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.setAlternatingRowColors(True)
        table.setWordWrap(True)
        table.horizontalHeader().setStretchLastSection(True)
        rowfit_timer = QTimer(dlg)
        rowfit_timer.setSingleShot(True)
        rowfit_timer.timeout.connect(lambda: auto_fit_table_rows(table, min_height=24, max_height=None))
        table.horizontalHeader().sectionResized.connect(lambda *_args: rowfit_timer.start(90))
        root.addWidget(table, 1)

        action_row = QHBoxLayout()
        btn_add = QPushButton("Add Website")
        btn_add.setObjectName("PrimaryButton")
        btn_edit = QPushButton("Edit Selected")
        btn_del = QPushButton("Delete Selected")
        btn_del.setObjectName("DangerButton")
        btn_refresh = QPushButton("Refresh")
        btn_close = QPushButton("Close")
        for b in (btn_add, btn_edit, btn_del, btn_refresh):
            action_row.addWidget(b)
        action_row.addStretch(1)
        action_row.addWidget(btn_close)
        root.addLayout(action_row)

        def set_row(row_idx, vals):
            for c, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                if c in (0, 1):
                    item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row_idx, c, item)

        def fill_table():
            table.setRowCount(0)
            websites = self.backend.get_websites()
            for idx, (sid, info) in enumerate(sorted(websites.items(), key=lambda x: int(x[0])), 1):
                row = table.rowCount()
                table.insertRow(row)
                set_row(row, [idx, sid, info["name"], info["url"], info["status_url"]])
            table.resizeColumnsToContents()
            auto_fit_table_rows(table, min_height=24, max_height=None)

        def selected_site_row():
            rows = table.selectionModel().selectedRows()
            if not rows:
                return None
            return rows[0].row()

        def load_selected_into_form():
            row = selected_site_row()
            if row is None:
                return
            e_name.setText(table.item(row, 2).text() if table.item(row, 2) else "")
            e_url.setText(table.item(row, 3).text() if table.item(row, 3) else "")
            e_surl.setText(table.item(row, 4).text() if table.item(row, 4) else "")

        def add_new():
            name = e_name.text().strip()
            url = e_url.text().strip()
            status_url = e_surl.text().strip()
            if not name or not url or not status_url:
                QMessageBox.warning(dlg, "Add Website", "All fields are required.")
                return
            if self.backend.add_website_logic(name, url, status_url):
                e_name.clear()
                e_url.clear()
                e_surl.clear()
                fill_table()
                self.refresh_sites()

        def edit_selected():
            row = selected_site_row()
            if row is None:
                QMessageBox.warning(dlg, "Edit Website", "Select a website first.")
                return
            sid_item = table.item(row, 1)
            sid = int(str(sid_item.text()).strip()) if sid_item else None
            if sid is None:
                QMessageBox.warning(dlg, "Edit Website", "Invalid selected website.")
                return
            name = e_name.text().strip()
            url = e_url.text().strip()
            status_url = e_surl.text().strip()
            if not name or not url or not status_url:
                QMessageBox.warning(dlg, "Edit Website", "All fields are required.")
                return
            conn = sqlite3.connect(core.DB_FILE)
            try:
                conn.execute("UPDATE websites SET name=?, url=?, status_url=? WHERE id=?", (name, url, status_url, sid))
                conn.commit()
            except Exception as e:
                QMessageBox.critical(dlg, "Edit Website", str(e))
            finally:
                conn.close()
            fill_table()
            self.refresh_sites()

        def delete_selected():
            row = selected_site_row()
            if row is None:
                QMessageBox.warning(dlg, "Delete Website", "Select a website first.")
                return
            sid_item = table.item(row, 1)
            name_item = table.item(row, 2)
            sid = int(str(sid_item.text()).strip()) if sid_item else None
            name = str(name_item.text()).strip() if name_item else ""
            if sid is None:
                QMessageBox.warning(dlg, "Delete Website", "Invalid selected website.")
                return
            ok = QMessageBox.question(dlg, "Confirm", f"Delete website '{name}' and all related scraper data?") == QMessageBox.Yes
            if not ok:
                return
            if self.backend.delete_website_logic(sid):
                fill_table()
                self.refresh_sites()

        table.itemSelectionChanged.connect(load_selected_into_form)
        btn_add.clicked.connect(add_new)
        btn_edit.clicked.connect(edit_selected)
        btn_del.clicked.connect(delete_selected)
        btn_refresh.clicked.connect(fill_table)
        btn_close.clicked.connect(dlg.close)

        fill_table()
        dlg.exec()

    def clear_saved_details_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Clear Saved Details")
        dlg.resize(460, 280)
        root = QVBoxLayout(dlg)
        root.addWidget(QLabel("Choose what to clear:"))

        chk_org = QCheckBox("1. Organization")
        chk_active = QCheckBox("2. Active Tenders")
        chk_arch = QCheckBox("3. Archive")
        chk_org.setChecked(True)
        chk_active.setChecked(True)
        chk_arch.setChecked(True)
        root.addWidget(chk_org)
        root.addWidget(chk_active)
        root.addWidget(chk_arch)
        note = QLabel("Websites are kept. Downloaded files on disk are not deleted.")
        note.setObjectName("SoftText")
        root.addWidget(note)
        root.addStretch(1)

        btns = QHBoxLayout()
        btn_clear = QPushButton("Clear Selected")
        btn_clear.setObjectName("DangerButton")
        btn_cancel = QPushButton("Cancel")
        btns.addWidget(btn_clear)
        btns.addStretch(1)
        btns.addWidget(btn_cancel)
        root.addLayout(btns)

        def run_clear():
            clear_orgs = chk_org.isChecked()
            clear_active = chk_active.isChecked()
            clear_archived = chk_arch.isChecked()
            if not (clear_orgs or clear_active or clear_archived):
                QMessageBox.warning(dlg, "Clear Saved Details", "Select at least one option.")
                return
            ok = QMessageBox.question(dlg, "Confirm", "Proceed with selected clear options?") == QMessageBox.Yes
            if not ok:
                return
            result = self.backend.clear_saved_scraper_details_logic(
                clear_orgs=clear_orgs,
                clear_active=clear_active,
                clear_archived=clear_archived,
            )
            if result is None:
                QMessageBox.critical(dlg, "Clear Saved Details", "Failed to clear saved details. Check logs.")
                return
            dlg.accept()
            self.refresh_sites()
            self.tabs.setCurrentIndex(3)
            summary = (
                f"Cleared: organizations={result['organizations']}, "
                f"active={result['active_tenders']}, archived={result['archived_tenders']}, "
                f"downloads={result['downloaded_files']}"
            )
            core.log_to_gui(summary)
            QMessageBox.information(self, "Clear Saved Details", summary)

        btn_clear.clicked.connect(run_clear)
        btn_cancel.clicked.connect(dlg.reject)
        dlg.exec()

    def add_selected_tenders_to_new_project(self):
        sid = self.get_selected_site_id()
        archived_mode = (self.tabs.currentIndex() == 2)
        conn = sqlite3.connect(core.DB_FILE)
        c = conn.cursor()
        where = ["COALESCE(is_downloaded,0)=1"]
        params = []
        if sid is not None:
            where.append("website_id=?")
            params.append(sid)
        if archived_mode:
            where.append("COALESCE(is_archived,0)=1")
        else:
            where.append("COALESCE(is_archived,0)=0")
        where_sql = " AND ".join(where)
        c.execute(
            f"""SELECT tender_id, title, org_chain, closing_date, website_id, tender_value, pre_bid_meeting_date, folder_path
                FROM tenders
                WHERE {where_sql}
                ORDER BY created_at DESC""",
            tuple(params),
        )
        rows = c.fetchall()
        conn.close()
        if not rows:
            QMessageBox.information(self, "Add to New Project", "No selected tenders found. Select tenders first.")
            return

        pending_rows = self._rows_not_in_projects(rows)
        if not pending_rows:
            QMessageBox.information(self, "Add Projects", "All selected tenders are already added as projects.")
            return

        first = pending_rows[0]
        first_tid = str(first[0] or "").strip()
        first_title = str(first[1] or "").strip()
        first_org = str(first[2] or "").strip()
        first_deadline = str(first[3] or "").strip()
        first_value = str(first[5] or "").strip()
        first_prebid = str(first[6] or "").strip()
        first_tender_folder = str(first[7] or "").strip()
        if len(pending_rows) > 1:
            pick = QMessageBox(self)
            pick.setWindowTitle("Add Projects")
            pick.setIcon(QMessageBox.Question)
            pick.setText(f"{len(pending_rows)} tenders are selected.")
            pick.setInformativeText(
                "Yes: create one project per selected tender.\n"
                "Edit: open editable prefilled forms with Previous/Next.\n"
                "No: open only a single prefilled project form."
            )
            yes_btn = pick.addButton("Yes", QMessageBox.YesRole)
            edit_btn = pick.addButton("Edit", QMessageBox.ActionRole)
            no_btn = pick.addButton("No", QMessageBox.NoRole)
            pick.setDefaultButton(yes_btn)
            pick.exec()
            clicked = pick.clickedButton()
            if clicked is yes_btn:
                created, skipped = self.create_projects_from_rows(pending_rows)
                QMessageBox.information(self, "Add Projects", f"Created {created} project(s). Skipped {skipped}.")
                self.controller.show_projects()
                return
            if clicked is edit_btn:
                prefills = [self._row_to_project_prefill(r) for r in pending_rows]
                dlg = CreateProjectDialog(self, prefill=prefills[0], batch_prefills=prefills, start_index=0)
                if dlg.exec() == QDialog.Accepted:
                    self.controller.show_projects()
                return
            if clicked is not no_btn:
                return

        default_title = first_tid or "Tender Project"
        prefill = {
            "title": default_title,
            "client_name": first_org,
            "deadline": first_deadline,
            "description": first_title,
            "tender_id": first_tid,
            "project_value": first_value,
            "prebid": first_prebid,
            "tender_folder_path": first_tender_folder,
        }
        dlg = CreateProjectDialog(self, prefill=prefill)
        if dlg.exec() == QDialog.Accepted:
            self.controller.show_projects()

    def _rows_not_in_projects(self, rows):
        ids = [str(r[0] or "").strip() for r in rows if str(r[0] or "").strip()]
        if not ids:
            return list(rows)
        conn = sqlite3.connect(core.DB_FILE)
        try:
            ph = ",".join(["?"] * len(ids))
            existing = conn.execute(
                f"SELECT COALESCE(source_tender_id,'') FROM projects WHERE COALESCE(source_tender_id,'') IN ({ph})",
                tuple(ids),
            ).fetchall()
        finally:
            conn.close()
        existing_ids = {str(x[0] or "").strip() for x in existing if x and str(x[0] or "").strip()}
        return [r for r in rows if str(r[0] or "").strip() not in existing_ids]

    def _row_to_project_prefill(self, row):
        tid = str(row[0] or "").strip()
        title = str(row[1] or "").strip()
        org = str(row[2] or "").strip()
        deadline = str(row[3] or "").strip()
        value = str(row[5] or "").strip()
        prebid = str(row[6] or "").strip()
        tender_folder = str(row[7] or "").strip()
        return {
            "title": tid or "Tender Project",
            "client_name": org,
            "deadline": deadline,
            "description": title,
            "tender_id": tid,
            "project_value": value,
            "prebid": prebid,
            "tender_folder_path": tender_folder,
            "status": "Active",
        }

    def create_projects_from_rows(self, rows):
        def safe_name(text):
            return "".join(c for c in str(text or "") if c.isalnum() or c in (" ", "_", "-")).strip() or "Tender Project"

        def ensure_unique_folder(base):
            path = os.path.join(core.ROOT_FOLDER, base)
            if not os.path.exists(path):
                return path
            i = 2
            while True:
                p = os.path.join(core.ROOT_FOLDER, f"{base}_{i}")
                if not os.path.exists(p):
                    return p
                i += 1

        raw = core.ScraperBackend.get_setting("project_client_options", "[]")
        try:
            client_opts = [str(x).strip() for x in json.loads(raw) if str(x).strip()]
        except Exception:
            client_opts = []

        conn = sqlite3.connect(core.DB_FILE)
        c = conn.cursor()
        created = 0
        skipped = 0
        try:
            for tid, title, org, closing, _wid, project_value, prebid, tender_folder_path in rows:
                project_title = str(tid or "").strip() or "Tender Project"
                source_tender_id = str(tid or "").strip()
                if source_tender_id:
                    exists = c.execute(
                        "SELECT 1 FROM projects WHERE source_tender_id=? LIMIT 1",
                        (source_tender_id,),
                    ).fetchone()
                    if exists:
                        skipped += 1
                        continue
                folder_path = ensure_unique_folder(safe_name(project_title))
                std_folders = core.ensure_project_standard_folders(folder_path)
                c.execute(
                    "INSERT INTO projects (title, client_name, deadline, description, folder_path, source_tender_id, project_value, prebid) VALUES (?,?,?,?,?,?,?,?)",
                    (
                        project_title,
                        str(org or "").strip(),
                        str(closing or "").strip(),
                        str(title or "").strip(),
                        folder_path,
                        source_tender_id or None,
                        str(project_value or "").strip(),
                        str(prebid or "").strip(),
                    ),
                )
                tender_src = str(tender_folder_path or "").strip()
                if tender_src and os.path.isdir(tender_src):
                    copied = core.copy_tree_contents(tender_src, std_folders["tender_docs"])
                    if copied:
                        core.log_to_gui(f"Copied {copied} tender item(s) to project Tender Docs: {project_title}")
                created += 1
                if org and str(org).strip() and str(org).strip() not in client_opts:
                    client_opts.append(str(org).strip())
            conn.commit()
        finally:
            conn.close()

        client_opts = sorted(set(client_opts), key=lambda x: x.lower())
        core.ScraperBackend.set_setting("project_client_options", json.dumps(client_opts))
        return created, skipped

    def open_advanced_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Advanced")
        dlg.resize(560, 220)
        root = QVBoxLayout(dlg)
        key = self._current_key()
        table_name = {
            "orgs": "Organizations",
            "tenders": "Active Tenders",
            "archived": "Archived Tenders",
            "logs": "Logs",
        }.get(key, "Current")
        root.addWidget(QLabel(f"Current Tab: {table_name}"))
        root.addWidget(QLabel("Column visibility and text filters are now available in Qt."))
        row = QHBoxLayout()
        btn_cols = QPushButton("Manage Columns")
        btn_filters = QPushButton("Filters")
        btn_export = QPushButton("Export Data")
        row.addWidget(btn_cols)
        row.addWidget(btn_filters)
        row.addWidget(btn_export)
        root.addLayout(row)
        root.addStretch(1)
        close_btn = QPushButton("Close")
        root.addWidget(close_btn, 0, Qt.AlignRight)

        btn_cols.clicked.connect(self.open_manage_columns_dialog)
        btn_filters.clicked.connect(self.open_filters_dialog)
        btn_export.clicked.connect(self.open_active_tender_export_dialog)
        close_btn.clicked.connect(dlg.close)
        dlg.exec()

    def open_manage_columns_dialog(self, key_override=None):
        key = key_override or self._current_key()
        if key not in ("orgs", "tenders", "archived"):
            QMessageBox.information(self, "Manage Columns", "Column settings are available for table tabs.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Manage Columns")
        dlg.resize(520, 620)
        root = QVBoxLayout(dlg)
        cols = self._columns_for_key(key)
        root.addWidget(QLabel("Drag to reorder columns, and check/uncheck to show or hide."))

        list_cols = QListWidget()
        list_cols.setDragDropMode(QAbstractItemView.InternalMove)
        list_cols.setDefaultDropAction(Qt.MoveAction)
        list_cols.setSelectionMode(QAbstractItemView.SingleSelection)
        root.addWidget(list_cols, 1)

        vis = self.column_visibility.get(key, {}) or {}
        base = [c for c in cols if c not in ("ID", "OrgID")]
        saved_order = self.column_order.get(key, []) if isinstance(self.column_order.get(key, []), list) else []
        ordered = [c for c in saved_order if c in base]
        for c in base:
            if c not in ordered:
                ordered.append(c)
        in_populate = {"value": False}
        auto_apply_timer = QTimer(dlg)
        auto_apply_timer.setSingleShot(True)

        def populate():
            in_populate["value"] = True
            list_cols.clear()
            for col in ordered:
                it = QListWidgetItem(col)
                it.setFlags(it.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsDragEnabled | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                it.setCheckState(Qt.Checked if bool(vis.get(col, True)) else Qt.Unchecked)
                list_cols.addItem(it)
            in_populate["value"] = False

        populate()

        btns = QHBoxLayout()
        apply_btn = QPushButton("Apply")
        clear_btn = QPushButton("Hide All")
        show_btn = QPushButton("Show All")
        close_btn = QPushButton("Close")
        btns.addWidget(apply_btn)
        btns.addWidget(clear_btn)
        btns.addWidget(show_btn)
        btns.addStretch(1)
        btns.addWidget(close_btn)
        root.addLayout(btns)

        def apply_now():
            cur = self.column_visibility.get(key, {})
            new_order = []
            for i in range(list_cols.count()):
                it = list_cols.item(i)
                col = str(it.text())
                cur[col] = (it.checkState() == Qt.Checked)
                new_order.append(col)
            for hidden in ("ID", "OrgID"):
                cur[hidden] = False
            self.column_visibility[key] = cur
            self.column_order[key] = new_order
            self.apply_column_visibility(key)
            self._restore_column_order(key)
            self._persist_column_visibility(key)
            self._save_json_setting(f"viewtenders_order_{key}", new_order)
            self._persist_column_order(key)

        auto_apply_timer.timeout.connect(apply_now)

        def schedule_apply():
            if in_populate["value"]:
                return
            auto_apply_timer.start(80)

        def hide_all():
            list_cols.blockSignals(True)
            for i in range(list_cols.count()):
                list_cols.item(i).setCheckState(Qt.Unchecked)
            list_cols.blockSignals(False)
            apply_now()

        def show_all():
            list_cols.blockSignals(True)
            for i in range(list_cols.count()):
                list_cols.item(i).setCheckState(Qt.Checked)
            list_cols.blockSignals(False)
            apply_now()

        apply_btn.clicked.connect(apply_now)
        clear_btn.clicked.connect(hide_all)
        show_btn.clicked.connect(show_all)
        list_cols.itemChanged.connect(lambda _it: schedule_apply())
        list_cols.model().rowsMoved.connect(lambda *_args: schedule_apply())
        close_btn.clicked.connect(dlg.close)
        dlg.exec()

    def open_filters_dialog(self, key_override=None):
        key = key_override or self._current_key()
        if key not in ("orgs", "tenders", "archived"):
            QMessageBox.information(self, "Filters", "Filters are available for table tabs.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Filters")
        dlg.resize(760, 620)
        root = QVBoxLayout(dlg)
        cols = self._columns_for_key(key)
        active_filters = dict(self.filter_map.get(key, {}) or {})

        top = QHBoxLayout()
        top.addWidget(QLabel("Column:"))
        col_combo = QComboBox()
        col_combo.addItems(cols)
        top.addWidget(col_combo, 1)
        top.addWidget(QLabel("Filter Type:"))
        mode_combo = QComboBox()
        mode_combo.addItems(["Values", "Text Contains", "Text Equals", "Number Comparison"])
        top.addWidget(mode_combo, 1)
        root.addLayout(top)

        body = QHBoxLayout()
        left = QVBoxLayout()
        right = QVBoxLayout()
        body.addLayout(left, 2)
        body.addLayout(right, 1)
        root.addLayout(body, 1)

        values_list = QListWidget()
        values_list.setSelectionMode(QAbstractItemView.NoSelection)
        left.addWidget(values_list, 1)
        values_btns = QHBoxLayout()
        btn_sel_all = QPushButton("Select All")
        btn_sel_none = QPushButton("Clear")
        values_btns.addWidget(btn_sel_all)
        values_btns.addWidget(btn_sel_none)
        left.addLayout(values_btns)

        op_row = QHBoxLayout()
        op_row.addWidget(QLabel("Operator:"))
        op_combo = QComboBox()
        op_combo.addItems(["=", "!=", ">", ">=", "<", "<="])
        op_row.addWidget(op_combo)
        value_edit = QLineEdit()
        value_edit.setPlaceholderText("value")
        op_row.addWidget(value_edit, 1)
        left.addLayout(op_row)
        helper = QLabel("")
        helper.setObjectName("SoftText")
        left.addWidget(helper)

        right.addWidget(QLabel("Current Filters"))
        list_filters = QListWidget()
        right.addWidget(list_filters, 1)

        def get_unique_values(col_name):
            table = self._table_for_key(key)
            if table is None or col_name not in cols:
                return []
            idx = cols.index(col_name)
            vals = set()
            for r in range(table.rowCount()):
                it = table.item(r, idx)
                vals.add("" if it is None else str(it.text()))
            return sorted(vals, key=lambda x: x.lower())

        def refresh_current_filters():
            list_filters.clear()
            for c, f in active_filters.items():
                if isinstance(f, str):
                    list_filters.addItem(f"{c}: contains '{f}'")
                    continue
                m = f.get("mode", "")
                if m == "values":
                    sel = f.get("selected", [])
                    list_filters.addItem(f"{c}: {len(sel)} value(s)")
                elif m == "contains":
                    list_filters.addItem(f"{c}: contains '{f.get('value','')}'")
                elif m == "equals":
                    list_filters.addItem(f"{c}: equals '{f.get('value','')}'")
                elif m == "number":
                    list_filters.addItem(f"{c}: {f.get('op','=')} {f.get('value','')}")

        def refresh_mode_ui():
            mode = mode_combo.currentText()
            col = col_combo.currentText()
            is_values = (mode == "Values")
            is_num = (mode == "Number Comparison")

            values_list.setVisible(is_values)
            btn_sel_all.setVisible(is_values)
            btn_sel_none.setVisible(is_values)
            op_combo.setVisible(is_num)

            values_list.clear()
            value_edit.clear()
            op_combo.setCurrentText("=")
            if is_values:
                existing = active_filters.get(col, {})
                selected_existing = set(existing.get("selected", [])) if isinstance(existing, dict) and existing.get("mode") == "values" else set()
                for v in get_unique_values(col):
                    it = QListWidgetItem(v)
                    it.setFlags(it.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                    it.setCheckState(Qt.Checked if v in selected_existing else Qt.Unchecked)
                    values_list.addItem(it)
                helper.setText("Select one or more values using checkboxes.")
            elif mode in ("Text Contains", "Text Equals"):
                existing = active_filters.get(col, {})
                if isinstance(existing, dict) and existing.get("mode") in ("contains", "equals"):
                    value_edit.setText(str(existing.get("value", "")))
                helper.setText("Enter text value.")
            else:
                existing = active_filters.get(col, {})
                if isinstance(existing, dict) and existing.get("mode") == "number":
                    op_combo.setCurrentText(str(existing.get("op", "=")))
                    value_edit.setText(str(existing.get("value", "")))
                helper.setText("Numeric comparison.")

        def set_all_checks(state):
            for i in range(values_list.count()):
                it = values_list.item(i)
                it.setCheckState(Qt.Checked if state else Qt.Unchecked)

        def apply_one():
            col = col_combo.currentText()
            mode = mode_combo.currentText()
            if not col:
                return
            if mode == "Values":
                sel = []
                for i in range(values_list.count()):
                    it = values_list.item(i)
                    if it.checkState() == Qt.Checked:
                        sel.append(it.text())
                if not sel:
                    active_filters.pop(col, None)
                else:
                    active_filters[col] = {"mode": "values", "selected": sel}
            elif mode == "Text Contains":
                txt = value_edit.text().strip()
                if not txt:
                    active_filters.pop(col, None)
                else:
                    active_filters[col] = {"mode": "contains", "value": txt}
            elif mode == "Text Equals":
                txt = value_edit.text().strip()
                if not txt:
                    active_filters.pop(col, None)
                else:
                    active_filters[col] = {"mode": "equals", "value": txt}
            else:
                txt = value_edit.text().strip()
                if not txt:
                    active_filters.pop(col, None)
                else:
                    active_filters[col] = {"mode": "number", "op": op_combo.currentText(), "value": txt}
            self.filter_map[key] = active_filters
            self._persist_filters(key)
            refresh_current_filters()
            self.refresh_current_table_view()

        def clear_one():
            col = col_combo.currentText()
            if col in active_filters:
                active_filters.pop(col, None)
                self.filter_map[key] = active_filters
                self._persist_filters(key)
                refresh_current_filters()
                self.refresh_current_table_view()
            refresh_mode_ui()

        def clear_all():
            active_filters.clear()
            self.filter_map[key] = {}
            self._persist_filters(key)
            refresh_current_filters()
            self.refresh_current_table_view()
            refresh_mode_ui()

        btns = QHBoxLayout()
        btn_apply = QPushButton("Apply Column Filter")
        btn_apply.setObjectName("PrimaryButton")
        btn_clear_one = QPushButton("Clear Column Filter")
        btn_clear_all = QPushButton("Clear All Filters")
        btn_close = QPushButton("Close")
        btns.addWidget(btn_apply)
        btns.addWidget(btn_clear_one)
        btns.addWidget(btn_clear_all)
        btns.addStretch(1)
        btns.addWidget(btn_close)
        root.addLayout(btns)

        col_combo.currentTextChanged.connect(lambda _x: refresh_mode_ui())
        mode_combo.currentTextChanged.connect(lambda _x: refresh_mode_ui())
        btn_sel_all.clicked.connect(lambda: set_all_checks(True))
        btn_sel_none.clicked.connect(lambda: set_all_checks(False))
        btn_apply.clicked.connect(apply_one)
        btn_clear_one.clicked.connect(clear_one)
        btn_clear_all.clicked.connect(clear_all)
        btn_close.clicked.connect(dlg.close)

        refresh_current_filters()
        refresh_mode_ui()
        dlg.exec()

    def _normalize_export_setting(self, value, fallback):
        if isinstance(value, str):
            try:
                value = json.loads(value)
            except Exception:
                return fallback
        return value if value is not None else fallback

    def _get_active_tender_export_rows(self, selected_cols):
        if not selected_cols:
            return []
        idx_map = {c: i for i, c in enumerate(self.tender_cols)}
        rows = []
        for r in range(self.table_tenders.rowCount()):
            row_out = []
            for col in selected_cols:
                idx = idx_map.get(col, -1)
                txt = ""
                if 0 <= idx < self.table_tenders.columnCount():
                    item = self.table_tenders.item(r, idx)
                    txt = item.text() if item else ""
                row_out.append(" ".join(str(txt).replace("\n", " ").split()))
            rows.append(row_out)
        return rows

    def _draw_pdf_table(self, out_path, cols, rows, width_map):
        if not getattr(core, "PDF_SUPPORT", False):
            raise RuntimeError("PDF export requires PyMuPDF (fitz).")
        fitz = core.fitz
        doc = fitz.open()
        margin = 28
        font_size = 8
        line_h = 12
        page_rect = fitz.paper_rect("a4-l")
        page = doc.new_page(width=page_rect.width, height=page_rect.height)
        table_w = page.rect.width - (2 * margin)
        header_bg = (0.90, 0.93, 0.97)

        weights = []
        for c in cols:
            try:
                w = int(width_map.get(c, 126))
            except Exception:
                w = 126
            weights.append(max(48, min(720, w)))
        total_w = float(sum(weights) or 1.0)
        col_w = [(w / total_w) * table_w for w in weights]
        x_edges = [margin]
        for w in col_w:
            x_edges.append(x_edges[-1] + w)

        def draw_header(y):
            hdr_h = line_h + 6
            for i, c in enumerate(cols):
                rect = fitz.Rect(x_edges[i], y, x_edges[i + 1], y + hdr_h)
                page.draw_rect(rect, color=(0, 0, 0), fill=header_bg, width=0.6)
                page.insert_textbox(rect, str(c), fontsize=font_size, align=fitz.TEXT_ALIGN_CENTER)
            return y + hdr_h

        def col_char_cap(i):
            return max(5, int(col_w[i] / 5.2))

        y = draw_header(margin)
        bottom = page.rect.height - margin
        for row in rows:
            wrapped_cells = []
            max_lines = 1
            for i, cell in enumerate(row):
                txt = str(cell or "")
                lines = core.textwrap.wrap(txt, width=col_char_cap(i)) or [""]
                wrapped_cells.append("\n".join(lines))
                max_lines = max(max_lines, len(lines))

            row_h = (max_lines * line_h) + 8
            if y + row_h > bottom:
                page = doc.new_page(width=page_rect.width, height=page_rect.height)
                y = draw_header(margin)

            for i, txt in enumerate(wrapped_cells):
                rect = fitz.Rect(x_edges[i], y, x_edges[i + 1], y + row_h)
                page.draw_rect(rect, color=(0, 0, 0), width=0.5)
                page.insert_textbox(rect, txt, fontsize=font_size, align=fitz.TEXT_ALIGN_LEFT)
            y += row_h

        doc.save(out_path)
        doc.close()

    def _export_active_tenders(self, export_kind, selected_cols, preview_widths):
        if not selected_cols:
            QMessageBox.warning(self, "Export", "Select at least one column.")
            return
        rows = self._get_active_tender_export_rows(selected_cols)
        if not rows:
            QMessageBox.warning(self, "Export", "No rows to export in Active Tenders.")
            return

        now_tag = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_kind == "excel":
            path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Active Tenders to Excel",
                f"Active_Tenders_{now_tag}.xlsx",
                "Excel Workbook (*.xlsx)",
            )
            if not path:
                return
            try:
                pd = importlib.import_module("pandas")
            except Exception:
                QMessageBox.critical(self, "Export Excel", "Pandas is not available. Install pandas/openpyxl.")
                return

            df = pd.DataFrame(rows, columns=selected_cols)
            try:
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Active Tenders")
                    try:
                        from openpyxl.utils import get_column_letter
                        from openpyxl.styles import Alignment
                        ws = writer.sheets.get("Active Tenders")
                        if ws is not None:
                            excel_width_map = {}
                            for i, c in enumerate(selected_cols, 1):
                                try:
                                    px = int(preview_widths.get(c, 126))
                                    w = max(8, min(80, int(px / 7)))
                                except Exception:
                                    w = 18
                                w = max(8, min(80, w))
                                excel_width_map[i] = w
                                ws.column_dimensions[get_column_letter(i)].width = w
                            for row_idx in range(2, ws.max_row + 1):
                                max_lines = 1
                                for col_idx in range(1, len(selected_cols) + 1):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    txt = "" if cell.value is None else str(cell.value)
                                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                                    col_w = float(excel_width_map.get(col_idx, 18))
                                    char_cap = max(6, int(col_w - 1))
                                    wrapped = core.textwrap.wrap(txt, width=char_cap, break_long_words=True, replace_whitespace=False) or [""]
                                    max_lines = max(max_lines, len(wrapped))
                                ws.row_dimensions[row_idx].height = max(15, (max_lines * 15))
                    except Exception:
                        pass
            except Exception as e:
                QMessageBox.critical(self, "Export Excel", f"Failed to export:\n{e}")
                return
            QMessageBox.information(self, "Export Excel", f"Exported successfully:\n{path}")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Active Tenders to PDF",
            f"Active_Tenders_{now_tag}.pdf",
            "PDF Document (*.pdf)",
        )
        if not path:
            return
        try:
            self._draw_pdf_table(path, selected_cols, rows, preview_widths)
        except Exception as e:
            QMessageBox.critical(self, "Export PDF", f"Failed to export:\n{e}")
            return
        QMessageBox.information(self, "Export PDF", f"Exported successfully:\n{path}")

    def open_active_tender_export_dialog(self):
        if self.tabs.currentIndex() != 1:
            QMessageBox.information(self, "Export", "Export is available in Active Tenders tab.")
            return

        key_cols = "viewtenders_export_cols"
        key_order = "viewtenders_export_order"
        saved_cols = self._normalize_export_setting(core.get_user_setting(key_cols, None), None)
        saved_order = self._normalize_export_setting(core.get_user_setting(key_order, None), None)

        if not isinstance(saved_cols, list):
            saved_cols = [c for c in self.tender_cols if c != "ID"]
        saved_cols = [c for c in saved_cols if c in self.tender_cols and c != "ID"]
        if not saved_cols:
            saved_cols = [c for c in self.tender_cols if c != "ID"]

        all_export_cols = [c for c in self.tender_cols if c != "ID"]
        ordered_cols = [c for c in saved_order if c in all_export_cols] if isinstance(saved_order, list) else []
        for c in all_export_cols:
            if c not in ordered_cols:
                ordered_cols.append(c)
        checked_cols = {c for c in saved_cols if c in all_export_cols}
        if not checked_cols:
            checked_cols = set(all_export_cols)

        dlg = QDialog(self)
        dlg.setWindowTitle("Export Active Tenders")
        dlg.resize(1180, 700)
        root = QHBoxLayout(dlg)

        left_wrap = QVBoxLayout()
        right_wrap = QVBoxLayout()
        root.addLayout(left_wrap, 0)
        root.addLayout(right_wrap, 1)

        left_wrap.addWidget(QLabel("Columns (check + drag to reorder)"))
        cols_list = QListWidget()
        cols_list.setDragDropMode(QAbstractItemView.InternalMove)
        cols_list.setDefaultDropAction(Qt.MoveAction)
        left_wrap.addWidget(cols_list, 1)

        preview = QTableWidget(0, 0)
        preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        preview.verticalHeader().setVisible(False)
        preview.setAlternatingRowColors(True)
        preview.setWordWrap(True)
        preview.horizontalHeader().setStretchLastSection(True)
        right_wrap.addWidget(QLabel("Preview (first 80 rows)"))
        right_wrap.addWidget(preview, 1)

        btns = QHBoxLayout()
        btn_excel = QPushButton("Export Excel")
        btn_excel.setObjectName("PrimaryButton")
        btn_pdf = QPushButton("Export PDF")
        btn_close = QPushButton("Close")
        btns.addWidget(btn_excel)
        btns.addWidget(btn_pdf)
        btns.addStretch(1)
        btns.addWidget(btn_close)
        right_wrap.addLayout(btns)

        def selected_cols():
            out = []
            for i in range(cols_list.count()):
                it = cols_list.item(i)
                if it.checkState() == Qt.Checked:
                    out.append(it.text())
            return out

        def save_export_prefs(cols_now):
            core.set_user_setting(key_cols, cols_now)
            core.set_user_setting(key_order, [cols_list.item(i).text() for i in range(cols_list.count())])

        def current_preview_widths(cols_now):
            widths = {}
            for c in cols_now:
                try:
                    idx = cols_now.index(c)
                    widths[c] = int(preview.columnWidth(idx))
                except Exception:
                    widths[c] = 126
            return widths

        def refresh_preview():
            cols_now = selected_cols()
            preview.setRowCount(0)
            preview.setColumnCount(len(cols_now))
            preview.setHorizontalHeaderLabels(cols_now)
            base_index = {c: i for i, c in enumerate(self.tender_cols)}
            for cidx, c in enumerate(cols_now):
                src_idx = base_index.get(c, -1)
                w = self.table_tenders.columnWidth(src_idx) if src_idx >= 0 else 126
                preview.setColumnWidth(cidx, max(50, w))
            rows = self._get_active_tender_export_rows(cols_now)[:80]
            for rv in rows:
                row = preview.rowCount()
                preview.insertRow(row)
                for cidx, val in enumerate(rv):
                    preview.setItem(row, cidx, QTableWidgetItem(str(val)))
            auto_fit_table_rows(preview, min_height=24, max_height=None)
            save_export_prefs(cols_now)

        def populate_columns():
            cols_list.clear()
            for c in ordered_cols:
                it = QListWidgetItem(c)
                it.setFlags(it.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsDragEnabled | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                it.setCheckState(Qt.Checked if c in checked_cols else Qt.Unchecked)
                cols_list.addItem(it)

        def run_export(kind):
            cols_now = selected_cols()
            save_export_prefs(cols_now)
            self._export_active_tenders(kind, cols_now, current_preview_widths(cols_now))

        cols_list.itemChanged.connect(lambda _it: refresh_preview())
        cols_list.model().rowsMoved.connect(lambda *_args: refresh_preview())
        btn_excel.clicked.connect(lambda: run_export("excel"))
        btn_pdf.clicked.connect(lambda: run_export("pdf"))
        btn_close.clicked.connect(dlg.close)

        populate_columns()
        refresh_preview()
        dlg.exec()

class AppSettingsPage(QWidget):
    def __init__(self, controller):
        super().__init__()
        self.controller = controller

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(12)

        title = QLabel("Settings")
        title.setObjectName("PageTitle")
        root.addWidget(title)

        form = QGridLayout()
        form.setColumnStretch(1, 1)
        root.addLayout(form)

        self.db_dir_edit = QLineEdit()
        self.projects_dir_edit = QLineEdit()
        self.download_dir_edit = QLineEdit()
        self.update_dir_edit = QLineEdit()
        self.update_manifest_url_edit = QLineEdit()
        self.backend_mode_combo = QComboBox()
        self.backend_mode_combo.addItem("Local", "local")
        self.backend_mode_combo.addItem("Remote (Railway/API)", "remote")
        if FRONTEND_REMOTE_ONLY:
            self.backend_mode_combo.setCurrentIndex(1)
            self.backend_mode_combo.setEnabled(False)
        self.backend_url_edit = QLineEdit()
        self.backend_api_key_edit = QLineEdit()
        self.backend_api_key_edit.setEchoMode(QLineEdit.Password)
        self.test_backend_btn = QPushButton("Test Backend")
        self.backend_test_status = QLabel("")
        self.backend_test_status.setObjectName("SoftText")
        self.backend_test_status.setWordWrap(True)
        self.backend_test_status.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        self.show_tender_info_chk = QCheckBox("Show tender info box in Project Details")
        self.scraper_auto_fetch_chk = QCheckBox("Enable auto fetch for selected organizations and tenders")
        self.scraper_auto_fetch_interval_edit = QLineEdit()
        self.scraper_auto_fetch_interval_edit.setPlaceholderText("Minutes")
        self.scraper_last_auto_fetch_label = QLabel("")
        self.scraper_last_auto_fetch_label.setObjectName("SoftText")
        self.scraper_run_auto_fetch_btn = QPushButton("Run Auto Fetch Now")

        btn_db = QPushButton("Browse")
        btn_projects = QPushButton("Browse")
        btn_download = QPushButton("Browse")
        btn_update_dir = QPushButton("Browse")

        form.addWidget(QLabel("Tender Database Folder:"), 0, 0)
        form.addWidget(self.db_dir_edit, 0, 1)
        form.addWidget(btn_db, 0, 2)

        form.addWidget(QLabel("My Tender Projects Folder:"), 1, 0)
        form.addWidget(self.projects_dir_edit, 1, 1)
        form.addWidget(btn_projects, 1, 2)

        form.addWidget(QLabel("Tender Downloads Folder:"), 2, 0)
        form.addWidget(self.download_dir_edit, 2, 1)
        form.addWidget(btn_download, 2, 2)

        form.addWidget(QLabel("Update Folder:"), 3, 0)
        form.addWidget(self.update_dir_edit, 3, 1)
        form.addWidget(btn_update_dir, 3, 2)

        form.addWidget(QLabel("Update Manifest URL:"), 4, 0)
        form.addWidget(self.update_manifest_url_edit, 4, 1, 1, 2)

        form.addWidget(QLabel("Backend Mode:"), 5, 0)
        form.addWidget(self.backend_mode_combo, 5, 1, 1, 2)

        form.addWidget(QLabel("Backend URL:"), 6, 0)
        form.addWidget(self.backend_url_edit, 6, 1, 1, 2)

        form.addWidget(QLabel("Backend API Key:"), 7, 0)
        form.addWidget(self.backend_api_key_edit, 7, 1)
        form.addWidget(self.test_backend_btn, 7, 2)

        form.addWidget(self.backend_test_status, 8, 1, 1, 2)
        form.addWidget(self.show_tender_info_chk, 9, 1, 1, 2)
        form.addWidget(QLabel("Online Tender Scraping:"), 10, 0)
        form.addWidget(self.scraper_auto_fetch_chk, 10, 1, 1, 2)
        form.addWidget(QLabel("Auto Fetch Interval (minutes):"), 11, 0)
        form.addWidget(self.scraper_auto_fetch_interval_edit, 11, 1, 1, 2)
        form.addWidget(QLabel("Last Auto Fetch:"), 12, 0)
        form.addWidget(self.scraper_last_auto_fetch_label, 12, 1)
        form.addWidget(self.scraper_run_auto_fetch_btn, 12, 2)

        hint = QLabel(
            "Database path uses 'tender_manager.db' inside the selected DB folder. "
            f"Current build: {core.APP_VERSION}"
        )
        hint.setObjectName("SoftText")
        root.addWidget(hint)

        updates_row = QHBoxLayout()
        self.check_update_btn = QPushButton("Check for Upgrade")
        self.install_update_btn = QPushButton("Install Upgrade")
        self.install_update_btn.setObjectName("PrimaryButton")
        self.install_update_btn.setEnabled(False)
        self.update_status = QLabel("")
        self.update_status.setObjectName("SoftText")
        updates_row.addWidget(self.check_update_btn)
        updates_row.addWidget(self.install_update_btn)
        updates_row.addWidget(self.update_status, 1)
        root.addLayout(updates_row)

        view_row = QHBoxLayout()
        self.projects_view_label = QLabel("")
        self.projects_view_label.setObjectName("SoftText")
        self.projects_view_toggle_btn = QPushButton("Switch View")
        view_row.addWidget(QLabel("Projects Create View:"))
        view_row.addWidget(self.projects_view_label, 1)
        view_row.addWidget(self.projects_view_toggle_btn)
        root.addLayout(view_row)

        root.addStretch(1)

        btns = QHBoxLayout()
        self.save_btn = QPushButton("Save")
        self.save_btn.setObjectName("PrimaryButton")
        reload_btn = QPushButton("Reload")
        btns.addWidget(self.save_btn)
        btns.addWidget(reload_btn)
        btns.addStretch(1)
        root.addLayout(btns)

        btn_db.clicked.connect(lambda: self._pick_dir_into(self.db_dir_edit))
        btn_projects.clicked.connect(lambda: self._pick_dir_into(self.projects_dir_edit))
        btn_download.clicked.connect(lambda: self._pick_dir_into(self.download_dir_edit))
        btn_update_dir.clicked.connect(lambda: self._pick_dir_into(self.update_dir_edit))
        self.save_btn.clicked.connect(self.save_settings)
        reload_btn.clicked.connect(self.reload_settings)
        self.check_update_btn.clicked.connect(self.check_for_upgrade)
        self.install_update_btn.clicked.connect(self.install_upgrade)
        self.projects_view_toggle_btn.clicked.connect(self.toggle_projects_create_view)
        self.test_backend_btn.clicked.connect(self.test_backend_connection)
        self.show_tender_info_chk.toggled.connect(self._on_show_tender_info_toggled)
        self.backend_mode_combo.currentIndexChanged.connect(self._apply_backend_mode_ui)
        self.scraper_run_auto_fetch_btn.clicked.connect(self._run_auto_fetch_now)

        self._pending_update_exe = ""
        self._pending_update_version = ""
        self._pending_update_exe_url = ""
        self._pending_update_installer_url = ""

        self.reload_settings()

    def _pick_dir_into(self, target_edit):
        start = str(target_edit.text() or "").strip() or core._safe_getcwd()
        picked = QFileDialog.getExistingDirectory(self, "Select Folder", start)
        if picked:
            target_edit.setText(picked)

    def reload_settings(self):
        self.db_dir_edit.setText(os.path.dirname(core._resolve_path(core.DB_FILE)) or core._safe_getcwd())
        self.projects_dir_edit.setText(core._resolve_path(core.ROOT_FOLDER))
        self.download_dir_edit.setText(core._resolve_path(core.BASE_DOWNLOAD_DIRECTORY))
        self.update_dir_edit.setText(str(core.get_user_setting("update_directory", "") or "").strip())
        self.update_manifest_url_edit.setText(str(core.get_user_setting("update_manifest_url", "") or "").strip())
        backend_mode = str(core.get_user_setting("backend_mode", "local") or "local").strip().lower()
        if FRONTEND_REMOTE_ONLY:
            backend_mode = "remote"
        idx = self.backend_mode_combo.findData("remote" if backend_mode == "remote" else "local")
        if idx >= 0:
            self.backend_mode_combo.setCurrentIndex(idx)
        self.backend_url_edit.setText(str(core.get_user_setting("backend_url", "") or "").strip())
        self.backend_api_key_edit.setText(str(core.get_user_setting("backend_api_key", "") or "").strip())
        self._set_backend_test_status("")
        self._apply_backend_mode_ui()
        self.install_update_btn.setEnabled(False)
        self._pending_update_exe = ""
        self._pending_update_version = ""
        self._pending_update_exe_url = ""
        self._pending_update_installer_url = ""
        self.update_status.setText("")
        mode = str(core.get_user_setting("projects_entry_mode", "inline") or "inline").strip().lower()
        if mode not in ("inline", "popup"):
            mode = "inline"
        self.projects_view_label.setText("Inline Form" if mode == "inline" else "Minimal Popup")
        self.projects_view_toggle_btn.setText("Switch to Popup" if mode == "inline" else "Switch to Inline")
        self.show_tender_info_chk.setChecked(bool(core.get_user_setting("project_details_show_tender_info", True)))
        self.scraper_auto_fetch_chk.setChecked(bool(core.get_user_setting("scraper_auto_fetch_enabled", False)))
        self.scraper_auto_fetch_interval_edit.setText(str(core.get_user_setting("scraper_auto_fetch_interval_minutes", 30) or 30))
        self.scraper_last_auto_fetch_label.setText(self._format_last_auto_fetch_text())

    def save_settings(self):
        db_dir = str(self.db_dir_edit.text() or "").strip()
        proj_dir = str(self.projects_dir_edit.text() or "").strip()
        down_dir = str(self.download_dir_edit.text() or "").strip()
        update_dir = str(self.update_dir_edit.text() or "").strip()
        update_manifest_url = str(self.update_manifest_url_edit.text() or "").strip()
        backend_mode = str(self.backend_mode_combo.currentData() or "local").strip().lower()
        if FRONTEND_REMOTE_ONLY:
            backend_mode = "remote"
        backend_url = str(self.backend_url_edit.text() or "").strip().rstrip("/")
        backend_api_key = str(self.backend_api_key_edit.text() or "").strip()
        show_tender_info = bool(self.show_tender_info_chk.isChecked())
        auto_fetch_enabled = bool(self.scraper_auto_fetch_chk.isChecked())
        try:
            auto_fetch_minutes = int(str(self.scraper_auto_fetch_interval_edit.text() or "").strip() or "30")
        except Exception:
            auto_fetch_minutes = 30
        auto_fetch_minutes = max(1, min(24 * 60, auto_fetch_minutes))
        if not db_dir or not proj_dir or not down_dir:
            QMessageBox.critical(self, "Settings", "All three paths are required.")
            return

        new_db_file = core._normalize_db_file(db_dir)
        new_root = core._resolve_path(proj_dir)
        new_down = core._resolve_path(down_dir)
        try:
            os.makedirs(os.path.dirname(new_db_file) or ".", exist_ok=True)
            os.makedirs(new_root, exist_ok=True)
            os.makedirs(new_down, exist_ok=True)
            core.save_app_paths_config(new_db_file, new_root, new_down, core.TEMPLATE_LIBRARY_FOLDER)

            old_db = core.DB_FILE
            core.DB_FILE = new_db_file
            core.ROOT_FOLDER = new_root
            core.BASE_DOWNLOAD_DIRECTORY = new_down
            core.set_user_setting("update_directory", update_dir)
            core.set_user_setting("update_manifest_url", update_manifest_url)
            core.set_user_setting("backend_mode", backend_mode)
            core.set_user_setting("backend_url", backend_url)
            core.set_user_setting("backend_api_key", backend_api_key)
            core.set_user_setting("project_details_show_tender_info", show_tender_info)
            core.set_user_setting("scraper_auto_fetch_enabled", auto_fetch_enabled)
            core.set_user_setting("scraper_auto_fetch_interval_minutes", auto_fetch_minutes)
            if core._resolve_path(old_db) != core._resolve_path(core.DB_FILE):
                core.init_db()
            if hasattr(self.controller, "projects_page") and self.controller.projects_page is not None:
                self.controller.projects_page.reload_entry_mode()
            if hasattr(self.controller, "project_details_page") and self.controller.project_details_page is not None:
                self.controller.project_details_page.apply_tender_info_visibility_setting()
            if hasattr(self.controller, "online_page") and self.controller.online_page is not None:
                self.controller.online_page.refresh_auto_fetch_settings()
            QMessageBox.information(self, "Settings", "Paths updated successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Settings", f"Failed to save settings:\n{e}")

    def _on_show_tender_info_toggled(self, checked):
        try:
            core.set_user_setting("project_details_show_tender_info", bool(checked))
            if hasattr(self.controller, "project_details_page") and self.controller.project_details_page is not None:
                self.controller.project_details_page.apply_tender_info_visibility_setting()
        except Exception:
            pass

    def _format_last_auto_fetch_text(self):
        raw = str(core.get_user_setting("scraper_last_auto_fetch_at", "") or "").strip()
        if not raw:
            return "Never"
        try:
            dt = datetime.datetime.fromisoformat(raw.replace("Z", "+00:00"))
            return dt.astimezone().strftime("%d-%b-%Y %I:%M:%S %p")
        except Exception:
            return raw

    def _run_auto_fetch_now(self):
        try:
            page = self.controller._ensure_online_page()
            page.run_auto_fetch_now()
            self.scraper_last_auto_fetch_label.setText(self._format_last_auto_fetch_text())
        except Exception as e:
            QMessageBox.critical(self, "Auto Fetch", f"Failed to start auto fetch:\n{e}")

    def _set_backend_test_status(self, message):
        txt = str(message or "").strip()
        if not txt:
            self.backend_test_status.clear()
            return
        wrapped = "<div style='white-space:normal; overflow-wrap:anywhere; word-break:break-word;'>%s</div>" % (
            html.escape("\n".join(textwrap.wrap(txt, width=90, break_long_words=False, replace_whitespace=False)) or txt)
        )
        self.backend_test_status.setText(wrapped)

    def _apply_backend_mode_ui(self):
        mode = str(self.backend_mode_combo.currentData() or "local").strip().lower()
        if FRONTEND_REMOTE_ONLY:
            mode = "remote"
        is_remote = mode == "remote"
        self.backend_url_edit.setEnabled(is_remote)
        self.backend_api_key_edit.setEnabled(is_remote)
        self.test_backend_btn.setEnabled(is_remote)
        if not is_remote:
            self._set_backend_test_status("Local backend is enabled. Remote connectivity test is not required.")
        elif "Local backend is enabled." in self.backend_test_status.text():
            self._set_backend_test_status("")

    def toggle_projects_create_view(self):
        mode = str(core.get_user_setting("projects_entry_mode", "inline") or "inline").strip().lower()
        if mode not in ("inline", "popup"):
            mode = "inline"
        new_mode = "popup" if mode == "inline" else "inline"
        core.set_user_setting("projects_entry_mode", new_mode)
        if hasattr(self.controller, "projects_page") and self.controller.projects_page is not None:
            self.controller.projects_page.reload_entry_mode()
        self.reload_settings()

    def test_backend_connection(self):
        mode = str(self.backend_mode_combo.currentData() or "local").strip().lower()
        if FRONTEND_REMOTE_ONLY:
            mode = "remote"
        if mode != "remote":
            self._set_backend_test_status("Local backend is enabled. Remote connectivity test is not required.")
            return
        url = str(self.backend_url_edit.text() or "").strip().rstrip("/")
        api_key = str(self.backend_api_key_edit.text() or "").strip()
        if not url or not api_key:
            self._set_backend_test_status("Backend URL and API key are required.")
            return
        try:
            client = BidApiClient(base_url=url, api_key=api_key, timeout_seconds=25)
            health = client.health()
            if str(health.get("status") or "").lower() == "ok":
                self._set_backend_test_status("Connected successfully.")
            else:
                self._set_backend_test_status("Connected, but health response was unexpected.")
        except Exception as e:
            self._set_backend_test_status(f"Connection failed: {e}")

    def check_for_upgrade(self):
        update_dir = str(self.update_dir_edit.text() or "").strip()
        update_manifest_url = str(self.update_manifest_url_edit.text() or "").strip()
        info = core.get_update_info(update_dir=update_dir, manifest_url=update_manifest_url)
        self.install_update_btn.setEnabled(False)
        self._pending_update_exe = ""
        self._pending_update_version = ""
        self._pending_update_exe_url = ""
        self._pending_update_installer_url = ""
        if not info.get("ok"):
            self.update_status.setText(str(info.get("message", "No update information found.")))
            return
        current = str(info.get("current_version", core.APP_VERSION))
        available = str(info.get("available_version", ""))
        if info.get("newer"):
            self._pending_update_exe = str(info.get("exe_path", ""))
            self._pending_update_exe_url = str(info.get("exe_url", ""))
            self._pending_update_installer_url = str(info.get("installer_url", ""))
            self._pending_update_version = available
            self.install_update_btn.setEnabled(True)
            src = str(info.get("source", "local"))
            self.update_status.setText(f"Upgrade available: {available} (current: {current}) [{src}]")
        else:
            self.update_status.setText(f"No upgrade found. Current build: {current}")

    def install_upgrade(self):
        if not self._pending_update_exe and not self._pending_update_exe_url and not self._pending_update_installer_url:
            QMessageBox.information(self, "Upgrade", "Please run 'Check for Upgrade' first.")
            return
        ans = QMessageBox.question(
            self,
            "Install Upgrade",
            f"Install build {self._pending_update_version} and restart now?",
        )
        if ans != QMessageBox.Yes:
            return
        ok = False
        msg = "Upgrade failed."
        if self._pending_update_installer_url:
            dl = core.download_remote_update_binary(self._pending_update_installer_url, suffix=".exe")
            if not dl.get("ok"):
                QMessageBox.critical(self, "Upgrade", str(dl.get("message", "Failed to download installer update.")))
                return
            ok, msg = core.launch_installer_update(str(dl.get("path", "")))
        elif self._pending_update_exe_url:
            dl = core.download_remote_update_binary(self._pending_update_exe_url, suffix=".exe")
            if not dl.get("ok"):
                QMessageBox.critical(self, "Upgrade", str(dl.get("message", "Failed to download update executable.")))
                return
            ok, msg = core.launch_self_update(str(dl.get("path", "")))
        else:
            ok, msg = core.launch_self_update(self._pending_update_exe)
        if not ok:
            QMessageBox.critical(self, "Upgrade", msg)
            return
        QMessageBox.information(self, "Upgrade", "Upgrade started. The app will close and relaunch.")
        QApplication.instance().quit()


class PlaceholderPage(QWidget):
    def __init__(self, title, text, button_text=None, on_click=None):
        super().__init__()
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)
        label = QLabel(title)
        label.setObjectName("PageTitle")
        body = QLabel(text)
        body.setWordWrap(True)
        body.setObjectName("SoftText")
        root.addWidget(label)
        root.addWidget(body)
        root.addStretch(1)
        if button_text and on_click:
            btn = QPushButton(button_text)
            btn.setObjectName("PrimaryButton")
            btn.clicked.connect(on_click)
            root.addWidget(btn, 0, Qt.AlignLeft)


class BidManagerQt(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"Tender & Bid Manager Pro - PySide6 - v{core.APP_VERSION}")
        self.resize(1360, 880)
        self.archive_job_running = False
        self._pending_online_refresh = False
        self.scraper_backend = BackendModeScraperProxy()

        wrapper = QWidget()
        self.setCentralWidget(wrapper)
        root = QHBoxLayout(wrapper)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        splitter = QSplitter(Qt.Horizontal)
        root.addWidget(splitter)

        self.sidebar = QFrame()
        self.sidebar.setObjectName("Sidebar")
        self.sidebar.setMinimumWidth(230)
        self.sidebar.setMaximumWidth(260)
        sbl = QVBoxLayout(self.sidebar)
        sbl.setContentsMargins(14, 18, 14, 18)
        sbl.setSpacing(8)

        logo = QLabel("BID MANAGER")
        logo.setObjectName("Brand")
        sbl.addWidget(logo)

        self.btn_projects = QPushButton("Projects")
        self.btn_online = QPushButton("Online Tenders")
        self.btn_templates = QPushButton("Templates")
        self.btn_settings = QPushButton("Settings")
        self.btn_settings.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        for btn in (self.btn_projects, self.btn_online, self.btn_templates):
            btn.setObjectName("NavButton")
            btn.setCursor(Qt.PointingHandCursor)
            sbl.addWidget(btn)

        sbl.addStretch(1)
        self.btn_settings.setObjectName("NavButton")
        self.btn_settings.setCursor(Qt.PointingHandCursor)
        sbl.addWidget(self.btn_settings)

        self.content = QWidget()
        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setContentsMargins(0, 0, 0, 0)

        self.projects_page = ProjectsPage(self)
        self.project_details_page = None
        self.online_page = None
        self.templates_page = None
        self.settings_page = None
        self._projects_last_subview = "list"
        self._projects_last_pid = None

        splitter.addWidget(self.sidebar)
        splitter.addWidget(self.content)
        splitter.setSizes([240, 1120])
        splitter.setCollapsible(0, False)

        self.btn_projects.clicked.connect(self.open_projects_section)
        self.btn_online.clicked.connect(self.show_online_section)
        self.btn_templates.clicked.connect(self.show_templates_section)
        self.btn_settings.clicked.connect(self.show_settings_section)

        last_main = str(core.get_user_setting("main_last_view", "projects") or "projects").strip().lower()
        if last_main == "online":
            self.show_online_section()
        elif last_main == "templates":
            self.show_templates_section()
        elif last_main == "settings":
            self.show_settings_section()
        else:
            self.open_projects_section()
        self._build_menu()
        self._poll_timer = None
        QTimer.singleShot(0, self._start_deferred_services)
        try:
            geom = str(core.get_user_setting("main_window_geometry", "") or "").strip()
            if geom:
                self.restoreGeometry(QByteArray.fromBase64(geom.encode("ascii")))
        except Exception:
            pass
        if bool(core.get_user_setting("main_window_maximized", False)):
            QTimer.singleShot(0, self.showMaximized)

    def _build_menu(self):
        self.menuBar().addMenu("Tools")

    def _start_deferred_services(self):
        if self._poll_timer is None:
            self._poll_timer = QTimer(self)
            self._poll_timer.timeout.connect(self._poll_legacy_queues)
            self._poll_timer.start(200)
        self.start_daily_archive_scheduler()

    def _ensure_project_details_page(self):
        if self.project_details_page is None:
            self.project_details_page = ProjectDetailsPage(self)
        return self.project_details_page

    def _ensure_online_page(self):
        if self.online_page is None:
            self.online_page = ViewTendersPage(self)
        return self.online_page

    def _ensure_templates_page(self):
        if self.templates_page is None:
            self.templates_page = TemplatesPage(self)
        return self.templates_page

    def _ensure_settings_page(self):
        if self.settings_page is None:
            self.settings_page = AppSettingsPage(self)
        return self.settings_page

    def set_page(self, page, active_button):
        if page is None:
            return
        if page is self.projects_page:
            self.projects_page.reload_entry_mode()
            self.projects_page.load_projects()
            self._projects_last_subview = "list"
            core.set_user_setting("main_last_view", "projects")
        elif self.project_details_page is not None and page is self.project_details_page:
            self._projects_last_subview = "details"
            try:
                pid = int(getattr(self.project_details_page, "project_id", 0) or 0)
            except Exception:
                pid = 0
            self._projects_last_pid = pid if pid > 0 else self._projects_last_pid
            core.set_user_setting("main_last_view", "projects")
        elif self.online_page is not None and page is self.online_page:
            core.set_user_setting("main_last_view", "online")
        elif self.templates_page is not None and page is self.templates_page:
            core.set_user_setting("main_last_view", "templates")
        elif self.settings_page is not None and page is self.settings_page:
            core.set_user_setting("main_last_view", "settings")
        while self.content_layout.count():
            item = self.content_layout.takeAt(0)
            if item.widget():
                item.widget().setParent(None)
        self.content_layout.addWidget(page)

        for btn in (self.btn_projects, self.btn_online, self.btn_templates, self.btn_settings):
            btn.setProperty("active", btn is active_button)
            btn.style().unpolish(btn)
            btn.style().polish(btn)

    def open_projects_section(self):
        if self._projects_last_subview == "details":
            pid = getattr(self, "_projects_last_pid", None)
            if isinstance(pid, int) and pid > 0:
                details_page = self._ensure_project_details_page()
                current_pid = int(getattr(details_page, "project_id", 0) or 0)
                if current_pid != pid:
                    details_page.load_project(pid)
                self.set_page(details_page, self.btn_projects)
                return
        self.set_page(self.projects_page, self.btn_projects)

    def show_projects(self):
        self.set_page(self.projects_page, self.btn_projects)

    def show_online_section(self):
        self.set_page(self._ensure_online_page(), self.btn_online)

    def show_templates_section(self):
        self.set_page(self._ensure_templates_page(), self.btn_templates)

    def show_settings_section(self):
        self.set_page(self._ensure_settings_page(), self.btn_settings)

    def open_project_details(self, pid):
        details_page = self._ensure_project_details_page()
        details_page.load_project(pid)
        try:
            self._projects_last_pid = int(pid)
        except Exception:
            pass
        self._projects_last_subview = "details"
        self.set_page(details_page, self.btn_projects)

    def closeEvent(self, event):
        try:
            if hasattr(self, "online_page") and self.online_page is not None:
                self.online_page.persist_all_column_layouts()
            try:
                geom = bytes(self.saveGeometry().toBase64()).decode("ascii")
                core.set_user_setting("main_window_geometry", geom)
            except Exception:
                pass
            core.set_user_setting("main_window_maximized", bool(self.isMaximized()))
        except Exception:
            pass
        super().closeEvent(event)

    def _poll_legacy_queues(self):
        while True:
            try:
                msg = core.log_queue.get_nowait()
            except py_queue.Empty:
                break
            try:
                if hasattr(self, "online_page") and self.online_page:
                    self.online_page.append_log(msg)
            except Exception:
                pass

        while True:
            try:
                img_data = core.captcha_req_queue.get_nowait()
            except py_queue.Empty:
                break
            self._open_captcha_dialog(img_data)
        if self._pending_online_refresh:
            self._pending_online_refresh = False
            try:
                self._ensure_online_page().on_site_changed()
            except Exception:
                pass

    def start_daily_archive_scheduler(self):
        # Match legacy behavior: first check after app has run for 1 hour.
        QTimer.singleShot(60 * 60 * 1000, self._archive_scheduler_tick)

    def _archive_scheduler_tick(self):
        self.run_daily_archive_if_due()
        QTimer.singleShot(60 * 60 * 1000, self._archive_scheduler_tick)

    def run_daily_archive_if_due(self):
        if self.archive_job_running:
            return
        last_run_raw = core.ScraperBackend.get_setting("last_auto_archive_utc")
        due = True
        if last_run_raw:
            try:
                last_run = datetime.datetime.fromisoformat(last_run_raw)
                if last_run.tzinfo is None:
                    last_run = last_run.replace(tzinfo=datetime.UTC)
                due = (datetime.datetime.now(datetime.UTC) - last_run) >= datetime.timedelta(hours=12)
            except Exception:
                due = True
        if due:
            self.archive_job_running = True
            threading.Thread(target=self._daily_archive_worker, daemon=True).start()

    def _daily_archive_worker(self):
        total = 0
        websites_count = 0
        try:
            websites = core.ScraperBackend.get_websites()
            websites_count = len(websites)
            for sid in websites.keys():
                total += int(core.ScraperBackend.archive_completed_tenders_logic(sid) or 0)
            core.ScraperBackend.set_setting("last_auto_archive_utc", datetime.datetime.now(datetime.UTC).isoformat())
            core.ScraperBackend.log_auto_archive_run(
                status="success",
                archived_count=total,
                archived_status_updated=0,
                websites_count=websites_count,
                notes="12-hour scheduled run",
            )
            core.log_to_gui(f"Auto-archive complete (12-hour schedule). Archived {total} tenders.")
            self._pending_online_refresh = True
        except Exception as e:
            core.ScraperBackend.log_auto_archive_run(
                status="failed",
                archived_count=total,
                archived_status_updated=0,
                websites_count=websites_count,
                notes=f"12-hour scheduled run failed: {e}",
            )
            core.log_to_gui(f"Auto-archive failed: {e}")
        finally:
            self.archive_job_running = False

    def _open_captcha_dialog(self, img_data):
        dlg = QDialog(self)
        dlg.setWindowTitle("Solve Captcha")
        dlg.resize(320, 280)
        lay = QVBoxLayout(dlg)
        img = QLabel("[captcha]")
        img.setAlignment(Qt.AlignCenter)
        try:
            from PySide6.QtGui import QPixmap
            px = QPixmap()
            px.loadFromData(img_data)
            if not px.isNull():
                img.setPixmap(px)
        except Exception:
            pass
        lay.addWidget(img)
        lay.addWidget(QLabel("Enter Captcha:"))
        cap = QLineEdit()
        lay.addWidget(cap)
        btns = QHBoxLayout()
        ok = QPushButton("Submit")
        cancel = QPushButton("Cancel")
        btns.addWidget(ok)
        btns.addWidget(cancel)
        lay.addLayout(btns)
        sent = {"done": False}

        def submit():
            sent["done"] = True
            core.captcha_res_queue.put(cap.text())
            dlg.accept()

        def on_cancel():
            sent["done"] = True
            core.captcha_res_queue.put(None)
            dlg.reject()

        def on_finished(result):
            if not sent["done"] and result != QDialog.Accepted:
                sent["done"] = True
                core.captcha_res_queue.put(None)

        ok.clicked.connect(submit)
        cancel.clicked.connect(on_cancel)
        dlg.finished.connect(on_finished)
        dlg.exec()


def apply_styles(app):
    app.setStyle("Fusion")
    app.setFont(QFont("Segoe UI", 10))

    app.setStyleSheet(
        """
        QMainWindow { background: #f6f8fb; }
        QMenuBar { background: #ffffff; border-bottom: 1px solid #d9e0ea; }
        #Sidebar {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #12344c, stop:1 #1e526f);
            border-right: 1px solid #294f63;
        }
        QLabel#Brand {
            color: #ffffff;
            font-size: 20px;
            font-weight: 700;
            letter-spacing: 0.5px;
            padding: 8px 6px 16px 6px;
        }
        QPushButton#NavButton {
            background: rgba(255, 255, 255, 0.08);
            color: #dce8ef;
            border: 1px solid rgba(255, 255, 255, 0.10);
            border-radius: 10px;
            padding: 10px 12px;
            text-align: left;
            font-weight: 600;
        }
        QPushButton#NavButton[active="true"] {
            background: #ffffff;
            color: #12344c;
            border: 1px solid #d6e2eb;
        }
        QLabel#PageTitle {
            font-size: 24px;
            font-weight: 700;
            color: #16364a;
            padding: 4px 0px;
        }
        QLabel#SoftText {
            color: #4d6170;
            font-size: 13px;
        }
        QFrame#ProjectsHeaderBar {
            background: #067a69;
            border: none;
        }
        QLabel#ProjectsHeaderTitle {
            color: #ffffff;
            font-size: 22px;
            font-weight: 600;
            padding: 2px 0px;
        }
        QFrame#ProjectsToolbarPanel {
            background: #f4f6fa;
            border: none;
        }
        QFrame#ProjectsTablePanel {
            background: #f4f6fa;
            border: none;
        }
        QFrame#ProjectTopBar {
            background: #344852;
            border: none;
        }
        QLabel#ProjectTitleBarText {
            color: #ffffff;
            font-size: 20px;
            font-weight: 700;
            padding: 2px 4px;
        }
        QFrame#ChecklistPanel {
            background: #f2f5f8;
            border: 1px solid #d4dce6;
            border-radius: 0px;
        }
        QFrame#PreviewBox {
            background: #f4f6f8;
            border: 1px solid #ccd6df;
            border-radius: 0px;
        }
        QPushButton#PreviewSideToggle {
            background: #7f878e;
            color: #ffffff;
            border: 1px solid #737b82;
            border-radius: 0px;
            padding: 0px;
            min-width: 16px;
            max-width: 16px;
        }
        QFrame#ScraperHeaderBar {
            background: #3f51b5;
            border: none;
        }
        QLabel#ScraperHeaderTitle {
            color: #ffffff;
            font-size: 21px;
            font-weight: 500;
            padding: 2px 0px;
        }
        QFrame#ScraperControls {
            background: #eef1fa;
            border: none;
        }
        QPushButton {
            background: #e8edf4;
            border: 1px solid #cdd8e3;
            border-radius: 8px;
            color: #1d3342;
            padding: 8px 12px;
            font-weight: 600;
        }
        QPushButton[compact="true"] {
            padding: 2px 9px;
            border-radius: 8px;
            font-weight: 500;
            min-height: 22px;
        }
        QPushButton:hover { background: #dde7f0; }
        QPushButton#PrimaryButton {
            background: #0a7a5b;
            border: 1px solid #086a4f;
            color: #ffffff;
        }
        QPushButton#PrimaryButton:hover { background: #096f53; }
        QPushButton#ClearActionButton {
            background: #ffb300;
            border: 1px solid #cc8f00;
            color: #1d2a33;
        }
        QPushButton#DangerButton {
            background: #c0392b;
            border: 1px solid #a32f24;
            color: #ffffff;
        }
        QPushButton#AccentBlueButton {
            background: #1f8de4;
            border: 1px solid #1976c2;
            color: #ffffff;
        }
        QPushButton#ProjectsOpenFolderButton {
            background: #d7e0e9;
            border: 1px solid #b8c7d6;
            color: #16364a;
        }
        QPushButton#ExportDataButton {
            padding: 1px 8px;
            min-height: 20px;
            max-height: 22px;
            border-radius: 8px;
            font-size: 8.5pt;
            font-weight: 500;
            background: #455a64;
            border: 1px solid #2f3e46;
            color: #ffffff;
        }
        QPushButton#ScraperManageButton {
            background: #9e9e9e;
            border: 1px solid #757575;
            color: #ffffff;
        }
        QPushButton#ScraperDownloadButton {
            background: #2e7d32;
            border: 1px solid #1b5e20;
            color: #ffffff;
        }
        QPushButton#ScraperClearButton {
            background: #b71c1c;
            border: 1px solid #7f1111;
            color: #ffffff;
        }
        QPushButton#ScraperFetchButton {
            background: #2196f3;
            border: 1px solid #0d6eb7;
            color: #ffffff;
        }
        QPushButton#ScraperGetButton {
            background: #ff9800;
            border: 1px solid #c97800;
            color: #ffffff;
        }
        QPushButton#ScraperDownloadSelectedButton {
            background: #4caf50;
            border: 1px solid #2e7d32;
            color: #ffffff;
        }
        QPushButton#ScraperDownloadResultsButton {
            background: #00897b;
            border: 1px solid #00695c;
            color: #ffffff;
        }
        QPushButton#ScraperCheckStatusButton {
            background: #9c27b0;
            border: 1px solid #6a1b9a;
            color: #ffffff;
        }
        QPushButton#ScraperAddProjectsButton {
            background: #607d8b;
            border: 1px solid #455a64;
            color: #ffffff;
        }
        QPushButton#ScraperFiltersButton {
            background: #5d4037;
            border: 1px solid #3e2723;
            color: #ffffff;
        }
        QPushButton#ProjectBackButton {
            background: #546e7a;
            border: 1px solid #455a64;
            color: #ffffff;
        }
        QPushButton#ProjectUpdateButton {
            background: #ffb300;
            border: 1px solid #cc8f00;
            color: #000000;
        }
        QPushButton#ProjectOpenExplorerButton {
            background: #00897b;
            border: 1px solid #00695c;
            color: #ffffff;
        }
        QPushButton#ProjectAddButton {
            background: #4caf50;
            border: 1px solid #2e7d32;
            color: #ffffff;
        }
        QPushButton#ProjectUpdateSelectedButton {
            background: #ff9800;
            border: 1px solid #c97800;
            color: #ffffff;
        }
        QPushButton#ProjectNewFolderButton {
            background: #607d8b;
            border: 1px solid #455a64;
            color: #ffffff;
        }
        QPushButton#ProjectManageFoldersButton {
            background: #5d4037;
            border: 1px solid #3e2723;
            color: #ffffff;
        }
        QPushButton#ProjectImportButton {
            background: #3949ab;
            border: 1px solid #283593;
            color: #ffffff;
        }
        QPushButton#ProjectImportTemplateButton {
            background: #5e6aa3;
            border: 1px solid #4a5688;
            color: #ffffff;
        }
        QPushButton#ProjectDownloadButton {
            background: #00897b;
            border: 1px solid #00695c;
            color: #ffffff;
        }
        QPushButton#ProjectSaveTemplateButton {
            background: #6d4c41;
            border: 1px solid #4e342e;
            color: #ffffff;
        }
        QPushButton#ProjectAttachButton {
            background: #2196f3;
            border: 1px solid #0d6eb7;
            color: #ffffff;
        }
        QPushButton#ProjectOpenFileButton {
            background: #607d8b;
            border: 1px solid #455a64;
            color: #ffffff;
        }
        QPushButton#ProjectDeleteButton {
            background: #f44336;
            border: 1px solid #c62828;
            color: #ffffff;
        }
        QPushButton[legacyProjectButton="true"] {
            min-height: 22px;
            max-height: 24px;
            padding: 2px 8px;
            border-radius: 8px;
            font-weight: 500;
        }
        QPushButton[legacyProjectTopButton="true"] {
            min-height: 18px;
            max-height: 20px;
            padding: 2px 7px;
            border-radius: 8px;
            font-weight: 500;
        }
        QLineEdit, QComboBox, QTextEdit, QDateTimeEdit {
            background: #ffffff;
            border: 1px solid #ccd6e1;
            border-radius: 8px;
            padding: 6px 8px;
            selection-background-color: #0f88c7;
        }
        QLineEdit[isSearchBar="true"] {
            min-height: 24px;
            padding: 4px 8px;
        }
        QTextBrowser#LogView {
            background: #000000;
            color: #00ff66;
            border: 1px solid #243124;
            border-radius: 6px;
            font-family: Consolas, "Courier New", monospace;
            font-size: 10.5pt;
            selection-background-color: #1b5e20;
        }
        QTabWidget::pane {
            border: 1px solid #d4dce6;
            top: -1px;
        }
        QTabBar::tab {
            min-height: 24px;
            padding: 2px 12px;
        }
        QTableWidget {
            background: #ffffff;
            border: 1px solid #d2dce8;
            border-radius: 10px;
            gridline-color: #e6edf5;
            alternate-background-color: #f9fbfe;
        }
        QHeaderView::section {
            background: #ebf1f7;
            color: #21384a;
            padding: 8px;
            border: none;
            border-right: 1px solid #d8e3ee;
            border-bottom: 1px solid #d8e3ee;
            font-weight: 700;
        }
        """
    )


def run():
    if core.install_runtime_exe_if_needed():
        return 0
    core.init_db()
    app = QApplication(sys.argv)
    apply_styles(app)
    win = BidManagerQt()
    win.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(run())
