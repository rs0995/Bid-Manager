# Start your Python code here
import requests
from bs4 import BeautifulSoup
import time # For adding a delay between requests to be polite
import os # Required for creating directories (though DOWNLOAD_DIR is removed, os might still be useful for other path operations)
from urllib.parse import urlparse, urljoin # To extract filename from URL and join relative URLs
import pandas as pd # For Excel operations
import openpyxl # Explicitly import openpyxl to access its functionalities for column resizing
from openpyxl.styles import Alignment, Font # For text wrapping, alignment, and font styling for hyperlinks
from openpyxl.styles.numbers import NumberFormat # Corrected import for NumberFormat
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
import json # For saving and loading global_organisations_data
import re # For regex-based Tender ID extraction
from datetime import datetime, timedelta, date # For logging and re-checking tenders
import google.generativeai as genai

# --- Selenium Imports for Download Functionality ---
from selenium import webdriver
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, StaleElementReferenceException, NoSuchElementException
from webdriver_manager.firefox import GeckoDriverManager
import shutil # For moving files after download
# --- End Selenium Imports ---

# --- NEW: Gemini CAPTCHA Solver Imports ---
try:
    import google.generativeai as genai
    from PIL import Image
    import io
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False
# --- End Gemini Imports ---

BASE_DOWNLOAD_DIRECTORY = "Tender_Downloads"
CSV_DATABASE_DIRECTORY = "csv_database" # For storing CSV data
# NEW: Your GOOGLE_API_KEY. For better security, consider loading this from an environment variable instead of hardcoding it.
GOOGLE_API_KEY = "AIzaSyD78VcTNJCh3qlSYN9ZcLl4MdA3Q88TXQU"
# --- End Configuration ---

# --- 🧠 AUTO-DETECT MODEL ---
def get_working_model():
    try:
        for m in genai.list_models():
            if 'generateContent' in m.supported_generation_methods:
                if 'flash' in m.name or 'pro' in m.name: return genai.GenerativeModel(m.name)
        return genai.GenerativeModel('gemini-pro')
    except:
        return genai.GenerativeModel('gemini-pro')

if GEMINI_AVAILABLE:
    genai.configure(api_key=GOOGLE_API_KEY)
    model = get_working_model()

# Initialize a session for persistent connections and cookies
# This helps maintain session state and bypass "Stale Session" errors
session = requests.Session()
# Add a User-Agent header to mimic a web browser
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWeb7Kit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'DNT': '1' # Do Not Track header
})

global_organisations_data = {}
ORG_DATA_FILE = "global_organizations_data.json" # File to store the organization data
DOWNLOAD_LOG_FILE = "download_log.json" # Log file for downloads
ACTIVITY_LOG_FILE = "activity_log.txt" # New log for user actions
WEBSITES_DATA_FILE = "websites_data.json" # NEW: File to store website options

# Define the available websites globally - This will be populated from a file
website_options = {}
excel_filename = "List of Organisations and Tenders.xlsx" # Consistent filename
EXPIRED_TENDERS_SHEET = "Expired Tenders" # New sheet name for expired tenders
CUSTOM_TENDERS_SHEET = "Custom Tenders" # New sheet for manually checked tenders
BOOKMARKED_TENDERS_FILENAME = "Tenders to Focus.xlsx" # New file for bookmarked tenders

# Global flag to track if CAPTCHA has been solved in the current Selenium session
captcha_solved_in_session = False

# --- Helper functions ---

def log_activity(message):
    """Appends a timestamped message to the activity log file."""
    try:
        with open(ACTIVITY_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(f"{datetime.now().isoformat()} - {message}\n")
    except Exception as e:
        print(f"Warning: Could not write to activity log: {e}")

def clean_currency_value(value_str):
    """
    Cleans a string to extract a numeric currency value.
    Handles currency symbols, commas, and common non-numeric text.
    Returns a float or None.
    """
    if not isinstance(value_str, str):
        # If it's already a number (e.g. from a CSV read), just return it.
        if isinstance(value_str, (int, float)):
            return float(value_str)
        return None

    # Use regex to find a number, which can include commas and a decimal part.
    match = re.search(r'[\d,.]+', value_str)
    if not match:
        return None

    # Take the matched part and remove commas to make it parsable by float()
    num_str = match.group(0).replace(',', '')
    try:
        return float(num_str)
    except (ValueError, TypeError):
        return None

# --- NEW: CSV and Data Management Functions ---

def get_csv_path(sheet_name):
    """Returns the full path for a sheet's corresponding CSV file."""
    os.makedirs(CSV_DATABASE_DIRECTORY, exist_ok=True)
    # Sanitize sheet_name to be a valid filename
    safe_sheet_name = re.sub(r'[\\/*?:"<>|]',"", sheet_name)
    return os.path.join(CSV_DATABASE_DIRECTORY, f"{safe_sheet_name}.csv")

def read_sheet_from_csv(sheet_name):
    """Reads a sheet's data from its CSV file into a DataFrame."""
    csv_path = get_csv_path(sheet_name)
    if os.path.exists(csv_path):
        try:
            # Keep_default_na=False helps in reading empty strings correctly
            return pd.read_csv(csv_path, keep_default_na=False)
        except pd.errors.EmptyDataError:
            return pd.DataFrame() # Return empty if CSV is empty
        except Exception as e:
            print(f"Warning: Could not read CSV '{csv_path}': {e}. Returning empty DataFrame.")
            return pd.DataFrame()
    return pd.DataFrame() # Return empty if CSV doesn't exist

def write_sheet_to_csv(df, sheet_name):
    """Writes a DataFrame to a sheet's corresponding CSV file."""
    csv_path = get_csv_path(sheet_name)
    try:
        df.to_csv(csv_path, index=False)
    except Exception as e:
        print(f"Error writing to CSV '{csv_path}': {e}")

def sync_excel_to_csv_on_startup(filename=excel_filename):
    """
    One-time migration. If Excel exists and CSVs don't, copy data from Excel to CSVs.
    This ensures that on the first run with the new system, existing Excel data is preserved.
    """
    if not os.path.exists(filename):
        return # No Excel file to migrate from

    os.makedirs(CSV_DATABASE_DIRECTORY, exist_ok=True)
    
    try:
        xls = pd.ExcelFile(filename)
        for sheet_name in xls.sheet_names:
            csv_path = get_csv_path(sheet_name)
            if not os.path.exists(csv_path) or os.path.getsize(csv_path) == 0:
                print(f"Migrating sheet '{sheet_name}' from Excel to CSV...")
                df = pd.read_excel(xls, sheet_name=sheet_name)
                
                # When reading from Excel, the folder hyperlink is lost. We need to capture it.
                if not df.empty and 'Tender Id' in df.columns and not sheet_name.endswith('-O'):
                    hyperlink_map = get_hyperlinks_from_sheet(filename, sheet_name)
                    if hyperlink_map:
                        df['__folder_hyperlink__'] = df['Tender Id'].astype(str).map(hyperlink_map)
                
                write_sheet_to_csv(df, sheet_name)
    except Exception as e:
        print(f"An error occurred during initial Excel-to-CSV migration: {e}")


# --- End NEW Functions ---


def save_organisations_data(data, filename=ORG_DATA_FILE):
    """Saves the global organizations data (including URLs) to a JSON file."""
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)
        print(f"Organization data saved to '{filename}'.")
    except Exception as e:
        print(f"Error saving organization data to file: {e}")

def load_organisations_data(filename=ORG_DATA_FILE):
    """Loads the global organizations data from a JSON file."""
    global global_organisations_data # Declare intent to modify global variable
    try:
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as f:
                global_organisations_data = json.load(f)
            return True
        else:
            return False
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON from '{filename}': {e}. File might be corrupt or empty.")
        global_organisations_data = {} # Reset to empty to prevent further errors
        return False
    except Exception as e:
        print(f"Error loading organization data from file: {e}")
        return False

def load_download_log(filename=DOWNLOAD_LOG_FILE):
    """Loads the download log from a JSON file."""
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError:
            return {} # Return empty dict if file is corrupt or empty
    return {}

def save_download_log(data, filename=DOWNLOAD_LOG_FILE):
    """Saves the download log to a JSON file."""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)

def save_websites_data(data, filename=WEBSITES_DATA_FILE):
    """Saves the website options to a JSON file."""
    try:
        # Convert keys to string for JSON compatibility
        data_to_save = {str(k): v for k, v in data.items()}
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data_to_save, f, indent=4)
        print(f"Website data saved to '{filename}'.")
    except Exception as e:
        print(f"Error saving website data to file: {e}")

def initialize_default_websites():
    """Initializes the default websites and saves them."""
    global website_options
    website_options = {
        1: {"name": "MahaTenders", "url": "https://mahatenders.gov.in/nicgep/app?page=FrontEndTendersByOrganisation&service=page", "status_url": "https://mahatenders.gov.in/nicgep/app?page=WebTenderStatusLists&service=page"},
        2: {"name": "ETenders", "url": "https://etenders.gov.in/eprocure/app?page=FrontEndTendersByOrganisation&service=page", "status_url": "https://etenders.gov.in/eprocure/app?page=WebTenderStatusLists&service=page"},
        3: {"name": "Eprocure", "url": "https://eprocure.gov.in/eprocure/app?page=FrontEndTendersByOrganisation&service=page", "status_url": "https://eprocure.gov.in/eprocure/app?page=WebTenderStatusLists&service=page"}
    }
    save_websites_data(website_options)

def load_websites_data(filename=WEBSITES_DATA_FILE):
    """Loads the website options from a JSON file, or creates it with defaults."""
    global website_options
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                loaded_data = json.load(f)
                # Convert keys back to int
                website_options = {int(k): v for k, v in loaded_data.items()}
            return True
        except (json.JSONDecodeError, ValueError) as e:
            print(f"Error decoding JSON from '{filename}': {e}. File might be corrupt. Initializing with defaults.")
            initialize_default_websites()
            return False
        except Exception as e:
            print(f"Error loading website data from file: {e}. Initializing with defaults.")
            initialize_default_websites()
            return False
    else:
        print(f"'{filename}' not found. Initializing with default websites and saving.")
        initialize_default_websites()
        return True

def add_website():
    """Prompts user to add a new website and saves it."""
    global website_options
    print("\n--- Add a New Website ---")
    name = input("Enter the website name (e.g., 'NewTenders'): ").strip()
    org_url = input("Enter the 'Tenders by Organisation' URL: ").strip()
    status_url = input("Enter the 'Tender Status' URL: ").strip()

    if not (name and org_url and status_url):
        print("All fields are required. Website not added.")
        return

    # Find the next available integer key
    new_key = max(website_options.keys()) + 1 if website_options else 1
    
    website_options[new_key] = {
        "name": name,
        "url": org_url,
        "status_url": status_url
    }
    
    save_websites_data(website_options)
    print(f"Successfully added '{name}' to the website list.")
    log_activity(f"Added new website: {name}")

def delete_website():
    """Prompts user to delete an existing website and re-keys the dictionary."""
    global website_options
    print("\n--- Delete a Website ---")
    if not website_options:
        print("No websites to delete.")
        return

    for key, value in sorted(website_options.items()):
        print(f"{key}. {value['name']}")
    
    choice_str = input("Enter the number of the website to delete (or press Enter to cancel): ").strip()
    if not choice_str.isdigit():
        print("Invalid input or cancellation. No website deleted.")
        return

    choice_key = int(choice_str)
    if choice_key not in website_options:
        print("Invalid choice. No website deleted.")
        return
    
    # Prevent deleting the last website
    if len(website_options) == 1:
        print("Cannot delete the last remaining website.")
        return

    deleted_name = website_options[choice_key]['name']
    del website_options[choice_key]
    
    # Re-key the dictionary to be sequential from 1
    rekeyed_options = {}
    # Sort by old key to maintain some order before re-keying
    for i, (old_key, value) in enumerate(sorted(website_options.items()), 1):
        rekeyed_options[i] = value
    website_options = rekeyed_options

    save_websites_data(website_options)
    print(f"Successfully deleted '{deleted_name}'. The website list has been re-ordered.")
    log_activity(f"Deleted website: {deleted_name}")

def fetch_organisations(url):
    """
    Fetches the list of organizations, their tender counts, and the URL to their tenders
    from the given URL by parsing an HTML table.
    Returns a dictionary mapping organization names to a dictionary of their details.
    """
    # Ensure global_organisations_data is loaded from file before updating
    load_organisations_data()

    organisations_this_run = {} # Store organizations fetched in this specific run
    print("Fetching organizations from table...")
    try:
        response = session.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        organisation_name_header_cell = soup.find('td', string='Organisation Name')

        if organisation_name_header_cell:
            table = organisation_name_header_cell.find_parent('table')

            if table:
                rows = table.find_all('tr')
                header_row_index = -1
                for i, row in enumerate(rows):
                    if organisation_name_header_cell in row.find_all('td') or organisation_name_header_cell in row.find_all('th'):
                        header_row_index = i
                        break

                for i, row in enumerate(rows):
                    if i <= header_row_index:
                        continue

                    cols = row.find_all('td')

                    if len(cols) > 2:
                        s_no = cols[0].text.strip()
                        org_name = cols[1].text.strip()
                        tender_count_cell = cols[2]
                        tender_count_text = tender_count_cell.text.strip()
                        tender_link_tag = tender_count_cell.find('a', href=True)

                        tender_url = None
                        if tender_link_tag:
                            relative_url = tender_link_tag['href']
                            tender_url = requests.compat.urljoin(url, relative_url)

                        if org_name:
                            organisations_this_run[org_name] = { # Populate this temp dict
                                'S.No': s_no,
                                'Organisation Name': org_name,
                                'Tender Count': tender_count_text,
                                'Tenders URL': tender_url
                            }
            else:
                print("Could not find the parent table containing the 'Organisation Name' cell.")
        else:
            print("Could not find the 'Organisation Name' cell (<td> tag with exact text).")
            print("Please inspect the HTML of the website very carefully to find the correct tag and its text/attributes.")

    except requests.exceptions.RequestException as e:
        print(f"Error fetching organizations: {e}")
    
    # Merge newly fetched organizations into the global data
    global global_organisations_data
    global_organisations_data.update(organisations_this_run)
    save_organisations_data(global_organisations_data) # Save the merged data
    return organisations_this_run # Return only this run's orgs for export_organizations_to_excel

def update_and_export_organizations(newly_fetched_orgs_dict, filename, org_sheet_name):
    """
    Exports fetched organizations to an Excel file with a 'Select' checkbox column.
    The 'Tenders URL' column is excluded from the Excel output.
    'S.No' and 'Tender Count' are converted to numeric.
    Columns are auto-resized to fit content.
    This updated version ensures existing 'Select' column data is preserved and new
    organizations are added without disturbing existing entries.
    """
    if not newly_fetched_orgs_dict:
        print(f"No new organizations were fetched for '{org_sheet_name}'. Syncing existing selections from Excel.")

    # --- NEW: Sync from Excel to CSV at the start ---
    # This ensures that any manual changes in the Excel file (like 'Select' or 'Fetch Results')
    # are read and become the baseline before merging with newly scraped data.
    try:
        df_from_excel = pd.read_excel(filename, sheet_name=org_sheet_name)
        write_sheet_to_csv(df_from_excel, org_sheet_name)
        print(f"Successfully synced user selections from Excel sheet '{org_sheet_name}'.")
    except (FileNotFoundError, ValueError):
        # This is fine if the sheet doesn't exist yet. The rest of the function will create it.
        print(f"Info: Org sheet '{org_sheet_name}' not found in Excel. Will be created if new organizations are found.")
    except Exception as e:
        print(f"An unexpected error occurred during org sheet sync: {e}")

    # 1. Read existing data from CSV (which is now a mirror of Excel) and prepare it
    existing_orgs_df = read_sheet_from_csv(org_sheet_name)
    if not existing_orgs_df.empty:
        # Ensure key is string for merging and set as index
        existing_orgs_df['Organisation Name'] = existing_orgs_df['Organisation Name'].astype(str)
        existing_orgs_df.set_index('Organisation Name', inplace=True)

    # 2. Prepare newly fetched data and prepare it
    df_newly_fetched_orgs = pd.DataFrame(list(newly_fetched_orgs_dict.values()))
    if df_newly_fetched_orgs.empty:
        return

    if 'Tenders URL' in df_newly_fetched_orgs.columns:
        df_newly_fetched_orgs = df_newly_fetched_orgs.drop(columns=['Tenders URL'])
    
    df_newly_fetched_orgs['S.No'] = pd.to_numeric(df_newly_fetched_orgs['S.No'], errors='coerce')
    df_newly_fetched_orgs['Tender Count'] = pd.to_numeric(df_newly_fetched_orgs['Tender Count'], errors='coerce')
    df_newly_fetched_orgs['Organisation Name'] = df_newly_fetched_orgs['Organisation Name'].astype(str)
    df_newly_fetched_orgs.set_index('Organisation Name', inplace=True)

    # 3. Combine data, starting with ONLY newly fetched ("live") organizations
    #    and preserving 'Select' status from the old data.
    
    # Start with the live organizations as the base.
    combined_df = df_newly_fetched_orgs.copy()

    if not existing_orgs_df.empty and 'Select (Type "Yes" to select)' in existing_orgs_df.columns:
        # Create a map of Organisation Name -> Select Status from the old data.
        # The index of existing_orgs_df is already 'Organisation Name'.
        select_status_map = existing_orgs_df['Select (Type "Yes" to select)'].to_dict()
        
        # Apply the saved 'Select' status to the live organizations.
        # The index of combined_df is 'Organisation Name' from the new fetch.
        combined_df['Select (Type "Yes" to select)'] = combined_df.index.map(select_status_map)

    # Reset index to make 'Organisation Name' a regular column before final structuring.
    combined_df.reset_index(inplace=True)

    # 4. Finalize DataFrame structure
    if 'Select (Type "Yes" to select)' not in combined_df.columns:
        # This will run if there was no existing data or it had no 'Select' column.
        combined_df['Select (Type "Yes" to select)'] = ''
    # For any orgs that were in the new fetch but not the old, their 'Select' will be NaN.
    # Fill these with an empty string.
    combined_df['Select (Type "Yes" to select)'].fillna('', inplace=True)

    final_columns = ['S.No', 'Organisation Name', 'Tender Count', 'Select (Type "Yes" to select)']
    combined_df = combined_df.reindex(columns=final_columns)
    # Sort by S.No, putting orgs no longer on the site (NaN S.No) at the bottom.
    combined_df = combined_df.sort_values(by='S.No', ascending=True, na_position='last', ignore_index=True)
    
    # 5. Write to CSV and then update Excel
    write_sheet_to_csv(combined_df, org_sheet_name)
    update_excel_sheet(combined_df, org_sheet_name, filename)

    print(f"\nSuccessfully updated data for '{org_sheet_name}'. Total organizations: {len(combined_df)}.")
    print("Please open the Excel file, type 'Yes' (case-insensitive) in the 'Select' column for organizations you want to process, and SAVE the file.")

def update_excel_sheet(df, sheet_name, filename):
    """
    Writes a DataFrame to a specific sheet in an Excel file. This function is designed
    to be robust against file corruption by reading all existing data, updating it
    in memory, and then writing a completely new file, which is safer than append mode.
    """
    try:
        # 1. Read all existing sheets from the file into memory.
        all_sheets = {}
        if os.path.exists(filename):
            try:
                xls = pd.ExcelFile(filename, engine='openpyxl')
                for s_name in xls.sheet_names:
                    if s_name != sheet_name:
                        all_sheets[s_name] = pd.read_excel(xls, sheet_name=s_name)
            except Exception as e:
                print(f"Warning: Could not read existing Excel file '{filename}' due to error: {e}.")
                print("The Excel file will be rebuilt from scratch with the current data.")
                all_sheets = {}

        # 2. Add or update the DataFrame for the target sheet.
        all_sheets[sheet_name] = df.copy()

        # 3. Write all sheets back to the file in one go (overwrite mode).
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            for s_name in sorted(all_sheets.keys()):
                df_to_write = all_sheets[s_name]
                df_to_write.to_excel(writer, sheet_name=s_name, index=False)
                worksheet = writer.sheets[s_name]

                # --- Apply Formatting ---
                if s_name.endswith('-O'):
                    for column in worksheet.columns:
                        max_length = 0
                        column_name = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                            except: pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column_name].width = min(max(adjusted_width, 10), 80)
                else: # Tender-like sheets
                    col_indices = {col_name: i + 1 for i, col_name in enumerate(df_to_write.columns)}
                    
                    if 'Tender Title' in col_indices and 'Tender URL' in col_indices:
                        for row_idx, row_data in df_to_write.iterrows():
                            cell = worksheet.cell(row=row_idx + 2, column=col_indices['Tender Title'])
                            if cell.value and pd.notna(row_data.get('Tender URL')):
                                cell.hyperlink = row_data.get('Tender URL')
                                cell.font = Font(color="0000FF", underline="single")
                    
                    if 'Tender Id' in col_indices and '__folder_hyperlink__' in col_indices:
                        for row_idx, row_data in df_to_write.iterrows():
                            cell = worksheet.cell(row=row_idx + 2, column=col_indices['Tender Id'])
                            hyperlink_val = row_data.get('__folder_hyperlink__')
                            if cell.value and pd.notna(hyperlink_val) and str(hyperlink_val).strip():
                                cell.hyperlink = hyperlink_val
                                cell.font = Font(color="0000FF", underline="single")

                    if 'Current Status' in col_indices and '__status_hyperlink__' in col_indices:
                        for row_idx, row_data in df_to_write.iterrows():
                            cell = worksheet.cell(row=row_idx + 2, column=col_indices['Current Status'])
                            if cell.value and pd.notna(row_data.get('__status_hyperlink__')):
                                cell.hyperlink = row_data.get('__status_hyperlink__')
                                if not cell.font or (cell.font.color and cell.font.color.rgb != "0000FF"):
                                    cell.font = Font(color="0000FF", underline="single")

                    specific_column_widths = {
                        'Sr. No.': 5, 'Tender Title': 40, 'Tender Category': 20, 'Organization Chain': 20,
                        'Location': 10, 'Tender Id': 25, 'Closing Date': 12, 'Closing Time': 12, 
                        'Pre bid meeting date': 18, 'Tender Value': 20, 'EMD': 15,
                        'Download (Type "Yes" to download)': 10, 'Bookmark': 10
                    }
                    text_wrap_columns = [
                        'Tender Title', 'Organization Chain', 'Location', 'Tender Id', 'Pre bid meeting date'
                    ]
                    
                    for col_name, col_idx in col_indices.items():
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        
                        if col_name in ['Tender URL', '__folder_hyperlink__', '__is_available_on_site__', '__status_hyperlink__']:
                            worksheet.column_dimensions[col_letter].hidden = True
                            continue

                        if col_name in specific_column_widths:
                            worksheet.column_dimensions[col_letter].width = specific_column_widths[col_name]
                        else:
                            max_len_val = df_to_write[col_name].astype(str).map(len).max()
                            max_length = max(max_len_val if pd.notna(max_len_val) else 0, len(col_name))
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[col_letter].width = min(max(adjusted_width, 10), 80)

                        for row_idx_in_sheet in range(1, df_to_write.shape[0] + 2):
                            cell = worksheet[f'{col_letter}{row_idx_in_sheet}']
                            if col_name in ['Tender Value', 'EMD']:
                                if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                                    cell.number_format = '#,##0.00'
                            elif col_name in ['Closing Date', 'Status Last Checked']:
                                if isinstance(cell.value, (datetime, pd.Timestamp, date)):
                                    cell.number_format = 'DD-MM-YYYY'
                            
                            alignment_horizontal = 'center' if col_name not in ['Tender Title'] else 'left'
                            alignment_vertical = 'top' if col_name in text_wrap_columns else 'center'
                            cell.alignment = Alignment(wrap_text=(col_name in text_wrap_columns), horizontal=alignment_horizontal, vertical=alignment_vertical)
                    
                    worksheet.row_dimensions[1].height = 30

        print(f"Successfully updated Excel sheet '{sheet_name}' in '{filename}'.")

    except Exception as e:
        print(f"FATAL ERROR updating Excel sheet '{sheet_name}': {e}")
        print("This may be due to file permissions or the file being open in another program.")

def export_organizations_to_excel(organizations, filename, org_sheet_name):
    """
    Exports fetched organizations to an Excel file with a 'Select' checkbox column.
    The 'Tenders URL' column is excluded from the Excel output.
    'S.No' and 'Tender Count' are converted to numeric.
    Columns are auto-resized to fit content.
    This updated version ensures existing 'Select' column data is preserved and new
    organizations are added without disturbing existing entries.
    """
    df_newly_fetched_orgs = pd.DataFrame(list(organizations.values()))
    
    if df_newly_fetched_orgs.empty:
        print(f"No organizations were newly fetched for '{org_sheet_name}'.")
        return

    if 'Tenders URL' in df_newly_fetched_orgs.columns:
        df_newly_fetched_orgs = df_newly_fetched_orgs.drop(columns=['Tenders URL'])
    
    df_newly_fetched_orgs['S.No'] = pd.to_numeric(df_newly_fetched_orgs['S.No'], errors='coerce')
    df_newly_fetched_orgs['Tender Count'] = pd.to_numeric(df_newly_fetched_orgs['Tender Count'], errors='coerce')

    try:
        existing_orgs_df = pd.DataFrame()
        
        if os.path.exists(filename):
            try:
                workbook = openpyxl.load_workbook(filename)
                if org_sheet_name in workbook.sheetnames:
                    existing_orgs_df = pd.read_excel(filename, sheet_name=org_sheet_name)
                workbook.close()
            except Exception as e:
                print(f"Warning: Could not load existing organizations from sheet '{org_sheet_name}': {e}. Starting with empty data for this sheet.")
                existing_orgs_df = pd.DataFrame()

        # Merge the new data with the old, preserving the 'Select' column
        if not existing_orgs_df.empty:
            # Preserve the 'Select' column from the existing data
            merged_df = pd.merge(df_newly_fetched_orgs, existing_orgs_df[['Organisation Name', 'Select (Type "Yes" to select)']], on='Organisation Name', how='left')
            merged_df['Select (Type "Yes" to select)'].fillna('', inplace=True)
        else:
            merged_df = df_newly_fetched_orgs
            merged_df['Select (Type "Yes" to select)'] = ''
        
        # Reorder columns to the desired format
        final_columns = ['S.No', 'Organisation Name', 'Tender Count', 'Select (Type "Yes" to select)']
        merged_df = merged_df[final_columns]
        
        # Sort by S.No for consistent order
        merged_df = merged_df.sort_values(by='S.No', ascending=True, ignore_index=True)
        
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            merged_df.to_excel(writer, sheet_name=org_sheet_name, index=False)
            
            worksheet = writer.sheets[org_sheet_name]
            
            # Auto-resize columns
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                if adjusted_width < 10:
                    adjusted_width = 10
                if adjusted_width > 80:
                    adjusted_width = 80
                worksheet.column_dimensions[column_name].width = adjusted_width

        print(f"\nSuccessfully updated '{org_sheet_name}' sheet in '{filename}'. Total organizations: {len(merged_df)}.")
        print("Please open the Excel file, type 'Yes' (case-insensitive) in the 'Select' column for organizations you want to process, and SAVE the file.")
    except Exception as e:
        print(f"Error exporting organizations to Excel: {e}")

def fetch_tenders_for_organisation(tenders_url):
    """
    Fetches all tenders for a given organisation using the provided tenders_url,
    handling basic pagination. This function extracts key tender details and their URLs.
    """
    all_tenders = []
    current_page_url = tenders_url
    base_tender_page_url = tenders_url.split('?')[0]

    column_headers = [
        'S.No', 'e-Published Date', 'Closing Date', 'Opening Date',
        'Title and Ref.No./Tender ID', 'Organisation Chain'
    ]

    while current_page_url:
        try:
            response = session.get(current_page_url)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')

            if "Stale Session" in soup.title.text or "Error" in soup.title.text:
                print(f"Detected a 'Stale Session' or error page for {current_page_url}. Attempting to re-initialize session and retry...")
                try:
                    main_org_page_url = current_page_url.split('app?')[0] + 'app?page=FrontEndTendersByOrganisation&service=page'
                    session.get(main_org_page_url) # Re-fetch a known good page to refresh session
                    print(f"Session re-initialized with {main_org_page_url}. Retrying tender fetch...")
                    response = session.get(current_page_url) # Retry the original tender URL
                    response.raise_for_status()
                    soup = BeautifulSoup(response.text, 'html.parser')
                    if "Stale Session" in soup.title.text or "Error" in soup.title.text:
                        print("Session still stale after retry. Cannot proceed with scraping from this URL.")
                        break # Give up on this URL if still stale
                except requests.exceptions.RequestException as e:
                    print(f"Failed to re-initialize session or retry tender fetch: {e}")
                    break # Give up if re-initialization fails

            # --- START Enhanced Table Finding Logic for Tenders ---
            tender_table = soup.find('table', {'id': 'table'})
            if not tender_table:
                tender_table = soup.find('table', {'class': 'list_table'})
            if not tender_table: # Broader search for any table that might contain the listing
                # Look for a table that contains 'S.No', 'e-Published Date', etc.
                header_tds = soup.find_all('td', string=lambda s: s and 'S.No' in s)
                for td in header_tds:
                    potential_table = td.find_parent('table')
                    if potential_table:
                        # Check if this table likely contains tender data by looking for other headers
                        if potential_table.find('td', string=lambda s: s and 'e-Published Date' in s):
                            tender_table = potential_table
                            break
            # --- END Enhanced Table Finding Logic for Tenders ---

            if not tender_table:
                print(f"Could not find the tender listing table on {current_page_url}. No more tenders found or parsing failed on this page.")
                break

            rows = tender_table.find_all('tr')
            header_row = None
            # Attempt to find the header row, usually has specific classes or content
            for row in rows:
                if 'class' in row.attrs and 'list_header' in row['class']:
                    header_row = row
                    break
                if row.find('td', string='S.No') or row.find('th', string='S.No'):
                    header_row = row
                    break

            if header_row:
                header_index = rows.index(header_row)
                data_rows = rows[header_index + 1:] # Data rows are after the header
            else:
                data_rows = rows # Fallback: assume all rows are data if header not found
                print("Warning: Could not reliably identify header row for tenders. Proceeding assuming all rows are data.")

            if not data_rows:
                print(f"No data rows found in the tender table on {current_page_url}.")

            # Process each data row
            for row in data_rows:
                cols = row.find_all('td')
                # Skip empty rows or rows that don't look like tender entries
                if not cols or all(c.text.strip() == '' for c in cols):
                    continue

                tender_details = {}
                for i, header in enumerate(column_headers):
                    if i < len(cols):
                        cell_element = cols[i]
                        cell_content = cell_element.text.strip()

                        if header == 'Title and Ref.No./Tender ID':
                            tender_details[header] = cell_content
                            link_tag = cell_element.find('a', href=True)
                            if link_tag:
                                tender_details['Tender URL'] = requests.compat.urljoin(current_page_url, link_tag['href'])
                        else:
                            tender_details[header] = cell_content
                    else:
                        tender_details[header] = "" # Ensure all headers have a value even if cell is missing
                
                print(f"Sr. No. {tender_details.get('S.No', 'N/A')}") # Live detail print
                all_tenders.append(tender_details)

            # Look for "Next" page link for pagination
            next_page_link = None
            next_link_element = soup.find('a', string=lambda text: text and 'Next' in text)
            if next_link_element and 'href' in next_link_element.attrs:
                next_page_link = next_link_element['href']

            if next_page_link:
                current_page_url = requests.compat.urljoin(base_tender_page_url, next_page_link)
                time.sleep(1) # Polite delay
            else:
                print("No more next pages found.")
                current_page_url = None # End loop

        except requests.exceptions.RequestException as e:
            print(f"Error fetching tenders from {current_page_url}: {e}")
            return None # Return None to indicate failure and prevent data loss
        except Exception as e:
            print(f"An unexpected error occurred while parsing tenders from {current_page_url}: {e}")
            return None # Return None to indicate failure and prevent data loss

    return all_tenders

def fetch_tender_details(tender_detail_url):
    """
    Fetches specific details of a tender from its dedicated detail page.
    This version now also extracts download links for internal use by the download function.
    Returns a dictionary of extracted details (for Excel) and a dictionary of download links.
    """
    tender_specific_details = {}
    download_links_info = {} # To store download links and their display names
    response = None

    # NEW: Retry logic for network errors
    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = session.get(tender_detail_url, timeout=30) # Increased timeout
            response.raise_for_status()
            break # Success, exit loop
        except requests.exceptions.RequestException as e:
            print(f"    Error fetching tender details from {tender_detail_url} (Attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(5) # Wait before retrying
            else:
                print(f"    Failed to fetch details for this tender after {max_retries} attempts. Skipping.")
                return {}, {} # Return empty dicts on complete failure

    try:
        soup = BeautifulSoup(response.text, 'html.parser')
        if "Stale Session" in soup.title.text or "Error" in soup.title.text:
            print(f"Stale session detected for detail page: {tender_detail_url}. Skipping detail fetch.")
            return tender_specific_details, download_links_info # Return empty if session is stale

        # Helper function to extract content by a label (e.g., "Tender ID")
        def get_detail_by_label(soup_obj, label_variations):
            for label_var in label_variations:
                caption_cells = soup_obj.find_all('td', class_='td_caption')
                for label_td in caption_cells:
                    if label_td.text and label_var.lower() in label_td.text.lower():
                        value_td = label_td.find_next_sibling('td', class_='td_field')
                        if value_td:
                            return value_td.text.strip()
                    bold_tag = label_td.find('b') # Sometimes labels are bolded
                    if bold_tag and bold_tag.text and label_var.lower() in bold_tag.text.lower():
                        value_td = label_td.find_next_sibling('td', class_='td_field')
                        if value_td:
                            return value_td.text.strip()
            return None

        # --- Extracting Standard Tender Details for Excel Export ---
        tender_specific_details["Organisation Chain"] = get_detail_by_label(soup, ["Organisation Chain"])
        
        tender_id_value = None
        id_regex_pattern = r'([A-Z0-9\-_/.]+)' # Regex to capture alphanumeric IDs with special chars
        tender_id_raw_label = get_detail_by_label(soup, ["Tender ID"])
        if tender_id_raw_label:
            match = re.search(id_regex_pattern, tender_id_raw_label, re.IGNORECASE)
            if match:
                tender_id_value = match.group(1).strip()
        tender_specific_details["Tender Id"] = tender_id_value if tender_id_value else None

        tender_title_td_label = soup.find('td', class_='td_caption', string=lambda s: s and 'Title' in s)
        if tender_title_td_label:
            tender_specific_details["Tender Title"] = tender_title_td_label.find_next_sibling('td', class_='td_field').text.strip()
        elif soup.title and soup.title.string and len(soup.title.string) < 200:
            tender_specific_details["Tender Title"] = soup.title.string.strip().replace("E-Procurement System :: ", "")

        tender_specific_details["EMD Amount"] = get_detail_by_label(soup, ["EMD Amount In ₹", "EMD Amount (in Rs.)", "EMD"])
        tender_specific_details["Tender Value"] = get_detail_by_label(soup, ["Tender Value In ₹", "Tender Value"])
        tender_specific_details["Pre-bid Meeting Date"] = get_detail_by_label(soup, ["Pre Bid Meeting Date", "Pre-Bid Meeting Date"])
        tender_specific_details["Bid Submission End Date"] = get_detail_by_label(soup, ["Bid Submission End Date"])
        tender_specific_details["Location"] = get_detail_by_label(soup, ["Location", "Work Location", "Place of Work"])
        tender_specific_details["Tender Category"] = get_detail_by_label(soup, ["Tender Category", "Tender Type"])

        # --- Download Links Extraction (for internal use by download function) ---
        # Find the 'Tender Notice' download link. This is typically 'DirectLink_8' which then reveals 'DirectLink_0'
        # Or sometimes DirectLink_0 is directly present.
        tender_notice_link_tag = soup.find('a', id=re.compile(r'DirectLink_[08]')) # Looks for DirectLink_0 or DirectLink_8
        if tender_notice_link_tag and 'href' in tender_notice_link_tag.attrs:
            download_links_info["Tender Notice Link"] = requests.compat.urljoin(tender_detail_url, tender_notice_link_tag['href'])
            # Extract filename from the link or a nearby text if possible
            parsed_url = urlparse(download_links_info["Tender Notice Link"])
            download_links_info["Tender Notice Display Name"] = os.path.basename(parsed_url.path) or f"Tendernotice_{tender_id_value}.pdf"
        else:
            download_links_info["Tender Notice Link"] = None
            download_links_info["Tender Notice Display Name"] = None

        # Find the 'Zip File' download link (often 'DirectLink_7')
        zip_file_link_tag = soup.find('a', id='DirectLink_7')
        if zip_file_link_tag and 'href' in zip_file_link_tag.attrs:
            download_links_info["Zip File Link"] = requests.compat.urljoin(tender_detail_url, zip_file_link_tag['href'])
        else:
            download_links_info["Zip File Link"] = None
        
    except requests.exceptions.RequestException as e:
        print(f"Error fetching tender details from {tender_detail_url}: {e}")
    except Exception as e:
        print(f"An unexpected error occurred while parsing tender details from {tender_detail_url}: {e}")

    return tender_specific_details, download_links_info # Return both dictionaries

def solve_captcha_with_gemini(image_data, max_retries=3):
    """
    Sends a CAPTCHA image to Google Gemini Vision for OCR and returns the text.
    Retries internally if the result is not 6 characters long.
    """
    if not GEMINI_AVAILABLE:
        print("  Error: 'google.generativeai' or 'Pillow' not installed. Cannot use auto-solver.")
        print("  Install them with: pip install google-generativeai pillow")
        return None

    if GOOGLE_API_KEY == "YOUR_GOOGLE_AI_API_KEY" or not model:
        print("  Warning: GOOGLE_API_KEY not set. Cannot solve CAPTCHA automatically.")
        return None
    try:
        img = Image.open(io.BytesIO(image_data))
    except Exception as e:
        print(f"  Error initializing Gemini or loading image: {e}")
        return None

    for attempt in range(max_retries):
        try:
            prompt = "Perform OCR on this image. The image contains a CAPTCHA that is exactly 6 characters long. Return only the 6 alphanumeric characters you see, with no other text, explanation, or formatting."
            
            # Send the image and prompt to the model
            response = model.generate_content([prompt, img], generation_config={"temperature": 0})
            
            # Clean up the response to get just the text
            captcha_text = response.text.strip()
            # Further cleaning to remove any potential markdown or extra words from the model
            captcha_text = re.sub(r'[^a-zA-Z0-9]', '', captcha_text)

            if captcha_text and len(captcha_text) == 6:
                print(f"  Gemini OCR result (valid): '{captcha_text}'")
                return captcha_text
            elif captcha_text:
                print(f"  Gemini returned an invalid length result: '{captcha_text}' (length {len(captcha_text)}). Retrying... ({attempt + 1}/{max_retries})")
            else:
                print(f"  Gemini returned an empty result. Retrying... ({attempt + 1}/{max_retries})")

        except Exception as e:
            print(f"  An error occurred while calling Gemini API on attempt {attempt + 1}: {e}")
        
        if attempt < max_retries - 1:
            time.sleep(1) # Wait a second before retrying the API call

    print(f"  Auto-solve failed after {max_retries} attempts.")
    return None


def handle_captcha_interaction(driver, context_for_log, submit_button_id="Submit", success_element_id=None):
    """
    Handles the CAPTCHA interaction on a page.
    Assumes CAPTCHA elements have IDs: 'captchaImage', 'captchaText'.
    Returns True on success, False on failure/skip.
    This function can be called for different contexts (e.g., downloads, status checks).
    :param submit_button_id: The ID of the submit button (e.g., "Submit" or "Search").
    :param success_element_id: Optional ID of an element that appears on success (e.g., a results table).
    """
    global captcha_solved_in_session
    if captcha_solved_in_session:
        print(f"  CAPTCHA previously solved. Skipping interaction for {context_for_log}.")
        return True
    
    # Step 1: Check if a CAPTCHA is even present on the page.
    # We wait for the captcha image. If it doesn't appear, we assume no captcha is needed.
    try:
        captcha_image_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "captchaImage")))
        print(f"\nCAPTCHA detected for {context_for_log}.")
    except TimeoutException:
        # No CAPTCHA image found after 10 seconds, assume we can proceed.
        print(f"  No CAPTCHA detected for {context_for_log}. Proceeding.")
        return True
    except Exception as e:
        print(f"  An error occurred while checking for CAPTCHA: {e}")
        return False

    # Step 2: If we are here, CAPTCHA is present. Try to solve it.
    max_retries = 3
    for attempt in range(max_retries):
        try:
            # Re-find elements in each attempt to avoid stale references
            captcha_text_input = driver.find_element(By.ID, "captchaText")
            submit_button = driver.find_element(By.ID, submit_button_id)

            # --- NEW: Capture and attempt to auto-solve ---
            image_data_for_ocr = None
            try:
                # Get image data in memory
                image_data_for_ocr = captcha_image_element.screenshot_as_png
            except Exception as e:
                print(f"  Warning: Could not capture CAPTCHA image: {e}")

            # --- Auto-solve logic with manual fallback ---
            captcha_solution = None
            # Always attempt to auto-solve if image data was captured.
            # The solver function itself will handle checks for API key and libraries.
            if image_data_for_ocr:
                print(f"  Attempting to auto-solve CAPTCHA (attempt {attempt + 1}/{max_retries})...")
                captcha_solution = solve_captcha_with_gemini(image_data_for_ocr)
            
            if not captcha_solution:
                # Fallback to manual input. Loop until input is provided.
                while not captcha_solution:
                    captcha_solution = input(f"Please enter CAPTCHA manually (attempt {attempt + 1}/{max_retries}, input is required): ").strip()
                    if not captcha_solution:
                        print("  CAPTCHA input is mandatory. Please try again.")
            
            captcha_text_input.clear()
            captcha_text_input.send_keys(captcha_solution)
            submit_button.click()
            time.sleep(5) # Wait for the page to process the CAPTCHA

            # --- NEW SUCCESS CHECK ---
            success = False
            if success_element_id:
                try:
                    # Check if the success element (e.g., the results table) is now present
                    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, success_element_id)))
                    success = True
                    print(f"  Success condition met: Element with ID '{success_element_id}' found.")
                except TimeoutException:
                    success = False # Element did not appear, CAPTCHA likely failed
            else:
                # Original logic: check if captcha image is gone
                if not driver.find_elements(By.ID, 'captchaImage'):
                    success = True
            
            if success:
                print(f"  CAPTCHA entered successfully for {context_for_log}!")
                captcha_solved_in_session = True
                return True
            else:
                print(f"  Incorrect CAPTCHA or page did not update. Retrying...")
                # The page might have reloaded with a new CAPTCHA image. We need to find it again.
                captcha_image_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "captchaImage")))
        except (NoSuchElementException, StaleElementReferenceException) as e:
            print(f"  CAPTCHA elements not found on attempt {attempt + 1}, might have been solved or page changed. Error: {e}")
            # Let's re-check if the image is gone. If so, it was a success.
            if not driver.find_elements(By.ID, 'captchaImage'):
                print("  CAPTCHA seems to be gone. Assuming success.")
                captcha_solved_in_session = True
                return True
            continue # Loop to try again
        except Exception as e:
            print(f"  An unexpected error occurred during CAPTCHA attempt: {e}")
    
    print(f"  Failed to pass CAPTCHA after {max_retries} attempts. Falling back to required manual input.")

    # Loop indefinitely for manual input until success or user aborts.
    manual_attempt_count = 0
    while True:
        manual_attempt_count += 1
        try:
            # Re-find elements to avoid stale references
            captcha_image_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "captchaImage")))
            captcha_text_input = driver.find_element(By.ID, "captchaText")
            submit_button = driver.find_element(By.ID, submit_button_id)

            # Ask for manual input
            captcha_solution = ""
            while not captcha_solution:
                captcha_solution = input(f"Please enter CAPTCHA manually (attempt {manual_attempt_count}, input is required): ").strip()
                if not captcha_solution:
                    print("  CAPTCHA input is mandatory. Please try again.")
            
            captcha_text_input.clear()
            captcha_text_input.send_keys(captcha_solution)
            submit_button.click()
            time.sleep(5) # Wait for the page to process the CAPTCHA

            # --- NEW SUCCESS CHECK (for manual loop) ---
            success = False
            if success_element_id:
                try:
                    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, success_element_id)))
                    success = True
                    print(f"  Success condition met: Element with ID '{success_element_id}' found.")
                except TimeoutException:
                    success = False
            else:
                if not driver.find_elements(By.ID, 'captchaImage'):
                    success = True

            if success:
                print(f"  CAPTCHA entered successfully for {context_for_log}!")
                captcha_solved_in_session = True
                return True
            else:
                print(f"  Incorrect CAPTCHA. Please try again...")
        except (NoSuchElementException, StaleElementReferenceException):
            if not driver.find_elements(By.ID, 'captchaImage'):
                print("  CAPTCHA seems to be gone. Assuming success.")
                captcha_solved_in_session = True
                return True
            print(f"  CAPTCHA elements not found on manual attempt {manual_attempt_count}. Retrying.")
            time.sleep(2) # Wait a bit before retrying the loop
        except Exception as e:
            print(f"  An unexpected error occurred during manual CAPTCHA entry: {e}")
            print("  Skipping task due to this error.")
            return False

def is_tender_available(tender_url):
    try:
        response = session.get(tender_url, timeout=10)
        return response.status_code == 200
    except requests.exceptions.RequestException:
        return False
    except Exception:
        return False

# --- Selenium Driver Initialization for Downloads ---
def initialize_selenium_driver():
    """
    Initializes and returns a Selenium Firefox WebDriver instance.
    """
    options = FirefoxOptions()
    # options.add_argument("--headless") # Keep headless commented as captcha requires interaction
    
    service = FirefoxService(GeckoDriverManager().install())
    driver = webdriver.Firefox(service=service, options=options)
    return driver

# --- NEW: Helper for Requests-based Download with Cookies (with progress) ---
def download_file_with_requests_session(url, file_path, cookies):
    """
    Downloads a file using requests.Session, populating it with cookies
    obtained from a Selenium driver session. Includes download progress.
    """
    try:
        req_session = requests.Session()
        for cookie in cookies:
            req_session.cookies.set(cookie['name'], cookie['value'], domain=cookie.get('domain'), path=cookie.get('path'))
        
        req_session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36'
        })

        response = req_session.get(url, stream=True, timeout=120)
        response.raise_for_status()

        total_size = int(response.headers.get('content-length', 0))
        
        with open(file_path, 'wb') as f:
            downloaded_size = 0
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    downloaded_size += len(chunk)
                    if total_size > 0:
                        percent = (downloaded_size / total_size) * 100
                        print(f"  Downloading {os.path.basename(file_path)}: {percent:.2f}%", end='\r')
        
        print(f"\n  Successfully downloaded {os.path.basename(file_path)}.")
        return True
    except requests.exceptions.RequestException as e:
        print(f"\n  Requests download failed for {os.path.basename(file_path)}: {e}")
        return False
    except Exception as e:
        print(f"\n  Unexpected error during requests download for {os.path.basename(file_path)}: {e}")
        return False

def download_all_tender_files(driver, tender_info, base_download_directory, download_mode='full'):
    """
    Manages the download process for all available files for a single tender using a hybrid
    Selenium (for interaction) and Requests (for download) approach.
    'download_mode' can be 'full' or 'update'. 'update' skips notice/zip files.
    """
    tender_id = str(tender_info.get('Tender Id', 'UNKNOWN'))
    tender_title = tender_info.get('Tender Title', 'Untitled')
    tender_detail_page_url = tender_info.get('Tender URL')

    print(f"\nProcessing Tender for downloads: ID: {tender_id}")

    if not tender_detail_page_url:
        print(f"  Skipping downloads for Tender ID {tender_id}: No Tender URL available.")
        return False

    tender_download_folder = os.path.join(base_download_directory, str(tender_id))
    os.makedirs(tender_download_folder, exist_ok=True)

    try:
        wait = WebDriverWait(driver, 30)

        if download_mode == 'full':
            # --- Download Tender Notice (PDF) ---
            notice_filename = f"Tendernotice_{tender_id}.pdf"
            final_notice_file_path = os.path.join(tender_download_folder, notice_filename)
            if not os.path.exists(final_notice_file_path):
                print("  Attempting to download Tender Notice (PDF)...")
                driver.get(tender_detail_page_url) # Re-navigate to the page to ensure a clean state
                time.sleep(3)
                if "Stale Session" in driver.title or "Error" in driver.title:
                    print(f"  Stale Session/Error page detected for {tender_id}. Skipping downloads.")
                    return False
                
                try:
                    # If captcha has been solved, we can directly look for DirectLink_0
                    if captcha_solved_in_session:
                        final_link_element = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_0")))
                        final_url = final_link_element.get_attribute("href")
                        if final_url:
                            # print(f"  Final download URL for Tender Notice: {final_url}")
                            selenium_cookies = driver.get_cookies()
                            download_file_with_requests_session(final_url, final_notice_file_path, selenium_cookies)
                        else:
                            print("  Final link 'DirectLink_0' found but its 'href' is missing.")
                    else:
                        try:
                            trigger_element = wait.until(EC.element_to_be_clickable((By.ID, "docDownload")))
                        except TimeoutException:
                            trigger_element = wait.until(EC.element_to_be_clickable((By.ID, "DirectLink_8")))
                        
                        driver.execute_script("arguments[0].click();", trigger_element)
                        
                        if handle_captcha_interaction(driver, "Tender Notice"):
                            try:
                                final_link_element = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_0")))
                                final_url = final_link_element.get_attribute("href")
                                if final_url:
                                    # print(f"  Final download URL for Tender Notice: {final_url}")
                                    selenium_cookies = driver.get_cookies()
                                    download_file_with_requests_session(final_url, final_notice_file_path, selenium_cookies)
                                else:
                                    print("  Final link 'DirectLink_0' found but its 'href' is missing.")
                            except TimeoutException:
                                print("  Could not find final download link 'DirectLink_0' for Tender Notice after CAPTCHA.")
                except Exception as e:
                    print(f"  An error occurred trying to download the Tender Notice: {e}")
            else:
                print(f"  Skipping Tender Notice: '{notice_filename}' already exists.")

            # --- Download Zip File ---
            zip_filename = f"{tender_id}.zip"
            final_zip_file_path = os.path.join(tender_download_folder, zip_filename)
            if not os.path.exists(final_zip_file_path):
                print("  Attempting to download Zip File...")
                driver.get(tender_detail_page_url) # Re-navigate for a clean state
                time.sleep(3)
                if "Stale Session" in driver.title or "Error" in driver.title:
                    print(f"  Stale Session/Error page detected for {tender_id}. Skipping downloads.")
                    return False
                try:
                    # More robust locator for the Zip file link.
                    # It might be identified by text or a specific ID.
                    zip_url = None
                    try:
                        # Strategy 1: Find by partial link text.
                        trigger_element = wait.until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Download as zip file")))
                        zip_url = trigger_element.get_attribute("href")
                    except TimeoutException:
                        # Strategy 2: Fallback to the original ID-based method.
                        print("    - Could not find 'Download as Zip' link by text, trying by ID 'DirectLink_8'...")
                        try:
                            trigger_element = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_8")))
                            zip_url = trigger_element.get_attribute("href")
                        except TimeoutException:
                            print("    - Could not find Zip download link by ID either. Skipping Zip file.")

                    if zip_url:
                        selenium_cookies = driver.get_cookies()
                        download_file_with_requests_session(zip_url, final_zip_file_path, selenium_cookies)
                    else:
                        print("  Could not find a URL for the Zip File.")
                except Exception as e:
                    print(f"  An error occurred trying to download the Zip File: {e}")
            else:
                print(f"  Skipping Zip File: '{zip_filename}' already exists.")
        else:
            print("  Update mode: Skipping Tender Notice and Zip file download.")

        # --- Download Pre-Bid Meeting File ---
        prebid_filename = f"PreBid_Meeting_{tender_id}.pdf"
        final_prebid_file_path = os.path.join(tender_download_folder, prebid_filename)
        if not os.path.exists(final_prebid_file_path):
            print("  Checking for Pre-Bid Meeting File...")
            driver.get(tender_detail_page_url) # Re-navigate for a clean state
            time.sleep(3)
            if "Stale Session" in driver.title or "Error" in driver.title:
                print(f"  Stale Session/Error page detected for {tender_id}. Skipping downloads.")
                return
                return False
            selenium_cookies = driver.get_cookies()
            try:
                # The link for pre-bid is typically DirectLink_2
                trigger_element = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_2")))
                prebid_url = trigger_element.get_attribute("href")
                
                if prebid_url:
                    # print(f"  Found URL for Pre-Bid Meeting File: {prebid_url}")
                    download_file_with_requests_session(prebid_url, final_prebid_file_path, selenium_cookies)
                else:
                    print("  Could not find a URL on the Pre-Bid Meeting File trigger link.")
            except TimeoutException:
                print("  No Pre-Bid Meeting File found for this tender.")
            except Exception as e:
                print(f"  An error occurred trying to download the Pre-Bid Meeting File: {e}")
        else:
            print(f"  Skipping Pre-Bid Meeting File: '{prebid_filename}' already exists.")

        # --- Download Corrigendum Files ---
        print("  Checking for Corrigendum Files...")
        driver.get(tender_detail_page_url) # Re-navigate for a clean state
        time.sleep(3)
        if "Stale Session" in driver.title or "Error" in driver.title:
            print(f"  Stale Session/Error page detected for {tender_id}. Skipping downloads.")
            return
            return False
        selenium_cookies = driver.get_cookies()
        try:
            corrigendum_links = driver.find_elements(By.XPATH, "//a[contains(@id, 'DirectLink_')]")
            for link in corrigendum_links:
                if "View Corrigendum History" in link.get_attribute("title"):
                    main_window = driver.current_window_handle
                    # Using JavaScript to click the link. This is more robust and avoids
                    # "ElementClickInterceptedError" which happens when another element
                    # (like a header, footer, or overlay) is physically covering the
                    # target element, preventing a standard click.
                    driver.execute_script("arguments[0].click();", link)
                    wait.until(EC.number_of_windows_to_be(2))
                    
                    for window_handle in driver.window_handles:
                        if window_handle != main_window:
                            driver.switch_to.window(window_handle)
                            break
                    
                    corrigendum_docs = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//a[contains(@id, 'DirectLink_')]")))
                    for doc in corrigendum_docs:
                        doc_url = doc.get_attribute("href")
                        doc_name = doc.text.strip() # Use strip() to remove leading/trailing whitespace
                        if doc_url and doc_name: # Check that both URL and a valid name exist
                            # print(f"    Found Corrigendum: {doc_name}")
                            file_path = os.path.join(tender_download_folder, doc_name)
                            if not os.path.exists(file_path):
                                download_file_with_requests_session(doc_url, file_path, driver.get_cookies())
                            else:
                                print(f"    Skipping existing Corrigendum: {doc_name}")
                        else:
                            # This handles cases where a link is found but has no text, preventing an error. We just skip it silently.
                            pass

                    driver.close()
                    driver.switch_to.window(main_window)

        except TimeoutException:
            print("  No Corrigendum Files found for this tender.")
        except Exception as e:
            print(f"  An error occurred trying to download Corrigendum Files: {e}")

    except Exception as e:
        print(f"An unexpected error occurred during downloads for Tender ID {tender_id}: {e}")
        return False

    return True

def get_hyperlinks_from_sheet(filename, sheet_name): # Kept for one-time migration
    """Reads an Excel sheet and returns a dict mapping Tender ID to its hyperlink target."""
    hyperlink_map = {}
    if not os.path.exists(filename):
        return hyperlink_map
    try:
        wb = openpyxl.load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return hyperlink_map
        ws = wb[sheet_name]
        
        header = {cell.value: cell.column for cell in ws[1]}
        if 'Tender Id' not in header:
            return hyperlink_map
        
        tender_id_col = header['Tender Id']
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=tender_id_col)
            if cell.hyperlink and cell.hyperlink.target:
                # Ensure key is a string for consistent matching
                hyperlink_map[str(cell.value)] = cell.hyperlink.target
        wb.close()
    except Exception as e:
        print(f"Warning: Could not read hyperlinks from '{sheet_name}': {e}")
    return hyperlink_map

def process_and_export_tenders(website_choice, filename=excel_filename):
    """
    Reads selected organizations from all relevant org sheets, fetches their tenders.
    Active tenders are saved to their respective website-specific sheets.
    Expired tenders are moved to a dedicated 'Expired Tenders' sheet.
    This function handles loading global organization data, preventing duplicates,
    and formatting the Excel output, including hyperlinking tender titles.
    It can be limited to a specific website or run for all.
    """
    global global_organisations_data # Declare intent to use global variable

    # Ensure global_organisations_data is loaded from the file before proceeding
    if not global_organisations_data:
        if not load_organisations_data():
            print("Error: Organization data (global_organizations_data.json) not found or could not be loaded. Please ensure you've run an 'Export Organizations' option (1, 2, or 3) at least once.")
            return

    # Determine which organization sheets to process based on user's website choice
    if website_choice in website_options.keys():
        org_sheet_names_to_process = [f"{website_options[website_choice]['name']}-O"]
    else: # All websites
        org_sheet_names_to_process = [f"{site_info['name']}-O" for site_info in website_options.values()]
    
    # Load existing expired tenders once at the start to avoid re-adding duplicates
    df_expired_all = read_sheet_from_csv(EXPIRED_TENDERS_SHEET)
    existing_expired_ids = set(df_expired_all['Tender Id'].astype(str)) if not df_expired_all.empty and 'Tender Id' in df_expired_all.columns else set()

    for current_org_sheet_name in org_sheet_names_to_process:
        current_tender_sheet_name = current_org_sheet_name.replace('-O', '') # e.g., "MahaTenders"

        # --- 1. Sync all user input from Excel to CSV database ---
        print(f"\n--- Syncing user selections from Excel for '{current_org_sheet_name}' and '{current_tender_sheet_name}' ---")
        
        # Sync org selection sheet first, as it's required for the scrape
        try:
            # Sync org selection sheet
            df_orgs_from_excel = pd.read_excel(filename, sheet_name=current_org_sheet_name)
            write_sheet_to_csv(df_orgs_from_excel, current_org_sheet_name)
            print(f"Successfully synced selections from '{current_org_sheet_name}'.")
        except (FileNotFoundError, ValueError) as e:
            print(f"Warning: Could not sync org sheet '{current_org_sheet_name}' from Excel: {e}. Will proceed with existing CSV data if available.")
        except Exception as e:
            print(f"An unexpected error occurred during org sheet sync: {e}")

        # Sync tender data sheet (to capture 'Download' selections and folder hyperlinks for preservation)
        try:
            # Sync tender data sheet (to capture 'Download' selections and folder hyperlinks)
            df_tenders_from_excel = pd.read_excel(filename, sheet_name=current_tender_sheet_name)
            hyperlink_map_from_excel = get_hyperlinks_from_sheet(filename, current_tender_sheet_name)
            if hyperlink_map_from_excel:
                # Ensure column exists before assigning
                if '__folder_hyperlink__' not in df_tenders_from_excel.columns:
                    df_tenders_from_excel['__folder_hyperlink__'] = None
                df_tenders_from_excel['__folder_hyperlink__'] = df_tenders_from_excel['Tender Id'].astype(str).map(hyperlink_map_from_excel).fillna(df_tenders_from_excel['__folder_hyperlink__'])

            write_sheet_to_csv(df_tenders_from_excel, current_tender_sheet_name)
            print(f"Successfully synced selections from '{current_tender_sheet_name}'.")
        except (FileNotFoundError, ValueError):
            # This is not an error if the sheet doesn't exist yet, just a warning.
            print(f"Info: Tender sheet '{current_tender_sheet_name}' not found in Excel. This is normal on first run. It will be created.")
        except Exception as e:
            print(f"An unexpected error occurred during tender sheet sync: {e}")

        # --- 2. Load data and prepare for scrape ---
        df_orgs = read_sheet_from_csv(current_org_sheet_name)
        selected_orgs = df_orgs[df_orgs['Select (Type "Yes" to select)'].astype(str).str.lower() == 'yes']
        if selected_orgs.empty:
            print(f"No organizations selected in sheet '{current_org_sheet_name}'. Skipping tender fetching for this sheet.")
            continue

        # This is our "old" data, the single source of truth for this run
        df_old_all_tenders = read_sheet_from_csv(current_tender_sheet_name)

        # --- 3. Scrape new data ---
        tenders_scraped_in_current_run_for_sheet = []
        failed_orgs = set() # Track orgs that failed to scrape
        for index, row in selected_orgs.iterrows():
            org_name_from_excel = row['Organisation Name'].strip() # Ensure no leading/trailing spaces
            
            # Debug: Check if organization exists in global_organisations_data and what its Tenders URL is
            org_data = global_organisations_data.get(org_name_from_excel, {})
            tenders_url = org_data.get('Tenders URL')

            if tenders_url:
                print(f"\nFetching tenders for '{org_name_from_excel}' from website: {current_tender_sheet_name}...")
                tenders_from_listing = fetch_tenders_for_organisation(tenders_url)
                
                if tenders_from_listing is None:
                    print(f"    Error fetching tenders for '{org_name_from_excel}'. Preserving existing data.")
                    failed_orgs.add(org_name_from_excel)
                    continue

                if tenders_from_listing:
                    for tender_listing_data in tenders_from_listing:
                        tender_detail_url = tender_listing_data.get('Tender URL')
                        
                        if not tender_detail_url:
                            print(f"    Skipping tender due to missing detail URL: {tender_listing_data.get('Title and Ref.No./Tender ID', 'N/A')}")
                            continue

                        # Fetch full details (excluding download links for Excel)
                        details_from_detail_page, _ = fetch_tender_details(tender_detail_url)
                        
                        # Extract tender ID (from details page if possible, fallback to listing)
                        tender_id_value = details_from_detail_page.get('Tender Id')
                        if not tender_id_value and tender_listing_data.get('Title and Ref.No./Tender ID'):
                            listing_title = tender_listing_data['Title and Ref.No./Tender ID']
                            id_regex_pattern = r'([A-Z0-9\-_/.]+)'
                            match = re.search(r'\[(' + id_regex_pattern[1:-1] + r')\]', listing_title, re.IGNORECASE)
                            if match:
                                tender_id_value = match.group(1).strip()
                            else:
                                match_direct_in_title = re.search(id_regex_pattern, listing_title, re.IGNORECASE)
                                if match_direct_in_title:
                                    tender_id_value = match_direct_in_title.group(1).strip()
                        if not tender_id_value:
                            tender_id_value = 'N/A' # Fallback for Tender ID if not found anywhere

                        # Construct base tender details from listing and detail page
                        # MODIFIED: Keep original string format for dates and times by splitting the scraped string.
                        # This prevents reformatting to a different date/time standard.
                        closing_date_time_str = tender_listing_data.get('Closing Date', 'N/A')
                        closing_date = 'N/A'
                        closing_time = 'N/A'
                        if closing_date_time_str and ' ' in closing_date_time_str:
                            parts = closing_date_time_str.split(' ', 1)
                            closing_date = parts[0]
                            if len(parts) > 1:
                                closing_time = parts[1]
                        else:
                            closing_date = closing_date_time_str

                        full_tender_info = {
                            'Sr. No.': pd.to_numeric(tender_listing_data.get('S.No'), errors='coerce'),
                            'Tender Title': details_from_detail_page.get('Tender Title') if details_from_detail_page and details_from_detail_page.get('Tender Title') else tender_listing_data.get('Title and Ref.No./Tender ID', 'N/A'),
                            'Tender Value': clean_currency_value(details_from_detail_page.get('Tender Value')) if details_from_detail_page else None,
                            'EMD': clean_currency_value(details_from_detail_page.get('EMD Amount')) if details_from_detail_page else None,
                            'Organization Chain': org_name_from_excel,
                            'Closing Date': closing_date,
                            'Closing Time': closing_time,
                            'Pre bid meeting date': details_from_detail_page.get('Pre-bid Meeting Date') if details_from_detail_page else None,
                            'Location': details_from_detail_page.get('Location', 'N/A (Not Scraped)') if details_from_detail_page else 'N/A (Not Scraped)',
                            'Tender Id': tender_id_value,
                            'Tender URL': tender_detail_url,
                            'Tender Category': details_from_detail_page.get('Tender Category', 'N/A')
                        }
                        full_tender_info['__folder_hyperlink__'] = None # Placeholder for folder link
                        
                        tenders_scraped_in_current_run_for_sheet.append(full_tender_info)

                else:
                    print(f"No tenders found from listing for {org_name_from_excel} or error fetching tenders.")
            else:
                print(f"No tender list URL found for organization {org_name_from_excel} in in-memory data. Skipping.")

        df_new_scraped_tenders = pd.DataFrame(tenders_scraped_in_current_run_for_sheet)

        # --- 4. Process and Merge ---
        if df_old_all_tenders.empty and df_new_scraped_tenders.empty:
            print(f"No existing tenders and no new tenders found for '{current_tender_sheet_name}'. Sheet is empty.")
            write_sheet_to_csv(pd.DataFrame([]), current_tender_sheet_name)
            update_excel_sheet(pd.DataFrame([]), current_tender_sheet_name, filename)
            continue

        # Ensure Tender ID is string for consistent matching
        if not df_old_all_tenders.empty:
            df_old_all_tenders['Tender Id'] = df_old_all_tenders['Tender Id'].astype(str)
        if not df_new_scraped_tenders.empty:
            df_new_scraped_tenders['Tender Id'] = df_new_scraped_tenders['Tender Id'].astype(str)

        # Split old data into "to be processed" and "to be preserved"
        selected_org_names = set(selected_orgs['Organisation Name'].str.strip())
        selected_org_names -= failed_orgs # Exclude failed orgs so they are preserved
        df_old_preserved = pd.DataFrame()
        df_old_to_process = pd.DataFrame()

        if not df_old_all_tenders.empty and 'Organization Chain' in df_old_all_tenders.columns and selected_org_names:
            df_old_all_tenders.dropna(subset=['Organization Chain'], inplace=True)
            pattern = '|'.join(re.escape(name) for name in selected_org_names)
            is_from_selected_org_mask = df_old_all_tenders['Organization Chain'].str.contains(pattern, regex=True, na=False)

            df_old_to_process = df_old_all_tenders[is_from_selected_org_mask]
            df_old_preserved = df_old_all_tenders[~is_from_selected_org_mask]
            print(f"Found {len(df_old_to_process)} existing tenders from selected orgs. Preserving {len(df_old_preserved)} tenders from unselected orgs.")
        else:
            df_old_to_process = df_old_all_tenders

        # --- 5. Identify and move expired tenders ---
        old_ids_to_process = set(df_old_to_process['Tender Id']) if not df_old_to_process.empty else set()
        new_scraped_ids = set(df_new_scraped_tenders['Tender Id']) if not df_new_scraped_tenders.empty else set()
        
        expired_ids_this_run = old_ids_to_process - new_scraped_ids # Tenders that were in the old list for selected orgs, but not in the new scrape.
        expired_ids_to_move = expired_ids_this_run - existing_expired_ids

        if expired_ids_to_move:
            df_to_move_to_expired = df_old_to_process[df_old_to_process['Tender Id'].isin(expired_ids_to_move)].copy()
            # Add the source website name to the expired tender record
            df_to_move_to_expired['Source Website'] = current_tender_sheet_name
            print(f"Found {len(df_to_move_to_expired)} expired/removed tenders to move to '{EXPIRED_TENDERS_SHEET}'.")
            df_expired_all = pd.concat([df_expired_all, df_to_move_to_expired], ignore_index=True)
            df_expired_all.drop_duplicates(subset=['Tender Id'], keep='last', inplace=True)
            existing_expired_ids.update(expired_ids_to_move)

        # --- 6. Merge new data with old, preserving user selections ---
        df_updated_for_selected_orgs = df_new_scraped_tenders
        
        if not df_updated_for_selected_orgs.empty:
            download_status_map, hyperlink_map, bookmark_status_map = {}, {}, {}
            if not df_old_all_tenders.empty:
                temp_df = df_old_all_tenders.drop_duplicates(subset=['Tender Id'], keep='first')
                if 'Download (Type "Yes" to download)' in temp_df.columns:
                    download_status_map = temp_df.set_index('Tender Id')['Download (Type "Yes" to download)'].to_dict()
                # NEW: Preserve bookmark status
                if 'Bookmark' in temp_df.columns:
                    bookmark_status_map = temp_df.set_index('Tender Id')['Bookmark'].to_dict()
                if '__folder_hyperlink__' in temp_df.columns:
                    hyperlink_map = temp_df.set_index('Tender Id')['__folder_hyperlink__'].to_dict()

            df_updated_for_selected_orgs['Download (Type "Yes" to download)'] = df_updated_for_selected_orgs['Tender Id'].map(download_status_map).fillna('')
            df_updated_for_selected_orgs['Bookmark'] = df_updated_for_selected_orgs['Tender Id'].map(bookmark_status_map).fillna('')
            df_updated_for_selected_orgs['__folder_hyperlink__'] = df_updated_for_selected_orgs['Tender Id'].map(hyperlink_map)

            print(f"Processed {len(df_updated_for_selected_orgs)} active tenders for selected organizations.")

        # --- 7. Combine and Finalize ---
        dfs_to_concat = [df for df in [df_old_preserved, df_updated_for_selected_orgs] if not df.empty]
        df_final_for_sheet = pd.concat(dfs_to_concat, ignore_index=True) if dfs_to_concat else pd.DataFrame()

        # --- 8. Write to files ---
        if not df_final_for_sheet.empty:
            if 'Work Description' in df_final_for_sheet.columns:
                df_final_for_sheet = df_final_for_sheet.drop(columns=['Work Description'])

            # Ensure numeric columns are of a numeric type before saving/writing.
            # This guarantees that Excel stores them as numbers, allowing for correct formatting.
            if 'Sr. No.' in df_final_for_sheet.columns:
                df_final_for_sheet['Sr. No.'] = pd.to_numeric(df_final_for_sheet['Sr. No.'], errors='coerce')
            if 'Tender Value' in df_final_for_sheet.columns:
                df_final_for_sheet['Tender Value'] = pd.to_numeric(df_final_for_sheet['Tender Value'], errors='coerce')
            if 'EMD' in df_final_for_sheet.columns:
                df_final_for_sheet['EMD'] = pd.to_numeric(df_final_for_sheet['EMD'], errors='coerce')

            if 'Organization Chain' in df_final_for_sheet.columns and 'Sr. No.' in df_final_for_sheet.columns:
                df_final_for_sheet.sort_values(by=['Organization Chain', 'Sr. No.'], ascending=[True, True], inplace=True, na_position='last', ignore_index=True)

            # Define the final order of columns for active tenders
            final_tender_cols_order = [
                'Sr. No.', 'Tender Title', 'Tender Value', 'EMD', 'Organization Chain',
                'Closing Date', 'Closing Time', 'Pre bid meeting date', 'Location', 'Tender Id', 'Tender URL',
                'Tender Category', 'Download (Type "Yes" to download)', 'Bookmark', '__folder_hyperlink__'
            ]
            existing_ordered_cols = [col for col in final_tender_cols_order if col in df_final_for_sheet.columns]
            df_final_for_sheet = df_final_for_sheet.reindex(columns=final_tender_cols_order).dropna(axis=1, how='all')
            
            write_sheet_to_csv(df_final_for_sheet, current_tender_sheet_name)
            update_excel_sheet(df_final_for_sheet, current_tender_sheet_name, filename)
        else:
            print(f"\nNo active tenders remain for '{current_tender_sheet_name}'. Clearing sheet.")
            write_sheet_to_csv(pd.DataFrame([]), current_tender_sheet_name)
            update_excel_sheet(pd.DataFrame([]), current_tender_sheet_name, filename)

    # --- After the loop, update the Expired Tenders sheet once with all collected data ---
    if not df_expired_all.empty:
        print(f"\nUpdating '{EXPIRED_TENDERS_SHEET}' sheet with all expired tenders...")

        # Ensure numeric columns are of a numeric type before saving/writing.
        if 'Tender Value' in df_expired_all.columns:
            df_expired_all['Tender Value'] = pd.to_numeric(df_expired_all['Tender Value'], errors='coerce')
        if 'EMD' in df_expired_all.columns:
            df_expired_all['EMD'] = pd.to_numeric(df_expired_all['EMD'], errors='coerce')

        df_expired_all.sort_values(by=['Closing Date', 'Organization Chain'], ascending=[False, True], inplace=True, na_position='last', ignore_index=True)
        
        write_sheet_to_csv(df_expired_all, EXPIRED_TENDERS_SHEET)
        update_excel_sheet(df_expired_all, EXPIRED_TENDERS_SHEET, filename)

def write_and_format_focus_sheet(writer, df, sheet_name):
    """
    A dedicated, safe function to write and format a sheet within the 'Tenders to Focus.xlsx' file.
    It avoids hiding columns to prevent the 'IndexError: At least one sheet must be visible'.
    """
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]

    # Apply basic formatting without hiding columns
    col_indices = {col_name: i + 1 for i, col_name in enumerate(df.columns)}

    # Apply hyperlinks to Tender Title
    if 'Tender Title' in col_indices and 'Tender URL' in col_indices:
        title_col_idx = col_indices['Tender Title']
        for row_idx, row_data in df.iterrows():
            cell = worksheet.cell(row=row_idx + 2, column=title_col_idx)
            url = row_data.get('Tender URL')
            if cell.value and pd.notna(url):
                cell.hyperlink = url
                cell.font = Font(color="0000FF", underline="single")
    
    # Define columns that need text wrapping
    text_wrap_columns = ['Tender Title', 'Organization Chain', 'Location']

    # Auto-fit columns for better readability
    for col_name, col_idx in col_indices.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        # Hide internal columns that are not needed in the final report
        if col_name in ['Tender URL', '__folder_hyperlink__']:
            worksheet.column_dimensions[col_letter].hidden = True
            continue
        
        # Set specific widths for some columns
        if col_name == 'Sr. No.':
            worksheet.column_dimensions[col_letter].width = 5
        elif col_name == 'Tender Title':
            worksheet.column_dimensions[col_letter].width = 45
        else: # Auto-fit for other columns
            max_len_val = df[col_name].astype(str).map(len).max()
            max_length = max(max_len_val if pd.notna(max_len_val) else 0, len(col_name))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[col_letter].width = min(max(adjusted_width, 12), 60)

        # Apply cell formatting
        for row_idx_in_sheet in range(1, df.shape[0] + 2):
            cell = worksheet[f'{col_letter}{row_idx_in_sheet}']
            if col_name in ['Tender Value', 'EMD']:
                if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            elif col_name == 'Closing Date':
                if isinstance(cell.value, (datetime, pd.Timestamp, date)):
                    cell.number_format = 'DD-MM-YYYY'
            
            # Apply text wrapping
            if col_name in text_wrap_columns:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')


def process_bookmarked_tenders(main_filename=excel_filename, focus_filename=BOOKMARKED_TENDERS_FILENAME):
    """
    Collects all tenders marked with 'Yes' in the 'Bookmark' column from all tender sheets
    in the main Excel file and copies them to a separate 'Tenders to Focus.xlsx' file,
    with individual sheets for each website.
    """
    print(f"\n--- Updating '{focus_filename}' ---")
    log_activity(f"Action: Update Bookmarked Tenders")

    if not os.path.exists(main_filename):
        print(f"Error: The main tender file '{main_filename}' was not found. Please run other options first.")
        return

    # This list will hold all bookmarked tender DataFrames from all sheets.
    all_bookmarked_tenders = []

    # Define which sheets to scan for bookmarked tenders
    sheets_to_scan = [info['name'] for info in website_options.values()]

    # Use pandas to read data from all relevant sheets directly from the main Excel file
    try:
        xls = pd.ExcelFile(main_filename)
        for sheet_name in sheets_to_scan:
            if sheet_name in xls.sheet_names:
                print(f"Reading '{sheet_name}' for bookmark selections...")
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if 'Bookmark' in df.columns:
                    # Filter for 'yes' and ensure the column exists before filtering
                    focus_df = df[df['Bookmark'].astype(str).str.lower() == 'yes'].copy()
                    if not focus_df.empty:
                        all_bookmarked_tenders.append(focus_df)
                        print(f"  Found {len(focus_df)} bookmarked tenders in '{sheet_name}'.")
    except Exception as e:
        print(f"An error occurred while reading the main Excel file: {e}")
        return

    if not all_bookmarked_tenders:
        print("No tenders were marked with 'Yes' in the 'Bookmark' column in any sheet.")
        # To ensure the focus file is clean, we'll write an empty dataframe to it.
        df_focus_final = pd.DataFrame()
    else:
        # Concatenate all found tenders into a single DataFrame
        df_focus_final = pd.concat(all_bookmarked_tenders, ignore_index=True)
        df_focus_final.drop_duplicates(subset=['Tender Id'], keep='last', inplace=True)
        
        # --- NEW: Re-number Sr. No. and drop unwanted columns ---
        df_focus_final.reset_index(drop=True, inplace=True)
        df_focus_final['Sr. No.'] = df_focus_final.index + 1
        
        columns_to_drop = ['Tender Category', 'Download (Type "Yes" to download)', 'Bookmark']
        df_focus_final.drop(columns=[col for col in columns_to_drop if col in df_focus_final.columns], inplace=True)
        
        print(f"Found {len(df_focus_final)} unique tenders to display in '{focus_filename}'.")
        # --- END NEW ---

    # Write the consolidated data to a single sheet in the 'Tenders to Focus.xlsx' file
    print(f"Writing all bookmarked tenders to a single sheet in '{focus_filename}'...")
    with pd.ExcelWriter(focus_filename, engine='openpyxl') as writer:
        # Use the dedicated function to write and format the consolidated sheet
        write_and_format_focus_sheet(writer, df_focus_final, "Tenders to Focus")
    print(f"--- Successfully created/updated '{focus_filename}' with a consolidated list of bookmarked tenders. ---")


def update_tender_hyperlink(tender_id, tender_download_folder_path, filename=excel_filename):
    """
    Finds a tender in the CSV database, updates its folder hyperlink,
    and triggers a refresh of the corresponding Excel sheet.
    """
    folder_uri = f"file:///{os.path.normpath(tender_download_folder_path).replace(os.sep, '/')}"

    # Identify sheets to search in (not org sheets or expired)
    sheets_to_search = [
        f"{site_info['name']}" for site_info in website_options.values()
    ]
    
    tender_found_and_updated = False
    for sheet_name in sheets_to_search:
        df = read_sheet_from_csv(sheet_name)
        if not df.empty and 'Tender Id' in df.columns:
            # Ensure comparison is robust (string vs string)
            mask = df['Tender Id'].astype(str) == str(tender_id)
            if mask.any():
                # Add column if it doesn't exist
                if '__folder_hyperlink__' not in df.columns:
                    df['__folder_hyperlink__'] = None
                
                # Update the hyperlink
                df.loc[mask, '__folder_hyperlink__'] = folder_uri
                
                # --- FIX: Re-apply numeric conversion before saving ---
                # This ensures that when the Excel sheet is updated, the numeric
                # columns are correctly typed, allowing formatting to be applied.
                if 'Sr. No.' in df.columns:
                    df['Sr. No.'] = pd.to_numeric(df['Sr. No.'], errors='coerce')
                if 'Tender Value' in df.columns:
                    df['Tender Value'] = pd.to_numeric(df['Tender Value'], errors='coerce')
                if 'EMD' in df.columns:
                    df['EMD'] = pd.to_numeric(df['EMD'], errors='coerce')

                # --- FIX: Preserve datetime objects for date columns ---
                # MODIFIED: Only convert columns that are generated by the script (like Status Last Checked), not scraped ones.
                for col in ['Status Last Checked']:
                    if col in df.columns:
                        df[col] = pd.to_datetime(df[col], errors='coerce')
                # Write back to CSV
                write_sheet_to_csv(df, sheet_name)
                
                # Refresh the Excel sheet
                update_excel_sheet(df, sheet_name, filename)
                
                print(f"  Successfully updated and hyperlinked Tender ID {tender_id} in data source and Excel sheet '{sheet_name}'.")
                tender_found_and_updated = True
                break
    
    if not tender_found_and_updated:
        print(f"  Warning: Could not find Tender ID {tender_id} in any active tender CSV to add hyperlink.")

def add_hyperlink_to_tender_id_in_excel(filename, tender_id, tender_download_folder_path):
    """
    Opens an Excel file, finds a specific tender ID, and adds a hyperlink
    to its cell pointing to the local download folder.
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        
        # Create the file URI for the hyperlink
        folder_uri = f"file:///{os.path.normpath(tender_download_folder_path).replace(os.sep, '/')}"
        
        # Identify sheets to search in (not org sheets or expired)
        sheets_to_search = [
            s for s in workbook.sheetnames
            if not s.endswith('-O') and s != EXPIRED_TENDERS_SHEET
        ]

        tender_found_and_linked = False
        for sheet_name in sheets_to_search:
            worksheet = workbook[sheet_name]
            
            tender_id_col_idx = None
            for col in range(1, worksheet.max_column + 1):
                if worksheet.cell(row=1, column=col).value == 'Tender Id':
                    tender_id_col_idx = col
                    break
            
            if not tender_id_col_idx:
                continue

            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=tender_id_col_idx)
                if str(cell.value) == str(tender_id):
                    cell.hyperlink = folder_uri
                    cell.font = Font(color="0000FF", underline="single")
                    print(f"  Successfully hyperlinked Tender ID {tender_id} in sheet '{sheet_name}'.")
                    tender_found_and_linked = True
                    break
            
            if tender_found_and_linked:
                break

        if tender_found_and_linked:
            workbook.save(filename)
    except Exception as e:
        print(f"  An unexpected error occurred while adding hyperlink for Tender ID {tender_id}: {e}")

def update_tender_status_hyperlink(tender_id, folder_uri, filename):
    """
    Finds a tender in the Expired or Custom tenders CSV, updates its status hyperlink,
    and triggers a refresh of the corresponding Excel sheet.
    """
    sheets_to_search = [EXPIRED_TENDERS_SHEET, CUSTOM_TENDERS_SHEET]
    tender_found_and_updated = False
    for sheet_name in sheets_to_search:
        df = read_sheet_from_csv(sheet_name)
        if not df.empty and 'Tender Id' in df.columns:
            mask = df['Tender Id'].astype(str) == str(tender_id)
            if mask.any():
                if '__status_hyperlink__' not in df.columns:
                    df['__status_hyperlink__'] = None
                
                df.loc[mask, '__status_hyperlink__'] = folder_uri

                # Re-apply numeric conversions before saving
                if 'Tender Value' in df.columns: df['Tender Value'] = pd.to_numeric(df['Tender Value'], errors='coerce')
                if 'EMD' in df.columns: df['EMD'] = pd.to_numeric(df['EMD'], errors='coerce')
                
                write_sheet_to_csv(df, sheet_name)
                update_excel_sheet(df, sheet_name, filename)
                print(f"       Successfully updated status hyperlink for Tender ID {tender_id} in sheet '{sheet_name}'.")
                tender_found_and_updated = True
                break # Stop after finding it
    if not tender_found_and_updated:
                print(f"       Warning: Could not find Tender ID {tender_id} in expired/custom sheets to add status hyperlink.")

def download_tender_result_files(driver, tender_id):
    """
    After a tender's status is confirmed, this function navigates through the
    status pages to download all available financial result documents.
    """
    print(f"    -> Starting result download process for {tender_id}...")
    # 1. Setup paths
    # Create the main tender folder and a "Financial Result" subfolder
    main_tender_folder = os.path.join(BASE_DOWNLOAD_DIRECTORY, str(tender_id))
    financial_result_folder = os.path.join(main_tender_folder, "Financial Result")
    os.makedirs(financial_result_folder, exist_ok=True) # Creates both if needed
    
    wait = WebDriverWait(driver, 20)
    main_window = driver.current_window_handle
    download_succeeded = False

    # 2. Click 'View' icon on the status results page.
    # This part is now handled in the calling function (bulk_check_tender_results)
    # This function now assumes the driver is already on the "Tender status" page for the specific tender.

    # 3. Click the summary link to open the popup
    try:
        summary_link = wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "view the summary details")))
        summary_link.click()
        wait.until(EC.number_of_windows_to_be(2))
        print("       Clicked summary link to open popup.")
    except Exception as e:
        print(f"       Could not find or click the 'view summary details' link for {tender_id}.")
        return False # Cannot proceed without the popup

    # 4. Switch to popup and download all document files
    popup_window = None
    for window_handle in driver.window_handles:
        if window_handle != main_window:
            popup_window = window_handle
            break

    if popup_window:
        try:
            driver.switch_to.window(popup_window)
            print("       Switched to popup window.")
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            time.sleep(2) # Extra wait for all content to render
            
            # Find all links in the popup
            all_links = driver.find_elements(By.TAG_NAME, "a")
            doc_links = []
            
            # Filter for links that point to .pdf or .xlsx files
            for link in all_links:
                href = link.get_attribute('href')
                if href and ('.pdf' in href.lower() or '.xlsx' in href.lower() or '.zip' in href.lower()):
                    doc_links.append(link)

            if not doc_links:
                print("       No downloadable PDF or Excel files found in the popup.")
            else:
                print(f"       Found {len(doc_links)} document(s) to download.")

            # Download each found document
            for doc_link in doc_links:
                try:
                    doc_url = doc_link.get_attribute('href')
                    # Sanitize filename from link text, or generate one if text is empty
                    doc_name = doc_link.text.strip()
                    if not doc_name:
                        doc_name = os.path.basename(urlparse(doc_url).path)
                    if not doc_name: # Fallback filename
                        doc_name = f"downloaded_file_{int(time.time())}"

                    filepath = os.path.join(financial_result_folder, doc_name)
                    if not os.path.exists(filepath):
                        if download_file_with_requests_session(doc_url, filepath, driver.get_cookies()):
                            download_succeeded = True
                    else:
                        print(f"       Skipping already downloaded file: {doc_name}")

                except Exception as e:
                    print(f"       Error processing a download link: {e}")

        finally:
            # 5. Clean up windows
            if driver.current_window_handle != main_window:
                driver.close()
            driver.switch_to.window(main_window)
            print("       Closed popup and switched back to main window.")
            return download_succeeded
    else:
        print("       Could not find the popup window for downloads.")
        return False

def interactive_select_by_id(driver, wait, dropdown_id, prompt_text):
    """
    Finds a dropdown by its ID, scrapes options, prompts user, and selects.
    Returns True if a selection was made, False otherwise.
    """
    try:
        dropdown_element = wait.until(EC.presence_of_element_located((By.ID, dropdown_id)))
        dropdown = Select(dropdown_element)
        time.sleep(2)
        
        options = dropdown.options
        if len(options) <= 1:
            print(f"    - No options available for '{prompt_text}'.")
            return False

        print(f"\n--- Please select a '{prompt_text}' ---")
        for i, option in enumerate(options):
            # Don't show the default "Select" option in the choices
            if option.get_attribute("value") and option.get_attribute("value") != "0":
                print(f"  {i}. {option.text}")
        
        while True:
            # Make the choice mandatory, as this function is for required fields
            choice_str = input(f"Enter the number for your choice: ").strip()
            if choice_str.isdigit() and 0 <= int(choice_str) < len(options):
                choice_idx = int(choice_str)
                # Double-check that the user isn't trying to select the placeholder
                if options[choice_idx].get_attribute("value") and options[choice_idx].get_attribute("value") != "0":
                    selected_option_text = options[choice_idx].text
                    print(f"    Selected: {selected_option_text}")
                    dropdown.select_by_index(choice_idx)
                    time.sleep(3) # Wait for dependent elements to load
                    return True
                else:
                    print("    Invalid selection. Please choose one of the numbered options.")
            else:
                print("    Invalid input. Please enter a valid number from the list.")

    except (TimeoutException, NoSuchElementException):
        print(f"    - Error: '{prompt_text}' dropdown with ID '{dropdown_id}' not found.")
        return False
    except Exception as e:
        print(f"    An error occurred while handling the '{prompt_text}' dropdown: {e}")
        return False

def select_dropdown_option_by_index(driver, wait, dropdown_id, option_index, option_name_for_log=""):
    """Finds a dropdown by ID and selects an option by its index as a default."""
    try:
        select_element = wait.until(EC.presence_of_element_located((By.ID, dropdown_id)))
        # Wait for options to be populated beyond the default "--Select--"
        wait.until(EC.presence_of_element_located((By.XPATH, f"//select[@id='{dropdown_id}']/option[2]")))
        time.sleep(1)

        dropdown = Select(select_element)
        options = dropdown.options
        if option_index < len(options):
            selected_option_text = options[option_index].text
            dropdown.select_by_index(option_index)
            log_text = option_name_for_log if option_name_for_log else selected_option_text
            print(f"    Defaulted Tender Status to '{log_text}' (option {option_index}).")
            time.sleep(2) # Wait for any dynamic content to load after selection
            return True
        else:
            print(f"    Could not select default option at index {option_index} in dropdown '{dropdown_id}': Index out of range. Skipping organization.")
            return False
    except (TimeoutException, NoSuchElementException):
       print(f"    Timeout or error while waiting for options in dropdown '{dropdown_id}'. Skipping organization.")
       return False
    except Exception as e:
        print(f"    Could not select default option in dropdown '{dropdown_id}': {e}. Skipping organization.")
        return False

def select_dropdown_option_by_text(driver, wait, dropdown_id, option_text):
    """Finds a dropdown by ID and selects an option by its visible text."""
    try:
        dropdown = Select(wait.until(EC.element_to_be_clickable((By.ID, dropdown_id))))
        dropdown.select_by_visible_text(option_text)
        print(f"    Selected '{option_text}' in dropdown '{dropdown_id}'.")
        time.sleep(2) # Wait for any dynamic content to load after selection
        return True
    except Exception as e:
        print(f"    Could not select '{option_text}' in dropdown '{dropdown_id}': {e}")
        return False

def scrape_and_select_from_dropdown_by_label(driver, wait, label_text):
    """
    Finds a dropdown menu by its preceding text label, scrapes its options,
    prompts the user for a selection in the console, and makes the selection.
    Returns True if a selection was made, False otherwise.
    """
    try:
        # ROBUST XPATH: Use contains() and normalize-space() to be flexible with whitespace and extra characters like '*'.
        xpath = f"//td[contains(normalize-space(.), '{label_text}')]/following-sibling::td[1]/select"
        dropdown_element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        dropdown = Select(dropdown_element)
        
        # Give a moment for options to be populated, especially if they are dynamic.
        time.sleep(2)
        
        options = dropdown.options
        if len(options) <= 1: # Only contains "Select" or is empty
            print(f"    - No options available for '{label_text}'.")
            return False

        print(f"\n--- Please select a '{label_text}' ---")
        for i, option in enumerate(options):
            if option.get_attribute("value"): # Don't show the default "Select" option
                print(f"  {i}. {option.text}")
        
        while True:
            choice_str = input(f"Enter the number for your choice (or press Enter to skip): ").strip()
            if not choice_str:
                print(f"    Skipping '{label_text}' selection.")
                return False
            
            if choice_str.isdigit() and 0 <= int(choice_str) < len(options):
                choice_idx = int(choice_str)
                selected_option_text = options[choice_idx].text
                print(f"    Selected: {selected_option_text}")
                dropdown.select_by_index(choice_idx)
                # Wait after selection to allow any subsequent dynamic dropdowns to load.
                time.sleep(3)
                return True
            else:
                print("    Invalid input. Please enter a valid number from the list.")

    except (TimeoutException, NoSuchElementException):
        # This is not an error, it just means this filter doesn't exist on the page
        # for the currently selected Organization.
        print(f"    - Info: '{label_text}' dropdown not found. Skipping.")
        return False
    except Exception as e:
        print(f"    An error occurred while handling the '{label_text}' dropdown: {e}")
        return False

def bulk_check_tender_results(website_choice, filename=excel_filename):
    """
    Performs a detailed, interactive bulk search for tender results based on user input for
    status, dates, and departments, then scrapes and saves qualifying tenders, handling pagination.
    """
    print("\n--- Starting Interactive Bulk Tender Result Check ---")
    log_activity(f"Action: Bulk Check Tender Results for website choice {website_choice}")

    # 1. Determine which organization sheets to process
    if website_choice in website_options.keys():
        org_sheet_names_to_process = [f"{website_options[website_choice]['name']}-O"]
    else: # All websites
        org_sheet_names_to_process = [f"{site_info['name']}-O" for site_info in website_options.values()]
    
    # 2. Find organizations selected for bulk check
    orgs_by_website = {}
    for org_sheet_name in org_sheet_names_to_process:
        website_name = org_sheet_name.replace('-O', '')
        try:
            df_orgs_from_excel = pd.read_excel(filename, sheet_name=org_sheet_name)
            if 'Fetch Results (Yes/No)' not in df_orgs_from_excel.columns:
                df_orgs_from_excel['Fetch Results (Yes/No)'] = ''
            write_sheet_to_csv(df_orgs_from_excel, org_sheet_name)
            
            df_orgs = read_sheet_from_csv(org_sheet_name)
            if 'Fetch Results (Yes/No)' in df_orgs.columns:
                selected_df = df_orgs[df_orgs['Fetch Results (Yes/No)'].astype(str).str.lower() == 'yes']
                if not selected_df.empty:
                    orgs_by_website[website_name] = selected_df['Organisation Name'].tolist()
        except (FileNotFoundError, ValueError):
            print(f"Info: Org sheet '{org_sheet_name}' not found. Skipping.")
        except Exception as e:
            print(f"An error occurred while processing org sheet '{org_sheet_name}': {e}")

    if not orgs_by_website:
        print("No organizations marked with 'Yes' in the 'Fetch Results (Yes/No)' column.")
        return

    # 3. Load existing data from Custom Tenders sheet
    df_custom_all = read_sheet_from_csv(CUSTOM_TENDERS_SHEET)
    all_scraped_tenders_data = [] # Temp list to hold data from this run

    # 4. Initialize Selenium and process
    driver = None
    global captcha_solved_in_session
    try:
        driver = initialize_selenium_driver()
        wait = WebDriverWait(driver, 20)

        for site_name, org_list in orgs_by_website.items():
            site_id = next((k for k, v in website_options.items() if v['name'] == site_name), None)
            if website_choice in website_options.keys() and site_id != website_choice: continue

            status_url = website_options.get(site_id, {}).get('status_url')
            if not status_url: continue

            print(f"\n--- Processing website: {site_name} ---")

            for org_name in org_list:
                print(f"\nConfiguring search for Organisation: '{org_name}'")
                try:
                    driver.get(status_url)
                    time.sleep(3)
                    captcha_solved_in_session = False
                    
                    # --- INTERACTIVE SEARCH SETUP ---
                    # 1. Set Tender Status to "Financial Bid Opening" by default (option 4)
                    if not select_dropdown_option_by_index(driver, wait, "tenderStatus", 4, "Financial Bid Opening"):
                        # The function already prints a detailed error, so we just continue.
                        continue
                    
                    # 2. Get Dates from user
                    from_date = input("Enter From Date (dd-mm-yyyy): ").strip()
                    to_date = input("Enter To Date (dd-mm-yyyy): ").strip()
                    if from_date: driver.find_element(By.ID, "fromDate").send_keys(from_date)
                    if to_date: driver.find_element(By.ID, "toDate").send_keys(to_date)

                    # NEW: Get Keyword from user
                    keyword = input("Enter Keyword (or press Enter to skip): ").strip()
                    if keyword:
                        driver.find_element(By.ID, "keyword").send_keys(keyword)

                    # 3. Select Organisation by ID (OrganName) and visible text
                    if not select_dropdown_option_by_text(driver, wait, "OrganName", org_name):
                        print(f"Could not select organization '{org_name}'. It might not be in the list. Skipping.")
                        continue
                    
                    # 4. Select dynamic dropdowns by label (using the improved function)
                    scrape_and_select_from_dropdown_by_label(driver, wait, "Department")
                    scrape_and_select_from_dropdown_by_label(driver, wait, "Division")
                    scrape_and_select_from_dropdown_by_label(driver, wait, "Sub Division")
                    
                    if not handle_captcha_interaction(driver, f"Bulk Check for {org_name}", submit_button_id="Search", success_element_id="tabList"):
                        print(f"    Could not solve CAPTCHA. Skipping this search.")
                        continue
                    
                    # --- NEW WORKFLOW: Process-as-you-go based on user input ---

                    # 1. Get total records and user input for how many to process
                    total_records = 0
                    try:
                        wait.until(EC.presence_of_element_located((By.ID, "tabList")))
                        try:
                            total_records_element = driver.find_element(By.ID, "If_41")
                        except NoSuchElementException:
                            print("    Could not find 'Total Records' by ID 'If_41', falling back to text search.")
                            total_records_element = driver.find_element(By.XPATH, "//*[contains(text(), 'Total Records')]")
                        match = re.search(r'\d+', total_records_element.text)
                        if match:
                            total_records = int(match.group(0))
                            print(f"    Website indicates Total Records: {total_records}")
                    except (NoSuchElementException, TimeoutException):
                        print("    Could not find 'Total Records' count on the page. Will proceed with scraping.")

                    num_to_fetch = 0
                    while True:
                        num_to_fetch_str = input(f"    How many of the most recent tenders do you want to process for results? (Enter a number): ").strip()
                        if num_to_fetch_str.isdigit() and int(num_to_fetch_str) > 0:
                            num_to_fetch = int(num_to_fetch_str)
                            break
                        else:
                            print("    Invalid input. Please enter a positive number.")

                    # 2. Loop through pages, processing tenders until the user's limit is reached.
                    processed_count = 0 # Counts tenders that are actually processed for download
                    tenders_considered = 0 # Counts all tenders seen from the top of the list
                    page_number = 1
                    target_stages = ["Financial Bid Opening", "Financial Evaluation", "AOC", "Concluded"]
                    scraping_active = True

                    while scraping_active and tenders_considered < num_to_fetch:
                        print(f"\n    --- Processing Page {page_number} ---")
                        try:
                            table = wait.until(EC.presence_of_element_located((By.ID, "tabList")))
                            # Re-find headers on each page/loop to be safe
                            header_cells = table.find_elements(By.CSS_SELECTOR, "tr.list_header td")
                            headers = [cell.text.strip() for cell in header_cells]
                            header_map = {header: i for i, header in enumerate(headers)}

                            rows_xpath = ".//tbody/tr[contains(@class, 'even') or contains(@class, 'odd')]"
                            row_count = len(table.find_elements(By.XPATH, rows_xpath))
                            print(f"    Found {row_count} tenders on this page. Filtering and processing...")

                            if row_count == 0:
                                print("    No more tenders found. Ending search for this organization.")
                                scraping_active = False
                                break

                            # Iterate by index to avoid stale elements after navigating back
                            for i in range(row_count):
                                if tenders_considered >= num_to_fetch:
                                    scraping_active = False; break
                                
                                tenders_considered += 1 # Count every tender we look at

                                table = wait.until(EC.presence_of_element_located((By.ID, "tabList")))
                                row = table.find_elements(By.XPATH, rows_xpath)[i]
                                
                                cells = row.find_elements(By.TAG_NAME, "td")
                                def get_cell_data(header_name):
                                    idx = header_map.get(header_name)
                                    return cells[idx].text.strip() if idx is not None and idx < len(cells) else None
                                tender_id_for_log = get_cell_data('Tender ID')
                                
                                tender_stage = get_cell_data('Tender Stage')
                                if tender_stage in target_stages:
                                    tender_id = get_cell_data('Tender ID')
                                    # FIX: Use a more robust locator for the view link based on its image, as the ID is dynamic.
                                    view_links = row.find_elements(By.XPATH, ".//a[img[contains(@src, 'view.png')]]")
                                    if not view_links:
                                        print(f"      - Skipping a row for Tender ID '{tender_id}' as it has no 'view' link.")
                                        continue

                                    view_url = view_links[0].get_attribute('href') # This is a javascript: link, but click() handles it.
                                    if not (tender_id and view_url): continue

                                    tender_data = {
                                        'Tender Id': tender_id, 'Tender Title': get_cell_data('Title and Ref.No.'), 'Organisation Chain': get_cell_data('Organisation Chain'), 
                                        'Current Status': tender_stage, 'Source Website': site_name, 
                                        'Status Last Checked': datetime.now(), 'view_url': view_url
                                    }

                                    print(f"\n    Processing download for Tender ID: {tender_id} ({tenders_considered}/{num_to_fetch})")
                                    try:
                                        view_links[0].click()
                                        time.sleep(3)
                                        
                                        if download_tender_result_files(driver, tender_id):
                                            main_tender_folder = os.path.join(BASE_DOWNLOAD_DIRECTORY, str(tender_id))
                                            abs_folder_path = os.path.abspath(main_tender_folder)
                                            folder_uri = f"file:///{os.path.normpath(abs_folder_path).replace(os.sep, '/')}"
                                            tender_data['__folder_hyperlink__'] = folder_uri
                                        
                                        all_scraped_tenders_data.append(tender_data)
                                        processed_count += 1
                                    except Exception as e:
                                        print(f"      Error during detail processing/download for {tender_id}: {e}")
                                        all_scraped_tenders_data.append(tender_data)
                                        processed_count += 1
                                    finally:
                                        print("      Navigating back to search results...")
                                        # The ID 'PageLink_16' is dynamic and unreliable. This is a more robust approach.
                                        try:
                                            # Strategy: Find the 'Back' button by its link destination ('href') or its text value, which is more stable than an ID.
                                            back_button_xpath = "//a[contains(@href, 'WebTenderStatusLists')] | //input[@value='Back']"
                                            back_button = wait.until(EC.element_to_be_clickable((By.XPATH, back_button_xpath)))
                                            
                                            driver.execute_script("arguments[0].click();", back_button)
                                            
                                            # Wait for the search results table to confirm successful navigation.
                                            wait.until(EC.presence_of_element_located((By.ID, "tabList")))
                                            time.sleep(8) # Extra pause for stability as requested.
                                        except Exception as e:
                                            print(f"      Could not find or click the 'Back' button to return to search results. Error: {e}. The script will likely fail.")
                                            scraping_active = False
                                            break
                                else:
                                    print(f"      - Skipping Tender ID '{tender_id_for_log}' (Stage: '{tender_stage}') - does not match target stages.")
                            if not scraping_active: break

                            # After processing all rows on the page, check if we need to paginate
                            if scraping_active and tenders_considered < num_to_fetch:
                                try:
                                    wait_short = WebDriverWait(driver, 5)
                                    next_button = wait_short.until(EC.element_to_be_clickable((By.ID, "loadNext")))

                                    # Get a reference to an element that will go stale after the AJAX call.
                                    try:
                                        first_row_before_click = driver.find_element(By.XPATH, "//table[@id='tabList']//tbody/tr[1]")
                                    except NoSuchElementException:
                                        first_row_before_click = None # No rows on page, can't check for staleness.

                                    print(f"    Navigating to next page...")
                                    driver.execute_script("arguments[0].click();", next_button)

                                    if first_row_before_click:
                                        # Wait for the old first row to become stale. This confirms the DOM is updating.
                                        print("      Waiting for page content to refresh...")
                                        wait.until(EC.staleness_of(first_row_before_click))
                                    
                                    # After the content is stale, wait for the table to be stable again.
                                    wait.until(EC.presence_of_element_located((By.ID, "tabList")))
                                    time.sleep(2) # A brief pause for safety after content refresh.
                                    page_number += 1
                                except TimeoutException:
                                    print("    'Next' button (ID: loadNext) not found. Reached the end of results.")
                                    scraping_active = False
                        except (TimeoutException, StaleElementReferenceException) as e:
                            print(f"    Error processing page {page_number}: {e}. Stopping search for this organization.")
                            scraping_active = False

                except Exception as e:
                    print(f"    An unexpected error occurred while processing this search for '{org_name}': {e}")

    except WebDriverException as e: print(f"A Selenium WebDriver error occurred: {e}")
    except Exception as e: print(f"An error occurred during the overall status check process: {e}")
    finally:
        if driver: driver.quit(); print("\nSelenium WebDriver closed.")

    # 5. Merge and save results to Custom Tenders sheet
    if all_scraped_tenders_data:
        print(f"\nMerging {len(all_scraped_tenders_data)} scraped tenders with the '{CUSTOM_TENDERS_SHEET}' sheet...")
        df_scraped = pd.DataFrame(all_scraped_tenders_data)
        
        # Ensure 'Tender Id' is string for merging
        if not df_custom_all.empty and 'Tender Id' in df_custom_all.columns:
            df_custom_all['Tender Id'] = df_custom_all['Tender Id'].astype(str)
        df_scraped['Tender Id'] = df_scraped['Tender Id'].astype(str)

        # Merge, keeping the latest scraped data for any duplicates
        df_final_custom = pd.concat([df_custom_all, df_scraped], ignore_index=True)
        df_final_custom.drop_duplicates(subset=['Tender Id'], keep='last', inplace=True)
        
        # Ensure date columns are in datetime format for correct Excel formatting
        if 'Status Last Checked' in df_final_custom.columns:
            df_final_custom['Status Last Checked'] = pd.to_datetime(df_final_custom['Status Last Checked'], errors='coerce')

        print("Updating data source and Excel file...")
        df_final_custom['Status Last Checked'] = pd.to_datetime(df_final_custom['Status Last Checked']).dt.strftime('%Y-%m-%d %H:%M:%S').replace('NaT', '')
        write_sheet_to_csv(df_final_custom, CUSTOM_TENDERS_SHEET)
        update_excel_sheet(df_final_custom, CUSTOM_TENDERS_SHEET, filename)
    else:
        print("\nNo new tender information was scraped that matched the criteria.")

    print("\n--- Bulk Tender Result Check Complete. ---")

def check_expired_tender_status(website_choice, filename=excel_filename):
    """
    Checks the live status of tenders marked in the 'Expired Tenders' sheet.
    """
    print("\n--- Starting Expired Tender Status Check Process ---")
    log_activity(f"Action: Check Expired Tender Status for website choice {website_choice}")

    # 1. Sync user input from Excel to CSV
    try:
        # Read from Excel to get latest user input
        df_expired_from_excel = pd.read_excel(filename, sheet_name=EXPIRED_TENDERS_SHEET)
        # Add "Check Status" column if it doesn't exist in the Excel file
        if 'Check Status' not in df_expired_from_excel.columns:
            df_expired_from_excel['Check Status'] = ''
        # Save the potentially modified DataFrame to CSV, making it the source of truth
        write_sheet_to_csv(df_expired_from_excel, EXPIRED_TENDERS_SHEET)
        print(f"Successfully synced selections from '{EXPIRED_TENDERS_SHEET}'.")
    except (FileNotFoundError, ValueError):
        print(f"Error: Sheet '{EXPIRED_TENDERS_SHEET}' not found in '{filename}'. Cannot check status.")
        return
    except Exception as e:
        print(f"An unexpected error occurred during expired sheet sync: {e}")
        return

    # 2. Load data from CSV and filter for tenders to check
    df_expired = read_sheet_from_csv(EXPIRED_TENDERS_SHEET)
    if 'Check Status' not in df_expired.columns:
        print(f"Warning: 'Check Status' column not found in data. No tenders will be checked.")
        tenders_to_check = pd.DataFrame()
    else:
        # Ensure the column is treated as string to handle non-string values gracefully
        tenders_to_check = df_expired[df_expired['Check Status'].astype(str).str.lower() == 'yes'].copy()

    if tenders_to_check.empty:
        print(f"No tenders marked with 'Yes' in the 'Check Status' column of the '{EXPIRED_TENDERS_SHEET}' sheet.")
        return

    print(f"Found {len(tenders_to_check)} expired tenders marked for status check.")

    # --- NEW: Infer 'Source Website' if missing, using the Tender URL ---
    if 'Source Website' not in tenders_to_check.columns:
        tenders_to_check['Source Website'] = ''
    
    # Create a reverse map from hostname to website name for inference
    hostname_to_name_map = {urlparse(v['url']).hostname: v['name'] for k, v in website_options.items()}

    for index, row in tenders_to_check.iterrows():
        # If 'Source Website' is empty or NaN for this row, try to infer it
        if pd.isna(row['Source Website']) or not row['Source Website']:
            tender_url = row.get('Tender URL')
            if tender_url and isinstance(tender_url, str):
                try:
                    hostname = urlparse(tender_url).hostname
                    if hostname in hostname_to_name_map:
                        inferred_site_name = hostname_to_name_map[hostname]
                        tenders_to_check.loc[index, 'Source Website'] = inferred_site_name
                        # Also update the main df_expired so it gets saved to CSV/Excel later
                        df_expired.loc[df_expired['Tender Id'] == row['Tender Id'], 'Source Website'] = inferred_site_name
                    else:
                        print(f"  Warning: Could not determine website for Tender ID {row['Tender Id']} from URL's hostname '{hostname}'. Skipping.")
                except Exception as e:
                    print(f"  Warning: Error parsing URL for Tender ID {row['Tender Id']}: {e}. Skipping.")
            else:
                print(f"  Warning: Missing 'Tender URL' for Tender ID {row['Tender Id']}. Cannot infer source website. Skipping.")
    
    # Filter out any rows where we still couldn't determine the source
    tenders_to_check.dropna(subset=['Source Website'], inplace=True)
    tenders_to_check = tenders_to_check[tenders_to_check['Source Website'] != '']

    if tenders_to_check.empty:
        print("Could not determine the source website for any of the selected tenders. Aborting status check.")
        return
    # --- END of new inference logic ---

    # Add/update columns for status tracking
    if 'Current Status' not in df_expired.columns:
        df_expired['Current Status'] = ''
    if 'Status Last Checked' not in df_expired.columns:
        df_expired['Status Last Checked'] = pd.NaT # Use NaT for proper date handling
    
    # 3. Initialize Selenium and process
    driver = None
    global captcha_solved_in_session
    try:
        driver = initialize_selenium_driver()
        wait = WebDriverWait(driver, 20)
        
        # The old error check is now replaced by the inference logic above.
        # The DataFrame is now ready for grouping.
        # Group by source website to be efficient
        for site_name, group in tenders_to_check.groupby('Source Website'):
            site_id = next((k for k, v in website_options.items() if v['name'] == site_name), None)

            if website_choice in website_options.keys() and site_id != website_choice:
                continue

            status_url = website_options.get(site_id, {}).get('status_url')
            if not status_url:
                print(f"Warning: No status check URL configured for '{site_name}'. Skipping {len(group)} tenders.")
                continue

            print(f"\nProcessing {len(group)} tenders for '{site_name}'...")

            for index, tender_row in group.iterrows():
                tender_id = tender_row['Tender Id']
                print(f"  Checking status for Tender ID: {tender_id}")

                try:
                    # For each tender, we must navigate to the status page, as each check is a separate form submission.
                    driver.get(status_url)
                    time.sleep(2)
                    # We must reset the CAPTCHA solved flag for each tender, as each form submission requires a new CAPTCHA.
                    captcha_solved_in_session = False

                    # Input the Tender ID into the form field.
                    tender_id_input = wait.until(EC.presence_of_element_located((By.ID, "tenderId")))
                    tender_id_input.clear()
                    tender_id_input.send_keys(tender_id)

                    # Solve the CAPTCHA and submit the form.
                    if not handle_captcha_interaction(driver, f"Status Check for {tender_id}", submit_button_id="Search", success_element_id="tabList"):
                        print(f"    Could not solve CAPTCHA for {tender_id}. Skipping.")
                        df_expired.loc[df_expired['Tender Id'] == tender_id, 'Current Status'] = "CAPTCHA Failed"
                        df_expired.loc[df_expired['Tender Id'] == tender_id, 'Status Last Checked'] = datetime.now()
                        continue
                    
                    # After submission, the page refreshes/updates. We now scrape the status from the results table.
                    status = "Status Not Found"
                    try:
                        table = wait.until(EC.presence_of_element_located((By.ID, "tabList")))
                        # Corrected header finding: The header row has a class 'list_header' and is inside the tbody.
                        header_cells = table.find_elements(By.CSS_SELECTOR, "tr.list_header td")
                        headers = [cell.text.strip() for cell in header_cells]
                        if not headers: # Fallback if class name changes or is not present
                            headers = [cell.text.strip() for cell in table.find_elements(By.XPATH, ".//tbody/tr[1]/td")]

                        stage_index = headers.index("Tender Stage") if "Tender Stage" in headers else -1
                        
                        # --- START: CORRECTED STATUS FETCHING LOGIC ---
                        if stage_index != -1:
                            # Find data rows by looking for class 'even' or 'odd', which filters out the header row.
                            data_rows = table.find_elements(By.XPATH, ".//tbody/tr[contains(@class, 'even') or contains(@class, 'odd')]")
                            if data_rows:
                                # The first data row contains the status we need.
                                cells_in_row = data_rows[0].find_elements(By.TAG_NAME, "td")
                                if stage_index < len(cells_in_row):
                                    status_cell = cells_in_row[stage_index]
                                    status = status_cell.text.strip()
                                    print(f"    Found Status: {status}")
                                else:
                                    status = "Status Column Index Out of Range"
                            else: 
                                print("    No data row found in the results table. Leaving status blank.")
                                status = "" # Leave blank to allow re-checking on next run
                        else: 
                            status = "Status Column Not Found"
                        # --- END: CORRECTED STATUS FETCHING LOGIC ---

                    except TimeoutException:
                        if "No Records Found" in driver.page_source:
                            print(f"    Tender ID {tender_id} not found on the portal. Leaving row blank.")
                            status = "" # Leave blank to be handled by logic below
                        else:
                            status = "Error or Not Found"
                    
                    # If no data row was found, only set status to blank for re-checking.
                    # Otherwise, update both status and the timestamp.
                    if status == "":
                        df_expired.loc[df_expired['Tender Id'] == tender_id, 'Current Status'] = ""
                    else:
                        df_expired.loc[df_expired['Tender Id'] == tender_id, 'Current Status'] = status
                        df_expired.loc[df_expired['Tender Id'] == tender_id, 'Status Last Checked'] = datetime.now()
                    
                    # --- NEW: Check if status requires result download ---
                    target_statuses = ["Financial Bid Opening", "Financial Evaluation", "AOC", "Concluded"]
                    if status in target_statuses:
                        print(f"    Status '{status}' found. Checking for result files...")
                        download_tender_result_files(driver, tender_id)


                except Exception as e:
                    print(f"    An unexpected error occurred while checking {tender_id}: {e}")
                    df_expired.loc[df_expired['Tender Id'] == tender_id, 'Current Status'] = "Processing Error"
                    df_expired.loc[df_expired['Tender Id'] == tender_id, 'Status Last Checked'] = datetime.now()
                    continue
    except WebDriverException as e: print(f"A Selenium WebDriver error occurred: {e}")
    except Exception as e: print(f"An error occurred during the overall status check process: {e}")
    finally:
        if driver: driver.quit(); print("\nSelenium WebDriver closed.")

    # Ensure date columns are in datetime format for correct Excel formatting
    if 'Status Last Checked' in df_expired.columns:
        df_expired['Status Last Checked'] = pd.to_datetime(df_expired['Status Last Checked'], errors='coerce')

    print("\nUpdating data source and Excel file with new statuses...")
    df_expired['Status Last Checked'] = pd.to_datetime(df_expired['Status Last Checked']).dt.strftime('%Y-%m-%d %H:%M:%S').replace('NaT', '')
    write_sheet_to_csv(df_expired, EXPIRED_TENDERS_SHEET)
    update_excel_sheet(df_expired, EXPIRED_TENDERS_SHEET, filename)
    print("\n--- Expired Tender Status Check Complete. ---")

def check_custom_tender_status(website_choice, filename=excel_filename):
    """
    Allows user to manually input a Tender ID to check its status and details.
    Saves the result to a 'Custom Tenders' sheet.
    """
    print("\n--- Custom Tender Status Check (Manual Input) ---")
    
    # 1. Get Tender ID from user
    tender_id = input("Enter the Tender ID to check: ").strip()
    if not tender_id:
        print("No Tender ID entered. Aborting.")
        return

    # 2. Determine which website to use
    site_choice = None
    if website_choice in website_options.keys():
        # If a specific website was chosen in the main menu, use it directly.
        site_choice = website_choice
    else: # This handles 'All Websites' or other non-key choices
        print("\nWhich website does this tender belong to?") #
        for key, value in website_options.items():
            print(f"{key}. {value['name']}")
        site_choice_str = input(f"Enter your choice (1-{len(website_options)}): ").strip()
        if not site_choice_str.isdigit() or int(site_choice_str) not in website_options:
            print("Invalid website choice. Aborting.")
            return
        site_choice = int(site_choice_str)
    site_info = website_options[site_choice]
    status_url = site_info['status_url']
    site_name = site_info['name']
    
    log_activity(f"Action: Custom Tender Check for ID {tender_id} on {site_name}")

    # 3. Initialize Selenium and perform check
    driver = None
    scraped_data = None
    global captcha_solved_in_session
    try:
        driver = initialize_selenium_driver()
        wait = WebDriverWait(driver, 20)
        
        driver.get(status_url)
        time.sleep(2)
        captcha_solved_in_session = False

        tender_id_input = wait.until(EC.presence_of_element_located((By.ID, "tenderId")))
        tender_id_input.clear()
        tender_id_input.send_keys(tender_id)

        if not handle_captcha_interaction(driver, f"Custom Status Check for {tender_id}", submit_button_id="Search", success_element_id="tabList"):
            print(f"Could not solve CAPTCHA for {tender_id}. Aborting.")
            return

        # 4. Scrape data from the results table
        try:
            table = wait.until(EC.presence_of_element_located((By.ID, "tabList")))
            header_cells = table.find_elements(By.CSS_SELECTOR, "tr.list_header td")
            headers = [cell.text.strip() for cell in header_cells]
            if not headers:
                headers = [cell.text.strip() for cell in table.find_elements(By.XPATH, ".//tbody/tr[1]/td")]

            header_map = {header: i for i, header in enumerate(headers)}
            
            data_rows = table.find_elements(By.XPATH, ".//tbody/tr[contains(@class, 'even') or contains(@class, 'odd')]")
            if data_rows:
                first_row_cells = data_rows[0].find_elements(By.TAG_NAME, "td")
                
                def get_cell_data(header_name):
                    if header_name in header_map:
                        idx = header_map[header_name]
                        if idx < len(first_row_cells): return first_row_cells[idx].text.strip()
                    return "N/A"

                scraped_data = {
                    'Tender Id': get_cell_data('Tender ID'), 'Tender Title': get_cell_data('Title and Ref.No.'),
                    'Organisation Chain': get_cell_data('Organisation Chain'), 'Current Status': get_cell_data('Tender Stage'),
                    'Source Website': site_name, 'Status Last Checked': datetime.now()
                }
                print("\n--- Scraped Details ---\n" + "\n".join([f"{k}: {v}" for k, v in scraped_data.items()]) + "\n-----------------------")
            else:
                print("No data rows found in the results table.")
                scraped_data = None # Ensure no data is saved if scraping fails

        except (TimeoutException, IndexError) as e:
            status_msg = "No Records Found" if "No Records Found" in driver.page_source else "Error or Not Found"
            print(f"Could not find tender details table or data. Status: {status_msg}. Error: {e}")
            scraped_data = None # Ensure no data is saved if scraping fails
        
        # --- NEW LOGIC: Save and download inside the 'try' block ---
        if scraped_data:
            # 1. Save the scraped data to CSV and Excel
            print("\nUpdating data source and Excel file with scraped info...")
            df_custom = read_sheet_from_csv(CUSTOM_TENDERS_SHEET)
            df_updated = pd.concat([df_custom, pd.DataFrame([scraped_data])], ignore_index=True)
            df_updated.drop_duplicates(subset=['Tender Id'], keep='last', inplace=True)
            # Ensure date columns are in datetime format for correct Excel formatting
            if 'Status Last Checked' in df_updated.columns:
                df_updated['Status Last Checked'] = pd.to_datetime(df_updated['Status Last Checked'], errors='coerce')

            df_updated['Status Last Checked'] = pd.to_datetime(df_updated['Status Last Checked']).dt.strftime('%Y-%m-%d %H:%M:%S').replace('NaT', '')
            write_sheet_to_csv(df_updated, CUSTOM_TENDERS_SHEET)
            update_excel_sheet(df_updated, CUSTOM_TENDERS_SHEET, filename)

            # 2. Now attempt to download results if applicable
            target_statuses = ["Financial Bid Opening", "Financial Evaluation", "AOC", "Concluded"]
            if scraped_data.get('Current Status') in target_statuses:
                print(f"    Status '{scraped_data['Current Status']}' found. Checking for result files...")
                download_tender_result_files(driver, tender_id)
        else:
            print("Failed to scrape any data for the tender.")

    except Exception as e: 
        print(f"An unexpected error occurred during the custom check: {e}")
    finally:
        if driver: driver.quit()

    print("\n--- Custom Tender Check Complete. ---")

# --- NEW: Function for checking unchecked tenders from Custom List ---
def check_custom_list_status(website_choice, filename=excel_filename):
    """
    Checks the live status of tenders listed in the 'Custom Tenders' sheet
    that have a Tender ID but have not yet been checked (i.e., 'Current Status' is empty).
    """
    print("\n--- Starting Custom Tender List Status Check Process ---")
    log_activity(f"Action: Check Custom List Tender Status for website choice {website_choice}")

    # 1. Sync data from Excel to CSV to ensure we have the latest user additions
    try:
        # Read from Excel to get latest user input
        df_custom_from_excel = pd.read_excel(filename, sheet_name=CUSTOM_TENDERS_SHEET)
        # Save to CSV, making it the source of truth
        write_sheet_to_csv(df_custom_from_excel, CUSTOM_TENDERS_SHEET)
        print(f"Successfully synced data from '{CUSTOM_TENDERS_SHEET}'.")
    except (FileNotFoundError, ValueError):
        print(f"Info: Sheet '{CUSTOM_TENDERS_SHEET}' not found in '{filename}' or is empty. Nothing to check.")
        # If the sheet doesn't exist, there's nothing to check.
        return
    except Exception as e:
        print(f"An unexpected error occurred during custom sheet sync: {e}")
        return

    # 2. Load data from CSV and filter for tenders to check
    df_custom = read_sheet_from_csv(CUSTOM_TENDERS_SHEET)
    if df_custom.empty:
        print(f"The '{CUSTOM_TENDERS_SHEET}' sheet is empty. Nothing to check.")
        return

    # Ensure required columns exist, adding them if they don't
    for col in ['Tender Id', 'Current Status', 'Tender Title', 'Organisation Chain', 'Source Website']:
        if col not in df_custom.columns:
            df_custom[col] = ''

    # Filter for rows that have a Tender ID but the status has not been filled in yet.
    tenders_to_check = df_custom[
        df_custom['Tender Id'].notna() &
        (df_custom['Tender Id'].astype(str).str.strip() != '') &
        (df_custom['Current Status'].isna() | (df_custom['Current Status'].astype(str).str.strip() == ''))
    ].copy()


    if tenders_to_check.empty:
        print(f"No new tenders to check in the '{CUSTOM_TENDERS_SHEET}' sheet (all existing tenders have a status).")
        return

    print(f"Found {len(tenders_to_check)} new tenders in the custom list for status check.")

    # --- Infer 'Source Website' if missing ---
    hostname_to_name_map = {urlparse(v['url']).hostname: v['name'] for k, v in website_options.items()}
    
    # Get the name of the selected website if a specific one was chosen in the main menu
    selected_site_name_from_menu = None
    if website_choice in website_options.keys():
        selected_site_name_from_menu = website_options[website_choice]['name']
    else:
        print("Info: 'All Websites' selected. Source for each tender must be specified in the sheet or be inferable from its URL.")

    for index, row in tenders_to_check.iterrows():
        if pd.isna(row['Source Website']) or not row['Source Website']:
            inferred_site_name = None
            # 1. Try to infer from 'Tender URL' column first, as it's the most specific.
            tender_url = row.get('Tender URL')
            if tender_url and isinstance(tender_url, str):
                try:
                    hostname = urlparse(tender_url).hostname
                    if hostname in hostname_to_name_map:
                        inferred_site_name = hostname_to_name_map[hostname]
                except Exception as e:
                    print(f"  Warning: Error parsing URL for Tender ID {row['Tender Id']}: {e}.")
            
            # 2. If URL inference failed, fall back to the user's menu choice (if a specific site was chosen).
            final_site_name = inferred_site_name or selected_site_name_from_menu

            if final_site_name:
                tenders_to_check.loc[index, 'Source Website'] = final_site_name

    # Filter out any rows where we still couldn't determine the source
    tenders_to_check.dropna(subset=['Source Website'], inplace=True)
    tenders_to_check = tenders_to_check[tenders_to_check['Source Website'] != '']

    if tenders_to_check.empty:
        print("Could not determine the source website for any of the new tenders. Aborting status check.")
        return
    # --- END of new inference logic ---

    # Add/update columns for status tracking
    if 'Status Last Checked' not in df_custom.columns:
        df_custom['Status Last Checked'] = pd.NaT
    
    # 3. Initialize Selenium and process
    driver = None
    global captcha_solved_in_session
    try:
        driver = initialize_selenium_driver()
        wait = WebDriverWait(driver, 20)
        
        for site_name, group in tenders_to_check.groupby('Source Website'):
            site_id = next((k for k, v in website_options.items() if v['name'] == site_name), None)

            if website_choice in website_options.keys() and site_id != website_choice:
                continue

            status_url = website_options.get(site_id, {}).get('status_url')
            if not status_url:
                print(f"Warning: No status check URL configured for '{site_name}'. Skipping {len(group)} tenders.")
                continue

            print(f"\nProcessing {len(group)} tenders for '{site_name}'...")

            for index, tender_row in group.iterrows():
                tender_id = tender_row['Tender Id']
                print(f"  Checking status for Tender ID: {tender_id}")

                try:
                    driver.get(status_url)
                    time.sleep(2)
                    captcha_solved_in_session = False

                    tender_id_input = wait.until(EC.presence_of_element_located((By.ID, "tenderId")))
                    tender_id_input.clear()
                    tender_id_input.send_keys(tender_id)

                    if not handle_captcha_interaction(driver, f"Status Check for {tender_id}", submit_button_id="Search", success_element_id="tabList"):
                        print(f"    Could not solve CAPTCHA for {tender_id}. Skipping.")
                        df_custom.loc[df_custom['Tender Id'] == tender_id, 'Current Status'] = "CAPTCHA Failed"
                        df_custom.loc[df_custom['Tender Id'] == tender_id, 'Status Last Checked'] = datetime.now()
                        continue
                    
                    try:
                        table = wait.until(EC.presence_of_element_located((By.ID, "tabList")))
                        header_cells = table.find_elements(By.CSS_SELECTOR, "tr.list_header td")
                        headers = [cell.text.strip() for cell in header_cells]
                        if not headers: headers = [cell.text.strip() for cell in table.find_elements(By.XPATH, ".//tbody/tr[1]/td")]
                        header_map = {header: i for i, header in enumerate(headers)}
                        
                        data_rows = table.find_elements(By.XPATH, ".//tbody/tr[contains(@class, 'even') or contains(@class, 'odd')]")
                        if data_rows:
                            first_row_cells = data_rows[0].find_elements(By.TAG_NAME, "td")
                            def get_cell_data(header_name):
                                if header_name in header_map:
                                    idx = header_map[header_name]; return first_row_cells[idx].text.strip() if idx < len(first_row_cells) else None
                                return None

                            scraped_status = get_cell_data('Tender Stage'); print(f"    Found Status: {scraped_status}")
                            row_mask = df_custom['Tender Id'] == tender_id
                            df_custom.loc[row_mask, 'Current Status'] = scraped_status
                            df_custom.loc[row_mask, 'Status Last Checked'] = datetime.now()
                            df_custom.loc[row_mask, 'Source Website'] = site_name
                            
                            if (df_custom.loc[row_mask, 'Tender Title'].isna() | (df_custom.loc[row_mask, 'Tender Title'] == '')).any(): df_custom.loc[row_mask, 'Tender Title'] = get_cell_data('Title and Ref.No.')
                            if (df_custom.loc[row_mask, 'Organisation Chain'].isna() | (df_custom.loc[row_mask, 'Organisation Chain'] == '')).any(): df_custom.loc[row_mask, 'Organisation Chain'] = get_cell_data('Organisation Chain')
                            
                            target_statuses = ["Financial Bid Opening", "Financial Evaluation", "AOC", "Concluded"]
                            if scraped_status in target_statuses: print(f"    Status '{scraped_status}' found. Checking for result files..."); download_tender_result_files(driver, tender_id)
                        else:
                            print("    No data row found in the results table. Leaving status blank.")
                            # Leave status blank to allow re-checking on the next run.
                            # Do not update any other fields, including 'Status Last Checked'.
                            df_custom.loc[df_custom['Tender Id'] == tender_id, 'Current Status'] = ""
                    except TimeoutException:
                        if "No Records Found" in driver.page_source:
                            print(f"    Tender ID {tender_id} not found on the portal. Leaving row blank.")
                            # Do not update the DataFrame for this tender ID, leaving it for the next run
                        else:
                            status = "Error or Not Found"
                            df_custom.loc[df_custom['Tender Id'] == tender_id, 'Current Status'] = status; df_custom.loc[df_custom['Tender Id'] == tender_id, 'Status Last Checked'] = datetime.now()
                except Exception as e: print(f"    An unexpected error occurred while checking {tender_id}: {e}"); df_custom.loc[df_custom['Tender Id'] == tender_id, 'Current Status'] = "Processing Error"; df_custom.loc[df_custom['Tender Id'] == tender_id, 'Status Last Checked'] = datetime.now(); continue
    except WebDriverException as e: print(f"A Selenium WebDriver error occurred: {e}")
    except Exception as e: print(f"An error occurred during the overall status check process: {e}")
    finally:
        if driver: driver.quit(); print("\nSelenium WebDriver closed.")

    # Ensure date columns are in datetime format for correct Excel formatting
    if 'Status Last Checked' in df_custom.columns:
        df_custom['Status Last Checked'] = pd.to_datetime(df_custom['Status Last Checked'], errors='coerce')

    print("\nUpdating data source and Excel file with new statuses...")
    df_custom['Status Last Checked'] = pd.to_datetime(df_custom['Status Last Checked']).dt.strftime('%Y-%m-%d %H:%M:%S').replace('NaT', '')
    write_sheet_to_csv(df_custom, CUSTOM_TENDERS_SHEET)
    update_excel_sheet(df_custom, CUSTOM_TENDERS_SHEET, filename)
    print("\n--- Custom List Tender Status Check Complete. ---")

def tender_status_submenu(website_choice, filename):
    """Displays a submenu for different ways to check tender status."""
    while True:
        print("\n--- Check Tender Status & Results Sub-Menu ---")
        print("1. From 'Expired Tenders' Sheet (tenders marked with 'Yes')")
        print("2. From 'Custom Tenders' Sheet (for unchecked tenders)")
        print("3. By Manual Tender ID Input")
        print("4. By Bulk Search (for selected org)")
        print("5. Back to Main Menu")
        
        choice = input("Enter your choice (1-5): ").strip()
        
        if choice == '1':
            check_expired_tender_status(website_choice, filename)
            break
        elif choice == '2':
            check_custom_list_status(website_choice, filename)
            break
        elif choice == '3':
            check_custom_tender_status(website_choice, filename)
            break
        elif choice == '4':
            bulk_check_tender_results(website_choice, filename)
            break
        elif choice == '5':
            break
        else:
            print("Invalid choice. Please try again.")

# --- NEW: Process and Download Tenders ---
def process_and_download_tenders(website_choice, filename=excel_filename):
    """
    Reads selected tenders from the dynamically identified tender sheets (where 'Download' is 'Yes'),
    and then downloads the tender notice and zip files for each.
    Manages a single Selenium session for efficiency and CAPTCHA handling.
    """
    global global_organisations_data
    global captcha_solved_in_session # Access the global flag

    if not global_organisations_data:
        if not load_organisations_data():
            print("Error: Organization data (global_organizations_data.json) not found or could not be loaded. Please ensure you've run an 'Export Organizations' option (1, 2, or 3) at least once.")
            return

    print(f"\n--- Starting Tender Download Process ---")
    os.makedirs(BASE_DOWNLOAD_DIRECTORY, exist_ok=True) # Ensure base download directory exists

    # Load all sheets from the Excel file to dynamically find tender lists
    all_tenders_for_download = []
    
    try:
        xls = pd.ExcelFile(filename)
        sheet_names = xls.sheet_names

        # Determine which tender sheets to read from based on website choice
        target_sheets = []
        if website_choice in website_options.keys():
            target_sheets = [website_options[website_choice]['name']]
        else: # All
            target_sheets = [site_info['name'] for site_info in sorted(website_options.values(), key=lambda x: x.get('name', ''))]

        tender_data_sheets = [
            s for s in sheet_names
            if s in target_sheets and not s.endswith('-O') and s != EXPIRED_TENDERS_SHEET
        ]

        if not tender_data_sheets:
            if website_choice in [1, 2, 3]:
                site_name = website_options[website_choice]['name']
                print(f"Error: Tender data sheet for '{site_name}' not found in '{filename}'. Please run the 'Process and Export Tender Details' option first.")
            else:
                print(f"Error: No tender data sheets found in '{filename}'. Please run the 'Process and Export Tender Details' option first.")
            return

        for sheet_name in tender_data_sheets:
            # Per requirement: First, read from Excel to get user input. Second, save to CSV.
            # Third, read from that CSV to continue processing.
            try:
                try:
                    current_sheet_df_from_excel = pd.read_excel(xls, sheet_name=sheet_name)
                except Exception as e:
                    if "Value must be either numerical" in str(e):
                        print(f"  Warning: Excel filter issue detected in '{sheet_name}'. Using fallback read.")
                        wb_ro = openpyxl.load_workbook(filename, read_only=True, data_only=True)
                        if sheet_name in wb_ro.sheetnames:
                            data = list(wb_ro[sheet_name].values)
                            if data:
                                current_sheet_df_from_excel = pd.DataFrame(data[1:], columns=data[0])
                            else:
                                current_sheet_df_from_excel = pd.DataFrame()
                        wb_ro.close()
                    else:
                        raise e
                current_sheet_df_from_excel = pd.read_excel(xls, sheet_name=sheet_name)

                # Also capture the folder hyperlinks from the Excel sheet to ensure they are preserved in the CSV
                hyperlink_map = get_hyperlinks_from_sheet(filename, sheet_name)
                if hyperlink_map and 'Tender Id' in current_sheet_df_from_excel.columns:
                    if '__folder_hyperlink__' not in current_sheet_df_from_excel.columns:
                        current_sheet_df_from_excel['__folder_hyperlink__'] = None
                    # Map the read hyperlinks to the dataframe.
                    current_sheet_df_from_excel['__folder_hyperlink__'] = current_sheet_df_from_excel['Tender Id'].astype(str).map(hyperlink_map).fillna(current_sheet_df_from_excel['__folder_hyperlink__'])

                print(f"Syncing user selections and folder hyperlinks from Excel sheet '{sheet_name}' to CSV database...")
                write_sheet_to_csv(current_sheet_df_from_excel, sheet_name)

                # Now, read from the newly synced CSV to process
                current_sheet_df = read_sheet_from_csv(sheet_name)
                
                if 'Download (Type "Yes" to download)' in current_sheet_df.columns:
                    selected_tenders_from_sheet = current_sheet_df[
                        current_sheet_df['Download (Type "Yes" to download)'].astype(str).str.lower() == 'yes'
                    ]
                    if not selected_tenders_from_sheet.empty:
                        all_tenders_for_download.extend(selected_tenders_from_sheet.to_dict(orient='records'))
            except Exception as e:
                print(f"Error syncing selections for sheet '{sheet_name}': {e}. Skipping this sheet.")
                continue
    except FileNotFoundError:
        print(f"Error: '{filename}' not found. Please run option 4 first to generate tender lists.")
        return
    except Exception as e:
        print(f"Error loading tenders from Excel sheets: {e}")
        return

    if not all_tenders_for_download:
        print("No tenders marked for download across all identified tender sheets.")
        return

    # Remove duplicates based on 'Tender Id' in case a tender appeared in multiple sheets (unlikely but safe)
    df_all_tenders_for_download = pd.DataFrame(all_tenders_for_download)
    df_all_tenders_for_download.drop_duplicates(subset=['Tender Id'], keep='first', inplace=True)
    all_tenders_for_download = df_all_tenders_for_download.to_dict(orient='records')


    print(f"\nIdentified {len(all_tenders_for_download)} unique tenders marked for download.")

    # Initialize Selenium driver once for all downloads
    driver = None
    try:
        driver = initialize_selenium_driver() # Initialize the simplified driver
        
        # Determine the initial URL to establish the session
        # If a specific website is chosen, use its URL. For 'All', default to MahaTenders.
        initial_url_choice = website_choice if website_choice in website_options.keys() else 1
        initial_url = website_options[initial_url_choice]['url']

        # Initial navigation to homepage to establish session
        print(f"Establishing initial Selenium session by navigating to: {initial_url}")
        driver.get(initial_url)
        time.sleep(5) # Give time for the page to load

        captcha_solved_in_session = False # Reset flag for this new session
        download_log = load_download_log()
        # six_hours_ago = datetime.now() - timedelta(hours=6) # 6-hour check disabled.

        for tender_info in all_tenders_for_download:
            tender_id = str(tender_info.get('Tender Id'))
            
            last_checked_str = download_log.get(tender_id)
            
            download_mode = 'full' # Default to full download
            if last_checked_str:
                # 6-hour skip logic is disabled. If a tender has been downloaded before,
                # always re-check it in 'update' mode for new files (like corrigendums).
                print(f"\nRe-checking Tender ID {tender_id} for updates (Pre-bid/Corrigendum)...")
                download_mode = 'update'
            
            download_success = download_all_tender_files(driver, tender_info, BASE_DOWNLOAD_DIRECTORY, download_mode)

            if download_success:
                # --- Add hyperlink to Excel after download attempt ---
                if tender_id and str(tender_id) != 'N/A':
                    tender_folder_path = os.path.join(BASE_DOWNLOAD_DIRECTORY, str(tender_id))
                    abs_folder_path = os.path.abspath(tender_folder_path)
                    update_tender_hyperlink(tender_id, abs_folder_path, filename)
                # --- End hyperlink logic ---

                download_log[tender_id] = datetime.now().isoformat()
                save_download_log(download_log)
            else:
                print(f"  Download incomplete/failed for Tender ID {tender_id}. Log and hyperlink not updated.")
        
    except Exception as e:
        print(f"An error occurred during the overall download process: {e}")
    finally:
        if driver:
            driver.quit()
            print("\nSelenium WebDriver closed.")
    
    print(f"\n--- Tender Download Process Complete. Check '{BASE_DOWNLOAD_DIRECTORY}' folder. ---")

def main():
    print("Welcome to the Government Tender Scraper!")

    # New: One-time migration from existing Excel to new CSV format on startup
    sync_excel_to_csv_on_startup()

    # NEW: Load websites from file, or create with defaults
    load_websites_data()

    try:
        while True:
            # --- Step 1: Select Website ---
            print("\n--- Step 1: Select Website ---")
            for key, value in sorted(website_options.items()):
                print(f"{key}. {value['name']}")
            
            all_websites_key = (max(website_options.keys()) if website_options else 0) + 1
            add_website_key = all_websites_key + 1
            delete_website_key = add_website_key + 1
            quit_key = delete_website_key + 1

            print(f"{all_websites_key}. All Websites")
            print(f"{add_website_key}. Add a new website")
            print(f"{delete_website_key}. Delete a website")
            print(f"{quit_key}. Quit")

            website_choice_str = input(f"Enter your choice (1-{quit_key}): ").strip()

            if not website_choice_str.isdigit():
                print("Invalid choice. Please enter a number.")
                continue
            
            website_choice = int(website_choice_str)

            if website_choice == quit_key:
                print("Exiting.")
                break
            elif website_choice == add_website_key:
                add_website()
                continue
            elif website_choice == delete_website_key:
                delete_website()
                continue
            
            # Check if it's a valid website choice (specific or 'All')
            if website_choice != all_websites_key and website_choice not in website_options:
                print("Invalid choice. Please try again.")
                continue

            # --- Step 2: Select Function ---
            print("\n--- Step 2: Select Function ---")
            print("1. Fetch Organisations")
            print("2. Process and Export Tender Details to Excel")
            print("3. Download Selected Tenders")
            print("4. Update Bookmarked Tenders File")
            print("5. Check Tender Status & Results")
            print("6. Back to Website Selection")

            action_choice_str = input("Enter your choice (1-6): ").strip()

            if action_choice_str == '6':
                continue # Go back to the start of the while loop

            if not action_choice_str.isdigit() or int(action_choice_str) not in range(1, 6):
                print("Invalid choice. Please try again.")
                continue

            action_choice = int(action_choice_str)

            # --- Execute Action ---
            if action_choice == 1: # Fetch Organisations
                log_activity(f"Action: Fetch Organisations, Website Choice: {website_choice}")
                if website_choice in website_options.keys():
                    selected_website_url = website_options[website_choice]['url']
                    selected_website_name = website_options[website_choice]['name']
                    print(f"\nFetching organisations for {selected_website_name}...")
                    organisations = fetch_organisations(selected_website_url)
                    update_and_export_organizations(organisations, excel_filename,
                                                    org_sheet_name=f"{selected_website_name}-O")
                elif website_choice == all_websites_key: # All
                    print("\nFetching organisations for all websites...")
                    for choice_int in sorted(website_options.keys()):
                        selected_website_url = website_options[choice_int]['url']
                        selected_website_name = website_options[choice_int]['name']
                        print(f"\n--- Processing: {selected_website_name} ---")
                        organisations = fetch_organisations(selected_website_url)
                        update_and_export_organizations(organisations, excel_filename,
                                                        org_sheet_name=f"{selected_website_name}-O")

            elif action_choice == 2: # Process and Export Tender Details
                log_activity(f"Action: Process and Export Tenders, Website Choice: {website_choice}")
                process_and_export_tenders(website_choice, filename=excel_filename)

            elif action_choice == 3: # Download Tenders
                log_activity(f"Action: Download Tenders, Website Choice: {website_choice}")
                process_and_download_tenders(website_choice)

            elif action_choice == 4: # Update Bookmarked Tenders
                process_bookmarked_tenders()

            elif action_choice == 5: # Check Tender Status & Results
                log_activity(f"Action: Check Tender Status & Results, Website Choice: {website_choice}")
                tender_status_submenu(website_choice, excel_filename)

            print("\nOperation complete. Returning to main menu.")
    except KeyboardInterrupt:
        print("\n\nProgram interrupted by user. Exiting gracefully...")


if __name__ == "__main__":
    main()
    print("\nScraper has shut down.")