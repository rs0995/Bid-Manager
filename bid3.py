import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkinter import font as tkfont
import sqlite3
import os
import sys
import shutil
import subprocess
import platform
import datetime
import threading
import queue
import time
import json
import re
import io
import csv
import zipfile
import hashlib
import textwrap
from urllib.parse import urlparse, urljoin, parse_qs, parse_qsl, urlencode, urlunparse

try:
    from tkcalendar import DateEntry
    CALENDAR_SUPPORT = True
except ImportError:
    DateEntry = None
    CALENDAR_SUPPORT = False

# --- External Libraries for Scraper ---
try:
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    from PIL import Image, ImageTk
    import google.generativeai as genai
    from selenium import webdriver
    from selenium.webdriver.firefox.service import Service as FirefoxService
    from selenium.webdriver.firefox.options import Options as FirefoxOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, WebDriverException, StaleElementReferenceException, NoSuchElementException
    from webdriver_manager.firefox import GeckoDriverManager
    
    SCRAPER_AVAILABLE = True
except ImportError as e:
    print(f"Scraper dependencies missing: {e}")
    SCRAPER_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# --- CONFIGURATION ---
APP_NAME = "BidManager"

def _get_app_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))

def _safe_getcwd():
    try:
        return os.getcwd()
    except Exception:
        return _get_app_base_dir()

def _resolve_path(raw, fallback=""):
    txt = str(raw or "").strip()
    if not txt:
        txt = str(fallback or "").strip()
    if not txt:
        return _get_app_base_dir()
    if os.path.isabs(txt):
        return os.path.normpath(txt)
    return os.path.normpath(os.path.join(_get_app_base_dir(), txt))

APP_CONFIG_DIR = _get_app_base_dir()
APP_PATHS_FILE = os.path.join(APP_CONFIG_DIR, "app_paths.json")
USER_SETTINGS_FILE = os.path.join(APP_CONFIG_DIR, "user_settings.json")
LEGACY_LOCALAPPDATA_DIR = os.path.join(
    os.getenv("LOCALAPPDATA") or os.path.expanduser("~"),
    APP_NAME
)
LEGACY_LOCALAPPDATA_USER_SETTINGS_FILE = os.path.join(LEGACY_LOCALAPPDATA_DIR, "user_settings.json")
DB_FILE = "tender_manager.db"
ROOT_FOLDER = "My_Tender_Projects"
BASE_DOWNLOAD_DIRECTORY = "Tender_Downloads"
GOOGLE_API_KEY = "AIzaSyD78VcTNJCh3qlSYN9ZcLl4MdA3Q88TXQU"  # API Key from tender_scraper.py

def _normalize_db_file(raw):
    txt = str(raw or "").strip()
    if not txt:
        return _resolve_path("tender_manager.db")
    if not txt.lower().endswith(".db"):
        return _resolve_path(os.path.join(txt, "tender_manager.db"))
    return _resolve_path(txt)

def load_app_paths_config():
    defaults = {
        "db_file": _normalize_db_file(DB_FILE),
        "root_folder": _resolve_path(ROOT_FOLDER),
        "download_folder": _resolve_path(BASE_DOWNLOAD_DIRECTORY),
    }
    if not os.path.exists(APP_PATHS_FILE):
        return defaults
    try:
        with open(APP_PATHS_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return {
            "db_file": _normalize_db_file(raw.get("db_file", defaults["db_file"])),
            "root_folder": _resolve_path(raw.get("root_folder", defaults["root_folder"])),
            "download_folder": _resolve_path(raw.get("download_folder", defaults["download_folder"])),
        }
    except Exception:
        return defaults

def save_app_paths_config(db_file, root_folder, download_folder):
    payload = {
        "db_file": _normalize_db_file(db_file),
        "root_folder": _resolve_path(root_folder),
        "download_folder": _resolve_path(download_folder),
    }
    os.makedirs(APP_CONFIG_DIR, exist_ok=True)
    with open(APP_PATHS_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

def _load_user_settings():
    candidates = [USER_SETTINGS_FILE, LEGACY_LOCALAPPDATA_USER_SETTINGS_FILE]
    for cand in candidates:
        if not os.path.exists(cand):
            continue
        try:
            with open(cand, "r", encoding="utf-8") as f:
                raw = json.load(f)
            if isinstance(raw, dict):
                return raw
        except Exception:
            continue
    return {}

def _save_user_settings(settings):
    os.makedirs(APP_CONFIG_DIR, exist_ok=True)
    tmp_file = USER_SETTINGS_FILE + ".tmp"
    with open(tmp_file, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=2)
    os.replace(tmp_file, USER_SETTINGS_FILE)

_user_settings_lock = threading.Lock()
_user_settings_cache = _load_user_settings()

def get_user_setting(key, default=None):
    with _user_settings_lock:
        return _user_settings_cache.get(key, default)

def set_user_setting(key, value):
    with _user_settings_lock:
        _user_settings_cache[key] = value
        try:
            _save_user_settings(_user_settings_cache)
        except Exception:
            pass

_path_cfg = load_app_paths_config()
DB_FILE = _path_cfg["db_file"]
ROOT_FOLDER = _path_cfg["root_folder"]
BASE_DOWNLOAD_DIRECTORY = _path_cfg["download_folder"]

try:
    os.makedirs(os.path.dirname(DB_FILE) or ".", exist_ok=True)
except Exception:
    DB_FILE = _resolve_path("tender_manager.db")
    os.makedirs(os.path.dirname(DB_FILE) or ".", exist_ok=True)

try:
    os.makedirs(ROOT_FOLDER, exist_ok=True)
except Exception:
    ROOT_FOLDER = _resolve_path("My_Tender_Projects")
    os.makedirs(ROOT_FOLDER, exist_ok=True)

try:
    os.makedirs(BASE_DOWNLOAD_DIRECTORY, exist_ok=True)
except Exception:
    BASE_DOWNLOAD_DIRECTORY = _resolve_path("Tender_Downloads")
    os.makedirs(BASE_DOWNLOAD_DIRECTORY, exist_ok=True)

# --- THREAD COMMUNICATION ---
log_queue = queue.Queue()
captcha_req_queue = queue.Queue()
captcha_res_queue = queue.Queue()

def log_to_gui(message):
    """Puts a message into the queue for the GUI to display."""
    log_queue.put(message)
    print(message)

# --- DATABASE LAYER ---
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    # Existing tables
    c.execute('''CREATE TABLE IF NOT EXISTS projects (id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT, client_name TEXT, deadline TEXT, description TEXT, folder_path TEXT, project_value TEXT, prebid TEXT, status TEXT DEFAULT 'Active', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS checklist_items (id INTEGER PRIMARY KEY AUTOINCREMENT, project_id INTEGER, sr_no INTEGER, req_file_name TEXT, description TEXT, subfolder TEXT DEFAULT 'Main', linked_file_path TEXT, status TEXT DEFAULT 'Pending', FOREIGN KEY(project_id) REFERENCES projects(id))''')
    
    # New tables for Scraper
    c.execute('''CREATE TABLE IF NOT EXISTS websites (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        url TEXT,
        status_url TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS organizations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        website_id INTEGER,
        name TEXT,
        tender_count INTEGER,
        tenders_url TEXT,
        is_selected INTEGER DEFAULT 0,
        FOREIGN KEY(website_id) REFERENCES websites(id),
        UNIQUE(website_id, name)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS tenders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        website_id INTEGER,
        org_chain TEXT,
        tender_ref_no TEXT,
        tender_id TEXT UNIQUE,
        title TEXT,
        tender_value TEXT,
        emd TEXT,
        closing_date TEXT,
        opening_date TEXT,
        tender_url TEXT,
        status TEXT,
        is_archived INTEGER DEFAULT 0,
        is_downloaded INTEGER DEFAULT 0,
        is_bookmarked INTEGER DEFAULT 0,
        folder_path TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        location TEXT,
        tender_category TEXT,
        pre_bid_meeting_date TEXT,
        work_description TEXT,
        normalized_tender_url TEXT,
        FOREIGN KEY(website_id) REFERENCES websites(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS downloaded_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tender_id TEXT,
        file_name TEXT,
        file_type TEXT DEFAULT 'document',
        source_url TEXT,
        local_path TEXT,
        downloaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(tender_id) REFERENCES tenders(tender_id)
    )''')

    # Seed Websites if empty
    c.execute("SELECT count(*) FROM websites")
    if c.fetchone()[0] == 0:
        websites = [
            ("MahaTenders", "https://mahatenders.gov.in/nicgep/app?page=FrontEndTendersByOrganisation&service=page", "https://mahatenders.gov.in/nicgep/app?page=WebTenderStatusLists&service=page"),
            ("ETenders", "https://etenders.gov.in/eprocure/app?page=FrontEndTendersByOrganisation&service=page", "https://etenders.gov.in/eprocure/app?page=WebTenderStatusLists&service=page"),
            ("Eprocure", "https://eprocure.gov.in/eprocure/app?page=FrontEndTendersByOrganisation&service=page", "https://eprocure.gov.in/eprocure/app?page=WebTenderStatusLists&service=page")
        ]
        c.executemany("INSERT INTO websites (name, url, status_url) VALUES (?, ?, ?)", websites)

    # Migrations for existing databases
    migrations = ["location", "tender_category", "pre_bid_meeting_date", "last_downloaded_at", "normalized_tender_url", "work_description"]
    for col in migrations:
        try:
            c.execute(f"ALTER TABLE tenders ADD COLUMN {col} TEXT")
        except sqlite3.OperationalError:
            pass
    try:
        c.execute("ALTER TABLE tenders ADD COLUMN is_archived INTEGER DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    # Backfill legacy archive marker stored in status.
    c.execute("UPDATE tenders SET is_archived=1 WHERE COALESCE(status,'')='Archived'")
    c.execute("UPDATE tenders SET status='' WHERE COALESCE(status,'')='Archived'")
    c.execute("UPDATE tenders SET status='' WHERE COALESCE(status,'')='Active'")
    c.execute("CREATE INDEX IF NOT EXISTS idx_tenders_site_normurl ON tenders(website_id, normalized_tender_url)")
    download_migrations = [
        ("file_type", "TEXT DEFAULT 'document'"),
        ("source_url", "TEXT"),
        ("local_path", "TEXT"),
    ]
    for col, ddl in download_migrations:
        try:
            c.execute(f"ALTER TABLE downloaded_files ADD COLUMN {col} {ddl}")
        except sqlite3.OperationalError:
            pass
    try:
        c.execute("ALTER TABLE projects ADD COLUMN source_tender_id TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE projects ADD COLUMN project_value TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE projects ADD COLUMN prebid TEXT")
    except sqlite3.OperationalError:
        pass
    c.execute(
        """CREATE UNIQUE INDEX IF NOT EXISTS idx_projects_source_tender_unique
           ON projects(source_tender_id)
           WHERE source_tender_id IS NOT NULL AND TRIM(source_tender_id) != ''"""
    )
    c.execute("DROP INDEX IF EXISTS idx_downloaded_file_unique")
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_downloaded_file_unique ON downloaded_files(tender_id, file_name, file_type)")
    c.execute('''CREATE TABLE IF NOT EXISTS app_settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS auto_archive_runs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        run_at_utc TEXT NOT NULL,
        status TEXT NOT NULL,
        archived_count INTEGER DEFAULT 0,
        archived_status_updated INTEGER DEFAULT 0,
        websites_count INTEGER DEFAULT 0,
        notes TEXT
    )''')
    c.execute("CREATE INDEX IF NOT EXISTS idx_auto_archive_runs_ts ON auto_archive_runs(run_at_utc)")

    conn.commit()
    conn.close()

def on_combo_key_cycling(event):
    if event.char.isalnum():
        char = event.char.lower()
        widget = event.widget
        values = widget['values']
        if not values: return
        current = widget.get()
        start_idx = values.index(current) + 1 if current in values else 0
        for i in range(start_idx, len(values)):
            if values[i].lower().startswith(char):
                widget.set(values[i]); widget.event_generate("<<ComboboxSelected>>"); return "break"
        for i in range(0, start_idx):
            if values[i].lower().startswith(char):
                widget.set(values[i]); widget.event_generate("<<ComboboxSelected>>"); return "break"

def clear_tree_rows(tree):
    for item in tree.get_children():
        tree.delete(item)

def ensure_project_standard_folders(project_root):
    root = resolve_project_folder_path(project_root, os.path.basename(str(project_root or "").strip()))
    try:
        os.makedirs(root, exist_ok=True)
    except Exception:
        leaf = os.path.basename(str(root or "").rstrip("\\/")) or "Project"
        leaf = re.sub(r'[<>:"/\\|?*]+', "_", leaf).strip(" .") or "Project"
        root = os.path.join(ROOT_FOLDER, leaf)
        os.makedirs(root, exist_ok=True)
    ready = os.path.join(root, "Ready Docs")
    tender = os.path.join(root, "Tender Docs")
    working = os.path.join(root, "Working Docs")
    for p in (ready, tender, working):
        os.makedirs(p, exist_ok=True)
    return {
        "project_root": root,
        "ready_docs": ready,
        "tender_docs": tender,
        "working_docs": working,
    }

def resolve_project_folder_path(saved_path, project_title=""):
    raw = str(saved_path or "").strip()
    title = str(project_title or "").strip()
    if not raw:
        base = re.sub(r'[<>:"/\\|?*]+', "_", title).strip(" .") or "Project"
        return os.path.join(ROOT_FOLDER, base)
    drive, _ = os.path.splitdrive(raw)
    if drive and not os.path.exists(drive + os.sep):
        # Saved path points to an unavailable drive; relocate under current ROOT_FOLDER.
        leaf = os.path.basename(raw.rstrip("\\/")) or title or "Project"
        leaf = re.sub(r'[<>:"/\\|?*]+', "_", str(leaf)).strip(" .") or "Project"
        return os.path.join(ROOT_FOLDER, leaf)
    return _resolve_path(raw)

def copy_tree_contents(src_dir, dest_dir, exclude_names=None):
    if not (src_dir and os.path.isdir(src_dir)):
        return 0
    os.makedirs(dest_dir, exist_ok=True)
    excluded = {str(x).strip().lower() for x in (exclude_names or []) if str(x).strip()}
    copied = 0
    for name in os.listdir(src_dir):
        if str(name).strip().lower() in excluded:
            continue
        src = os.path.join(src_dir, name)
        dst = os.path.join(dest_dir, name)
        if os.path.isdir(src):
            shutil.copytree(src, dst, dirs_exist_ok=True)
            copied += 1
        elif os.path.isfile(src):
            shutil.copy2(src, dst)
            copied += 1
    return copied

def resolve_tk_font(font_cfg, fallback=("Segoe UI", 10)):
    cfg = font_cfg or fallback
    if isinstance(cfg, str):
        try:
            return tkfont.nametofont(cfg)
        except Exception:
            return tkfont.Font(font=cfg)
    return tkfont.Font(font=cfg)

def wrap_tree_row_values(tree, cols, row_vals, skip_cols=None, fallback_wrap=None, width_bias=None):
    skip_cols = set(skip_cols or [])
    fallback_wrap = fallback_wrap or {}
    width_bias = width_bias or {}
    wrapped = []
    dcols = tree["displaycolumns"]
    if dcols == "#all":
        visible_cols = set(cols)
    elif isinstance(dcols, str):
        visible_cols = {dcols}
    else:
        visible_cols = set(dcols)
    avg_char_px = 7.0
    try:
        style_name = tree.cget("style") or "Treeview"
        font_cfg = ttk.Style().lookup(style_name, "font") or ("Segoe UI", 10)
        fnt = resolve_tk_font(font_cfg)
        sample = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 "
        avg_char_px = max(6.0, fnt.measure(sample) / len(sample))
    except Exception:
        pass
    for idx, col in enumerate(cols):
        val = row_vals[idx] if idx < len(row_vals) else ""
        if col in skip_cols:
            wrapped.append(val)
            continue
        text = " ".join(str(val or "").split())
        if col not in visible_cols:
            wrapped.append(text)
            continue
        try:
            px = int(tree.column(col, "width"))
            width = max(8, int(((px - 14) / avg_char_px) * width_bias.get(col, 1.0)))
        except Exception:
            width = max(8, int(fallback_wrap.get(col, 24)))
        wrapped.append(textwrap.fill(text, width=width, break_long_words=True, break_on_hyphens=True))
    return tuple(wrapped)

# --- COLUMN MANAGER ---
class ColumnManager:
    def __init__(self, tree, parent_window, on_apply_callback=None):
        self.tree = tree
        self.parent = parent_window
        self.on_apply_callback = on_apply_callback
        self.tree.bind("<Button-3>", self.on_right_click) # Windows/Linux
        if platform.system() == "Darwin":
            self.tree.bind("<Button-2>", self.on_right_click)

    def on_right_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region == "heading":
            menu = tk.Menu(self.parent, tearoff=0)
            menu.add_command(label="Column Settings...", command=self.open_settings)
            menu.post(event.x_root, event.y_root)

    def open_settings(self):
        win = tk.Toplevel(self.parent)
        win.title("Column Settings")
        win.geometry("460x560")

        all_cols = list(self.tree['columns'])
        display_cols = self.tree['displaycolumns']
        if display_cols == '#all' or (isinstance(display_cols, (tuple, list)) and '#all' in display_cols):
            display_cols = all_cols
        else:
            display_cols = list(display_cols)

        # Keep current on-screen order stable across repeated edits.
        display_cols = [c for c in display_cols if c in all_cols]
        hidden_cols = [c for c in all_cols if c not in display_cols]
        ordered_cols = display_cols + hidden_cols

        current_order = []
        visible_set = set(display_cols)
        for col in ordered_cols:
            current_order.append({'id': col, 'visible': col in visible_set, 'name': self.tree.heading(col, 'text')})

        frame_list = tk.Frame(win)
        frame_list.pack(fill="both", expand=True, padx=5, pady=5)
        canvas = tk.Canvas(frame_list)
        scrollbar = ttk.Scrollbar(frame_list, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _on_mousewheel(event):
            delta = event.delta
            if platform.system() == "Darwin":
                canvas.yview_scroll(int(-delta), "units")
            else:
                canvas.yview_scroll(int(-delta / 120), "units")
            return "break"

        def _on_linux_scroll_up(event):
            canvas.yview_scroll(-3, "units")
            return "break"

        def _on_linux_scroll_down(event):
            canvas.yview_scroll(3, "units")
            return "break"

        canvas.bind("<Enter>", lambda e: canvas.focus_set())
        canvas.bind("<MouseWheel>", _on_mousewheel)
        canvas.bind("<Button-4>", _on_linux_scroll_up)
        canvas.bind("<Button-5>", _on_linux_scroll_down)

        row_frames = []
        drag_state = {"from_idx": None}

        def swap(i, j):
            current_order[i], current_order[j] = current_order[j], current_order[i]
            refresh_ui()

        def move_item(i, j):
            if i == j or i < 0 or j < 0 or i >= len(current_order) or j >= len(current_order):
                return
            item = current_order.pop(i)
            current_order.insert(j, item)
            refresh_ui()

        def get_drop_index():
            if not row_frames:
                return 0
            win.update_idletasks()
            py = win.winfo_pointery()
            for idx, rf in enumerate(row_frames):
                mid = rf.winfo_rooty() + (rf.winfo_height() // 2)
                if py < mid:
                    return idx
            return len(row_frames) - 1

        def on_drag_start(idx, _event=None):
            drag_state["from_idx"] = idx

        def on_drag_release(_event=None):
            src = drag_state.get("from_idx")
            drag_state["from_idx"] = None
            if src is None:
                return
            dst = get_drop_index()
            move_item(src, dst)

        def refresh_ui():
            nonlocal row_frames
            for w in scrollable_frame.winfo_children():
                w.destroy()
            row_frames = []
            tk.Label(
                scrollable_frame,
                text="Use [Drag] to reorder columns with mouse",
                anchor="w",
                fg="#444"
            ).pack(fill="x", padx=4, pady=(0, 4))
            for i, item in enumerate(current_order):
                row = tk.Frame(scrollable_frame)
                row.pack(fill="x", pady=2)
                row_frames.append(row)
                var = tk.BooleanVar(value=item['visible'])

                def on_check(col_id=item['id'], v=var):
                    for x in current_order:
                        if x['id'] == col_id:
                            x['visible'] = v.get()
                            break

                tk.Checkbutton(row, variable=var, command=on_check).pack(side="left")
                tk.Label(row, text=item['name'], anchor="w").pack(side="left", fill="x", expand=True, padx=(2, 4))

                drag_lbl = tk.Label(row, text="[Drag]", fg="#1565C0", cursor="fleur")
                drag_lbl.pack(side="right", padx=(4, 2))
                drag_lbl.bind("<ButtonPress-1>", lambda e, idx=i: on_drag_start(idx, e))
                drag_lbl.bind("<ButtonRelease-1>", on_drag_release)

                if i > 0:
                    tk.Button(row, text="Up", command=lambda idx=i: swap(idx, idx - 1), width=3).pack(side="right")
                if i < len(current_order) - 1:
                    tk.Button(row, text="Dn", command=lambda idx=i: swap(idx, idx + 1), width=3).pack(side="right")

        refresh_ui()
        win.bind("<ButtonRelease-1>", on_drag_release, add="+")

        def apply():
            self.tree['displaycolumns'] = [x['id'] for x in current_order if x['visible']]
            if self.on_apply_callback:
                self.on_apply_callback()
            win.destroy()

        tk.Button(win, text="Apply Changes", command=apply, bg="#4CAF50", fg="white").pack(fill="x", padx=10, pady=10)

# --- SCRAPER BACKEND ---
class ScraperBackend:
    captcha_solved_in_session = False
    gemini_model = None
    gemini_model_name = None
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWeb7Kit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'DNT': '1'
    })

    @staticmethod
    def add_website_logic(name, url, status_url):
        conn = sqlite3.connect(DB_FILE)
        try:
            conn.execute("INSERT INTO websites (name, url, status_url) VALUES (?, ?, ?)", (name, url, status_url))
            conn.commit()
            return True
        except Exception as e:
            log_to_gui(f"Error adding website: {e}")
            return False
        finally:
            conn.close()

    @staticmethod
    def delete_website_logic(website_id):
        conn = sqlite3.connect(DB_FILE)
        try:
            conn.execute("DELETE FROM downloaded_files WHERE tender_id IN (SELECT tender_id FROM tenders WHERE website_id=?)", (website_id,))
            conn.execute("DELETE FROM tenders WHERE website_id=?", (website_id,))
            conn.execute("DELETE FROM organizations WHERE website_id=?", (website_id,))
            conn.execute("DELETE FROM websites WHERE id=?", (website_id,))
            conn.commit()
            return True
        except Exception as e:
            log_to_gui(f"Error deleting website: {e}")
            return False
        finally:
            conn.close()

    @staticmethod
    def clear_saved_scraper_details_logic(clear_orgs=False, clear_active=False, clear_archived=False):
        """
        Clears selected scraper data while preserving configured websites.
        - clear_orgs: organizations list/selections
        - clear_active: non-archived tenders (+ their download logs)
        - clear_archived: archived tenders (+ their download logs)
        """
        conn = sqlite3.connect(DB_FILE)
        try:
            c = conn.cursor()
            result = {
                "organizations": 0,
                "active_tenders": 0,
                "archived_tenders": 0,
                "downloaded_files": 0,
            }

            if clear_orgs:
                result["organizations"] = c.execute("SELECT COUNT(*) FROM organizations").fetchone()[0]
                c.execute("DELETE FROM organizations")

            if clear_active:
                active_ids = [r[0] for r in c.execute(
                    "SELECT tender_id FROM tenders WHERE COALESCE(is_archived,0)=0"
                ).fetchall() if r and r[0]]
                result["active_tenders"] = c.execute(
                    "SELECT COUNT(*) FROM tenders WHERE COALESCE(is_archived,0)=0"
                ).fetchone()[0]
                if active_ids:
                    q = ",".join("?" for _ in active_ids)
                    result["downloaded_files"] += c.execute(
                        f"SELECT COUNT(*) FROM downloaded_files WHERE tender_id IN ({q})",
                        tuple(active_ids)
                    ).fetchone()[0]
                    c.execute(f"DELETE FROM downloaded_files WHERE tender_id IN ({q})", tuple(active_ids))
                c.execute("DELETE FROM tenders WHERE COALESCE(is_archived,0)=0")

            if clear_archived:
                arch_ids = [r[0] for r in c.execute(
                    "SELECT tender_id FROM tenders WHERE COALESCE(is_archived,0)=1"
                ).fetchall() if r and r[0]]
                result["archived_tenders"] = c.execute(
                    "SELECT COUNT(*) FROM tenders WHERE COALESCE(is_archived,0)=1"
                ).fetchone()[0]
                if arch_ids:
                    q = ",".join("?" for _ in arch_ids)
                    result["downloaded_files"] += c.execute(
                        f"SELECT COUNT(*) FROM downloaded_files WHERE tender_id IN ({q})",
                        tuple(arch_ids)
                    ).fetchone()[0]
                    c.execute(f"DELETE FROM downloaded_files WHERE tender_id IN ({q})", tuple(arch_ids))
                c.execute("DELETE FROM tenders WHERE COALESCE(is_archived,0)=1")

            conn.commit()
            return result
        except Exception as e:
            log_to_gui(f"Error clearing saved scraper details: {e}")
            return None
        finally:
            conn.close()

    @staticmethod
    def get_websites():
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT id, name, url, status_url FROM websites")
        rows = c.fetchall()
        conn.close()
        return {row[0]: {"name": row[1], "url": row[2], "status_url": row[3]} for row in rows}

    @staticmethod
    def safe_request(url):
        """Performs a request with automatic session refreshing if stale."""
        try:
            response = ScraperBackend.session.get(url, timeout=30)
            
            # Check for stale session using BeautifulSoup to be specific (like tender_scraper.py)
            is_stale = False
            try:
                soup = BeautifulSoup(response.text, 'html.parser')
                if soup.title and ("Stale Session" in soup.title.text or "Error" in soup.title.text):
                    is_stale = True
            except:
                if "Stale Session" in response.text:
                    is_stale = True

            if is_stale:
                log_to_gui("Stale session detected. Refreshing...")
                # Attempt to refresh session by hitting the specific page that resets the session correctly
                if 'app?' in url:
                    base_part = url.split('app?')[0]
                    refresh_url = base_part + 'app?page=FrontEndTendersByOrganisation&service=page'
                else:
                    parsed = urlparse(url)
                    refresh_url = f"{parsed.scheme}://{parsed.netloc}/nicgep/app?page=FrontEndTendersByOrganisation&service=page"
                
                ScraperBackend.session.get(refresh_url, timeout=30)
                time.sleep(2)
                # Retry original request
                response = ScraperBackend.session.get(url, timeout=30)
            return response
        except Exception as e:
            log_to_gui(f"Request failed: {e}")
            # Reset session on connection error to recover for subsequent requests
            ScraperBackend.session = requests.Session()
            ScraperBackend.session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWeb7Kit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'DNT': '1'
            })
            return None

    @staticmethod
    def get_working_gemini_model():
        if not SCRAPER_AVAILABLE:
            return None
        if ScraperBackend.gemini_model is not None:
            return ScraperBackend.gemini_model
        try:
            genai.configure(api_key=GOOGLE_API_KEY)
        except Exception as e:
            log_to_gui(f"Gemini config error: {e}")
            return None

        discovered = []
        try:
            for m in genai.list_models():
                methods = getattr(m, "supported_generation_methods", []) or []
                if "generateContent" in methods:
                    mname = getattr(m, "name", None)
                    if mname:
                        discovered.append(mname)
        except Exception as e:
            log_to_gui(f"Gemini list_models warning: {e}")

        def norm(name):
            return name.split("/", 1)[1] if name.startswith("models/") else name

        discovered_norm = [norm(x) for x in discovered]
        discovered_norm.sort(key=lambda n: (0 if "flash" in n.lower() else 1, 0 if "pro" in n.lower() else 1, n))
        fallback = [
            "gemini-2.0-flash",
            "gemini-2.0-flash-lite",
            "gemini-1.5-flash-latest",
            "gemini-1.5-pro-latest",
            "gemini-1.5-flash",
            "gemini-pro",
        ]
        candidates = []
        seen = set()
        for name in discovered_norm + fallback:
            if name and name not in seen:
                seen.add(name)
                candidates.append(name)

        last_error = None
        for name in candidates:
            try:
                model = genai.GenerativeModel(name)
                # lightweight capability probe
                _ = model.generate_content("ok", generation_config={"temperature": 0})
                ScraperBackend.gemini_model = model
                ScraperBackend.gemini_model_name = name
                log_to_gui(f"Gemini model selected: {name}")
                return model
            except Exception as e:
                last_error = e
                continue

        if last_error:
            log_to_gui(f"Gemini model detection failed: {last_error}")
        return None

    @staticmethod
    def solve_captcha_with_gemini(image_data):
        if not SCRAPER_AVAILABLE: return None
        try:
            model = ScraperBackend.get_working_gemini_model()
            if model is None:
                return None
            img = Image.open(io.BytesIO(image_data))
            response = model.generate_content(["Extract the 6 alphanumeric characters from this CAPTCHA image. Return ONLY the text.", img])
            text = re.sub(r'[^a-zA-Z0-9]', '', response.text.strip())
            return text if len(text) == 6 else None
        except Exception as e:
            log_to_gui(f"Gemini Error: {e}")
            return None

    @staticmethod
    def handle_captcha_interaction(driver, context, submit_id="Submit"):
        def _status_table_visible():
            try:
                table = driver.find_element(By.ID, "tabList")
                if not table.is_displayed():
                    return False
                rows = table.find_elements(By.XPATH, ".//tr[contains(@class, 'even') or contains(@class, 'odd')]")
                return len(rows) > 0
            except Exception:
                return False

        # If session CAPTCHA was solved already, try submitting directly first.
        if ScraperBackend.captcha_solved_in_session:
            try:
                btn = driver.find_element(By.ID, submit_id)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(2)
                if _status_table_visible():
                    return True
                # Session may have expired; fall through to solve CAPTCHA again.
                if len(driver.find_elements(By.ID, "captchaImage")) > 0:
                    ScraperBackend.captcha_solved_in_session = False
            except Exception:
                ScraperBackend.captcha_solved_in_session = False

        # If result table is already visible, treat as solved for this session.
        if _status_table_visible():
            ScraperBackend.captcha_solved_in_session = True
            return True
        try:
            captcha_img = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "captchaImage")))
        except:
            log_to_gui(f"No CAPTCHA found for {context}.")
            return True

        max_retries = 3
        for attempt in range(max_retries):
            try:
                captcha_input = driver.find_element(By.ID, "captchaText")
                btn = driver.find_element(By.ID, submit_id)
                img_data = captcha_img.screenshot_as_png

                # Try Gemini first (auto attempts).
                solution = ScraperBackend.solve_captcha_with_gemini(img_data)
                if not solution:
                    log_to_gui(f"CAPTCHA auto-solve failed for {context}. Retrying ({attempt + 1}/{max_retries})...")
                    try:
                        captcha_img = driver.find_element(By.ID, "captchaImage")
                    except Exception:
                        pass
                    continue

                captcha_input.clear()
                captcha_input.send_keys(solution)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(3)
                
                # Check success: either result table loaded or CAPTCHA image disappeared.
                if _status_table_visible() or len(driver.find_elements(By.ID, "captchaImage")) == 0:
                    log_to_gui("CAPTCHA Solved!")
                    ScraperBackend.captcha_solved_in_session = True
                    return True
                else:
                    log_to_gui("CAPTCHA Failed. Retrying...")
                    captcha_img = driver.find_element(By.ID, "captchaImage") # Re-find
            except Exception as e:
                log_to_gui(f"Captcha Error: {e}")
        # Manual fallback after 3 failed auto attempts.
        try:
            for manual_try in range(3):
                try:
                    captcha_input = driver.find_element(By.ID, "captchaText")
                    btn = driver.find_element(By.ID, submit_id)
                    img_data = driver.find_element(By.ID, "captchaImage").screenshot_as_png
                except Exception:
                    log_to_gui(f"Could not load CAPTCHA image for manual input ({context}).")
                    return False

                log_to_gui(f"Requesting Manual CAPTCHA for {context} (manual try {manual_try + 1}/3)...")
                captcha_req_queue.put(img_data)
                solution = captcha_res_queue.get()
                if not solution:
                    return False  # user cancelled

                captcha_input.clear()
                captcha_input.send_keys(solution)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(3)
                if _status_table_visible() or len(driver.find_elements(By.ID, "captchaImage")) == 0:
                    log_to_gui("CAPTCHA Solved (manual)!")
                    ScraperBackend.captcha_solved_in_session = True
                    return True
                log_to_gui("Manual CAPTCHA incorrect. Retrying manual input...")
            return False
        except Exception as e:
            log_to_gui(f"Captcha manual fallback error: {e}")
            return False

    @staticmethod
    def ensure_download_tables():
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS downloaded_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tender_id TEXT,
            file_name TEXT,
            file_type TEXT DEFAULT 'document',
            source_url TEXT,
            local_path TEXT,
            downloaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(tender_id) REFERENCES tenders(tender_id)
        )''')
        download_migrations = [
            ("file_type", "TEXT DEFAULT 'document'"),
            ("source_url", "TEXT"),
            ("local_path", "TEXT"),
        ]
        for col, ddl in download_migrations:
            try:
                c.execute(f"ALTER TABLE downloaded_files ADD COLUMN {col} {ddl}")
            except sqlite3.OperationalError:
                pass
        c.execute("DROP INDEX IF EXISTS idx_downloaded_file_unique")
        c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_downloaded_file_unique ON downloaded_files(tender_id, file_name, file_type)")
        conn.commit()
        conn.close()

    @staticmethod
    def get_downloaded_file_log(tender_id):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT file_name FROM downloaded_files WHERE tender_id=?", (str(tender_id),))
        rows = c.fetchall()
        conn.close()
        return {r[0] for r in rows}

    @staticmethod
    def log_downloaded_file(tender_id, file_name, file_type="document", source_url=None, local_path=None):
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "INSERT OR IGNORE INTO downloaded_files (tender_id, file_name, file_type, source_url, local_path) VALUES (?, ?, ?, ?, ?)",
            (str(tender_id), str(file_name), str(file_type), source_url, local_path)
        )
        conn.commit()
        conn.close()

    @staticmethod
    def should_skip_file(tender_id, file_name, file_path):
        existing_log = ScraperBackend.get_downloaded_file_log(tender_id)
        file_name = str(file_name)
        if os.path.exists(file_path) and file_name not in existing_log:
            # Backfill legacy files into DB log on first encounter.
            ScraperBackend.log_downloaded_file(tender_id, file_name)
            existing_log.add(file_name)
        return os.path.exists(file_path) and file_name in existing_log

    @staticmethod
    def refresh_selenium_session(driver, init_url, tender_url):
        try:
            driver.get(init_url)
            time.sleep(2)
            driver.get(tender_url)
            time.sleep(2)
            title = (driver.title or "").lower()
            return not ("stale session" in title or title.strip() == "error")
        except Exception:
            return False

    @staticmethod
    def open_tender_page_with_recovery(driver, init_url, tender_url):
        driver.get(tender_url)
        time.sleep(2)
        title = (driver.title or "").lower()
        if "stale session" in title or title.strip() == "error":
            log_to_gui("Stale session detected. Reinitializing Selenium session...")
            if not ScraperBackend.refresh_selenium_session(driver, init_url, tender_url):
                return False
        return True

    @staticmethod
    def fetch_organisations_logic(website_id):
        websites = ScraperBackend.get_websites()
        if website_id not in websites: return
        site_data = websites[website_id]
        url = site_data['url']
        
        log_to_gui(f"Fetching organizations from {url}...")
        try:
            response = ScraperBackend.safe_request(url)
            if not response: return False
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Logic from tender_scraper.py
            org_name_header = soup.find('td', string='Organisation Name')
            if not org_name_header:
                log_to_gui("Could not find Organisation Name header.")
                return

            table = org_name_header.find_parent('table')
            rows = table.find_all('tr')
            
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            
            count = 0
            for row in rows:
                cols = row.find_all('td')
                if len(cols) > 2 and cols[0].text.strip().isdigit():
                    org_name = cols[1].text.strip()
                    tender_count = cols[2].text.strip()
                    link = cols[2].find('a')['href'] if cols[2].find('a') else ""
                    full_link = urljoin(url, link)
                    
                    # Insert or Ignore (to preserve selection status if exists)
                    # We use INSERT OR IGNORE then UPDATE to update details but keep selection
                    c.execute("SELECT id FROM organizations WHERE website_id=? AND name=?", (website_id, org_name))
                    exists = c.fetchone()
                    
                    if exists:
                        c.execute("UPDATE organizations SET tender_count=?, tenders_url=? WHERE id=?", (tender_count, full_link, exists[0]))
                    else:
                        c.execute("INSERT INTO organizations (website_id, name, tender_count, tenders_url) VALUES (?, ?, ?, ?)", 
                                  (website_id, org_name, tender_count, full_link))
                    count += 1
            
            conn.commit()
            conn.close()
            log_to_gui(f"Updated {count} organizations for {site_data['name']}")
            return True
        except Exception as e:
            log_to_gui(f"Error fetching orgs: {e}")
            return False

    @staticmethod
    def get_detail_by_label(soup_obj, label_variations, allow_contains=True):
        def norm(s):
            return " ".join((s or "").replace(":", " ").split()).lower()

        targets = [norm(v) for v in label_variations]
        caption_cells = soup_obj.find_all('td', class_=lambda c: c and 'td_caption' in c)

        # Primary path: exact label match in caption td -> immediate field sibling td.
        for label_td in caption_cells:
            label_text = norm(label_td.get_text(" ", strip=True))
            if label_text in targets:
                value_td = label_td.find_next_sibling('td')
                if value_td:
                    val = value_td.get_text(" ", strip=True)
                    if val:
                        return val

        # Secondary path: exact label match in explicit <b>/<strong> label nodes.
        for tag in soup_obj.find_all(['b', 'strong']):
            text = norm(tag.get_text(" ", strip=True))
            if text in targets:
                label_td = tag.find_parent('td')
                if label_td:
                    value_td = label_td.find_next_sibling('td')
                    if value_td:
                        val = value_td.get_text(" ", strip=True)
                        if val:
                            return val

        if allow_contains:
            # Contains-based fallback for portals that append extra caption text.
            for label_td in caption_cells:
                label_text = norm(label_td.get_text(" ", strip=True))
                if any(t in label_text for t in targets):
                    value_td = label_td.find_next_sibling('td')
                    if value_td:
                        val = value_td.get_text(" ", strip=True)
                        if val:
                            return val
            # Generic td fallback as last resort.
            for td in soup_obj.find_all('td'):
                text = norm(td.get_text(" ", strip=True))
                if any(t in text for t in targets):
                    value_td = td.find_next_sibling('td')
                    if value_td:
                        val = value_td.get_text(" ", strip=True)
                        if val:
                            return val
        return None

    @staticmethod
    def normalize_tender_id(raw_text):
        if not raw_text:
            return None
        txt = " ".join(str(raw_text).split())

        # 1) Explicit bracket format: [ID]
        m = re.search(r'\[([A-Z0-9\-_/.]+)\]', txt, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip()

        # 2) Label format: Tender ID: <id> / Ref No: <id>
        label_patterns = [
            r'(?i)tender\s*id\s*[:\-]?\s*([A-Z0-9\-_/.]+)',
            r'(?i)ref\.?\s*no\.?\s*[:\-]?\s*([A-Z0-9\-_/.]+)',
        ]
        for pat in label_patterns:
            m = re.search(pat, txt)
            if m:
                return m.group(1).strip()

        # 3) Generic token extraction fallback
        tokens = re.findall(r'[A-Z0-9][A-Z0-9\-_/.]{3,}', txt.upper())
        banned = {"TENDER", "TENDERID", "TENDERNO", "REF", "NO", "NIT", "ID"}
        ranked = []
        for t in tokens:
            if t in banned:
                continue
            score = 0
            if any(ch.isdigit() for ch in t):
                score += 2
            if any(ch.isalpha() for ch in t):
                score += 2
            if "/" in t or "-" in t or "_" in t or "." in t:
                score += 1
            ranked.append((score, len(t), t))
        if ranked:
            ranked.sort(reverse=True)
            return ranked[0][2]
        return None

    @staticmethod
    def derive_tender_id(title_text, detail_soup, tender_url):
        # Priority 1: Tender ID from detail page label (same intent as tender_scraper.py)
        if detail_soup is not None:
            raw_detail_id = ScraperBackend.get_detail_by_label(detail_soup, ["Tender ID", "Tender Id"])
            t_id = ScraperBackend.normalize_tender_id(raw_detail_id)
            if t_id:
                return t_id

        # Priority 2: parse from listing title/ref text
        t_id = ScraperBackend.normalize_tender_id(title_text)
        if t_id:
            return t_id

        # Priority 3: try URL query keys commonly used by portals
        try:
            qs = parse_qs(urlparse(tender_url).query)
            for key in ("tenderId", "tenderid", "tid", "id", "refNo"):
                if key in qs and qs[key]:
                    t_id = ScraperBackend.normalize_tender_id(qs[key][0])
                    if t_id:
                        return t_id
        except Exception:
            pass

        # Final fallback: deterministic URL-based id to avoid collisions/overwrites.
        return f"URL_{hashlib.md5(str(tender_url).encode('utf-8')).hexdigest()[:12].upper()}"

    @staticmethod
    def derive_tender_title(listing_title_text, detail_soup):
        """
        Resolve Tender Title from detail page first (same intent as tender_scraper.py),
        then fallback to listing title.
        """
        def clean_title(raw):
            txt = " ".join(str(raw or "").split())
            if not txt:
                return "N/A"
            # Remove trailing ref/id bracket blocks: [...][...]
            prev = None
            while prev != txt:
                prev = txt
                txt = re.sub(r"\s*\[[^\]]{3,}\]\s*$", "", txt).strip()
            if txt.startswith("[") and txt.endswith("]") and len(txt) > 2:
                txt = txt[1:-1].strip()
            return txt or "N/A"

        if detail_soup is not None:
            try:
                # Prefer exact "Title" in Work Item Details.
                for td in detail_soup.find_all("td", class_=lambda c: c and "td_caption" in c):
                    label = " ".join(td.get_text(" ", strip=True).replace(":", " ").split()).lower()
                    if label == "title":
                        value_td = td.find_next_sibling("td", class_="td_field") or td.find_next_sibling("td")
                        if value_td:
                            v = value_td.get_text(" ", strip=True)
                            if v:
                                return clean_title(v)
            except Exception:
                pass
            try:
                # Fallback to broader labels seen on NIC pages.
                v = ScraperBackend.get_detail_by_label(
                    detail_soup,
                    ["Title and Ref.No./Tender ID", "Title and Ref.No.", "Title"]
                )
                if v:
                    return clean_title(v)
            except Exception:
                pass
            try:
                if detail_soup.title and detail_soup.title.string:
                    v = detail_soup.title.string.strip().replace("E-Procurement System :: ", "")
                    if v and len(v) < 220:
                        return clean_title(v)
            except Exception:
                pass
        return clean_title(listing_title_text)

    @staticmethod
    def normalize_tender_url(raw_url):
        if not raw_url:
            return ""
        try:
            parsed = urlparse(raw_url)
            path = parsed.path or ""
            # Remove path params like ;jsessionid=...
            if ";" in path:
                path = path.split(";", 1)[0]
            skip_keys = {
                "session", "sessionid", "jsessionid", "sid", "phpsessid", "_", "ts", "timestamp"
            }
            q_items = []
            for k, v in parse_qsl(parsed.query, keep_blank_values=True):
                if k.lower() in skip_keys:
                    continue
                q_items.append((k, v))
            q_items.sort(key=lambda x: (x[0].lower(), x[1]))
            clean_query = urlencode(q_items, doseq=True)
            return urlunparse((parsed.scheme.lower(), parsed.netloc.lower(), path, "", clean_query, ""))
        except Exception:
            return str(raw_url)

    @staticmethod
    def upsert_tender_row(conn, tender_row):
        (
            website_id, org_chain, tender_id, title, tender_value, emd, closing_date, opening_date,
            tender_url, location, tender_category, pre_bid_meeting_date, work_description
        ) = tender_row
        norm_url = ScraperBackend.normalize_tender_url(tender_url)
        c = conn.cursor()

        existing = c.execute(
            """SELECT id, is_downloaded, is_bookmarked, folder_path, last_downloaded_at
               FROM tenders
               WHERE website_id=? AND normalized_tender_url=?
               ORDER BY id DESC LIMIT 1""",
            (website_id, norm_url)
        ).fetchone()
        if not existing and tender_id:
            existing = c.execute(
                """SELECT id, is_downloaded, is_bookmarked, folder_path, last_downloaded_at
                   FROM tenders
                   WHERE website_id=? AND tender_id=?
                   ORDER BY id DESC LIMIT 1""",
                (website_id, tender_id)
            ).fetchone()
        if not existing:
            existing = c.execute(
                """SELECT id, is_downloaded, is_bookmarked, folder_path, last_downloaded_at
                   FROM tenders
                   WHERE website_id=? AND org_chain=? AND title=? AND closing_date=?
                   ORDER BY id DESC LIMIT 1""",
                (website_id, org_chain, title, closing_date)
            ).fetchone()

        if existing:
            row_id, is_downloaded, is_bookmarked, folder_path, last_downloaded_at = existing
            c.execute(
                """UPDATE tenders
                   SET org_chain=?, tender_id=?, title=?, tender_value=?, emd=?, closing_date=?, opening_date=?,
                       tender_url=?, location=?, tender_category=?, pre_bid_meeting_date=?, normalized_tender_url=?,
                       work_description=?, status=CASE WHEN COALESCE(status,'')='Archived' THEN '' ELSE COALESCE(status,'') END,
                       is_archived=0, is_downloaded=?, is_bookmarked=?, folder_path=?, last_downloaded_at=?
                   WHERE id=?""",
                (
                    org_chain, tender_id, title, tender_value, emd, closing_date, opening_date,
                    tender_url, location, tender_category, pre_bid_meeting_date, norm_url, work_description,
                    is_downloaded, is_bookmarked, folder_path, last_downloaded_at, row_id
                )
            )
            return "updated"

        try:
            c.execute(
                """INSERT INTO tenders
                   (website_id, org_chain, tender_id, title, tender_value, emd, closing_date, opening_date,
                    tender_url, location, tender_category, pre_bid_meeting_date, work_description, status, is_archived, normalized_tender_url)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', 0, ?)""",
                (
                    website_id, org_chain, tender_id, title, tender_value, emd, closing_date, opening_date,
                    tender_url, location, tender_category, pre_bid_meeting_date, work_description, norm_url
                )
            )
            return "inserted"
        except sqlite3.IntegrityError:
            # Unique tender_id collisions: update the row for same website+tender_id if present.
            row = c.execute(
                "SELECT id FROM tenders WHERE website_id=? AND tender_id=? ORDER BY id DESC LIMIT 1",
                (website_id, tender_id)
            ).fetchone()
            if row:
                c.execute(
                    """UPDATE tenders
                       SET org_chain=?, title=?, tender_value=?, emd=?, closing_date=?, opening_date=?,
                           tender_url=?, location=?, tender_category=?, pre_bid_meeting_date=?,
                           normalized_tender_url=?, work_description=?,
                           status=CASE WHEN COALESCE(status,'')='Archived' THEN '' ELSE COALESCE(status,'') END,
                           is_archived=0
                       WHERE id=?""",
                    (
                        org_chain, title, tender_value, emd, closing_date, opening_date,
                        tender_url, location, tender_category, pre_bid_meeting_date, norm_url, work_description, row[0]
                    )
                )
                return "updated"
            raise

    @staticmethod
    def dedupe_tenders_for_website(conn, website_id):
        c = conn.cursor()
        rows = c.execute(
            """SELECT id, tender_id, tender_url, normalized_tender_url, title, closing_date,
                      is_downloaded, is_bookmarked, folder_path, last_downloaded_at
               FROM tenders
               WHERE website_id=?
               ORDER BY id DESC""",
            (website_id,)
        ).fetchall()

        keep_for_key = {}
        to_delete = []

        def key_of(r):
            # Prefer normalized URL (most stable for same tender across runs)
            if r[3]:
                return ("url", r[3])
            if r[1]:
                return ("id", r[1])
            return ("meta", f"{r[4]}|{r[5]}")

        for r in rows:
            rid = r[0]
            key = key_of(r)
            if key not in keep_for_key:
                keep_for_key[key] = r
            else:
                keep = keep_for_key[key]
                # Preserve user state on kept row.
                if (keep[6] or 0) == 0 and (r[6] or 0) == 1:
                    c.execute("UPDATE tenders SET is_downloaded=1 WHERE id=?", (keep[0],))
                if (keep[7] or 0) == 0 and (r[7] or 0) == 1:
                    c.execute("UPDATE tenders SET is_bookmarked=1 WHERE id=?", (keep[0],))
                if (not keep[8]) and r[8]:
                    c.execute("UPDATE tenders SET folder_path=? WHERE id=?", (r[8], keep[0]))
                if (not keep[9]) and r[9]:
                    c.execute("UPDATE tenders SET last_downloaded_at=? WHERE id=?", (r[9], keep[0]))
                to_delete.append(rid)

        if to_delete:
            c.executemany("DELETE FROM tenders WHERE id=?", [(x,) for x in to_delete])
        return len(to_delete)

    @staticmethod
    def fetch_tenders_logic(website_id):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT name, tenders_url FROM organizations WHERE website_id=? AND is_selected=1", (website_id,))
        selected_orgs = c.fetchall()
        conn.close()

        if not selected_orgs:
            log_to_gui("No organizations selected. Please select organizations first.")
            return

        scraped_org_seen_ids = {}
        failed_orgs = set()

        for org_name, url in selected_orgs:
            log_to_gui(f"Scraping tenders for: {org_name}")
            current_url = url
            org_seen_ids = set()
            org_scrape_ok = False
            
            while current_url:
                try:
                    res = ScraperBackend.safe_request(current_url)
                    if not res: break
                    soup = BeautifulSoup(res.text, 'html.parser')
                    
                    # Enhanced table finding logic
                    table = soup.find('table', {'id': 'table'})
                    if not table:
                        table = soup.find('table', {'class': 'list_table'})
                    if not table:
                        header_tds = soup.find_all('td', string=lambda s: s and 'S.No' in s)
                        for td in header_tds:
                            potential_table = td.find_parent('table')
                            if potential_table and potential_table.find('td', string=lambda s: s and 'e-Published Date' in s):
                                table = potential_table
                                break
                    if not table: 
                        log_to_gui(f"No tender table found on {current_url}")
                        break
                    org_scrape_ok = True
                    
                    rows = table.find_all('tr')
                    tenders_to_save = []
                    
                    for tr in rows:
                        cols = tr.find_all('td')
                        if len(cols) > 4:
                            # Assuming standard NIC structure
                            # Col 4 is usually Title/Ref No
                            title_col = cols[4]
                            link_tag = title_col.find('a')
                            if link_tag:
                                full_link = urljoin(current_url, link_tag['href'])
                                listing_title_text = title_col.get_text(" ", strip=True)
                                closing_date = cols[2].text.strip()
                                opening_date = cols[3].text.strip()
                                t_id = None

                                # Fetch Details
                                emd = "N/A"
                                val = "N/A"
                                loc = "N/A"
                                cat = "N/A"
                                prebid = "N/A"
                                work_desc = "N/A"
                                d_soup = None
                                
                                try:
                                    d_res = ScraperBackend.safe_request(full_link)
                                    if d_res:
                                        d_soup = BeautifulSoup(d_res.text, 'html.parser')
                                        # Keep mappings aligned with tender_scraper.py label strategy.
                                        emd = ScraperBackend.get_detail_by_label(
                                            d_soup,
                                            ["EMD Amount In ₹", "EMD Amount (in Rs.)", "EMD Amount In", "EMD Amount", "EMD"]
                                        ) or "N/A"
                                        val = ScraperBackend.get_detail_by_label(
                                            d_soup,
                                            ["Tender Value In ₹", "Tender Value In Rs.", "Tender Value In", "Tender Value"]
                                        ) or "N/A"
                                        loc = ScraperBackend.get_detail_by_label(
                                            d_soup,
                                            ["Location", "Work Location", "Place of Work"]
                                        ) or "N/A"
                                        # Strict: Tender Category should come from Tender Category label only.
                                        cat = ScraperBackend.get_detail_by_label(
                                            d_soup,
                                            ["Tender Category"],
                                            allow_contains=False
                                        ) or "N/A"
                                        prebid = ScraperBackend.get_detail_by_label(
                                            d_soup,
                                            ["Pre Bid Meeting Date", "Pre-Bid Meeting Date"]
                                        ) or "N/A"
                                        work_desc = ScraperBackend.get_detail_by_label(
                                            d_soup,
                                            ["Work Description", "Description of Work", "Work Desc"]
                                        ) or "N/A"
                                except: pass

                                title_text = ScraperBackend.derive_tender_title(listing_title_text, d_soup)
                                t_id = ScraperBackend.derive_tender_id(listing_title_text, d_soup, full_link)
                                log_to_gui(f"  > {t_id}")
                                t_id_norm = str(t_id or "").strip()
                                if t_id_norm and t_id_norm.upper() != "N/A":
                                    org_seen_ids.add(t_id_norm)
                                
                                tenders_to_save.append((website_id, org_name, t_id, title_text, val, emd, closing_date, opening_date, full_link, loc, cat, prebid, work_desc))

                    # Save batch
                    if tenders_to_save:
                        conn = sqlite3.connect(DB_FILE)
                        inserted_count = 0
                        updated_count = 0
                        for t in tenders_to_save:
                            try:
                                result = ScraperBackend.upsert_tender_row(conn, t)
                                if result == "inserted":
                                    inserted_count += 1
                                else:
                                    updated_count += 1
                            except Exception as e:
                                log_to_gui(f"Upsert failed for tender '{t[2]}': {e}")
                        removed = ScraperBackend.dedupe_tenders_for_website(conn, website_id)
                        conn.commit()
                        conn.close()
                        log_to_gui(f"Saved page: inserted={inserted_count}, updated={updated_count}, deduped={removed}")
                    else:
                        log_to_gui("No tenders found in table rows.")
                    
                    # Pagination
                    next_link = soup.find('a', string=lambda t: t and 'Next' in t)
                    current_url = urljoin(current_url, next_link['href']) if next_link else None
                    time.sleep(1)
                    
                except Exception as e:
                    log_to_gui(f"Error scraping {org_name}: {e}")
                    break
            if org_scrape_ok:
                scraped_org_seen_ids[org_name] = org_seen_ids
            else:
                failed_orgs.add(org_name)
        try:
            conn = sqlite3.connect(DB_FILE)
            removed = ScraperBackend.dedupe_tenders_for_website(conn, website_id)
            archived_missing = 0
            for org_name, seen_ids in scraped_org_seen_ids.items():
                archived_missing += ScraperBackend.archive_missing_tenders_for_org(
                    conn, website_id, org_name, seen_ids
                )
            conn.commit()
            conn.close()
            if removed:
                log_to_gui(f"Post-scrape dedupe removed {removed} duplicate rows.")
            if archived_missing:
                log_to_gui(f"Auto-archived {archived_missing} stale tenders no longer present in latest scrape.")
            log_to_gui(
                f"Stale-archive check completed for selected orgs: total={len(selected_orgs)}, "
                f"processed={len(scraped_org_seen_ids)}, failed={len(failed_orgs)}."
            )
            if failed_orgs:
                log_to_gui(f"Skipped stale-archive for failed org scrape(s): {', '.join(sorted(failed_orgs))}")
        except Exception as e:
            log_to_gui(f"Post-scrape dedupe error: {e}")
        log_to_gui("Tender fetching complete.")
        return True

    @staticmethod
    def archive_missing_tenders_for_org(conn, website_id, org_name, seen_tender_ids):
        c = conn.cursor()
        clean_ids = sorted({str(x).strip() for x in (seen_tender_ids or set()) if str(x).strip()})
        sql = (
            "UPDATE tenders SET is_archived=1 "
            "WHERE website_id=? AND org_chain=? AND COALESCE(is_archived,0)=0"
        )
        params = [website_id, org_name]
        if clean_ids:
            placeholders = ",".join("?" for _ in clean_ids)
            sql += f" AND TRIM(COALESCE(tender_id,'')) NOT IN ({placeholders})"
            params.extend(clean_ids)
        c.execute(sql, tuple(params))
        return int(c.rowcount or 0)

    @staticmethod
    def parse_closing_datetime(value):
        txt = " ".join(str(value or "").replace(",", " ").split()).strip()
        if not txt or txt.upper() in {"N/A", "NA", "NONE", "-"}:
            return None
        formats = (
            "%d-%b-%Y %I:%M %p",
            "%d-%b-%Y %H:%M",
            "%d-%m-%Y %I:%M %p",
            "%d-%m-%Y %H:%M",
            "%d/%m/%Y %I:%M %p",
            "%d/%m/%Y %H:%M",
            "%d-%b-%Y",
            "%d-%m-%Y",
            "%d/%m/%Y",
        )
        for fmt in formats:
            try:
                dt = datetime.datetime.strptime(txt, fmt)
                if "%H:%M" not in fmt and "%I:%M" not in fmt:
                    dt = dt.replace(hour=23, minute=59, second=59)
                return dt
            except Exception:
                continue
        m = re.match(r"^(\d{1,2}-[A-Za-z]{3}-\d{4})\s+(\d{1,2}:\d{2}\s*[APMapm]{2})$", txt)
        if m:
            try:
                return datetime.datetime.strptime(
                    f"{m.group(1)} {m.group(2).upper().replace(' ', '')}",
                    "%d-%b-%Y %I:%M%p"
                )
            except Exception:
                return None
        return None

    @staticmethod
    def download_file_with_requests(url, file_path, cookies, tender_id=None, file_type="document"):
        try:
            s = requests.Session()
            for c in cookies:
                s.cookies.set(c['name'], c['value'], domain=c.get('domain'), path=c.get('path'))
            s.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWeb7Kit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36'
            })
            r = s.get(url, stream=True, timeout=120)
            r.raise_for_status()
            with open(file_path, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            if tender_id:
                ScraperBackend.log_downloaded_file(
                    tender_id,
                    os.path.basename(file_path),
                    file_type=file_type,
                    source_url=url,
                    local_path=file_path
                )
            return True
        except Exception as e:
            log_to_gui(f"Download failed for {os.path.basename(file_path)}: {e}")
            return False

    @staticmethod
    def _extract_status_and_row(driver):
        try:
            table = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "tabList")))
        except Exception:
            return None, None
        headers = [c.text.strip() for c in table.find_elements(By.CSS_SELECTOR, "tr.list_header td")]
        if not headers:
            headers = [c.text.strip() for c in table.find_elements(By.XPATH, ".//tbody/tr[1]/td")]
        idx = headers.index("Tender Stage") if "Tender Stage" in headers else -1
        if idx < 0:
            return None, None
        rows = table.find_elements(By.XPATH, ".//tbody/tr[contains(@class, 'even') or contains(@class, 'odd')]")
        if not rows:
            return None, None
        first_row = rows[0]
        cells = first_row.find_elements(By.TAG_NAME, "td")
        if idx >= len(cells):
            return None, first_row
        return cells[idx].text.strip(), first_row

    @staticmethod
    def _download_result_docs_from_popup(driver, tender_id, base_folder):
        wait = WebDriverWait(driver, 20)
        main_window = driver.current_window_handle
        downloaded_any = False
        result_folder = os.path.join(base_folder, "Financial Result")
        os.makedirs(result_folder, exist_ok=True)
        try:
            summary_link = wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "summary details")))
            driver.execute_script("arguments[0].click();", summary_link)
            wait.until(EC.number_of_windows_to_be(2))
            popup = None
            for wh in driver.window_handles:
                if wh != main_window:
                    popup = wh
                    break
            if not popup:
                return False
            driver.switch_to.window(popup)
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            links = driver.find_elements(By.TAG_NAME, "a")
            for link in links:
                href = link.get_attribute("href")
                if not href:
                    continue
                low = href.lower()
                if ".pdf" not in low and ".xlsx" not in low and ".zip" not in low:
                    continue
                name = (link.text or "").strip()
                if not name:
                    name = os.path.basename(urlparse(href).path) or f"result_{int(time.time())}"
                safe_name = re.sub(r'[\\/*?:"<>|]', "", name)
                fpath = os.path.join(result_folder, safe_name)
                if ScraperBackend.should_skip_file(tender_id, safe_name, fpath):
                    continue
                if ScraperBackend.download_file_with_requests(href, fpath, driver.get_cookies(), tender_id=tender_id, file_type="result"):
                    downloaded_any = True
                    log_to_gui(f"    Downloaded Result File: {safe_name}")
            driver.close()
            driver.switch_to.window(main_window)
        except Exception as e:
            log_to_gui(f"    Result popup processing error: {e}")
            try:
                if driver.current_window_handle != main_window:
                    driver.close()
                driver.switch_to.window(main_window)
            except Exception:
                pass
        return downloaded_any

    @staticmethod
    def check_tender_status_logic(website_id, archived_only=False):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        # Get status URL
        site_url = c.execute("SELECT status_url FROM websites WHERE id=?", (website_id,)).fetchone()
        if not site_url:
            conn.close()
            return 0
        status_url = site_url[0]

        if archived_only:
            tenders = c.execute(
                """SELECT id, tender_id
                   FROM tenders
                   WHERE website_id=?
                     AND COALESCE(is_archived,0)=1
                     AND COALESCE(is_downloaded,0)=1
                     AND TRIM(COALESCE(tender_id,''))<>''""",
                (website_id,)
            ).fetchall()
        else:
            # Keep status blank for unselected active tenders, and check only selected active tenders.
            c.execute(
                "UPDATE tenders SET status='' WHERE website_id=? AND COALESCE(is_archived,0)=0 AND COALESCE(is_downloaded,0)=0",
                (website_id,)
            )
            conn.commit()
            tenders = c.execute(
                """SELECT id, tender_id
                   FROM tenders
                   WHERE website_id=?
                     AND COALESCE(is_archived,0)=0
                     AND COALESCE(is_downloaded,0)=1
                     AND TRIM(COALESCE(tender_id,''))<>''""",
                (website_id,)
            ).fetchall()
        conn.close()

        if not tenders:
            if archived_only:
                log_to_gui("No selected archived tenders to check status.")
            else:
                log_to_gui("No selected tenders to check status.")
            return 0

        mode_txt = "selected archived" if archived_only else "selected"
        log_to_gui(f"Checking status for {len(tenders)} {mode_txt} tenders...")
        
        options = FirefoxOptions()
        # options.add_argument("--headless") 
        service = FirefoxService(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service, options=options)
        # Solve once per Selenium session; retry only if portal asks again.
        ScraperBackend.captcha_solved_in_session = False

        updated_count = 0
        try:
            for db_id, tid in tenders:
                log_to_gui(f"Checking: {tid}")
                driver.get(status_url)
                time.sleep(2)
                
                try:
                    # Input Tender ID
                    tid_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "tenderId")))
                    tid_input.clear()
                    tid_input.send_keys(tid)
                    
                    # Solve Captcha
                    if ScraperBackend.handle_captcha_interaction(driver, f"Status {tid}", "Search"):
                        new_status, _ = ScraperBackend._extract_status_and_row(driver)
                        if new_status:
                            conn = sqlite3.connect(DB_FILE)
                            conn.execute("UPDATE tenders SET status=? WHERE id=?", (new_status, db_id))
                            conn.commit()
                            conn.close()
                            updated_count += 1
                            log_to_gui(f"Updated status for {tid}: {new_status}")
                        else:
                            log_to_gui(f"No Tender Stage found for {tid}")
                    else:
                        log_to_gui(f"CAPTCHA failed for {tid}. Skipping this tender.")
                except Exception as e:
                    log_to_gui(f"Error checking {tid}: {e}")
        finally:
            driver.quit()
        log_to_gui("Status check complete.")
        return updated_count

    @staticmethod
    def download_tender_results_logic(website_id):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        site_row = c.execute("SELECT status_url FROM websites WHERE id=?", (website_id,)).fetchone()
        if not site_row:
            conn.close()
            log_to_gui("Status URL not configured for this website.")
            return
        status_url = site_row[0]
        c.execute("SELECT id, tender_id, folder_path FROM tenders WHERE website_id=? AND is_downloaded=1", (website_id,))
        targets = c.fetchall()
        conn.close()
        if not targets:
            log_to_gui("No tenders marked for download/result check.")
            return

        log_to_gui(f"Starting result file checks for {len(targets)} tenders...")
        options = FirefoxOptions()
        service = FirefoxService(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service, options=options)
        ScraperBackend.captcha_solved_in_session = False
        target_statuses = {"Financial Bid Opening", "Financial Evaluation", "AOC", "Concluded"}
        try:
            for db_id, tender_id, folder_path in targets:
                try:
                    log_to_gui(f"Checking result status: {tender_id}")
                    driver.get(status_url)
                    time.sleep(2)
                    inp = WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.ID, "tenderId")))
                    inp.clear()
                    inp.send_keys(tender_id)
                    if not ScraperBackend.handle_captcha_interaction(driver, f"Result Check {tender_id}", submit_id="Search"):
                        log_to_gui(f"  CAPTCHA failed for {tender_id}.")
                        continue
                    status, row = ScraperBackend._extract_status_and_row(driver)
                    if status:
                        conn = sqlite3.connect(DB_FILE)
                        conn.execute("UPDATE tenders SET status=? WHERE id=?", (status, db_id))
                        conn.commit()
                        conn.close()
                    if status not in target_statuses:
                        log_to_gui(f"  Status '{status}' not eligible for result docs.")
                        continue

                    view_links = row.find_elements(By.XPATH, ".//a[img[contains(@src, 'view.png')]]") if row else []
                    if not view_links and row:
                        view_links = row.find_elements(By.TAG_NAME, "a")
                    if not view_links:
                        log_to_gui("  Result view link not found.")
                        continue
                    driver.execute_script("arguments[0].click();", view_links[0])
                    time.sleep(2)

                    safe_id = re.sub(r'[\\/*?:"<>|]', "", str(tender_id))
                    existing_folder = (folder_path or "").strip()
                    if existing_folder and os.path.isdir(existing_folder):
                        tender_folder = existing_folder
                    else:
                        tender_folder = os.path.join(BASE_DOWNLOAD_DIRECTORY, safe_id)
                        os.makedirs(tender_folder, exist_ok=True)
                    conn = sqlite3.connect(DB_FILE)
                    conn.execute("UPDATE tenders SET folder_path=? WHERE id=?", (tender_folder, db_id))
                    conn.commit()
                    conn.close()
                    got = ScraperBackend._download_result_docs_from_popup(driver, tender_id, tender_folder)
                    if got:
                        conn = sqlite3.connect(DB_FILE)
                        conn.execute("UPDATE tenders SET folder_path=?, last_downloaded_at=CURRENT_TIMESTAMP WHERE id=?", (tender_folder, db_id))
                        conn.commit()
                        conn.close()
                except Exception as e:
                    log_to_gui(f"  Result download error for {tender_id}: {e}")
        finally:
            driver.quit()
        log_to_gui("Result file check complete.")

    @staticmethod
    def download_tenders_logic(website_id, target_db_ids=None, forced_mode=None):
        ScraperBackend.ensure_download_tables()
        ids = []
        if target_db_ids:
            try:
                ids = [int(x) for x in target_db_ids]
            except Exception:
                ids = []
        mode_override = str(forced_mode or "").strip().lower()
        if mode_override not in {"full", "update"}:
            mode_override = ""
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        where_sql = "website_id=? AND COALESCE(is_downloaded,0)=1 AND COALESCE(is_archived,0)=0"
        params = [website_id]
        if ids:
            placeholders = ",".join("?" for _ in ids)
            where_sql += f" AND id IN ({placeholders})"
            params.extend(ids)
        c.execute(
            f"SELECT id, tender_id, title, tender_url, last_downloaded_at, folder_path "
            f"FROM tenders WHERE {where_sql}",
            tuple(params)
        )
        to_download = c.fetchall()
        conn.close()
        
        if not to_download:
            log_to_gui("No tenders marked for download.")
            return
        
        # Get base URL for session refresh
        conn = sqlite3.connect(DB_FILE)
        site_url_row = conn.execute("SELECT url FROM websites WHERE id=?", (website_id,)).fetchone()
        conn.close()
        base_url = site_url_row[0] if site_url_row else "https://mahatenders.gov.in/nicgep/app?page=FrontEndTendersByOrganisation&service=page"

        log_to_gui(f"Starting download for {len(to_download)} tenders...")
        
        options = FirefoxOptions()
        service = FirefoxService(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service, options=options)
        wait = WebDriverWait(driver, 20)
        ScraperBackend.captcha_solved_in_session = False
        # Establish Selenium session once (same pattern as tender_scraper.py).
        try:
            driver.get(base_url)
            time.sleep(4)
        except Exception as e:
            log_to_gui(f"Failed to initialize Selenium session: {e}")
            driver.quit()
            return
        
        for db_id, t_id, title, url, last_dl, existing_folder in to_download:
            download_mode = mode_override if mode_override else ('update' if last_dl else 'full')
            log_to_gui(f"Processing: {t_id} (Mode: {download_mode})...")
            
            safe_id = re.sub(r'[\\/*?:"<>|]',"", t_id)
            preferred_dir = os.path.join(BASE_DOWNLOAD_DIRECTORY, safe_id)
            existing = str(existing_folder or "").strip()
            if existing:
                try:
                    ex_abs = os.path.normcase(os.path.abspath(existing))
                    pref_abs = os.path.normcase(os.path.abspath(preferred_dir))
                    save_dir = existing if ex_abs == pref_abs else preferred_dir
                except Exception:
                    save_dir = preferred_dir
            else:
                save_dir = preferred_dir
            if not os.path.exists(save_dir): os.makedirs(save_dir)
            
            try:
                if not ScraperBackend.open_tender_page_with_recovery(driver, base_url, url):
                    log_to_gui(f"Could not recover session for {t_id}. Skipping.")
                    continue

                # --- 1. Tender Notice (Full Mode Only) ---
                if download_mode == 'full':
                    try:
                        notice_filename = f"Tendernotice_{safe_id}.pdf"
                        notice_path = os.path.join(save_dir, notice_filename)
                        if not ScraperBackend.should_skip_file(t_id, notice_filename, notice_path):
                            log_to_gui("  Checking Tender Notice...")
                            if not ScraperBackend.open_tender_page_with_recovery(driver, base_url, url):
                                log_to_gui("  Could not open tender page for Tender Notice.")
                            else:
                                downloaded_notice = False
                                # If captcha already solved in this session, final link is often directly available.
                                if ScraperBackend.captcha_solved_in_session:
                                    try:
                                        final_link = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_0")))
                                        href = final_link.get_attribute('href')
                                        if href and ScraperBackend.download_file_with_requests(href, notice_path, driver.get_cookies(), t_id, file_type="notice"):
                                            log_to_gui("  Downloaded Tender Notice.")
                                            downloaded_notice = True
                                    except Exception:
                                        pass
                                if not downloaded_notice:
                                    trigger = None
                                    try:
                                        trigger = wait.until(EC.element_to_be_clickable((By.ID, "docDownload")))
                                    except Exception:
                                        try:
                                            trigger = wait.until(EC.element_to_be_clickable((By.ID, "DirectLink_8")))
                                        except Exception:
                                            pass
                                    if trigger:
                                        driver.execute_script("arguments[0].click();", trigger)
                                        if ScraperBackend.handle_captcha_interaction(driver, "Tender Notice"):
                                            try:
                                                final_link = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_0")))
                                                href = final_link.get_attribute('href')
                                                if href and ScraperBackend.download_file_with_requests(href, notice_path, driver.get_cookies(), t_id, file_type="notice"):
                                                    log_to_gui("  Downloaded Tender Notice.")
                                                    downloaded_notice = True
                                                else:
                                                    log_to_gui("  Final Tender Notice link missing/invalid.")
                                            except Exception:
                                                log_to_gui("  Could not find final Notice link.")
                                if not downloaded_notice:
                                    log_to_gui("  Tender Notice not downloaded.")
                        else:
                            log_to_gui("  Skipping Tender Notice (already logged and file exists).")
                    except Exception as e:
                        log_to_gui(f"  Notice download error: {e}")

                # --- 2. Zip File (Full Mode Only) ---
                if download_mode == 'full':
                    try:
                        zip_filename = f"{safe_id}.zip"
                        zip_path = os.path.join(save_dir, zip_filename)
                        if not ScraperBackend.should_skip_file(t_id, zip_filename, zip_path):
                            log_to_gui("  Checking Zip File...")
                            if ScraperBackend.open_tender_page_with_recovery(driver, base_url, url):
                                zip_href = None
                                try:
                                    zip_elem = wait.until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Download as zip file")))
                                    zip_href = zip_elem.get_attribute('href')
                                except Exception:
                                    try:
                                        zip_elem = wait.until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Download as zip")))
                                        zip_href = zip_elem.get_attribute('href')
                                    except Exception:
                                        try:
                                            zip_elem = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_7")))
                                            zip_href = zip_elem.get_attribute('href')
                                        except Exception:
                                            try:
                                                zip_elem = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_8")))
                                                zip_href = zip_elem.get_attribute('href')
                                            except Exception:
                                                pass
                                if zip_href and ScraperBackend.download_file_with_requests(zip_href, zip_path, driver.get_cookies(), t_id, file_type="zip"):
                                    log_to_gui("  Downloaded Zip File.")
                                else:
                                    log_to_gui("  Zip link not found.")
                        else:
                            log_to_gui("  Skipping Zip file (already logged and file exists).")
                    except Exception as e:
                        log_to_gui(f"  Zip download error: {e}")

                # --- 3. Pre-Bid Meeting (Always Check) ---
                try:
                    prebid_filename = f"PreBid_Meeting_{safe_id}.pdf"
                    prebid_path = os.path.join(save_dir, prebid_filename)
                    if not ScraperBackend.should_skip_file(t_id, prebid_filename, prebid_path):
                        if ScraperBackend.open_tender_page_with_recovery(driver, base_url, url):
                            try:
                                pb_link = wait.until(EC.presence_of_element_located((By.ID, "DirectLink_2")))
                                href = pb_link.get_attribute('href')
                                if href and ScraperBackend.download_file_with_requests(href, prebid_path, driver.get_cookies(), t_id, file_type="prebid"):
                                    log_to_gui("  Downloaded Pre-Bid File.")
                                else:
                                    log_to_gui("  Pre-Bid file link not found.")
                            except Exception:
                                log_to_gui("  No Pre-Bid file found.")
                    else:
                        log_to_gui("  Skipping Pre-Bid file (already logged and file exists).")
                except Exception as e:
                    log_to_gui(f"  Pre-bid error: {e}")

                # --- 4. Corrigendums (Always Check) ---
                try:
                    log_to_gui("  Checking Corrigendums...")
                    if ScraperBackend.open_tender_page_with_recovery(driver, base_url, url):
                        main_window = driver.current_window_handle
                        corr_links = driver.find_elements(By.XPATH, "//a[contains(@title, 'View Corrigendum History')]")
                        if corr_links:
                            driver.execute_script("arguments[0].click();", corr_links[0])
                            wait.until(EC.number_of_windows_to_be(2))
                            for handle in driver.window_handles:
                                if handle != main_window:
                                    driver.switch_to.window(handle)
                                    break

                            docs = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//a[contains(@id, 'DirectLink_')]")))
                            for doc in docs:
                                href = doc.get_attribute('href')
                                name = (doc.text or "").strip()
                                if href and name:
                                    safe_name = re.sub(r'[\\/*?:"<>|]', "", name)
                                    fpath = os.path.join(save_dir, safe_name)
                                    if ScraperBackend.should_skip_file(t_id, safe_name, fpath):
                                        continue
                                    if ScraperBackend.download_file_with_requests(href, fpath, driver.get_cookies(), t_id, file_type="corrigendum"):
                                        log_to_gui(f"    Downloaded Corrigendum: {safe_name}")
                            driver.close()
                            driver.switch_to.window(main_window)
                except Exception as e:
                    log_to_gui(f"  Corrigendum error: {e}")
                    try:
                        if 'main_window' in locals():
                            driver.switch_to.window(main_window)
                    except Exception:
                        pass

                # Update DB
                conn = sqlite3.connect(DB_FILE)
                conn.execute("UPDATE tenders SET folder_path=?, last_downloaded_at=CURRENT_TIMESTAMP WHERE id=?", (save_dir, db_id))
                conn.commit()
                conn.close()

            except Exception as e:
                log_to_gui(f"Error accessing {url}: {e}")
        
        driver.quit()
        log_to_gui("Download process finished.")

    @staticmethod
    def download_docs_for_tender_to_folder(source_tender_id, destination_folder):
        tender_id = str(source_tender_id or "").strip()
        dest = str(destination_folder or "").strip()
        if not tender_id:
            log_to_gui("Download Docs: missing source tender id.")
            return False
        if not dest:
            log_to_gui("Download Docs: missing destination folder.")
            return False
        os.makedirs(dest, exist_ok=True)

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        row = c.execute(
            "SELECT id, website_id FROM tenders WHERE TRIM(COALESCE(tender_id,''))=? LIMIT 1",
            (tender_id,)
        ).fetchone()
        if not row:
            conn.close()
            log_to_gui(f"Download Docs: tender not found for id '{tender_id}'.")
            return False
        target_db_id, website_id = row
        backups = c.execute(
            "SELECT id, COALESCE(is_downloaded,0), COALESCE(folder_path,'') FROM tenders WHERE website_id=?",
            (website_id,)
        ).fetchall()
        conn.close()

        success = False
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("UPDATE tenders SET is_downloaded=0 WHERE website_id=?", (website_id,))
            c.execute(
                "UPDATE tenders SET is_downloaded=1, folder_path=? WHERE id=?",
                (dest, target_db_id)
            )
            conn.commit()
            conn.close()

            ScraperBackend.download_tenders_logic(website_id)
            success = True
        finally:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            for tid, was_selected, old_folder in backups:
                c.execute(
                    "UPDATE tenders SET is_downloaded=?, folder_path=? WHERE id=?",
                    (int(was_selected or 0), str(old_folder or "").strip(), tid)
                )
            conn.commit()
            conn.close()
        return success

    @staticmethod
    def download_single_tender_logic(tender_db_id, mode):
        try:
            target_id = int(tender_db_id)
        except Exception:
            log_to_gui("Download: invalid tender selection.")
            return False
        mode_txt = str(mode or "").strip().lower()
        if mode_txt not in {"full", "update"}:
            mode_txt = ""
        conn = sqlite3.connect(DB_FILE)
        row = conn.execute(
            "SELECT website_id, COALESCE(is_archived,0) FROM tenders WHERE id=?",
            (target_id,)
        ).fetchone()
        conn.close()
        if not row:
            log_to_gui("Download: selected tender not found.")
            return False
        website_id, is_archived = row
        if int(is_archived or 0) != 0:
            log_to_gui("Download: selected tender is archived; only active tenders are allowed.")
            return False
        ScraperBackend.download_tenders_logic(website_id, target_db_ids=[target_id], forced_mode=mode_txt)
        return True

    @staticmethod
    def archive_completed_tenders_logic(website_id):
        completed = ("AOC", "Concluded", "Cancelled", "Withdrawn", "Terminated")
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        placeholders = ",".join("?" for _ in completed)
        c.execute(
            f"UPDATE tenders SET is_archived=1 WHERE website_id=? AND status IN ({placeholders})",
            (website_id, *completed)
        )
        changed_by_status = int(c.rowcount or 0)
        now_dt = datetime.datetime.now()
        overdue_ids = []
        rows = c.execute(
            "SELECT id, closing_date FROM tenders WHERE website_id=? AND COALESCE(is_archived,0)=0",
            (website_id,)
        ).fetchall()
        for tid, closing_raw in rows:
            closing_dt = ScraperBackend.parse_closing_datetime(closing_raw)
            if closing_dt and closing_dt < now_dt:
                overdue_ids.append(tid)
        changed_by_due = 0
        if overdue_ids:
            c.executemany("UPDATE tenders SET is_archived=1 WHERE id=?", [(tid,) for tid in overdue_ids])
            changed_by_due = int(c.rowcount or 0)
        changed = changed_by_status + changed_by_due
        conn.commit()
        conn.close()
        log_to_gui(
            f"Archived {changed} tenders (status-based={changed_by_status}, overdue={changed_by_due})."
        )
        return changed

    @staticmethod
    def get_download_log_rows(website_id=None, limit=500):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        if website_id:
            c.execute(
                """SELECT d.tender_id, d.file_name, COALESCE(d.file_type,'document'), d.downloaded_at
                   FROM downloaded_files d
                   JOIN tenders t ON t.tender_id = d.tender_id
                   WHERE t.website_id=?
                   ORDER BY d.downloaded_at DESC LIMIT ?""",
                (website_id, limit)
            )
        else:
            c.execute(
                """SELECT tender_id, file_name, COALESCE(file_type,'document'), downloaded_at
                   FROM downloaded_files
                   ORDER BY downloaded_at DESC LIMIT ?""",
                (limit,)
            )
        rows = c.fetchall()
        conn.close()
        return rows

    @staticmethod
    def archive_tender_logic(tender_db_id):
        conn = sqlite3.connect(DB_FILE)
        conn.execute("UPDATE tenders SET is_archived=1 WHERE id=?", (tender_db_id,))
        conn.commit()
        conn.close()
        log_to_gui("Tender archived.")

    @staticmethod
    def get_setting(key, default=None):
        conn = sqlite3.connect(DB_FILE)
        row = conn.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
        conn.close()
        return row[0] if row else default

    @staticmethod
    def set_setting(key, value):
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, str(value))
        )
        conn.commit()
        conn.close()

    @staticmethod
    def log_auto_archive_run(status, archived_count=0, archived_status_updated=0, websites_count=0, notes=""):
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            """INSERT INTO auto_archive_runs
               (run_at_utc, status, archived_count, archived_status_updated, websites_count, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                datetime.datetime.now(datetime.UTC).isoformat(),
                str(status or ""),
                int(archived_count or 0),
                int(archived_status_updated or 0),
                int(websites_count or 0),
                str(notes or ""),
            )
        )
        conn.commit()
        conn.close()

# --- VIEW 3: ONLINE TENDERS (NEW INTEGRATION) ---
class ViewTenders(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.ui_scale = getattr(controller, "ui_scale", 1.0)
        self.column_width_scale = getattr(controller, "column_width_scale", 1.0)
        self.backend = ScraperBackend
        self.org_cols = ["Sr", "OrgID", "Website", "Name", "Count", "Select"]
        self.tender_cols = ["Sr", "ID", "Website", "Tender ID", "Title", "Work Description", "Value", "EMD", "Org Chain", "Closing Date", "Closing Time", "Pre-Bid", "Location", "Category", "Select", "Download"]
        self.archived_cols = ["Sr", "ID", "Website", "Tender ID", "Title", "Work Description", "Value", "EMD", "Org Chain", "Closing Date", "Closing Time", "Pre-Bid", "Location", "Category", "Status", "Select", "Download"]
        self.filter_map = {
            "orgs": self._load_json_setting("viewtenders_filters_orgs", {}),
            "tenders": self._load_json_setting("viewtenders_filters_tenders", {}),
            "archived": self._load_json_setting("viewtenders_filters_archived", {}),
        }
        self.quick_search_map = {
            "orgs": get_user_setting("viewtenders_search_orgs", self.backend.get_setting("viewtenders_search_orgs", "")) or "",
            "tenders": get_user_setting("viewtenders_search_tenders", self.backend.get_setting("viewtenders_search_tenders", "")) or "",
            "archived": get_user_setting("viewtenders_search_archived", self.backend.get_setting("viewtenders_search_archived", "")) or "",
        }
        self._search_syncing = False
        self.sort_map = {
            "orgs": self._load_json_setting("viewtenders_sort_orgs", {}),
            "tenders": self._load_json_setting("viewtenders_sort_tenders", {}),
            "archived": self._load_json_setting("viewtenders_sort_archived", {}),
        }
        self._org_reflow_job = None
        self._org_reflowing = False
        
        # Header
        header = tk.Frame(self, bg="#3F51B5", height=60)
        header.pack(fill="x")
        tk.Label(header, text="Online Tender Scraper", bg="#3F51B5", fg="white", font=("Segoe UI", 18)).pack(side="left", padx=20)
        tk.Button(header, text="Open Folder", bg="#E0E0E0", fg="black", command=self.open_download_root_folder).pack(side="right", padx=20, pady=10)
        
        # Toolbar
        toolbar = tk.Frame(self, bg="#E8EAF6", pady=5, padx=10)
        toolbar.pack(fill="x")

        top_row = tk.Frame(toolbar, bg="#E8EAF6")
        top_row.pack(fill="x")
        tk.Label(top_row, text="Website:", bg="#E8EAF6", font=("Segoe UI", 10)).pack(side="left")

        self.cb_sites = ttk.Combobox(top_row, width=20, state="readonly")
        self.cb_sites.pack(side="left", padx=5)
        self.cb_sites.bind("<<ComboboxSelected>>", self.on_site_changed)

        tk.Label(top_row, text="Search:", bg="#E8EAF6", font=("Segoe UI", 10)).pack(side="left", padx=(10, 4))
        self.search_var = tk.StringVar()
        self.e_search = tk.Entry(top_row, textvariable=self.search_var, width=24)
        self.e_search.pack(side="left", padx=(0, 8))
        self.search_var.trace_add("write", self.on_search_changed)

        self.btn_manage_websites = tk.Button(top_row, text="Manage Websites", bg="#9E9E9E", fg="white", command=self.manage_websites_popup)
        self.btn_manage_websites.pack(side="left", padx=5)
        self.btn_download_one = tk.Button(top_row, text="Download", bg="#2E7D32", fg="white", command=self.run_single_download)
        self.btn_download_one.pack(side="left", padx=5)
        self.btn_clear = tk.Button(top_row, text="Clear Data", bg="#B71C1C", fg="white", command=self.clear_saved_details_popup)
        self.btn_clear.pack(side="right", padx=5)

        self.actions_row = tk.Frame(toolbar, bg="#E8EAF6")
        self.actions_row.pack(fill="x", pady=(6, 0))

        # Buttons
        self.btn_fetch_orgs = tk.Button(self.actions_row, text="Fetch Organizations", bg="#2196F3", fg="white", command=self.run_fetch_orgs)
        self.btn_fetch_orgs.pack(side="left", padx=6)
        self.btn_get_tenders = tk.Button(self.actions_row, text="Get Tenders", bg="#FF9800", fg="white", command=self.run_fetch_tenders)
        self.btn_get_tenders.pack(side="left", padx=6)
        self.btn_download_selected = tk.Button(self.actions_row, text="Download Selected", bg="#4CAF50", fg="white", command=self.run_download)
        self.btn_download_selected.pack(side="left", padx=6)
        self.btn_select_all = tk.Button(self.actions_row, text="Select All", bg="#607D8B", fg="white", command=self.select_all_tenders)
        self.btn_select_all.pack(side="left", padx=6)
        self.btn_download_results = tk.Button(self.actions_row, text="Download Results", bg="#00897B", fg="white", command=self.run_download_results)
        self.btn_download_results.pack(side="left", padx=6)
        self.btn_check_status = tk.Button(self.actions_row, text="Check Status", bg="#9C27B0", fg="white", command=self.run_status_check)
        self.btn_check_status.pack(side="left", padx=6)
        self.btn_add_to_project = tk.Button(self.actions_row, text="Add Projects", bg="#607D8B", fg="white", command=self.add_selected_tenders_to_new_project)
        self.btn_add_to_project.pack(side="left", padx=6)
        self.btn_manage_columns = tk.Button(self.actions_row, text="Manage Columns", bg="#455A64", fg="white", command=self.open_active_table_column_settings)
        self.btn_manage_columns.pack(side="left", padx=6)
        self.btn_filters = tk.Button(self.actions_row, text="Filters", bg="#5D4037", fg="white", command=self.open_active_table_filter_settings)
        self.btn_filters.pack(side="left", padx=6)
        # Content - Notebook
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        self.tab_end_actions = tk.Frame(self.notebook, bg="#d9d5cf")
        self.btn_export_data = tk.Button(self.tab_end_actions, text="Export Data", bg="#455A64", fg="white", font=("Segoe UI", 8), padx=6, pady=0, bd=1, relief="raised", command=self.open_active_tender_export_popup)
        self.btn_export_data.pack(side="left", pady=1)
        
        # Tab 1: Organizations
        self.tab_orgs = tk.Frame(self.notebook)
        self.notebook.add(self.tab_orgs, text="Organizations")
        self.tree_orgs = self.create_tree(self.tab_orgs, self.org_cols, tree_key="orgs")
        self.tree_orgs.configure(style="Org.Treeview")
        # Organizations tab requested as center aligned for all columns.
        for c in self.org_cols:
            self.tree_orgs.column(c, anchor="center")
        self.tree_orgs.bind("<Double-1>", self.toggle_org_selection)
        self.tree_orgs.bind("<space>", self.toggle_selected_org_rows)
        self.tree_orgs.bind("<Configure>", lambda e: self.schedule_org_table_reflow(), add="+")
        self.tree_orgs.bind("<ButtonRelease-1>", self.on_org_tree_click_release, add="+")
        
        # Tab 2: Tenders
        self.tab_tenders = tk.Frame(self.notebook)
        self.notebook.add(self.tab_tenders, text="Active Tenders")
        self.tree_tenders = self.create_tree(self.tab_tenders, self.tender_cols, fast_horizontal=True, tree_key="tenders")
        self.tree_tenders.configure(style="ActiveTender.Treeview")
        self.tree_tenders.bind("<ButtonRelease-1>", self.on_tender_click)
        self.tree_tenders.bind("<Configure>", lambda e: self.schedule_tender_table_reflow(False), add="+")
        self.tree_tenders.bind("<space>", self.toggle_selected_active_rows)
        self.tree_tenders.bind("<Shift-MouseWheel>", self.fast_horizontal_scroll)
        self.tree_tenders.bind("<Button-4>", lambda e: self.tree_tenders.yview_scroll(-3, "units"))
        self.tree_tenders.bind("<Button-5>", lambda e: self.tree_tenders.yview_scroll(3, "units"))

        # Tab 3: Archived Tenders
        self.tab_archived = tk.Frame(self.notebook)
        self.notebook.add(self.tab_archived, text="Archived Tenders")
        self.tree_archived = self.create_tree(self.tab_archived, self.archived_cols, fast_horizontal=True, tree_key="archived")
        self.tree_archived.configure(style="ArchivedTender.Treeview")
        self.tree_archived.bind("<ButtonRelease-1>", self.on_archived_tender_click)
        self.tree_archived.bind("<Configure>", lambda e: self.schedule_tender_table_reflow(True), add="+")
        self.tree_archived.bind("<space>", self.toggle_selected_archived_rows)
        self.tree_archived.bind("<Shift-MouseWheel>", self.fast_horizontal_scroll_archived)
        self.tree_archived.bind("<Button-4>", lambda e: self.tree_archived.yview_scroll(-3, "units"))
        self.tree_archived.bind("<Button-5>", lambda e: self.tree_archived.yview_scroll(3, "units"))

        # Tab 4: Logs
        self.tab_logs = tk.Frame(self.notebook)
        self.notebook.add(self.tab_logs, text="Live Logs")
        self.txt_log = tk.Text(self.tab_logs, bg="black", fg="#00FF00", font=("Consolas", 10))
        self.txt_log.pack(fill="both", expand=True)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_notebook_tab_changed)
        self.restore_notebook_last_tab()

        self.refresh_sites()
        self.sync_search_bar_for_current_tab()
        self.update_toolbar_for_current_tab()
        self.after_idle(self._stabilize_visible_table_wrap)

    def on_notebook_tab_changed(self, event=None):
        key = self.get_current_table_key()
        if key:
            set_user_setting("viewtenders_last_tab", key)
        self.sync_search_bar_for_current_tab()
        self.refresh_current_table_view()
        self.update_toolbar_for_current_tab()
        self.after_idle(self._stabilize_visible_table_wrap)

    def _stabilize_visible_table_wrap(self, second_pass=False):
        # First paint can happen before final widget widths settle.
        # Run a visible-tab reflow, then a short delayed pass to lock wrapping.
        try:
            self.update_idletasks()
        except Exception:
            pass
        key = self.get_current_table_key()
        if key == "orgs":
            self.apply_org_table_reflow()
        elif key == "tenders":
            self.apply_tender_table_reflow(False)
        elif key == "archived":
            self.apply_tender_table_reflow(True)
        if not second_pass:
            self.after(220, lambda: self._stabilize_visible_table_wrap(second_pass=True))

    def restore_notebook_last_tab(self):
        last_tab = str(get_user_setting("viewtenders_last_tab", "orgs") or "orgs").strip().lower()
        tab_map = {
            "orgs": self.tab_orgs,
            "tenders": self.tab_tenders,
            "archived": self.tab_archived,
            "logs": self.tab_logs,
        }
        target = tab_map.get(last_tab)
        if target is not None:
            self.notebook.select(target)

    def get_current_table_key(self):
        current_tab = self.notebook.select()
        if current_tab == str(self.tab_orgs):
            return "orgs"
        if current_tab == str(self.tab_tenders):
            return "tenders"
        if current_tab == str(self.tab_archived):
            return "archived"
        return None

    def sync_search_bar_for_current_tab(self):
        key = self.get_current_table_key()
        val = self.quick_search_map.get(key, "") if key else ""
        self._search_syncing = True
        self.search_var.set(val)
        self._search_syncing = False

    def on_search_changed(self, *args):
        if self._search_syncing:
            return
        key = self.get_current_table_key()
        if not key:
            return
        self.quick_search_map[key] = self.search_var.get()
        set_user_setting(f"viewtenders_search_{key}", self.quick_search_map[key])
        self.refresh_current_table_view()

    def row_matches_quick_search(self, table_key, row_values):
        q = (self.quick_search_map.get(table_key, "") or "").strip().lower()
        if not q:
            return True
        hay = " | ".join(str(v) for v in row_values).lower()
        return q in hay

    def refresh_current_table_view(self):
        key = self.get_current_table_key()
        if key == "orgs":
            self.load_org_tree()
        elif key == "tenders":
            self.load_tender_tree()
        elif key == "archived":
            self.load_archived_tender_tree()

    def update_toolbar_for_current_tab(self):
        current_tab = self.notebook.select()
        is_orgs = (current_tab == str(self.tab_orgs))
        is_tenders = (current_tab == str(self.tab_tenders))
        is_archived = (current_tab == str(self.tab_archived))
        is_logs = (current_tab == str(self.tab_logs))
        dynamic_buttons = [
            self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected,
            self.btn_download_results, self.btn_check_status, self.btn_select_all, self.btn_add_to_project,
            self.btn_manage_columns, self.btn_filters
        ]
        for btn in dynamic_buttons:
            if btn.winfo_manager():
                btn.pack_forget()

        if is_logs:
            return
        if is_archived:
            # Keep Select All as the last action button in Archived tab.
            for btn in (self.btn_download_results, self.btn_check_status, self.btn_select_all, self.btn_add_to_project, self.btn_manage_columns, self.btn_filters):
                btn.pack(side="left", padx=6)
            self.tab_end_actions.place_forget()
            return
        if is_orgs:
            for btn in (self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected, self.btn_add_to_project, self.btn_manage_columns, self.btn_filters):
                btn.pack(side="left", padx=6)
            self.tab_end_actions.place_forget()
            return
        if is_tenders:
            for btn in (self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected, self.btn_add_to_project, self.btn_manage_columns, self.btn_filters):
                btn.pack(side="left", padx=6)
            self.tab_end_actions.place(in_=self.notebook, relx=1.0, x=-10, y=2, anchor="ne")
            return
        for btn in (self.btn_fetch_orgs, self.btn_get_tenders, self.btn_download_selected, self.btn_add_to_project, self.btn_manage_columns, self.btn_filters):
            btn.pack(side="left", padx=6)
        self.tab_end_actions.place_forget()

    def create_tree(self, parent, cols, fast_horizontal=False, tree_key=None):
        frame = tk.Frame(parent)
        frame.pack(fill="both", expand=True)
        
        # Scrollbars
        v_scroll = ttk.Scrollbar(frame, orient="vertical")
        h_scroll = ttk.Scrollbar(frame, orient="horizontal")
        
        tree = ttk.Treeview(
            frame,
            columns=cols,
            show="headings",
            style=("Org.Treeview" if tree_key == "orgs" else "Wrapped.Treeview"),
            selectmode="extended",
            yscrollcommand=v_scroll.set,
            xscrollcommand=h_scroll.set
        )
        
        v_scroll.config(command=tree.yview)
        def fast_xview(*args):
            if not args:
                return
            if args[0] == "scroll" and len(args) >= 3:
                try:
                    step = int(args[1])
                except Exception:
                    step = 1
                unit = args[2]
                if unit == "units":
                    tree.xview_scroll(step * 5, "units")
                elif unit == "pages":
                    tree.xview_scroll(step * 2, "pages")
                else:
                    tree.xview(*args)
            else:
                tree.xview(*args)
        h_scroll.config(command=fast_xview)
        
        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")
        tree.pack(side="left", fill="both", expand=True)

        def scaled_col(px):
            return max(1, int(round(px * self.ui_scale * self.column_width_scale)))
        
        for col in cols:
            if tree_key:
                tree.heading(col, text=col, command=lambda c=col, k=tree_key: self.on_tree_header_click(k, c))
            else:
                tree.heading(col, text=col)
            if col == "ID" or col == "OrgID": width = 0; stretch=False; minw=0 # Hidden IDs
            elif col == "Sr": width = 50; stretch=False; minw=40
            elif col == "Select": width = 90; stretch=False; minw=70
            elif col == "Download": width = 120; stretch=False; minw=90
            else: width = 100; stretch=(not fast_horizontal); minw=80
            
            if col == "Name" or col == "Title": width = 400; stretch=True; minw=150
            if col == "Work Description": width = 420; stretch=True; minw=180
            if col == "Org Chain": width = 190; stretch=True; minw=140
            if col == "Location": width = 220; stretch=True; minw=140
            if col == "Closing Date": width = 140; stretch=False; minw=110
            if col == "Closing Time": width = 120; stretch=False; minw=96
            if col == "Website": width = 160; stretch=True; minw=120
            if col == "Downloaded At": width = 160; stretch=True; minw=140
            # Keep tender tables readable when columns are manually shrunk.
            if col == "Tender ID": minw = max(minw, 170)
            if col == "Title": minw = max(minw, 260)
            if col == "Work Description": minw = max(minw, 280)
            if col == "Org Chain": minw = max(minw, 220)
            if col == "Location": minw = max(minw, 180)
            if col == "Category": minw = max(minw, 130)
            if col == "Status": minw = max(minw, 120)
            if col == "Value" or col == "EMD": minw = max(minw, 120)
            if col == "Pre-Bid": minw = max(minw, 140)
            if fast_horizontal and col not in ("ID", "OrgID", "Sr"):
                stretch = False
            text_left_cols = {"Name", "Title", "Work Description", "Org Chain", "Location", "Website"}
            anchor = "w" if col in text_left_cols else "center"
            tree.column(col, width=scaled_col(width), minwidth=scaled_col(minw), stretch=stretch, anchor=anchor)
        
        tree._column_manager = ColumnManager(tree, self, on_apply_callback=(lambda k=tree_key, t=tree: self.on_tree_columns_applied(k, t)))
        if tree_key:
            self.restore_tree_columnwidths(tree_key, tree)
            self.restore_tree_displaycolumns(tree_key, tree)
        return tree

    def _load_json_setting(self, key, default):
        raw = get_user_setting(key, None)
        if raw is None:
            raw = self.backend.get_setting(key)
        if not raw:
            return default
        try:
            return json.loads(raw)
        except Exception:
            return default

    def _save_json_setting(self, key, value):
        try:
            set_user_setting(key, json.dumps(value))
        except Exception:
            pass

    def persist_tree_displaycolumns(self, tree_key, tree):
        if not tree_key:
            return
        dcols = tree['displaycolumns']
        if dcols == '#all':
            dcols = list(tree['columns'])
        elif isinstance(dcols, str):
            dcols = [dcols]
        else:
            dcols = list(dcols)
        self._save_json_setting(f"viewtenders_cols_{tree_key}", dcols)

    def persist_tree_columnwidths(self, tree_key, tree):
        if not tree_key:
            return
        payload = {}
        for c in list(tree["columns"]):
            try:
                payload[c] = int(tree.column(c, "width"))
            except Exception:
                continue
        self._save_json_setting(f"viewtenders_widths_{tree_key}", payload)

    def on_tree_columns_applied(self, tree_key, tree):
        self.persist_tree_displaycolumns(tree_key, tree)
        self.persist_tree_columnwidths(tree_key, tree)
        if tree_key == "tenders":
            self.schedule_tender_table_reflow(False)
        elif tree_key == "archived":
            self.schedule_tender_table_reflow(True)
        elif tree_key == "orgs":
            self.schedule_org_table_reflow()

    def restore_tree_displaycolumns(self, tree_key, tree):
        cols = self._load_json_setting(f"viewtenders_cols_{tree_key}", None)
        if not cols:
            return
        valid = [c for c in cols if c in tree['columns']]
        # Preserve hidden/unchecked columns as hidden until user re-checks them.
        if valid:
            tree['displaycolumns'] = valid

    def restore_tree_columnwidths(self, tree_key, tree):
        widths = self._load_json_setting(f"viewtenders_widths_{tree_key}", None)
        if not isinstance(widths, dict):
            return
        for c in list(tree["columns"]):
            if c not in widths:
                continue
            try:
                w = int(widths.get(c, 0))
            except Exception:
                continue
            # Keep explicit hidden ID columns hidden.
            if c in ("ID", "OrgID"):
                w = 0
            else:
                try:
                    minw = int(tree.column(c, "minwidth"))
                except Exception:
                    minw = 36
                w = max(minw, w)
            try:
                tree.column(c, width=w)
            except Exception:
                pass

    def _compute_uniform_rowheight_from_content(self, tree, base_px, max_px, line_px=18):
        # ttk.Treeview supports a single rowheight per widget (not per-row heights).
        # Fit the tallest wrapped row so text is not clipped after column resize.
        dcols = tree["displaycolumns"]
        if dcols == "#all":
            visible_cols = list(tree["columns"])
        elif isinstance(dcols, str):
            visible_cols = [dcols]
        else:
            visible_cols = list(dcols)
        col_idx = {c: i for i, c in enumerate(list(tree["columns"]))}
        visible_idx = [col_idx[c] for c in visible_cols if c in col_idx]

        line_counts = []
        for iid in tree.get_children():
            vals = tree.item(iid, "values")
            row_lines = 1
            for i in visible_idx:
                if i < len(vals):
                    txt = str(vals[i] or "")
                    row_lines = max(row_lines, txt.count("\n") + 1)
            line_counts.append(row_lines)

        if not line_counts:
            return max(base_px, min(max_px, int(line_px + 4)))

        target_lines = max(1, int(max(line_counts)))
        pad_px = max(2, int(round(line_px * 0.20)))
        return max(base_px, min(max_px, int((target_lines * line_px) + pad_px)))

    def wrap_tender_row(self, cols, row_vals):
        tree = self.tree_tenders if cols == self.tender_cols else self.tree_archived
        fallback_wrap = {
            "Website": 18,
            "Tender ID": 24,
            "Title": 32,
            "Work Description": 36,
            "Value": 14,
            "EMD": 14,
            "Org Chain": 20,
            "Closing Date": 16,
            "Closing Time": 12,
            "Pre-Bid": 16,
            "Location": 20,
            "Category": 16,
            "Status": 14,
        }
        skip_cols = {"Sr", "ID", "OrgID", "Select", "Download"}
        width_bias = {
            "Title": 1.14,
            "Org Chain": 1.18,
        }
        return wrap_tree_row_values(
            tree,
            cols,
            row_vals,
            skip_cols=skip_cols,
            fallback_wrap=fallback_wrap,
            width_bias=width_bias
        )

    def split_date_time_text(self, value):
        txt = " ".join(str(value or "").split())
        if not txt:
            return "N/A", "N/A"
        if txt.upper() == "N/A":
            return "N/A", "N/A"
        # Typical NIC pattern: 20-Feb-2026 03:00 PM
        m = re.match(r'^(\d{1,2}-[A-Za-z]{3}-\d{4})\s+(\d{1,2}:\d{2}\s*[APMapm]{2})$', txt)
        if m:
            return m.group(1), m.group(2).upper().replace(" ", "")
        # Generic split fallback.
        parts = txt.split(" ", 1)
        if len(parts) == 2:
            return parts[0], parts[1]
        return txt, "N/A"

    def open_download_root_folder(self):
        path = os.path.abspath(BASE_DOWNLOAD_DIRECTORY)
        if not os.path.exists(path):
            os.makedirs(path, exist_ok=True)
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Open Folder", f"Could not open folder:\n{e}")

    def open_active_table_filter_settings(self):
        current_tab = self.notebook.select()
        if current_tab == str(self.tab_orgs):
            self.open_filter_dialog("orgs", self.org_cols)
        elif current_tab == str(self.tab_tenders):
            self.open_filter_dialog("tenders", self.tender_cols)
        elif current_tab == str(self.tab_archived):
            self.open_filter_dialog("archived", self.archived_cols)
        else:
            messagebox.showinfo("Filters", "Filters are available for table tabs.")

    def open_filter_dialog(self, table_key, cols):
        win = tk.Toplevel(self)
        win.title("Column Filters")
        win.geometry("620x560")

        active_filters = self.filter_map.get(table_key, {})

        top = tk.Frame(win)
        top.pack(fill="x", padx=8, pady=8)
        tk.Label(top, text="Column:").pack(side="left")
        col_var = tk.StringVar(value=cols[0] if cols else "")
        cb_col = ttk.Combobox(top, state="readonly", values=cols, textvariable=col_var, width=24)
        cb_col.pack(side="left", padx=6)

        tk.Label(top, text="Filter Type:").pack(side="left", padx=(10, 0))
        mode_var = tk.StringVar(value="Values")
        mode_options = ["Values", "Text Contains", "Text Equals", "Number Comparison"]
        cb_mode = ttk.Combobox(top, state="readonly", values=mode_options, textvariable=mode_var, width=20)
        cb_mode.pack(side="left", padx=6)

        body = tk.Frame(win)
        body.pack(fill="both", expand=True, padx=8, pady=4)
        left = tk.LabelFrame(body, text="Options")
        left.pack(side="left", fill="both", expand=True, padx=(0, 6))
        right = tk.LabelFrame(body, text="Current Filters")
        right.pack(side="right", fill="y", padx=(6, 0))

        values_wrap = tk.Frame(left)
        values_canvas = tk.Canvas(values_wrap, highlightthickness=0)
        values_scroll = ttk.Scrollbar(values_wrap, orient="vertical", command=values_canvas.yview)
        values_inner = tk.Frame(values_canvas)
        values_inner.bind("<Configure>", lambda e: values_canvas.configure(scrollregion=values_canvas.bbox("all")))
        values_canvas.create_window((0, 0), window=values_inner, anchor="nw")
        values_canvas.configure(yscrollcommand=values_scroll.set)
        values_canvas.pack(side="left", fill="both", expand=True)
        values_scroll.pack(side="right", fill="y")
        value_checks = {}
        def _on_values_mousewheel(event):
            delta = event.delta
            if platform.system() == "Darwin":
                values_canvas.yview_scroll(int(-delta), "units")
            else:
                values_canvas.yview_scroll(int(-delta / 120), "units")
            return "break"
        values_canvas.bind("<MouseWheel>", _on_values_mousewheel)
        values_canvas.bind("<Button-4>", lambda e: (values_canvas.yview_scroll(-3, "units"), "break")[1])
        values_canvas.bind("<Button-5>", lambda e: (values_canvas.yview_scroll(3, "units"), "break")[1])

        values_btns = tk.Frame(left)
        def select_all_values():
            for v in value_checks.values():
                v.set(1)
        def clear_values():
            for v in value_checks.values():
                v.set(0)
        tk.Button(values_btns, text="Select All", command=select_all_values).pack(side="left", padx=4)
        tk.Button(values_btns, text="Clear", command=clear_values).pack(side="left", padx=4)

        op_row = tk.Frame(left)
        op_row.pack(fill="x", padx=8, pady=4)
        tk.Label(op_row, text="Operator:").pack(side="left")
        op_var = tk.StringVar(value="=")
        op_cb = ttk.Combobox(op_row, state="readonly", values=["=", "!=", ">", ">=", "<", "<="], textvariable=op_var, width=6)
        op_cb.pack(side="left", padx=6)
        value_var = tk.StringVar()
        value_entry = tk.Entry(op_row, textvariable=value_var)
        value_entry.pack(side="left", fill="x", expand=True)

        helper_lbl = tk.Label(left, text="", anchor="w", fg="#555")
        helper_lbl.pack(fill="x", padx=8, pady=(0, 8))

        list_filters = tk.Listbox(right, height=18, width=32)
        list_filters.pack(fill="both", expand=True, padx=8, pady=8)

        def refresh_current_filters():
            list_filters.delete(0, tk.END)
            for c, f in active_filters.items():
                if isinstance(f, str):
                    list_filters.insert(tk.END, f"{c}: contains '{f}'")
                    continue
                m = f.get("mode", "")
                if m == "values":
                    sel = f.get("selected", [])
                    list_filters.insert(tk.END, f"{c}: {len(sel)} value(s)")
                elif m == "contains":
                    list_filters.insert(tk.END, f"{c}: contains '{f.get('value','')}'")
                elif m == "equals":
                    list_filters.insert(tk.END, f"{c}: equals '{f.get('value','')}'")
                elif m == "number":
                    list_filters.insert(tk.END, f"{c}: {f.get('op','=')} {f.get('value','')}")

        def get_tree_for_key(k):
            return self.tree_orgs if k == "orgs" else (self.tree_tenders if k == "tenders" else self.tree_archived)

        def get_unique_values(col_name):
            vals = set()
            tree = get_tree_for_key(table_key)
            if col_name not in cols:
                return []
            idx = cols.index(col_name)
            for iid in tree.get_children():
                row = tree.item(iid, "values")
                if idx < len(row):
                    vals.add(str(row[idx]))
            return sorted(vals, key=lambda x: x.lower())

        def refresh_mode_ui(*_):
            mode = mode_var.get()
            col = col_var.get()
            if mode == "Values":
                for w in values_inner.winfo_children():
                    w.destroy()
                value_checks.clear()
                existing = active_filters.get(col, {})
                selected_existing = set(existing.get("selected", [])) if isinstance(existing, dict) and existing.get("mode") == "values" else set()
                for v in get_unique_values(col):
                    var = tk.IntVar(value=1 if v in selected_existing else 0)
                    value_checks[v] = var
                    cb = tk.Checkbutton(values_inner, text=v, variable=var, anchor="w")
                    cb.pack(fill="x", padx=4, pady=1)
                values_wrap.pack(fill="both", expand=True, padx=8, pady=8)
                values_btns.pack(fill="x", padx=8, pady=(0, 6))
                op_row.pack_forget()
                helper_lbl.config(text="Select one or more values using checkboxes.")
            elif mode in ("Text Contains", "Text Equals"):
                values_wrap.pack_forget()
                values_btns.pack_forget()
                op_row.pack(fill="x", padx=8, pady=4)
                op_cb.pack_forget()
                value_entry.pack(side="left", fill="x", expand=True)
                helper_lbl.config(text="Enter text value.")
            else:
                values_wrap.pack_forget()
                values_btns.pack_forget()
                op_row.pack(fill="x", padx=8, pady=4)
                op_cb.pack(side="left", padx=6)
                value_entry.pack(side="left", fill="x", expand=True)
                helper_lbl.config(text="Numeric comparison.")

        def apply_one():
            col = col_var.get()
            mode = mode_var.get()
            if not col:
                return
            if mode == "Values":
                sel = [k for k, var in value_checks.items() if var.get() == 1]
                if not sel:
                    active_filters.pop(col, None)
                else:
                    active_filters[col] = {"mode": "values", "selected": sel}
            elif mode == "Text Contains":
                txt = value_var.get().strip()
                if not txt:
                    active_filters.pop(col, None)
                else:
                    active_filters[col] = {"mode": "contains", "value": txt}
            elif mode == "Text Equals":
                txt = value_var.get().strip()
                if not txt:
                    active_filters.pop(col, None)
                else:
                    active_filters[col] = {"mode": "equals", "value": txt}
            else:
                txt = value_var.get().strip()
                if not txt:
                    active_filters.pop(col, None)
                else:
                    active_filters[col] = {"mode": "number", "op": op_var.get(), "value": txt}
            self.filter_map[table_key] = active_filters
            self._save_json_setting(f"viewtenders_filters_{table_key}", active_filters)
            refresh_current_filters()
            self.on_site_changed()

        def clear_one():
            col = col_var.get()
            if col in active_filters:
                active_filters.pop(col, None)
                self.filter_map[table_key] = active_filters
                self._save_json_setting(f"viewtenders_filters_{table_key}", active_filters)
                refresh_current_filters()
                self.on_site_changed()

        def clear_all():
            self.filter_map[table_key] = {}
            self._save_json_setting(f"viewtenders_filters_{table_key}", {})
            active_filters.clear()
            refresh_current_filters()
            self.on_site_changed()

        btns = tk.Frame(win)
        btns.pack(fill="x", padx=8, pady=8)
        tk.Button(btns, text="Apply Column Filter", bg="#4CAF50", fg="white", command=apply_one).pack(side="left", padx=4)
        tk.Button(btns, text="Clear Column Filter", bg="#F57C00", fg="white", command=clear_one).pack(side="left", padx=4)
        tk.Button(btns, text="Clear All Filters", bg="#F44336", fg="white", command=clear_all).pack(side="left", padx=4)
        tk.Button(btns, text="Close", command=win.destroy).pack(side="right", padx=4)

        cb_col.bind("<<ComboboxSelected>>", refresh_mode_ui)
        cb_mode.bind("<<ComboboxSelected>>", refresh_mode_ui)
        refresh_current_filters()
        refresh_mode_ui()

    def _normalize_export_setting(self, value, fallback):
        if isinstance(value, str):
            try:
                value = json.loads(value)
            except Exception:
                return fallback
        return value if value is not None else fallback

    def _get_active_tender_export_rows(self, selected_cols):
        if not selected_cols:
            return []
        idx_map = {c: i for i, c in enumerate(self.tender_cols)}
        rows = []
        for iid in self.tree_tenders.get_children():
            vals = list(self.tree_tenders.item(iid, "values"))
            row_out = []
            for col in selected_cols:
                idx = idx_map.get(col, -1)
                v = vals[idx] if 0 <= idx < len(vals) else ""
                row_out.append(" ".join(str(v).replace("\n", " ").split()))
            rows.append(row_out)
        return rows

    def _draw_pdf_table(self, out_path, cols, rows, width_map):
        if not PDF_SUPPORT:
            raise RuntimeError("PDF export requires PyMuPDF (fitz).")

        doc = fitz.open()
        margin = 28
        font_size = 8
        line_h = 12
        page_rect = fitz.paper_rect("a4-l")
        page = doc.new_page(width=page_rect.width, height=page_rect.height)
        table_w = page.rect.width - (2 * margin)
        header_bg = (0.90, 0.93, 0.97)

        weights = []
        for c in cols:
            try:
                w = int(width_map.get(c, 126))
            except Exception:
                w = 126
            weights.append(max(48, min(720, w)))
        total_w = float(sum(weights) or 1.0)
        col_w = [(w / total_w) * table_w for w in weights]
        x_edges = [margin]
        for w in col_w:
            x_edges.append(x_edges[-1] + w)

        def draw_header(y):
            hdr_h = line_h + 6
            for i, c in enumerate(cols):
                r = fitz.Rect(x_edges[i], y, x_edges[i + 1], y + hdr_h)
                page.draw_rect(r, color=(0, 0, 0), fill=header_bg, width=0.6)
                page.insert_textbox(r, str(c), fontsize=font_size, align=fitz.TEXT_ALIGN_CENTER)
            return y + hdr_h

        def col_char_cap(i):
            # Conservative wrap width to avoid clipping in narrow columns.
            return max(5, int(col_w[i] / 5.2))

        y = draw_header(margin)
        bottom = page.rect.height - margin

        for row in rows:
            wrapped_cells = []
            max_lines = 1
            for i, cell in enumerate(row):
                txt = str(cell or "")
                lines = textwrap.wrap(txt, width=col_char_cap(i)) or [""]
                wrapped_cells.append("\n".join(lines))
                max_lines = max(max_lines, len(lines))

            row_h = (max_lines * line_h) + 8
            if y + row_h > bottom:
                page = doc.new_page(width=page_rect.width, height=page_rect.height)
                y = draw_header(margin)

            for i, txt in enumerate(wrapped_cells):
                r = fitz.Rect(x_edges[i], y, x_edges[i + 1], y + row_h)
                page.draw_rect(r, color=(0, 0, 0), width=0.5)
                page.insert_textbox(r, txt, fontsize=font_size, align=fitz.TEXT_ALIGN_LEFT)
            y += row_h

        doc.save(out_path)
        doc.close()

    def _export_active_tenders(self, export_kind, selected_cols, preview_widths):
        if not selected_cols:
            messagebox.showwarning("Export", "Select at least one column.")
            return

        rows = self._get_active_tender_export_rows(selected_cols)
        if not rows:
            messagebox.showwarning("Export", "No rows to export in Active Tenders.")
            return

        now_tag = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_kind == "excel":
            path = filedialog.asksaveasfilename(
                parent=self,
                title="Export Active Tenders to Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel Workbook", "*.xlsx")],
                initialfile=f"Active_Tenders_{now_tag}.xlsx"
            )
            if not path:
                return
            if "pd" not in globals():
                messagebox.showerror("Export Excel", "Pandas is not available. Install pandas/openpyxl.")
                return
            df = pd.DataFrame(rows, columns=selected_cols)
            try:
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Active Tenders")
                    try:
                        import textwrap as _tw
                        from openpyxl.utils import get_column_letter
                        from openpyxl.styles import Alignment
                        ws = writer.sheets.get("Active Tenders")
                        if ws is not None:
                            excel_width_map = {}
                            for i, c in enumerate(selected_cols, 1):
                                try:
                                    px = int(preview_widths.get(c, 126))
                                    w = max(8, min(80, int(px / 7)))
                                except Exception:
                                    w = 18
                                w = max(8, min(80, w))
                                excel_width_map[i] = w
                                ws.column_dimensions[get_column_letter(i)].width = w

                            # Wrap text and set dynamic row height based on wrapped line count.
                            for row_idx in range(2, ws.max_row + 1):
                                max_lines = 1
                                for col_idx in range(1, len(selected_cols) + 1):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    txt = "" if cell.value is None else str(cell.value)
                                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                                    col_w = float(excel_width_map.get(col_idx, 18))
                                    char_cap = max(6, int(col_w - 1))
                                    wrapped = _tw.wrap(txt, width=char_cap, break_long_words=True, replace_whitespace=False) or [""]
                                    max_lines = max(max_lines, len(wrapped))
                                ws.row_dimensions[row_idx].height = max(15, (max_lines * 15))
                    except Exception:
                        pass
            except Exception as e:
                messagebox.showerror("Export Excel", f"Failed to export:\n{e}")
                return
            messagebox.showinfo("Export Excel", f"Exported successfully:\n{path}")
            return

        path = filedialog.asksaveasfilename(
            parent=self,
            title="Export Active Tenders to PDF",
            defaultextension=".pdf",
            filetypes=[("PDF Document", "*.pdf")],
            initialfile=f"Active_Tenders_{now_tag}.pdf"
        )
        if not path:
            return
        try:
            self._draw_pdf_table(path, selected_cols, rows, preview_widths)
        except Exception as e:
            messagebox.showerror("Export PDF", f"Failed to export:\n{e}")
            return
        messagebox.showinfo("Export PDF", f"Exported successfully:\n{path}")

    def open_active_tender_export_popup(self):
        if self.notebook.select() != str(self.tab_tenders):
            messagebox.showinfo("Export", "Export is available in Active Tenders tab.")
            return

        key_cols = "viewtenders_export_cols"
        key_order = "viewtenders_export_order"

        saved_cols = self._normalize_export_setting(get_user_setting(key_cols, None), None)
        saved_order = self._normalize_export_setting(get_user_setting(key_order, None), None)

        if not isinstance(saved_cols, list):
            dcols = self.tree_tenders["displaycolumns"]
            if dcols == "#all":
                saved_cols = [c for c in self.tender_cols if c != "ID"]
            elif isinstance(dcols, str):
                saved_cols = [dcols] if dcols != "ID" else []
            else:
                saved_cols = [c for c in list(dcols) if c != "ID"]
        saved_cols = [c for c in saved_cols if c in self.tender_cols and c != "ID"]

        if not saved_cols:
            saved_cols = [c for c in self.tender_cols if c not in ("ID",)]

        all_export_cols = [c for c in self.tender_cols if c != "ID"]
        if isinstance(saved_order, list):
            ordered_cols = [c for c in saved_order if c in all_export_cols]
        else:
            ordered_cols = []
        for c in all_export_cols:
            if c not in ordered_cols:
                ordered_cols.append(c)
        checked_cols = {c for c in saved_cols if c in all_export_cols}
        if not checked_cols:
            checked_cols = set(all_export_cols)

        win = tk.Toplevel(self)
        win.title("Export Active Tenders")
        win.geometry("1180x700")
        win.minsize(980, 560)
        win.transient(self)
        win.grab_set()

        outer = tk.Frame(win, padx=10, pady=10)
        outer.pack(fill="both", expand=True)
        outer.grid_columnconfigure(1, weight=1)
        outer.grid_rowconfigure(0, weight=1)

        left = tk.LabelFrame(outer, text="Columns", padx=6, pady=6)
        left.grid(row=0, column=0, sticky="nsw")
        right_mid = tk.LabelFrame(outer, text="Preview (first 80 rows)", padx=6, pady=6)
        right_mid.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        right_mid.grid_rowconfigure(0, weight=1)
        right_mid.grid_columnconfigure(0, weight=1)

        cols_wrap = tk.Frame(left)
        cols_wrap.pack(fill="both", expand=True)
        cols_list = tk.Listbox(cols_wrap, height=28, exportselection=False)
        cols_scroll = ttk.Scrollbar(cols_wrap, orient="vertical", command=cols_list.yview)
        cols_list.configure(yscrollcommand=cols_scroll.set)
        cols_list.pack(side="left", fill="both", expand=True)
        cols_scroll.pack(side="right", fill="y")

        pv_wrap = tk.Frame(right_mid)
        pv_wrap.grid(row=0, column=0, sticky="nsew")
        pv_wrap.grid_rowconfigure(0, weight=1)
        pv_wrap.grid_columnconfigure(0, weight=1)
        pv_tree = ttk.Treeview(pv_wrap, show="headings")
        pv_y = ttk.Scrollbar(pv_wrap, orient="vertical", command=pv_tree.yview)
        pv_x = ttk.Scrollbar(pv_wrap, orient="horizontal", command=pv_tree.xview)
        pv_tree.configure(yscrollcommand=pv_y.set, xscrollcommand=pv_x.set)
        pv_tree.grid(row=0, column=0, sticky="nsew")
        pv_y.grid(row=0, column=1, sticky="ns")
        pv_x.grid(row=1, column=0, sticky="ew")

        def selected_cols():
            return [c for c in ordered_cols if c in checked_cols]

        def save_export_prefs(cols_now):
            set_user_setting(key_cols, cols_now)
            set_user_setting(key_order, ordered_cols)

        def current_preview_widths(cols_now):
            widths = {}
            for c in cols_now:
                try:
                    widths[c] = int(pv_tree.column(c, "width"))
                except Exception:
                    widths[c] = 126
            return widths

        def render_columns_list():
            cols_list.delete(0, tk.END)
            for c in ordered_cols:
                mark = "[x]" if c in checked_cols else "[ ]"
                cols_list.insert(tk.END, f"{mark}  {c}")

        def refresh_preview():
            cols_now = selected_cols()
            pv_tree.delete(*pv_tree.get_children())
            pv_tree["columns"] = cols_now
            for c in cols_now:
                pv_tree.heading(c, text=c)
                try:
                    base_w = int(self.tree_tenders.column(c, "width"))
                except Exception:
                    base_w = 126
                pv_tree.column(c, width=base_w, minwidth=50, stretch=False, anchor="w")
            for row in self._get_active_tender_export_rows(cols_now)[:80]:
                pv_tree.insert("", "end", values=row)
            save_export_prefs(cols_now)

        drag_state = {"from_idx": None}

        def on_columns_press(event):
            if not ordered_cols:
                return "break"
            idx = int(cols_list.nearest(event.y))
            if idx < 0 or idx >= len(ordered_cols):
                return "break"
            cols_list.selection_clear(0, tk.END)
            cols_list.selection_set(idx)
            if event.x <= 24:
                col = ordered_cols[idx]
                if col in checked_cols:
                    checked_cols.discard(col)
                else:
                    checked_cols.add(col)
                render_columns_list()
                cols_list.selection_set(idx)
                refresh_preview()
                return "break"
            drag_state["from_idx"] = idx

        def on_columns_drag(event):
            src = drag_state.get("from_idx")
            if src is None or not ordered_cols:
                return "break"
            dst = int(cols_list.nearest(event.y))
            if dst < 0:
                dst = 0
            if dst >= len(ordered_cols):
                dst = len(ordered_cols) - 1
            if dst == src:
                return "break"
            moved = ordered_cols.pop(src)
            ordered_cols.insert(dst, moved)
            drag_state["from_idx"] = dst
            render_columns_list()
            cols_list.selection_set(dst)
            refresh_preview()
            return "break"

        def on_columns_release(event):
            drag_state["from_idx"] = None
            return "break"

        cols_list.bind("<Button-1>", on_columns_press)
        cols_list.bind("<B1-Motion>", on_columns_drag)
        cols_list.bind("<ButtonRelease-1>", on_columns_release)

        def run_export(export_kind):
            cols_now = selected_cols()
            save_export_prefs(cols_now)
            self._export_active_tenders(export_kind, cols_now, current_preview_widths(cols_now))

        btns = tk.Frame(win, padx=10, pady=8)
        btns.pack(fill="x")
        tk.Button(btns, text="Export Excel", bg="#2E7D32", fg="white", command=lambda: run_export("excel")).pack(side="left", padx=4)
        tk.Button(btns, text="Export PDF", bg="#455A64", fg="white", command=lambda: run_export("pdf")).pack(side="left", padx=4)

        render_columns_list()
        refresh_preview()

    def row_matches_filters(self, table_key, cols, row_values):
        filters = self.filter_map.get(table_key, {})
        if not filters:
            return True
        value_map = {c: str(row_values[i]) if i < len(row_values) else "" for i, c in enumerate(cols)}
        for col, needle in filters.items():
            hay_raw = value_map.get(col, "")
            hay = hay_raw.lower()
            if isinstance(needle, str):
                if needle.lower() not in hay:
                    return False
                continue
            mode = needle.get("mode")
            if mode == "values":
                selected = [str(x) for x in needle.get("selected", [])]
                if selected and str(hay_raw) not in selected:
                    return False
            elif mode == "equals":
                if hay != str(needle.get("value", "")).lower():
                    return False
            elif mode == "contains":
                if str(needle.get("value", "")).lower() not in hay:
                    return False
            elif mode == "number":
                def to_num(v):
                    try:
                        vv = str(v).replace(",", "").replace("Rs.", "").replace("INR", "").strip()
                        return float(vv)
                    except Exception:
                        return None
                lhs = to_num(hay_raw)
                rhs = to_num(needle.get("value"))
                if lhs is None or rhs is None:
                    return False
                op = needle.get("op", "=")
                ok = ((op == "=" and lhs == rhs) or
                      (op == "!=" and lhs != rhs) or
                      (op == ">" and lhs > rhs) or
                      (op == ">=" and lhs >= rhs) or
                      (op == "<" and lhs < rhs) or
                      (op == "<=" and lhs <= rhs))
                if not ok:
                    return False
        return True

    def on_tree_header_click(self, table_key, col):
        state = self.sort_map.get(table_key, {}) or {}
        if state.get("column") == col:
            ascending = not bool(state.get("ascending", True))
        else:
            ascending = True
        new_state = {"column": col, "ascending": ascending}
        self.sort_map[table_key] = new_state
        self._save_json_setting(f"viewtenders_sort_{table_key}", new_state)
        self.on_site_changed()

    def apply_sort(self, table_key, cols, display_rows):
        state = self.sort_map.get(table_key, {}) or {}
        col = state.get("column")
        if not col or col not in cols:
            return display_rows
        idx = cols.index(col)
        asc = bool(state.get("ascending", True))

        # Custom Value sorting:
        # Asc: NA first, then 0, then increasing numbers.
        # Desc: highest numbers first, NA at end.
        if col == "Value":
            def to_num(v):
                try:
                    vv = str(v).replace(",", "").replace("Rs.", "").replace("INR", "").strip()
                    return float(vv)
                except Exception:
                    return None

            if asc:
                def key_value_asc(row):
                    if idx >= len(row):
                        return (0, 0.0)
                    raw = row[idx]
                    s = str(raw).strip()
                    n = to_num(raw)
                    is_na = (not s) or (s.upper() == "NA") or (n is None)
                    return (0, 0.0) if is_na else (1, n)
                return sorted(display_rows, key=key_value_asc)
            else:
                def key_value_desc(row):
                    if idx >= len(row):
                        return (1, 0.0)
                    raw = row[idx]
                    s = str(raw).strip()
                    n = to_num(raw)
                    is_na = (not s) or (s.upper() == "NA") or (n is None)
                    return (1, 0.0) if is_na else (0, -n)
                return sorted(display_rows, key=key_value_desc)

        def key_fn(row):
            if idx >= len(row):
                return (3, "")
            val = row[idx]
            s = str(val).strip()
            # numeric sort first if possible
            try:
                n = float(s.replace(",", "").replace("Rs.", "").replace("INR", "").strip())
                return (0, n)
            except Exception:
                pass
            # datetime-ish sort
            for fmt in ("%d-%b-%Y %I:%M %p", "%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return (1, datetime.datetime.strptime(s, fmt))
                except Exception:
                    continue
            return (2, s.lower())

        return sorted(display_rows, key=key_fn, reverse=not asc)

    def open_active_table_column_settings(self):
        current_tab = self.notebook.select()
        if current_tab == str(self.tab_orgs):
            self.tree_orgs._column_manager.open_settings()
        elif current_tab == str(self.tab_tenders):
            self.tree_tenders._column_manager.open_settings()
        elif current_tab == str(self.tab_archived):
            self.tree_archived._column_manager.open_settings()
        else:
            messagebox.showinfo("Manage Columns", "Column settings are available for table tabs.")

    def fast_horizontal_scroll(self, event):
        step = -8 if event.delta > 0 else 8
        self.tree_tenders.xview_scroll(step, "units")
        return "break"

    def fast_horizontal_scroll_archived(self, event):
        step = -8 if event.delta > 0 else 8
        self.tree_archived.xview_scroll(step, "units")
        return "break"

    def refresh_sites(self):
        sites = self.backend.get_websites()
        values = ["ALL: All Websites"] + [f"{k}: {v['name']}" for k, v in sites.items()]
        self.cb_sites['values'] = values
        if values:
            preferred = get_user_setting("viewtenders_selected_site", self.backend.get_setting("viewtenders_selected_site", "ALL"))
            idx = 0
            for i, val in enumerate(values):
                if val.split(":")[0].strip().upper() == str(preferred).upper():
                    idx = i
                    break
            self.cb_sites.current(idx)
            self.on_site_changed()

    def get_selected_site_id(self):
        val = self.cb_sites.get()
        if not val: return None
        raw = val.split(":")[0].strip()
        if raw.upper() == "ALL":
            return None
        return int(raw)

    def get_target_site_ids(self):
        sid = self.get_selected_site_id()
        if sid is not None:
            return [sid]
        return sorted(self.backend.get_websites().keys())

    def append_log(self, text):
        self.txt_log.insert(tk.END, text + "\n")
        self.txt_log.see(tk.END)

    def manage_websites_popup(self):
        win = tk.Toplevel(self)
        win.title("Manage Websites")
        win.geometry("980x560")
        win.minsize(760, 420)

        top = tk.LabelFrame(win, text="Manage Website", padx=8, pady=8)
        top.pack(fill="x", padx=8, pady=(8, 4))
        bottom = tk.LabelFrame(win, text="Existing Websites", padx=8, pady=8)
        bottom.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        cols = ("Sr", "ID", "Name", "Tenders URL", "Status URL")
        tree = ttk.Treeview(bottom, columns=cols, show="headings", height=14, style="Wrapped.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            if c == "Sr":
                tree.column(c, width=45, anchor="center")
            elif c == "ID":
                tree.column(c, width=55, anchor="center")
            elif c == "Name":
                tree.column(c, width=230, anchor="w")
            else:
                tree.column(c, width=380, anchor="w")
        ysb = ttk.Scrollbar(bottom, orient="vertical", command=tree.yview)
        xsb = ttk.Scrollbar(bottom, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        tree.pack(side="left", fill="both", expand=True, pady=(0, 6))
        ysb.pack(side="right", fill="y")
        xsb.pack(side="bottom", fill="x")

        # Fast/smooth scrolling for website table (Windows/macOS/Linux).
        def on_mousewheel(event):
            delta = event.delta
            if platform.system() == "Darwin":
                step = int(-delta) if delta != 0 else 0
            else:
                step = int(-delta / 120) if delta != 0 else 0
            if step != 0:
                tree.yview_scroll(step * 4, "units")
            return "break"

        def on_shift_mousewheel(event):
            delta = event.delta
            if platform.system() == "Darwin":
                step = int(-delta) if delta != 0 else 0
            else:
                step = int(-delta / 120) if delta != 0 else 0
            if step != 0:
                tree.xview_scroll(step * 6, "units")
            return "break"

        tree.bind("<MouseWheel>", on_mousewheel)
        tree.bind("<Shift-MouseWheel>", on_shift_mousewheel)
        tree.bind("<Button-4>", lambda e: (tree.yview_scroll(-4, "units"), "break")[1])
        tree.bind("<Button-5>", lambda e: (tree.yview_scroll(4, "units"), "break")[1])

        def fill_tree():
            for i in tree.get_children():
                tree.delete(i)
            websites = self.backend.get_websites()
            for idx, (sid, info) in enumerate(sorted(websites.items(), key=lambda x: int(x[0])), 1):
                row_vals = wrap_tree_row_values(
                    tree,
                    cols,
                    (idx, sid, info["name"], info["url"], info["status_url"]),
                    skip_cols={"Sr", "ID"},
                    fallback_wrap={"Name": 20, "Tenders URL": 36, "Status URL": 36}
                )
                tree.insert("", "end", values=row_vals)

        fill_tree()

        def delete_selected():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Delete Website", "Select a website first.")
                return
            vals = tree.item(sel[0], "values")
            sid = int(vals[1])
            name = vals[2]
            if not messagebox.askyesno("Confirm", f"Delete website '{name}' and all related scraper data?"):
                return
            if self.backend.delete_website_logic(sid):
                fill_tree()
                self.refresh_sites()

        form = tk.Frame(top)
        form.pack(fill="x")
        form.grid_columnconfigure(1, weight=1)
        form.grid_columnconfigure(3, weight=1)
        form.grid_columnconfigure(5, weight=1)

        tk.Label(form, text="Name:").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=(0, 8))
        e_name = tk.Entry(form)
        e_name.grid(row=0, column=1, sticky="ew", padx=(0, 14), pady=(0, 8))
        tk.Label(form, text="Tenders URL:").grid(row=0, column=2, sticky="w", padx=(0, 6), pady=(0, 8))
        e_url = tk.Entry(form)
        e_url.grid(row=0, column=3, sticky="ew", padx=(0, 14), pady=(0, 8))
        tk.Label(form, text="Status URL:").grid(row=0, column=4, sticky="w", padx=(0, 6), pady=(0, 8))
        e_surl = tk.Entry(form)
        e_surl.grid(row=0, column=5, sticky="ew", pady=(0, 8))

        def load_selected_into_form(event=None):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            e_name.delete(0, tk.END); e_name.insert(0, vals[2])
            e_url.delete(0, tk.END); e_url.insert(0, vals[3])
            e_surl.delete(0, tk.END); e_surl.insert(0, vals[4])
        tree.bind("<<TreeviewSelect>>", load_selected_into_form)

        def add_new():
            name = e_name.get().strip()
            url = e_url.get().strip()
            status_url = e_surl.get().strip()
            if not name or not url or not status_url:
                messagebox.showwarning("Add Website", "All fields are required.")
                return
            if self.backend.add_website_logic(name, url, status_url):
                e_name.delete(0, tk.END)
                e_url.delete(0, tk.END)
                e_surl.delete(0, tk.END)
                fill_tree()
                self.refresh_sites()

        def edit_selected():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Edit Website", "Select a website first.")
                return
            vals = tree.item(sel[0], "values")
            sid = int(vals[1])
            name = e_name.get().strip()
            url = e_url.get().strip()
            status_url = e_surl.get().strip()
            if not name or not url or not status_url:
                messagebox.showwarning("Edit Website", "All fields are required.")
                return
            conn = sqlite3.connect(DB_FILE)
            try:
                conn.execute("UPDATE websites SET name=?, url=?, status_url=? WHERE id=?", (name, url, status_url, sid))
                conn.commit()
            except Exception as e:
                messagebox.showerror("Edit Website", str(e))
            finally:
                conn.close()
            fill_tree()
            self.refresh_sites()

        actions = tk.Frame(top)
        actions.pack(fill="x")
        tk.Button(actions, text="Add Website", bg="#2e7d32", fg="white", command=add_new).pack(side="left", padx=3)
        tk.Button(actions, text="Edit Selected", bg="#1565C0", fg="white", command=edit_selected).pack(side="left", padx=3)
        tk.Button(actions, text="Delete Selected", bg="#d32f2f", fg="white", command=delete_selected).pack(side="left", padx=3)
        tk.Button(actions, text="Refresh", command=fill_tree).pack(side="left", padx=8)
        tk.Button(actions, text="Close", command=win.destroy).pack(side="right", padx=3)

    def clear_saved_details_popup(self):
        win = tk.Toplevel(self)
        win.title("Clear Saved Details")
        win.geometry("430x270")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="Choose what to clear:", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16, pady=(14, 8))
        var_org = tk.IntVar(value=1)
        var_active = tk.IntVar(value=1)
        var_arch = tk.IntVar(value=1)
        tk.Checkbutton(win, text="1. Organization", variable=var_org).pack(anchor="w", padx=20, pady=2)
        tk.Checkbutton(win, text="2. Active Tenders", variable=var_active).pack(anchor="w", padx=20, pady=2)
        tk.Checkbutton(win, text="3. Archive", variable=var_arch).pack(anchor="w", padx=20, pady=2)
        tk.Label(
            win,
            text="Websites are kept. Downloaded files on disk are not deleted.",
            fg="#555"
        ).pack(anchor="w", padx=20, pady=(8, 0))

        btns = tk.Frame(win)
        btns.pack(fill="x", side="bottom", padx=14, pady=14)

        def run_clear():
            clear_orgs = bool(var_org.get())
            clear_active = bool(var_active.get())
            clear_archived = bool(var_arch.get())
            if not (clear_orgs or clear_active or clear_archived):
                messagebox.showwarning("Clear Saved Details", "Select at least one option.", parent=win)
                return
            if not messagebox.askyesno("Confirm", "Proceed with selected clear options?", icon="warning", parent=win):
                return
            result = self.backend.clear_saved_scraper_details_logic(
                clear_orgs=clear_orgs,
                clear_active=clear_active,
                clear_archived=clear_archived,
            )
            if result is None:
                messagebox.showerror("Clear Saved Details", "Failed to clear saved details. Check logs.", parent=win)
                return
            win.destroy()
            self.refresh_sites()
            self.notebook.select(self.tab_logs)
            summary = (
                f"Cleared: organizations={result['organizations']}, "
                f"active={result['active_tenders']}, archived={result['archived_tenders']}, "
                f"downloads={result['downloaded_files']}"
            )
            log_to_gui(summary)
            messagebox.showinfo("Clear Saved Details", summary)

        tk.Button(btns, text="Clear Selected", bg="#B71C1C", fg="white", command=run_clear).pack(side="left")
        tk.Button(btns, text="Cancel", command=win.destroy).pack(side="right")

    def on_site_changed(self, event=None):
        sid = self.get_selected_site_id()
        set_user_setting("viewtenders_selected_site", "ALL" if sid is None else str(sid))
        self.load_org_tree()
        self.load_tender_tree()
        self.load_archived_tender_tree()
        self.after_idle(self._stabilize_visible_table_wrap)

    # --- Actions ---
    def run_fetch_orgs(self):
        self.notebook.select(self.tab_logs)
        threading.Thread(target=self.thread_fetch_orgs, daemon=True).start()

    def thread_fetch_orgs(self):
        ok_any = False
        for sid in self.get_target_site_ids():
            ok = self.backend.fetch_organisations_logic(sid)
            ok_any = ok_any or bool(ok)
        if ok_any:
            self.after(0, self.on_site_changed)

    def load_org_tree(self, auto_reflow=True):
        clear_tree_rows(self.tree_orgs)
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        sid = self.get_selected_site_id()
        if sid is None:
            c.execute("""SELECT o.id, w.name, o.name, o.tender_count, o.is_selected
                         FROM organizations o
                         JOIN websites w ON w.id=o.website_id
                         ORDER BY w.name, o.id""")
        else:
            c.execute("""SELECT o.id, w.name, o.name, o.tender_count, o.is_selected
                         FROM organizations o
                         JOIN websites w ON w.id=o.website_id
                         WHERE o.website_id=?
                         ORDER BY o.id""", (sid,))
        rows = c.fetchall()
        conn.close()
        display_rows = []
        for r in rows:
            row_vals = (0, r[0], r[1], r[2], r[3], ("Yes" if r[4] else "No"))
            if self.row_matches_filters("orgs", self.org_cols, row_vals):
                if self.row_matches_quick_search("orgs", row_vals):
                    display_rows.append(row_vals)
        display_rows = self.apply_sort("orgs", self.org_cols, display_rows)
        for i, rv in enumerate(display_rows, 1):
            rv = list(rv)
            rv[0] = i
            rv = list(wrap_tree_row_values(
                self.tree_orgs,
                self.org_cols,
                tuple(rv),
                skip_cols={"Sr", "OrgID", "Select"},
                fallback_wrap={"Website": 18, "Name": 26, "Count": 10},
                width_bias={"Website": 0.95, "Name": 0.92}
            ))
            self.tree_orgs.insert("", "end", values=tuple(rv))
        if auto_reflow:
            self.schedule_org_table_reflow()

    def on_org_tree_click_release(self, event):
        region = self.tree_orgs.identify("region", event.x, event.y)
        if region in ("separator", "heading"):
            self.persist_tree_columnwidths("orgs", self.tree_orgs)
            self.schedule_org_table_reflow()

    def schedule_org_table_reflow(self):
        if self._org_reflowing:
            return
        prev = getattr(self, "_org_reflow_job", None)
        if prev:
            try:
                self.after_cancel(prev)
            except Exception:
                pass
        self._org_reflow_job = self.after(150, self.apply_org_table_reflow)

    def apply_org_table_reflow(self):
        tree = self.tree_orgs
        self._org_reflowing = True
        try:
            self.load_org_tree(auto_reflow=False)
            base_h = int(round(30 * self.ui_scale))
            max_h = int(round(140 * self.ui_scale))
            line_h = int(round(16 * self.ui_scale))
            row_h = self._compute_uniform_rowheight_from_content(tree, base_h, max_h, line_px=line_h)
            ttk.Style().configure("Org.Treeview", rowheight=row_h, font=("Segoe UI", 10))
        finally:
            self._org_reflowing = False

    def toggle_org_selection(self, event):
        sel = self.tree_orgs.selection()
        if not sel: return
        item = sel[0]
        vals = list(self.tree_orgs.item(item, "values"))
        new_val = 1 if vals[5] == "No" else 0
        vals[5] = "Yes" if new_val else "No"
        self.tree_orgs.item(item, values=vals)
        
        conn = sqlite3.connect(DB_FILE)
        conn.execute("UPDATE organizations SET is_selected=? WHERE id=?", (new_val, vals[1]))
        conn.commit()
        conn.close()

    def toggle_selected_org_rows(self, event=None):
        selected = self.tree_orgs.selection()
        if not selected:
            return "break"
        rows = [(item, list(self.tree_orgs.item(item, "values"))) for item in selected]
        target = 0 if all(str(vals[5]) == "Yes" for _, vals in rows) else 1
        conn = sqlite3.connect(DB_FILE)
        for item, vals in rows:
            vals[5] = "Yes" if target else "No"
            self.tree_orgs.item(item, values=vals)
            conn.execute("UPDATE organizations SET is_selected=? WHERE id=?", (target, vals[1]))
        conn.commit()
        conn.close()
        return "break"

    def run_fetch_tenders(self):
        self.notebook.select(self.tab_logs)
        threading.Thread(target=self.thread_fetch_tenders, daemon=True).start()

    def thread_fetch_tenders(self):
        ok_any = False
        for sid in self.get_target_site_ids():
            ok = self.backend.fetch_tenders_logic(sid)
            ok_any = ok_any or bool(ok)
        if ok_any:
            self.after(0, lambda: self.load_tender_tree(switch_tab=True))

    def load_tender_tree(self, switch_tab=False, auto_reflow=True):
        clear_tree_rows(self.tree_tenders)
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        sid = self.get_selected_site_id()
        where_parts = []
        params = []
        if sid is not None:
            where_parts.append("t.website_id=?")
            params.append(sid)
        where_parts.append("COALESCE(t.is_archived,0)=0")
        where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
        c.execute(
            f"""SELECT t.id, w.name, t.tender_id, t.title, t.work_description, t.tender_value, t.emd, t.org_chain, t.closing_date,
                       t.pre_bid_meeting_date, t.location, t.tender_category, t.status, t.is_downloaded, t.folder_path
                FROM tenders t
                JOIN websites w ON w.id=t.website_id
                {where_sql}
                ORDER BY t.created_at DESC""",
            tuple(params)
        )
        rows = c.fetchall()
        conn.close()
        display_rows = []
        for r in rows:
            closing_date_text, closing_time_text = self.split_date_time_text(r[8])
            download_action = self.get_download_action_label(r[14], r[2])
            row_vals = (
                0, r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], closing_date_text, closing_time_text, r[9], r[10], r[11],
                ("Yes" if r[13] else "No"), download_action
            )
            row_vals = self.wrap_tender_row(self.tender_cols, row_vals)
            if self.row_matches_filters("tenders", self.tender_cols, row_vals):
                if self.row_matches_quick_search("tenders", row_vals):
                    display_rows.append(row_vals)
        display_rows = self.apply_sort("tenders", self.tender_cols, display_rows)
        for i, rv in enumerate(display_rows, 1):
            rv = list(rv)
            rv[0] = i
            self.tree_tenders.insert("", "end", values=tuple(rv))
        if auto_reflow:
            self.schedule_tender_table_reflow(False)
        if switch_tab:
            self.notebook.select(self.tab_tenders)

    def load_archived_tender_tree(self, auto_reflow=True):
        for i in self.tree_archived.get_children():
            self.tree_archived.delete(i)
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        sid = self.get_selected_site_id()
        if sid is None:
            c.execute(
                """SELECT t.id, w.name, t.tender_id, t.title, t.work_description, t.tender_value, t.emd, t.org_chain, t.closing_date,
                          t.pre_bid_meeting_date, t.location, t.tender_category, t.status, t.is_downloaded, t.folder_path
                   FROM tenders t
                   JOIN websites w ON w.id=t.website_id
                   WHERE COALESCE(t.is_archived,0)=1
                   ORDER BY t.created_at DESC"""
            )
        else:
            c.execute(
                """SELECT t.id, w.name, t.tender_id, t.title, t.work_description, t.tender_value, t.emd, t.org_chain, t.closing_date,
                          t.pre_bid_meeting_date, t.location, t.tender_category, t.status, t.is_downloaded, t.folder_path
                   FROM tenders t
                   JOIN websites w ON w.id=t.website_id
                   WHERE t.website_id=? AND COALESCE(t.is_archived,0)=1
                   ORDER BY t.created_at DESC""",
                (sid,)
            )
        rows = c.fetchall()
        conn.close()
        display_rows = []
        for r in rows:
            closing_date_text, closing_time_text = self.split_date_time_text(r[8])
            download_action = self.get_download_action_label(r[14], r[2])
            row_vals = (
                0, r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], closing_date_text, closing_time_text, r[9], r[10], r[11], r[12],
                ("Yes" if r[13] else "No"), download_action
            )
            row_vals = self.wrap_tender_row(self.archived_cols, row_vals)
            if self.row_matches_filters("archived", self.archived_cols, row_vals):
                if self.row_matches_quick_search("archived", row_vals):
                    display_rows.append(row_vals)
        display_rows = self.apply_sort("archived", self.archived_cols, display_rows)
        for i, rv in enumerate(display_rows, 1):
            rv = list(rv)
            rv[0] = i
            self.tree_archived.insert("", "end", values=tuple(rv))
        if auto_reflow:
            self.schedule_tender_table_reflow(True)

    def on_tender_click(self, event):
        region = self.tree_tenders.identify("region", event.x, event.y)
        if region in ("separator", "heading"):
            self.persist_tree_columnwidths("tenders", self.tree_tenders)
            self.schedule_tender_table_reflow(False)
            return
        if region != "cell": return
        
        col_name = self.get_clicked_column_name(self.tree_tenders, event.x)
        item = self.tree_tenders.identify_row(event.y)
        if not item: return
        
        vals = list(self.tree_tenders.item(item, "values"))

        if col_name == "Select":
            self.toggle_download_status(item, vals)
        elif col_name == "Download":
            action_idx = self.tender_cols.index("Download")
            if str(vals[action_idx]) in ("Open", "Unzip & Open"):
                self.unzip_and_open_folder(vals[1])

    def on_archived_tender_click(self, event):
        region = self.tree_archived.identify("region", event.x, event.y)
        if region in ("separator", "heading"):
            self.persist_tree_columnwidths("archived", self.tree_archived)
            self.schedule_tender_table_reflow(True)
            return
        if region != "cell":
            return
        col_name = self.get_clicked_column_name(self.tree_archived, event.x)
        item = self.tree_archived.identify_row(event.y)
        if not item:
            return
        vals = list(self.tree_archived.item(item, "values"))
        if col_name == "Select":
            now_ms = int(getattr(event, "time", 0) or 0)
            last_item, last_ms = getattr(self, "_archive_select_last_click", ("", 0))
            if item == last_item and now_ms and last_ms and (now_ms - last_ms) <= 350:
                self._archive_select_last_click = ("", 0)
                return
            self._archive_select_last_click = (item, now_ms)
            self.toggle_download_status(item, vals, archived=True)
        elif col_name == "Download":
            action_idx = self.archived_cols.index("Download")
            if str(vals[action_idx]) in ("Open", "Unzip & Open"):
                self.unzip_and_open_folder(vals[1])

    def schedule_tender_table_reflow(self, archived=False):
        job_attr = "_archived_reflow_job" if archived else "_active_reflow_job"
        prev = getattr(self, job_attr, None)
        if prev:
            try:
                self.after_cancel(prev)
            except Exception:
                pass
        setattr(self, job_attr, self.after(150, lambda a=archived: self.apply_tender_table_reflow(a)))

    def apply_tender_table_reflow(self, archived=False):
        tree = self.tree_archived if archived else self.tree_tenders
        style_name = "ArchivedTender.Treeview" if archived else "ActiveTender.Treeview"
        if archived:
            self.load_archived_tender_tree(auto_reflow=False)
        else:
            self.load_tender_tree(auto_reflow=False)
        base_h = int(round(28 * self.ui_scale))
        max_h = int(round(640 * self.ui_scale))
        line_h = int(round(16 * self.ui_scale))
        row_h = self._compute_uniform_rowheight_from_content(tree, base_h, max_h, line_px=line_h)
        ttk.Style().configure(style_name, rowheight=row_h, font=("Segoe UI", 10))

    def get_clicked_column_name(self, tree, x):
        col = tree.identify_column(x)
        if not col or not str(col).startswith("#"):
            return None
        try:
            idx = int(str(col)[1:]) - 1
        except Exception:
            return None
        dcols = tree["displaycolumns"]
        if dcols == "#all":
            dcols = list(tree["columns"])
        elif isinstance(dcols, str):
            dcols = [dcols]
        else:
            dcols = list(dcols)
        if idx < 0 or idx >= len(dcols):
            return None
        return dcols[idx]

    def toggle_selected_active_rows(self, event=None):
        return self.toggle_selected_tender_rows(archived=False)

    def toggle_selected_archived_rows(self, event=None):
        return self.toggle_selected_tender_rows(archived=True)

    def toggle_selected_tender_rows(self, archived=False):
        tree = self.tree_archived if archived else self.tree_tenders
        select_idx = self.archived_cols.index("Select") if archived else self.tender_cols.index("Select")
        selected = tree.selection()
        if not selected:
            return "break"
        rows = [(item, list(tree.item(item, "values"))) for item in selected]
        target = 0 if all(str(vals[select_idx]) == "Yes" for _, vals in rows) else 1
        conn = sqlite3.connect(DB_FILE)
        for item, vals in rows:
            vals[select_idx] = "Yes" if target else "No"
            tree.item(item, values=vals)
            conn.execute("UPDATE tenders SET is_downloaded=? WHERE id=?", (target, vals[1]))
        conn.commit()
        conn.close()
        return "break"

    def unzip_and_open_folder(self, db_id):
        conn = sqlite3.connect(DB_FILE)
        row = conn.execute("SELECT folder_path, tender_id FROM tenders WHERE id=?", (db_id,)).fetchone()
        conn.close()

        if not row:
            log_to_gui(f"Error: Could not find tender with DB ID {db_id}.")
            return
        
        folder_path, tender_id = row

        if not (folder_path and os.path.exists(folder_path)):
            log_to_gui(f"Error: Folder path '{folder_path}' does not exist.")
            return

        archive_paths = self.get_archive_paths(folder_path, tender_id)
        if archive_paths and not self.is_already_extracted(folder_path, archive_paths):
            for ap in archive_paths:
                log_to_gui(f"Extracting {os.path.basename(ap)}...")
                ok = self.extract_archive(ap, folder_path)
                if not ok:
                    messagebox.showwarning("Extract Warning", f"Could not extract:\n{ap}\n\nOpening folder anyway.")

        log_to_gui(f"Opening folder: {folder_path}")
        try:
            if platform.system() == "Windows":
                os.startfile(folder_path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", folder_path])
            else:
                subprocess.Popen(["xdg-open", folder_path])
            # Refresh action labels (Unzip & Open -> Open) after extraction.
            self.after(0, self.on_site_changed)
        except Exception as e:
            log_to_gui(f"Error opening folder: {e}")
            messagebox.showerror("Error", f"Could not open the folder:\n{e}")

    def toggle_download_status(self, item, vals, archived=False):
        select_idx = self.archived_cols.index("Select") if archived else self.tender_cols.index("Select")
        new_val = 1 if vals[select_idx] == "No" else 0
        vals[select_idx] = "Yes" if new_val else "No"
        if archived:
            self.tree_archived.item(item, values=vals)
        else:
            self.tree_tenders.item(item, values=vals)
        conn = sqlite3.connect(DB_FILE)
        conn.execute("UPDATE tenders SET is_downloaded=? WHERE id=?", (new_val, vals[1]))
        conn.commit()
        conn.close()

    def get_archive_paths(self, folder_path, tender_id):
        safe_id = re.sub(r'[\\/*?:"<>|]', "", str(tender_id or ""))
        preferred = []
        if safe_id:
            preferred.append(os.path.join(folder_path, f"{safe_id}.rar"))
            preferred.append(os.path.join(folder_path, f"{safe_id}.zip"))
        existing = [p for p in preferred if os.path.exists(p)]
        if existing:
            return existing
        generic = []
        try:
            for name in os.listdir(folder_path):
                lower = name.lower()
                if lower.endswith(".rar") or lower.endswith(".zip"):
                    generic.append(os.path.join(folder_path, name))
        except Exception:
            return []
        return sorted(generic)

    def is_already_extracted(self, folder_path, archive_paths):
        archive_names = {os.path.basename(p).lower() for p in archive_paths}
        try:
            for name in os.listdir(folder_path):
                low = name.lower()
                if low in archive_names:
                    continue
                if low.endswith(".crdownload") or low.endswith(".part"):
                    continue
                return True
        except Exception:
            return False
        return False

    def get_download_action_label(self, folder_path, tender_id):
        if not folder_path or not os.path.exists(folder_path):
            return ""
        archive_paths = self.get_archive_paths(folder_path, tender_id)
        if not archive_paths:
            return "Open"
        return "Open" if self.is_already_extracted(folder_path, archive_paths) else "Unzip & Open"

    def extract_archive(self, archive_path, dest_dir):
        low = archive_path.lower()
        try:
            if low.endswith(".zip"):
                with zipfile.ZipFile(archive_path, "r") as zf:
                    zf.extractall(dest_dir)
                return True
            if low.endswith(".rar"):
                try:
                    import patoolib
                    patoolib.extract_archive(archive_path, outdir=dest_dir, verbosity=-1)
                    return True
                except Exception:
                    pass
                try:
                    proc = subprocess.run(["7z", "x", "-y", f"-o{dest_dir}", archive_path], capture_output=True, text=True)
                    return proc.returncode == 0
                except Exception:
                    return False
        except Exception as e:
            log_to_gui(f"Extract error for {archive_path}: {e}")
            return False
        return False

    def select_all_tenders(self):
        sid = self.get_selected_site_id()
        conn = sqlite3.connect(DB_FILE)
        current_tab = self.notebook.select()
        archived_mode = (current_tab == str(self.tab_archived))

        where_parts = []
        params = []
        if sid is not None:
            where_parts.append("website_id=?")
            params.append(sid)
        if archived_mode:
            where_parts.append("COALESCE(is_archived,0)=1")
        else:
            where_parts.append("COALESCE(is_archived,0)=0")
        where_sql = " AND ".join(where_parts)

        total = conn.execute(f"SELECT COUNT(*) FROM tenders WHERE {where_sql}", tuple(params)).fetchone()[0]
        selected = conn.execute(
            f"SELECT COUNT(*) FROM tenders WHERE {where_sql} AND COALESCE(is_downloaded,0)=1",
            tuple(params)
        ).fetchone()[0]
        target = 0 if total > 0 and selected == total else 1

        conn.execute(
            f"UPDATE tenders SET is_downloaded=? WHERE {where_sql}",
            (target, *params)
        )
        conn.commit()
        conn.close()
        self.on_site_changed()

    def run_download(self):
        self.notebook.select(self.tab_logs)
        threading.Thread(target=self.thread_download, daemon=True).start()

    def run_single_download(self):
        if self.notebook.select() != str(self.tab_tenders):
            return messagebox.showwarning("Download", "Switch to Active Tenders and select one tender.")
        sel = self.tree_tenders.selection()
        item = sel[0] if sel else self.tree_tenders.focus()
        if not item:
            return messagebox.showwarning("Download", "Select/highlight one active tender first.")
        vals = list(self.tree_tenders.item(item, "values") or [])
        if len(vals) < 2:
            return messagebox.showwarning("Download", "Unable to identify selected tender.")
        try:
            tender_db_id = int(str(vals[1]).strip())
        except Exception:
            return messagebox.showwarning("Download", "Unable to identify selected tender.")
        ans = messagebox.askyesnocancel(
            "Download Mode",
            "Choose mode for selected tender:\n\nYes = Full\nNo = Update\nCancel = Abort"
        )
        if ans is None:
            return
        mode = "full" if ans else "update"
        self.notebook.select(self.tab_logs)
        threading.Thread(target=self.thread_single_download, args=(tender_db_id, mode), daemon=True).start()

    def thread_single_download(self, tender_db_id, mode):
        self.backend.download_single_tender_logic(tender_db_id, mode)
        self.after(0, self.on_site_changed)

    def run_status_check(self):
        self.notebook.select(self.tab_logs)
        threading.Thread(target=self.thread_status_check, daemon=True).start()

    def thread_download(self):
        for sid in self.get_target_site_ids():
            self.backend.download_tenders_logic(sid)
        self.after(0, self.on_site_changed)

    def thread_status_check(self):
        for sid in self.get_target_site_ids():
            self.backend.check_tender_status_logic(sid)
        self.after(0, self.on_site_changed)

    def run_download_results(self):
        self.notebook.select(self.tab_logs)
        threading.Thread(target=self.thread_download_results, daemon=True).start()

    def thread_download_results(self):
        for sid in self.get_target_site_ids():
            self.backend.download_tender_results_logic(sid)
        self.after(0, self.on_site_changed)

    def add_selected_tenders_to_new_project(self):
        sid = self.get_selected_site_id()
        current_tab = self.notebook.select()
        archived_mode = (current_tab == str(self.tab_archived))
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        where = ["COALESCE(is_downloaded,0)=1"]
        params = []
        if sid is not None:
            where.append("website_id=?")
            params.append(sid)
        if archived_mode:
            where.append("COALESCE(is_archived,0)=1")
        else:
            where.append("COALESCE(is_archived,0)=0")
        where_sql = " AND ".join(where)
        c.execute(
            f"""SELECT tender_id, title, org_chain, closing_date, website_id, tender_value, pre_bid_meeting_date, folder_path
                FROM tenders
                WHERE {where_sql}
                ORDER BY created_at DESC""",
            tuple(params)
        )
        rows = c.fetchall()
        conn.close()
        if not rows:
            messagebox.showinfo("Add to New Project", "No selected tenders found. Select tenders first.")
            return

        first = rows[0]
        first_tid = str(first[0] or "").strip()
        first_title = str(first[1] or "").strip()
        first_org = str(first[2] or "").strip()
        first_deadline = str(first[3] or "").strip()
        first_value = str(first[5] or "").strip()
        first_prebid = str(first[6] or "").strip()
        first_tender_folder = str(first[7] or "").strip()
        if len(rows) > 1:
            bulk = messagebox.askyesno(
                "Add Projects",
                f"{len(rows)} tenders are selected.\n\n"
                "Click Yes to create one project per selected tender.\n"
                "Click No to open a single prefilled project form."
            )
            if bulk:
                created, skipped = self.create_projects_from_rows(rows)
                self.controller.show_frame("Projects")
                return

        default_title = first_tid or "Tender Project"
        prefill = {
            "title": default_title,
            "client_name": first_org,
            "deadline": first_deadline,
            "description": first_title,
            "tender_id": first_tid,
            "project_value": first_value,
            "prebid": first_prebid,
            "tender_folder_path": first_tender_folder,
        }
        self.controller.open_create_project_window(prefill=prefill)

    def create_projects_from_rows(self, rows):
        def safe_name(text):
            return "".join(c for c in str(text or "") if c.isalnum() or c in (" ", "_", "-")).strip() or "Tender Project"

        def ensure_unique_folder(base):
            path = os.path.join(ROOT_FOLDER, base)
            if not os.path.exists(path):
                return path
            i = 2
            while True:
                p = os.path.join(ROOT_FOLDER, f"{base}_{i}")
                if not os.path.exists(p):
                    return p
                i += 1

        raw = ScraperBackend.get_setting("project_client_options", "[]")
        try:
            client_opts = [str(x).strip() for x in json.loads(raw) if str(x).strip()]
        except Exception:
            client_opts = []

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        created = 0
        skipped = 0
        try:
            for tid, title, org, closing, _wid, project_value, prebid, tender_folder_path in rows:
                project_title = str(tid or "").strip() or "Tender Project"
                source_tender_id = str(tid or "").strip()
                if source_tender_id:
                    exists = c.execute(
                        "SELECT 1 FROM projects WHERE source_tender_id=? LIMIT 1",
                        (source_tender_id,)
                    ).fetchone()
                    if exists:
                        skipped += 1
                        continue
                folder_path = ensure_unique_folder(safe_name(project_title))
                std_folders = ensure_project_standard_folders(folder_path)
                c.execute(
                    "INSERT INTO projects (title, client_name, deadline, description, folder_path, source_tender_id, project_value, prebid) VALUES (?,?,?,?,?,?,?,?)",
                    (
                        project_title,
                        str(org or "").strip(),
                        str(closing or "").strip(),
                        str(title or "").strip(),
                        folder_path,
                        source_tender_id or None,
                        str(project_value or "").strip(),
                        str(prebid or "").strip()
                    )
                )
                tender_src = str(tender_folder_path or "").strip()
                if tender_src and os.path.isdir(tender_src):
                    copied = copy_tree_contents(tender_src, std_folders["tender_docs"])
                    if copied:
                        log_to_gui(f"Copied {copied} tender item(s) to project Tender Docs: {project_title}")
                created += 1
                if org and str(org).strip() and str(org).strip() not in client_opts:
                    client_opts.append(str(org).strip())
            conn.commit()
        finally:
            conn.close()

        client_opts = sorted(set(client_opts), key=lambda x: x.lower())
        ScraperBackend.set_setting("project_client_options", json.dumps(client_opts))
        return created, skipped

# --- EXISTING DASHBOARD & DETAILS CLASSES ---
class Dashboard(tk.Frame):
    PROJECT_SELECT_SQL = "SELECT id, title, description, client_name, project_value, prebid, deadline, status FROM projects"

    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self._dashboard_reflow_job = None
        self._dashboard_reflowing = False
        header = tk.Frame(self, bg="#00695C", height=60)
        header.pack(fill="x")
        tk.Label(header, text="Projects", bg="#00695C", fg="white", font=("Segoe UI", 18)).pack(side="left", padx=20)
        bar = tk.Frame(self, pady=10, padx=20, bg="#f0f0f0")
        bar.pack(fill="x")
        tk.Button(bar, text="Delete Project", bg="#F44336", fg="white", font=("Segoe UI", 10, "bold"), command=self.delete_selected).pack(side="right", padx=5)
        tk.Button(bar, text="Open Project", bg="#2196F3", fg="white", font=("Segoe UI", 10, "bold"), command=self.open_selected).pack(side="right", padx=5)
        tk.Label(bar, text="Search Projects:", bg="#f0f0f0").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.filter_projects)
        tk.Entry(bar, textvariable=self.search_var, width=30).pack(side="left", padx=10)
        tk.Button(
            bar,
            text="+ New Project",
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=self.controller.open_create_project_window
        ).pack(side="left", padx=(4, 0))
        tree_frame = tk.Frame(self, padx=20, pady=10)
        tree_frame.pack(fill="both", expand=True)
        cols = ("Sr", "ID", "Tender Id", "Description", "Client", "Value", "Prebid", "Deadline", "Status")
        self.project_cols = cols
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", style="Dashboard.Treeview", selectmode="extended")
        self.tree.heading("Sr", text="Sr. No."); self.tree.column("Sr", width=50, minwidth=40, stretch=False, anchor="center")
        self.tree.heading("ID", text="ID"); self.tree.column("ID", width=0, stretch=False)
        self.tree.heading("Tender Id", text="Tender Id"); self.tree.column("Tender Id", width=120, minwidth=80, stretch=True, anchor="center")
        self.tree.heading("Description", text="Description"); self.tree.column("Description", width=330, minwidth=180, stretch=True, anchor="w")
        self.tree.heading("Client", text="Client"); self.tree.column("Client", width=200, minwidth=100, stretch=True, anchor="center")
        self.tree.heading("Value", text="Value"); self.tree.column("Value", width=130, minwidth=90, stretch=True, anchor="center")
        self.tree.heading("Prebid", text="Prebid"); self.tree.column("Prebid", width=140, minwidth=100, stretch=True, anchor="center")
        self.tree.heading("Deadline", text="Deadline"); self.tree.column("Deadline", width=120, minwidth=80, stretch=True, anchor="center")
        self.tree.heading("Status", text="Status"); self.tree.column("Status", width=100, minwidth=80, stretch=True, anchor="center")
        scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self.open_selected)
        self.tree.bind("<Configure>", lambda e: self.schedule_dashboard_table_reflow(), add="+")
        self.tree.bind("<ButtonRelease-1>", self.on_dashboard_tree_click_release, add="+")
        ColumnManager(self.tree, self)

    def _fetch_project_rows(self):
        conn = sqlite3.connect(DB_FILE)
        try:
            return conn.cursor().execute(self.PROJECT_SELECT_SQL).fetchall()
        finally:
            conn.close()

    def load_projects(self):
        clear_tree_rows(self.tree)
        rows = self._fetch_project_rows()
        for i, r in enumerate(rows, 1):
            row_vals = self.wrap_dashboard_row((i,) + r)
            self.tree.insert("", tk.END, values=row_vals)
        if not self._dashboard_reflowing:
            self.after_idle(self.schedule_dashboard_table_reflow)

    def filter_projects(self, *args):
        query = self.search_var.get().lower()
        clear_tree_rows(self.tree)
        rows = self._fetch_project_rows()
        count = 1
        for r in rows:
            title = str(r[1] or "").lower()
            desc = str(r[2] or "").lower()
            client = str(r[3] or "").lower()
            val_txt = str(r[4] or "").lower()
            prebid_txt = str(r[5] or "").lower()
            if query in title or query in desc or query in client or query in val_txt or query in prebid_txt:
                row_vals = self.wrap_dashboard_row((count,) + r)
                self.tree.insert("", tk.END, values=row_vals)
                count += 1
        if not self._dashboard_reflowing:
            self.after_idle(self.schedule_dashboard_table_reflow)

    def wrap_dashboard_row(self, row_vals):
        fallback_wrap = {
            "Tender Id": 24,
            "Description": 34,
            "Client": 24,
            "Value": 14,
            "Prebid": 16,
            "Deadline": 18,
            "Status": 12,
        }
        width_bias = {
            "Description": 1.14,
            "Client": 1.18,
            "Value": 1.0,
            "Prebid": 1.0,
        }
        wrapped = []
        avg_char_px = 7.0
        try:
            font_cfg = ttk.Style().lookup("Dashboard.Treeview", "font") or ("Segoe UI", 10)
            fnt = resolve_tk_font(font_cfg)
            sample = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 "
            avg_char_px = max(6.0, fnt.measure(sample) / len(sample))
        except Exception:
            pass
        for idx, col in enumerate(self.project_cols):
            val = row_vals[idx] if idx < len(row_vals) else ""
            if col in {"Sr", "ID", "Status"}:
                wrapped.append(val)
                continue
            text = " ".join(str(val or "").split())
            if col == "Tender Id":
                # Allow wrap points for long IDs with underscores/hyphens.
                text = text.replace("_", "_ ").replace("-", "- ")
            try:
                px = int(self.tree.column(col, "width"))
                width = max(8, int(((px - 14) / avg_char_px) * width_bias.get(col, 1.0)))
            except Exception:
                width = fallback_wrap.get(col, 24)
            wrapped.append(textwrap.fill(text, width=width, break_long_words=False, break_on_hyphens=False))
        return tuple(wrapped)

    def on_dashboard_tree_click_release(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region in ("separator", "heading"):
            self.schedule_dashboard_table_reflow()

    def schedule_dashboard_table_reflow(self):
        prev = getattr(self, "_dashboard_reflow_job", None)
        if prev:
            try:
                self.after_cancel(prev)
            except Exception:
                pass
        self._dashboard_reflow_job = self.after(150, self.apply_dashboard_table_reflow)

    def apply_dashboard_table_reflow(self):
        tree = self.tree
        tree.update_idletasks()
        ui_scale = getattr(self.controller, "ui_scale", 1.0)
        row_h = int(round(36 * ui_scale))
        try:
            desc_w = int(tree.column("Description", "width"))
            client_w = int(tree.column("Client", "width"))
            tid_w = int(tree.column("Tender Id", "width"))
            if desc_w < 340:
                row_h = max(row_h, int(round(46 * ui_scale)))
            if desc_w < 260:
                row_h = max(row_h, int(round(58 * ui_scale)))
            if client_w < 200:
                row_h = max(row_h, int(round(44 * ui_scale)))
            if tid_w < 120:
                row_h = max(row_h, int(round(42 * ui_scale)))
        except Exception:
            pass
        # Keep row height dynamic based on actual wrapped line count.
        try:
            font_cfg = ttk.Style().lookup("Dashboard.Treeview", "font") or ("Segoe UI", 10)
            fnt = resolve_tk_font(font_cfg)
            line_px = max(14, int(fnt.metrics("linespace")))
            max_lines = 1
            for iid in tree.get_children():
                vals = tree.item(iid, "values")
                for idx, col in enumerate(self.project_cols):
                    if col in {"Sr", "ID", "Status"}:
                        continue
                    txt = str(vals[idx]) if idx < len(vals) else ""
                    max_lines = max(max_lines, txt.count("\n") + 1)
            # Add a marginal dynamic buffer to prevent glyph clipping.
            content_h = int(max_lines * line_px + round(12 * ui_scale))
            row_h = max(row_h, content_h)
        except Exception:
            pass
        ttk.Style().configure("Dashboard.Treeview", rowheight=row_h, font=("Segoe UI", 10))
        self._dashboard_reflowing = True
        try:
            if (self.search_var.get() or "").strip():
                self.filter_projects()
            else:
                self.load_projects()
        finally:
            self._dashboard_reflowing = False

    def open_selected(self, event=None):
        sel = self.tree.selection()
        if not sel: return
        pid = self.tree.item(sel[0])['values'][1]
        self.controller.show_frame("ProjectDetails", pid)

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel: return
        if len(sel) == 1:
            if not messagebox.askyesno("Confirm", "Delete this project?"): return
        else:
            if not messagebox.askyesno("Confirm", f"Delete {len(sel)} selected projects?"): return

        pids = []
        for item_id in sel:
            vals = self.tree.item(item_id).get('values', [])
            if len(vals) >= 2:
                pids.append(vals[1])
        if not pids:
            return

        conn = sqlite3.connect(DB_FILE)
        for pid in pids:
            conn.execute("DELETE FROM checklist_items WHERE project_id=?", (pid,))
            conn.execute("DELETE FROM projects WHERE id=?", (pid,))
        conn.commit()
        conn.close()
        self.load_projects()

class ProjectDetails(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.project_id = None
        self.folder_path = ""
        self.source_tender_id = ""
        self.paned_root = tk.PanedWindow(self, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=4)
        self.paned_root.pack(fill="both", expand=True)
        self.left_pane = tk.Frame(self.paned_root)
        self.paned_root.add(self.left_pane, minsize=600, stretch="always")
        self.header = tk.Frame(self.left_pane, bg="#37474F", height=50)
        self.header.pack(fill="x")
        tk.Button(self.header, text="< Back", bg="#546E7A", fg="white", relief="flat", command=lambda: controller.show_frame("Projects")).pack(side="left", padx=10, pady=10)
        self.lbl_title = tk.Label(self.header, text="Project Details", bg="#37474F", fg="white", font=("Segoe UI", 14, "bold"))
        self.lbl_title.pack(side="left", padx=10)
        tk.Button(self.header, text="Open in Explorer", bg="#00897B", fg="white", command=self.open_explorer).pack(side="right", padx=10)
        tk.Button(self.header, text="Update", bg="#FFB300", fg="black", command=self.update_project_table_from_folder).pack(side="right", padx=(0, 8), pady=8)
        self.action_frame = tk.LabelFrame(self.left_pane, text="Manage Document Checklist", padx=10, pady=10, bg="#f5f5f5")
        self.action_frame.pack(fill="x", padx=10, pady=10)
        input_row = tk.Frame(self.action_frame, bg="#f5f5f5")
        input_row.pack(fill="x", pady=5)
        tk.Label(input_row, text="Document Name:", bg="#f5f5f5").pack(side="left")
        self.e_doc_name = tk.Entry(input_row, width=30)
        self.e_doc_name.pack(side="left", padx=5)
        tk.Label(input_row, text="Description:", bg="#f5f5f5").pack(side="left", padx=(10,0))
        self.e_desc = tk.Entry(input_row, width=30)
        self.e_desc.pack(side="left", padx=5)
        tk.Label(input_row, text="Location (Folder):", bg="#f5f5f5").pack(side="left", padx=(10,0))
        self.folder_var = tk.StringVar()
        self.cb_folder = ttk.Combobox(input_row, textvariable=self.folder_var, width=20)
        self.cb_folder.pack(side="left", padx=5)
        self.cb_folder.bind("<Key>", on_combo_key_cycling)
        btn_row = tk.Frame(self.action_frame, bg="#f5f5f5")
        btn_row.pack(fill="x", pady=(10,0))
        tk.Button(btn_row, text="Add Item", bg="#4CAF50", fg="white", command=self.add_item).pack(side="left")
        tk.Button(btn_row, text="Update Selected", bg="#FF9800", fg="white", command=self.update_item).pack(side="left", padx=10)
        tk.Button(btn_row, text="+ New Folder", bg="#607D8B", fg="white", command=self.create_new_folder_popup).pack(side="left", padx=10)
        tk.Button(btn_row, text="Manage Folders", bg="#5D4037", fg="white", command=self.open_manage_folders_popup).pack(side="left", padx=10)
        tk.Button(btn_row, text="Import Docs", bg="#3949AB", fg="white", command=self.import_tender_docs).pack(side="left", padx=10)
        tk.Button(btn_row, text="Download Docs", bg="#00897B", fg="white", command=self.download_tender_docs).pack(side="left", padx=10)
        tk.Button(btn_row, text="Delete File", bg="#F44336", fg="white", command=self.delete_item).pack(side="right")
        tk.Button(btn_row, text="Open File", bg="#607D8B", fg="white", command=self.open_file).pack(side="right", padx=10)
        tk.Button(btn_row, text="Attach File", bg="#2196F3", fg="white", command=self.upload_file).pack(side="right")
        tree_container = tk.Frame(self.left_pane, padx=10, pady=10)
        tree_container.pack(fill="both", expand=True)
        cols = ("Document Name", "Description", "Status", "Attachment Name", "ID") 
        self.checklist_cols = cols
        self.tree = ttk.Treeview(tree_container, columns=cols, style="Wrapped.Treeview")
        self.tree.heading("#0", text="Sr. No.")
        self.tree.column("#0", width=80, minwidth=50, stretch=False, anchor="center")
        self.tree.heading("Document Name", text="Document Name"); self.tree.column("Document Name", width=350, minwidth=150, stretch=True, anchor="center")
        self.tree.heading("Description", text="Description"); self.tree.column("Description", width=250, minwidth=150, stretch=True, anchor="center")
        self.tree.heading("Status", text="Status"); self.tree.column("Status", width=100, minwidth=80, stretch=True, anchor="center")
        self.tree.heading("Attachment Name", text="Attachment Name"); self.tree.column("Attachment Name", width=200, minwidth=100, stretch=True, anchor="center")
        self.tree.heading("ID", text="ID"); self.tree.column("ID", width=0, stretch=False)
        self.tree.tag_configure('folder', background='#eceff1', font=('Segoe UI', 10, 'bold'))
        self.tree.tag_configure('Completed', foreground='green')
        self.tree.tag_configure('Pending', foreground='#d32f2f')
        scroll = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        ColumnManager(self.tree, self)
        self.preview_pane = tk.Frame(self.paned_root, bg="#546E7A", bd=1, relief=tk.SUNKEN)
        self.preview_visible = False
        self.pv_label = tk.Label(self.preview_pane, text="Preview Area\n(Select a file)", bg="white")
        self.pv_label.pack(fill="both", expand=True)
        pv_ctrl = tk.Frame(self.preview_pane, bg="#eceff1", height=40)
        pv_ctrl.pack(fill="x", side="bottom")
        tk.Button(pv_ctrl, text="<", command=self.prev_page).pack(side="left", padx=5, pady=5)
        self.pv_page = tk.Entry(pv_ctrl, width=5, justify="center"); self.pv_page.insert(0, "1"); self.pv_page.pack(side="left", padx=5)
        tk.Label(pv_ctrl, text="/", bg="#eceff1").pack(side="left")
        self.pv_total = tk.Label(pv_ctrl, text="1", bg="#eceff1"); self.pv_total.pack(side="left", padx=5)
        tk.Button(pv_ctrl, text=">", command=self.next_page).pack(side="left", padx=5)
        self.btn_toggle = tk.Canvas(self.left_pane, width=15, height=60, bg=self.left_pane.cget("bg"), highlightthickness=0)
        self.btn_toggle.place(relx=1.0, rely=0.5, anchor="e")
        self.btn_toggle.bind("<Button-1>", lambda e: self.toggle_preview())
        self.draw_toggle_button("open")

    def load_project(self, pid):
        self.project_id = pid
        clear_tree_rows(self.tree)
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        p = c.execute("SELECT title, folder_path, COALESCE(source_tender_id,'') FROM projects WHERE id=?", (pid,)).fetchone()
        self.lbl_title.config(text=f"{p[0]}")
        repaired_folder = resolve_project_folder_path(p[1], p[0])
        self.folder_path = repaired_folder
        try:
            old_folder = str(p[1] or "").strip()
            if os.path.normcase(old_folder) != os.path.normcase(repaired_folder):
                c.execute("UPDATE projects SET folder_path=? WHERE id=?", (repaired_folder, pid))
                conn.commit()
                log_to_gui(f"Project path repaired to: {repaired_folder}")
        except Exception:
            pass
        self.source_tender_id = str(p[2] or "").strip()
        std_folders = ensure_project_standard_folders(self.folder_path)
        effective_folder = std_folders.get("project_root", self.folder_path)
        if os.path.normcase(effective_folder) != os.path.normcase(self.folder_path):
            self.folder_path = effective_folder
            try:
                c.execute("UPDATE projects SET folder_path=? WHERE id=?", (self.folder_path, pid))
                conn.commit()
                log_to_gui(f"Project path fallback applied: {self.folder_path}")
            except Exception:
                pass
        self.refresh_folder_list()
        items = c.execute("SELECT id, sr_no, req_file_name, description, status, linked_file_path, subfolder FROM checklist_items WHERE project_id=? ORDER BY subfolder, sr_no", (pid,)).fetchall()
        item_map = {}
        linked_file_abs = set()
        rendered_linked_abs = set()
        for item in items:
            sf = self._normalize_subfolder(item[6] if item[6] else "Main")
            desc_txt = str(item[3] or "").strip().lower()
            lp = str(item[5] or "").strip()
            lp_abs = ""
            if lp:
                try:
                    lp_abs = os.path.normcase(os.path.abspath(lp))
                except Exception:
                    lp_abs = ""
            # Hide stale auto-import rows whose file no longer exists.
            if desc_txt == "auto-imported" and lp_abs and not os.path.isfile(lp_abs):
                continue
            # Prevent duplicate render of multiple DB rows pointing to same file.
            if lp_abs and lp_abs in rendered_linked_abs:
                continue
            if sf not in item_map: item_map[sf] = []
            item_map[sf].append(item)
            if lp_abs:
                linked_file_abs.add(lp_abs)
                rendered_linked_abs.add(lp_abs)

        fs_files_map = {}
        all_folders = set(item_map.keys())
        if os.path.exists(self.folder_path):
            for root, dirs, files in os.walk(self.folder_path):
                rel_root = os.path.relpath(root, self.folder_path)
                rel_root = self._normalize_subfolder(rel_root)
                all_folders.add(rel_root)
                for d in dirs:
                    full_path = os.path.join(root, d)
                    rel_path = self._normalize_subfolder(os.path.relpath(full_path, self.folder_path))
                    all_folders.add(rel_path)
                for fn in files:
                    f_abs = os.path.normcase(os.path.abspath(os.path.join(root, fn)))
                    if f_abs in linked_file_abs:
                        continue
                    fs_files_map.setdefault(rel_root, []).append(fn)
        all_folders.add("Main")
        sorted_folders = sorted({self._normalize_subfolder(x) for x in all_folders}, key=lambda x: x.lower())
        folder_nodes = {}
        folder_nodes["Main"] = ""
        folder_counters = {}
        item_seq = 1
        for folder_name in sorted_folders:
            if folder_name == "Main": current_node_id = ""
            else:
                parent_path = os.path.dirname(folder_name)
                if not parent_path or parent_path == ".": parent_path = "Main"
                parent_node_id = folder_nodes.get(parent_path, "")
                if parent_path not in folder_counters: folder_counters[parent_path] = 0
                folder_counters[parent_path] += 1
                sr_text = str(folder_counters[parent_path]) if parent_path == "Main" else f"({folder_counters[parent_path]})"
                display_name = os.path.basename(folder_name)
                folder_vals = wrap_tree_row_values(
                    self.tree,
                    self.checklist_cols,
                    (display_name, "", "", "", ""),
                    skip_cols={"ID"},
                    fallback_wrap={"Document Name": 24, "Description": 24, "Attachment Name": 20}
                )
                folder_node = self.tree.insert(parent_node_id, "end", text=sr_text, open=True, values=folder_vals, tags=('folder',))
                folder_nodes[folder_name] = folder_node
                current_node_id = folder_node
            if folder_name in item_map:
                for item in item_map[folder_name]:
                    fname = os.path.basename(item[5]) if item[5] else "-"
                    tag = "Completed" if item[4] == "Completed" else "Pending"
                    row_vals = wrap_tree_row_values(
                        self.tree,
                        self.checklist_cols,
                        (item[2], item[3], item[4], fname, item[0]),
                        skip_cols={"ID", "Status"},
                        fallback_wrap={"Document Name": 24, "Description": 24, "Attachment Name": 20}
                    )
                    self.tree.insert(current_node_id, "end", text=f"{item_seq}.", values=row_vals, tags=(tag,))
                    item_seq += 1
            if folder_name in fs_files_map:
                for fn in sorted(fs_files_map[folder_name], key=lambda x: x.lower()):
                    row_vals = wrap_tree_row_values(
                        self.tree,
                        self.checklist_cols,
                        (fn, "Imported File", "Completed", fn, ""),
                        skip_cols={"ID", "Status"},
                        fallback_wrap={"Document Name": 24, "Description": 24, "Attachment Name": 20}
                    )
                    self.tree.insert(current_node_id, "end", text=f"{item_seq}.", values=row_vals, tags=("Completed",))
                    item_seq += 1
        conn.close()

    def _project_tender_docs_path(self):
        return ensure_project_standard_folders(self.folder_path)["tender_docs"]

    def import_tender_docs(self):
        if not self.project_id:
            return messagebox.showwarning("Import Docs", "Open a project first.")
        tender_id = str(self.source_tender_id or "").strip()
        if not tender_id:
            return messagebox.showwarning("Import Docs", "Open project is not linked to a Tender ID.")

        conn = sqlite3.connect(DB_FILE)
        row = conn.execute(
            """SELECT COALESCE(tender_id,''), COALESCE(folder_path,''), COALESCE(is_downloaded,0), COALESCE(is_archived,0)
               FROM tenders
               WHERE TRIM(COALESCE(tender_id,''))=TRIM(?)
               ORDER BY created_at DESC
               LIMIT 1""",
            (tender_id,)
        ).fetchone()
        if not row:
            conn.close()
            return messagebox.showinfo("Import Docs", "Tender ID of this project was not found in Active Tenders.")
        _, db_folder_path, is_downloaded, is_archived = row
        if int(is_archived or 0) != 0:
            conn.close()
            return messagebox.showinfo("Import Docs", "Tender ID of this project is not in Active Tenders.")

        log_count = conn.execute(
            "SELECT COUNT(*) FROM downloaded_files WHERE TRIM(COALESCE(tender_id,''))=TRIM(?)",
            (tender_id,)
        ).fetchone()[0]
        conn.close()
        if int(is_downloaded or 0) != 1 and int(log_count or 0) == 0:
            return messagebox.showinfo("Import Docs", "No documents were downloaded for this tender.")

        safe_id = re.sub(r'[\\/*?:"<>|]', "", tender_id)
        src = os.path.join(BASE_DOWNLOAD_DIRECTORY, safe_id)
        if not os.path.isdir(src):
            alt = str(db_folder_path or "").strip()
            if alt and os.path.isdir(alt):
                src = alt
            else:
                return messagebox.showinfo("Import Docs", "No documents were downloaded for this tender.")
        try:
            if not os.listdir(src):
                return messagebox.showinfo("Import Docs", "No documents were downloaded for this tender.")
        except Exception:
            return messagebox.showinfo("Import Docs", "No documents were downloaded for this tender.")
        dest = self._project_tender_docs_path()
        copied = copy_tree_contents(
            src,
            dest,
            exclude_names={"Ready Docs", "Tender Docs", "Working Docs"}
        )
        if copied:
            log_to_gui(f"Imported {copied} item(s) into project Tender Docs for tender {tender_id}.")
            messagebox.showinfo("Import Docs", f"Imported {copied} item(s) into Tender Docs.")
        else:
            messagebox.showinfo("Import Docs", "No documents were downloaded for this tender.")
        self.load_project(self.project_id)

    def download_tender_docs(self):
        if not self.project_id:
            return messagebox.showwarning("Download Docs", "Open a project first.")
        tender_id = str(self.source_tender_id or "").strip()
        if not tender_id:
            return messagebox.showwarning("Download Docs", "This project is not linked to a Tender ID.")
        dest = self._project_tender_docs_path()
        self.controller.show_frame("ProjectDetails", self.project_id)
        threading.Thread(
            target=self._download_tender_docs_worker,
            args=(tender_id, dest),
            daemon=True
        ).start()

    def _download_tender_docs_worker(self, tender_id, dest):
        try:
            ok = ScraperBackend.download_docs_for_tender_to_folder(tender_id, dest)
            if ok:
                log_to_gui(f"Download Docs complete for tender {tender_id}.")
            else:
                log_to_gui(f"Download Docs failed for tender {tender_id}.")
        except Exception as e:
            log_to_gui(f"Download Docs error for tender {tender_id}: {e}")
        finally:
            if self.project_id:
                self.after(0, lambda: self.load_project(self.project_id))

    def refresh_folder_list(self):
        folder_list = ["Main"]
        try:
            for root, dirs, files in os.walk(self.folder_path):
                for d in dirs:
                    full_path = os.path.join(root, d)
                    rel_path = os.path.relpath(full_path, self.folder_path)
                    folder_list.append(rel_path)
            self.cb_folder['values'] = sorted(list(set(folder_list)))
            self.cb_folder.current(0)
        except: self.cb_folder['values'] = ["Main"]

    def update_project_table_from_folder(self):
        if not self.project_id:
            return messagebox.showwarning("Update", "Open a project first.")
        ensure_project_standard_folders(self.folder_path)
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        rows = c.execute(
            """SELECT id, COALESCE(req_file_name,''), COALESCE(description,''), COALESCE(status,'Pending'),
                      COALESCE(linked_file_path,''), COALESCE(subfolder,'Main')
               FROM checklist_items
               WHERE project_id=?""",
            (self.project_id,)
        ).fetchall()

        # Current filesystem snapshot for this project.
        fs_index = {}
        for root, _dirs, files in os.walk(self.folder_path):
            rel_root = self._normalize_subfolder(os.path.relpath(root, self.folder_path))
            for fn in files:
                full = os.path.join(root, fn)
                try:
                    f_abs = os.path.normcase(os.path.abspath(full))
                except Exception:
                    continue
                fs_index[f_abs] = (rel_root, fn, full)

        to_delete_ids = set()
        linked_by_path = {}
        represented_paths = set()

        # Reconcile existing linked rows against filesystem.
        for rid, req_name, desc, status, linked_file_path, subfolder in rows:
            lp = str(linked_file_path or "").strip()
            if not lp:
                continue
            try:
                lp_abs = os.path.normcase(os.path.abspath(lp))
            except Exception:
                to_delete_ids.add(rid)
                continue

            if lp_abs in linked_by_path:
                to_delete_ids.add(rid)
                continue

            if lp_abs not in fs_index:
                to_delete_ids.add(rid)
                continue

            linked_by_path[lp_abs] = rid
            represented_paths.add(lp_abs)
            rel_root, fn, full = fs_index[lp_abs]

            needs_update = (
                str(req_name or "") != fn or
                self._normalize_subfolder(subfolder) != rel_root or
                str(lp) != str(full) or
                str(status or "") != "Completed"
            )
            if needs_update:
                c.execute(
                    """UPDATE checklist_items
                       SET req_file_name=?, subfolder=?, linked_file_path=?, status='Completed'
                       WHERE id=?""",
                    (fn, rel_root, full, rid)
                )

        removed = 0
        if to_delete_ids:
            q = ",".join(["?"] * len(to_delete_ids))
            c.execute(f"DELETE FROM checklist_items WHERE id IN ({q})", tuple(to_delete_ids))
            removed = len(to_delete_ids)

        # Add files that exist on disk but are not represented in checklist_items.
        count_row = c.execute("SELECT COUNT(*) FROM checklist_items WHERE project_id=?", (self.project_id,)).fetchone()
        next_sr = int((count_row[0] if count_row else 0) or 0) + 1
        inserted = 0
        for f_abs, (rel_root, fn, full) in sorted(fs_index.items(), key=lambda kv: kv[1][1].lower()):
            if f_abs in represented_paths:
                continue
            c.execute(
                """INSERT INTO checklist_items
                   (project_id, sr_no, req_file_name, description, subfolder, linked_file_path, status)
                   VALUES (?, ?, ?, ?, ?, ?, 'Completed')""",
                (self.project_id, next_sr, fn, "Auto-imported", rel_root, full)
            )
            next_sr += 1
            inserted += 1
        conn.commit()
        conn.close()
        self.load_project(self.project_id)
        if inserted or removed:
            messagebox.showinfo("Update", f"Synced project table. Added {inserted}, removed {removed}.")
        else:
            messagebox.showinfo("Update", "Project table already mirrors the project folder.")

    def _normalize_subfolder(self, rel):
        txt = str(rel or "").strip()
        if not txt or txt == "." or txt.lower() == "main":
            return "Main"
        txt = txt.replace("/", os.sep).replace("\\", os.sep)
        txt = os.path.normpath(txt).strip()
        if txt in ("", ".", os.sep):
            return "Main"
        if txt.startswith(os.sep):
            txt = txt.lstrip(os.sep)
        return txt

    def _list_project_subfolders(self):
        out = []
        if not os.path.isdir(self.folder_path):
            return out
        for root, dirs, _files in os.walk(self.folder_path):
            for d in dirs:
                rel = os.path.relpath(os.path.join(root, d), self.folder_path)
                rel = self._normalize_subfolder(rel)
                if rel != "Main":
                    out.append(rel)
        return sorted(set(out), key=lambda x: x.lower())

    def _is_same_or_child_folder(self, parent_rel, candidate_rel):
        p = self._normalize_subfolder(parent_rel)
        c = self._normalize_subfolder(candidate_rel)
        if p == "Main":
            return True
        return c == p or c.startswith(p + os.sep)

    def _apply_folder_mapping_for_checklist(self, old_rel, new_rel=None, delete_mode=False):
        old_rel = self._normalize_subfolder(old_rel)
        if old_rel == "Main":
            return
        new_rel = self._normalize_subfolder(new_rel) if new_rel is not None else "Main"
        old_abs = os.path.normcase(os.path.abspath(os.path.join(self.folder_path, old_rel)))
        new_abs = os.path.normcase(os.path.abspath(os.path.join(self.folder_path, new_rel)))

        conn = sqlite3.connect(DB_FILE)
        rows = conn.execute(
            """SELECT id, COALESCE(subfolder,'Main'), COALESCE(linked_file_path,''), COALESCE(status,'Pending')
               FROM checklist_items
               WHERE project_id=?""",
            (self.project_id,)
        ).fetchall()
        updates = []
        for rid, subfolder, linked_file_path, status in rows:
            sf = self._normalize_subfolder(subfolder)
            if not self._is_same_or_child_folder(old_rel, sf):
                continue
            if delete_mode:
                updates.append(( "Main", "", "Pending", rid ))
                continue

            rest = ""
            if sf != old_rel:
                rest = sf[len(old_rel):].lstrip("\\/")
            mapped_sf = new_rel if new_rel != "Main" else "Main"
            if rest:
                mapped_sf = os.path.join(mapped_sf, rest) if mapped_sf != "Main" else rest
            mapped_sf = self._normalize_subfolder(mapped_sf)

            mapped_link = str(linked_file_path or "").strip()
            if mapped_link:
                lp_abs = os.path.normcase(os.path.abspath(mapped_link))
                if lp_abs == old_abs or lp_abs.startswith(old_abs + os.sep):
                    suffix = lp_abs[len(old_abs):].lstrip("\\/")
                    mapped_link = os.path.join(new_abs, suffix) if suffix else new_abs
            updates.append((mapped_sf, mapped_link, status, rid))

        if updates:
            conn.executemany(
                "UPDATE checklist_items SET subfolder=?, linked_file_path=?, status=? WHERE id=?",
                updates
            )
            conn.commit()
        conn.close()

    def open_manage_folders_popup(self):
        if not self.project_id:
            return messagebox.showwarning("Manage Folders", "Open a project first.")
        if hasattr(self, "manage_folders_win") and self.manage_folders_win and self.manage_folders_win.winfo_exists():
            self.manage_folders_win.lift()
            return

        win = tk.Toplevel(self)
        win.title("Manage Folders")
        win.geometry("620x430")
        win.transient(self)
        self.manage_folders_win = win

        left = tk.Frame(win, padx=10, pady=10)
        left.pack(side="left", fill="both", expand=True)
        right = tk.Frame(win, padx=10, pady=10)
        right.pack(side="right", fill="y")

        tk.Label(left, text="Project Folders", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        lb = tk.Listbox(left, height=18)
        lb.pack(fill="both", expand=True, pady=(6, 6))

        add_name_var = tk.StringVar()
        add_parent_var = tk.StringVar(value="Main")
        rename_var = tk.StringVar()
        move_parent_var = tk.StringVar(value="Main")

        def refresh_lists():
            folders = self._list_project_subfolders()
            lb.delete(0, tk.END)
            for f in folders:
                lb.insert(tk.END, f)
            choices = ["Main"] + folders
            cb_add_parent["values"] = choices
            cb_move_parent["values"] = choices
            if add_parent_var.get() not in choices:
                add_parent_var.set("Main")
            if move_parent_var.get() not in choices:
                move_parent_var.set("Main")

        def get_selected():
            sel = lb.curselection()
            if not sel:
                return None
            return str(lb.get(sel[0]))

        def add_folder():
            name = str(add_name_var.get() or "").strip()
            parent = self._normalize_subfolder(add_parent_var.get())
            if not name:
                return messagebox.showwarning("Manage Folders", "Folder name is required.", parent=win)
            target_rel = self._normalize_subfolder(name if parent == "Main" else os.path.join(parent, name))
            target_abs = os.path.join(self.folder_path, target_rel)
            if os.path.exists(target_abs):
                return messagebox.showwarning("Manage Folders", "Folder already exists.", parent=win)
            os.makedirs(target_abs, exist_ok=True)
            add_name_var.set("")
            refresh_lists()
            self.load_project(self.project_id)

        def rename_folder():
            old_rel = get_selected()
            if not old_rel:
                return messagebox.showwarning("Manage Folders", "Select a folder to rename.", parent=win)
            new_name = str(rename_var.get() or "").strip()
            if not new_name:
                return messagebox.showwarning("Manage Folders", "New name is required.", parent=win)
            parent_rel = self._normalize_subfolder(os.path.dirname(old_rel))
            if parent_rel == ".":
                parent_rel = "Main"
            new_rel = self._normalize_subfolder(new_name if parent_rel == "Main" else os.path.join(parent_rel, new_name))
            if new_rel == old_rel:
                return
            old_abs = os.path.join(self.folder_path, old_rel)
            new_abs = os.path.join(self.folder_path, new_rel)
            if os.path.exists(new_abs):
                return messagebox.showwarning("Manage Folders", "Target folder name already exists.", parent=win)
            os.makedirs(os.path.dirname(new_abs), exist_ok=True)
            shutil.move(old_abs, new_abs)
            self._apply_folder_mapping_for_checklist(old_rel, new_rel=new_rel, delete_mode=False)
            rename_var.set("")
            refresh_lists()
            self.load_project(self.project_id)

        def move_folder():
            old_rel = get_selected()
            if not old_rel:
                return messagebox.showwarning("Manage Folders", "Select a folder to move.", parent=win)
            dest_parent = self._normalize_subfolder(move_parent_var.get())
            if self._is_same_or_child_folder(old_rel, dest_parent):
                return messagebox.showwarning("Manage Folders", "Cannot move folder inside itself.", parent=win)
            base = os.path.basename(old_rel)
            new_rel = self._normalize_subfolder(base if dest_parent == "Main" else os.path.join(dest_parent, base))
            if new_rel == old_rel:
                return
            old_abs = os.path.join(self.folder_path, old_rel)
            new_abs = os.path.join(self.folder_path, new_rel)
            if os.path.exists(new_abs):
                return messagebox.showwarning("Manage Folders", "Target location already has this folder.", parent=win)
            os.makedirs(os.path.dirname(new_abs), exist_ok=True)
            shutil.move(old_abs, new_abs)
            self._apply_folder_mapping_for_checklist(old_rel, new_rel=new_rel, delete_mode=False)
            refresh_lists()
            self.load_project(self.project_id)

        def delete_folder():
            old_rel = get_selected()
            if not old_rel:
                return messagebox.showwarning("Manage Folders", "Select a folder to delete.", parent=win)
            if not messagebox.askyesno(
                "Delete Folder",
                f"Delete folder '{old_rel}' and all its contents?\nChecklist items will be moved to Main as Pending.",
                parent=win
            ):
                return
            old_abs = os.path.join(self.folder_path, old_rel)
            if os.path.isdir(old_abs):
                shutil.rmtree(old_abs, ignore_errors=False)
            self._apply_folder_mapping_for_checklist(old_rel, delete_mode=True)
            refresh_lists()
            self.load_project(self.project_id)

        add_box = tk.LabelFrame(right, text="Add Folder", padx=8, pady=8)
        add_box.pack(fill="x", pady=(0, 8))
        tk.Entry(add_box, textvariable=add_name_var).pack(fill="x", pady=(0, 6))
        cb_add_parent = ttk.Combobox(add_box, textvariable=add_parent_var, state="readonly")
        cb_add_parent.pack(fill="x", pady=(0, 6))
        tk.Button(add_box, text="Add", bg="#2E7D32", fg="white", command=add_folder).pack(fill="x")

        rename_box = tk.LabelFrame(right, text="Rename Selected", padx=8, pady=8)
        rename_box.pack(fill="x", pady=(0, 8))
        tk.Entry(rename_box, textvariable=rename_var).pack(fill="x", pady=(0, 6))
        tk.Button(rename_box, text="Rename", bg="#1565C0", fg="white", command=rename_folder).pack(fill="x")

        move_box = tk.LabelFrame(right, text="Move Selected", padx=8, pady=8)
        move_box.pack(fill="x", pady=(0, 8))
        cb_move_parent = ttk.Combobox(move_box, textvariable=move_parent_var, state="readonly")
        cb_move_parent.pack(fill="x", pady=(0, 6))
        tk.Button(move_box, text="Move", bg="#6D4C41", fg="white", command=move_folder).pack(fill="x")

        tk.Button(right, text="Delete Selected", bg="#C62828", fg="white", command=delete_folder).pack(fill="x", pady=(8, 6))
        tk.Button(right, text="Refresh", command=refresh_lists).pack(fill="x", pady=(0, 6))
        tk.Button(right, text="Close", command=win.destroy).pack(fill="x")

        refresh_lists()

    def create_new_folder_popup(self):
        if hasattr(self, 'folder_popup') and self.folder_popup and self.folder_popup.winfo_exists(): self.folder_popup.lift(); return
        win = tk.Toplevel(self)
        win.title("Create New Folder"); win.geometry("400x200")
        self.folder_popup = win
        tk.Label(win, text="Folder Name:").pack(pady=5)
        e_name = tk.Entry(win); e_name.pack(pady=5)
        tk.Label(win, text="Create Inside (Parent):").pack(pady=5)
        cb_parent = ttk.Combobox(win, values=self.cb_folder['values']); cb_parent.set("Main"); cb_parent.pack(pady=5)
        def save_folder():
            name = e_name.get(); parent = cb_parent.get()
            if not name: return
            new_path = os.path.join(self.folder_path, name) if parent == "Main" else os.path.join(self.folder_path, parent, name)
            try: os.makedirs(new_path); messagebox.showinfo("Success", "Folder Created"); self.load_project(self.project_id); win.destroy()
            except Exception as e: messagebox.showerror("Error", str(e))
        tk.Button(win, text="Create", bg="#4CAF50", fg="white", command=save_folder).pack(pady=20)

    def on_select(self, event):
        sel = self.tree.selection()
        if not sel: return
        sel_iid = sel[0]
        item = self.tree.item(sel_iid)
        if not item['values'] or item['values'][4] == "":
            self.folder_var.set(item['values'][0]); self.e_doc_name.delete(0, tk.END); self.e_desc.delete(0, tk.END)
        else:
            self.e_doc_name.delete(0, tk.END); self.e_doc_name.insert(0, item['values'][0])
            self.e_desc.delete(0, tk.END); self.e_desc.insert(0, item['values'][1])
            parent_id = self.tree.parent(sel_iid); parent_vals = self.tree.item(parent_id)['values']
            if parent_vals: self.folder_var.set(parent_vals[0])
            self.load_preview(item['values'][4])

    def add_item(self):
        name = self.e_doc_name.get(); desc = self.e_desc.get(); folder = self.folder_var.get()
        if not name: return messagebox.showerror("Error", "Document Name is required")
        folder_full_path = self.folder_path if folder == "Main" else os.path.join(self.folder_path, folder)
        if not os.path.exists(folder_full_path): os.makedirs(folder_full_path, exist_ok=True)
        conn = sqlite3.connect(DB_FILE)
        count = conn.execute("SELECT count(*) FROM checklist_items WHERE project_id=?", (self.project_id,)).fetchone()[0]
        conn.execute("INSERT INTO checklist_items (project_id, sr_no, req_file_name, description, subfolder) VALUES (?,?,?,?,?)", (self.project_id, count+1, name, desc, folder))
        conn.commit(); conn.close(); self.load_project(self.project_id); self.e_doc_name.delete(0, tk.END); self.e_desc.delete(0, tk.END)

    def update_item(self):
        sel = self.tree.selection()
        if not sel: return
        sel_iid = sel[0]
        item_data = self.tree.item(sel_iid)
        if not item_data['values'][4]: return
        conn = sqlite3.connect(DB_FILE)
        conn.execute("UPDATE checklist_items SET req_file_name=?, description=?, subfolder=? WHERE id=?", (self.e_doc_name.get(), self.e_desc.get(), self.folder_var.get(), item_data['values'][4]))
        conn.commit(); conn.close(); self.load_project(self.project_id)

    def upload_file(self):
        sel = self.tree.selection()
        if not sel: return
        sel_iid = sel[0]
        item_data = self.tree.item(sel_iid)
        if not item_data['values'][4]: return
        parent_id = self.tree.parent(sel_iid); parent_vals = self.tree.item(parent_id)['values']
        subfolder = parent_vals[0] if parent_vals else "Main"
        src_path = filedialog.askopenfilename()
        if src_path:
            dest_folder = self.folder_path if subfolder == "Main" else os.path.join(self.folder_path, subfolder)
            if not os.path.exists(dest_folder): os.makedirs(dest_folder)
            dest_path = os.path.join(dest_folder, os.path.basename(src_path))
            try:
                shutil.copy(src_path, dest_path)
                conn = sqlite3.connect(DB_FILE)
                conn.execute("UPDATE checklist_items SET linked_file_path=?, status='Completed' WHERE id=?", (dest_path, item_data['values'][4]))
                conn.commit(); conn.close(); self.load_project(self.project_id)
            except Exception as e: messagebox.showerror("Error", str(e))

    def open_file(self):
        sel = self.tree.selection()
        if not sel: return
        item_data = self.tree.item(sel[0])
        if not item_data['values'][4]: return
        conn = sqlite3.connect(DB_FILE)
        row = conn.execute("SELECT linked_file_path FROM checklist_items WHERE id=?", (item_data['values'][4],)).fetchone()
        conn.close()
        if row and row[0] and os.path.exists(row[0]):
            if platform.system() == "Windows": os.startfile(row[0])
            elif platform.system() == "Darwin": subprocess.Popen(["open", row[0]])
            else: subprocess.Popen(["xdg-open", row[0]])
        else: messagebox.showinfo("Info", "No file attached.")

    def delete_item(self):
        sel = self.tree.selection()
        if not sel: return
        sel_iid = sel[0]
        item_data = self.tree.item(sel_iid)
        vals = item_data.get("values", [])
        if not vals:
            return
        if "folder" in set(item_data.get("tags", ())):
            return messagebox.showwarning("Warning", "Cannot delete folders from here.")
        if not messagebox.askyesno("Confirm", "Delete selected file/item?"):
            return

        # Close current PDF before deletion to avoid Windows file-lock issues.
        try:
            if getattr(self, "current_pdf", None):
                self.current_pdf.close()
        except Exception:
            pass
        self.current_pdf = None

        linked_id = str(vals[4] or "").strip() if len(vals) > 4 else ""
        doc_name = str(vals[0] or "").strip() if len(vals) > 0 else ""
        attachment_name = str(vals[3] or "").strip() if len(vals) > 3 else ""
        target_file = ""

        parent_iid = self.tree.parent(sel_iid)
        parent_vals = self.tree.item(parent_iid).get("values", []) if parent_iid else []
        subfolder = str(parent_vals[0] or "").strip() if parent_vals else "Main"
        folder_abs = self.folder_path if subfolder in ("", "Main") else os.path.join(self.folder_path, subfolder)

        conn = sqlite3.connect(DB_FILE)
        try:
            if linked_id:
                row = conn.execute("SELECT linked_file_path FROM checklist_items WHERE id=?", (linked_id,)).fetchone()
                if row and row[0]:
                    target_file = str(row[0]).strip()
                conn.execute("DELETE FROM checklist_items WHERE id=?", (linked_id,))
                conn.commit()
            else:
                fname = attachment_name or doc_name
                if fname:
                    target_file = os.path.join(folder_abs, fname)
        finally:
            conn.close()

        if target_file:
            try:
                if os.path.isfile(target_file):
                    os.remove(target_file)
            except Exception as e:
                messagebox.showwarning("Warning", f"Item was removed from table, but file delete failed:\n{e}")
        self.load_project(self.project_id)

    def open_explorer(self):
        if platform.system() == "Windows": os.startfile(self.folder_path)
        elif platform.system() == "Darwin": subprocess.Popen(["open", self.folder_path])
        else: subprocess.Popen(["xdg-open", self.folder_path])

    def toggle_preview(self):
        if self.preview_visible: self.animate_close()
        else: self.animate_open()

    def draw_toggle_button(self, state):
        self.btn_toggle.delete("all")
        w, h = 15, 60
        points = [0, 10, w, 0, w, h, 0, h-10]
        self.btn_toggle.create_polygon(points, fill="#808080", outline="#606060")
        text = ">" if state == "close" else "<"
        self.btn_toggle.create_text(w/2, h/2, text=text, fill="white", font=("Segoe UI", 10, "bold"))

    def animate_open(self):
        self.paned_root.add(self.preview_pane, minsize=0, stretch="never"); self.paned_root.update_idletasks()
        w = self.paned_root.winfo_width(); self.paned_root.sash_place(0, w, 0); self.target_sash = w - 416; self.preview_visible = True
        self.draw_toggle_button("close"); self.animate_step(-25)

    def animate_close(self):
        self.target_sash = self.paned_root.winfo_width(); self.preview_visible = False
        self.draw_toggle_button("open"); self.animate_step(25)

    def animate_step(self, step):
        try: curr = self.paned_root.sash_coord(0)[0]
        except: return
        new_pos = curr + step
        if (step < 0 and new_pos <= self.target_sash) or (step > 0 and new_pos >= self.target_sash):
            self.paned_root.sash_place(0, self.target_sash, 0)
            if step > 0: self.paned_root.forget(self.preview_pane)
            return
        self.paned_root.sash_place(0, new_pos, 0); self.after(10, lambda: self.animate_step(step))

    def load_preview(self, item_id):
        if not self.preview_visible: return
        conn = sqlite3.connect(DB_FILE)
        path = conn.execute("SELECT linked_file_path FROM checklist_items WHERE id=?", (item_id,)).fetchone()
        conn.close()
        try:
            if getattr(self, "current_pdf", None):
                self.current_pdf.close()
        except Exception:
            pass
        self.current_pdf = None
        self.pv_label.config(image='', text="No file attached")
        if path and path[0] and os.path.exists(path[0]):
            fpath = path[0]; ext = os.path.splitext(fpath)[1].lower()
            if ext == ".pdf":
                if not PDF_SUPPORT: return self.pv_label.config(text="Install 'pymupdf' to view PDFs.")
                try:
                    self.current_pdf = fitz.open(fpath); self.current_page = 0
                    self.pv_total.config(text=str(len(self.current_pdf))); self.show_pdf_page()
                except Exception as e: self.pv_label.config(text=f"Error reading PDF:\n{e}")
            else: self.pv_label.config(text=f"Preview not available for {ext} files.")

    def show_pdf_page(self):
        if not self.current_pdf: return
        try:
            page = self.current_pdf.load_page(self.current_page)
            zoom = 464 / page.rect.width
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            tk_img = ImageTk.PhotoImage(img)
            self.pv_label.config(image=tk_img, text=""); self.pv_label.image = tk_img
            self.pv_page.delete(0, tk.END); self.pv_page.insert(0, str(self.current_page + 1))
        except Exception as e: self.pv_label.config(text=f"Error rendering page: {e}")

    def prev_page(self):
        if self.current_pdf and self.current_page > 0: self.current_page -= 1; self.show_pdf_page()
    def next_page(self):
        if self.current_pdf and self.current_page < len(self.current_pdf) - 1: self.current_page += 1; self.show_pdf_page()

# --- MAIN APP ---
class TenderApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tender & Bid Manager Pro")
        self.ui_scale = max(1.0, float(self.winfo_fpixels("1i")) / 96.0)
        scr_w = max(1024, self.winfo_screenwidth())
        if scr_w >= 2560:
            self.column_width_scale = 1.28
        elif scr_w >= 1920:
            self.column_width_scale = 1.18
        elif scr_w >= 1600:
            self.column_width_scale = 1.10
        else:
            self.column_width_scale = 1.0

        base_w = int(round(1280 * self.ui_scale))
        base_h = int(round(850 * self.ui_scale))
        init_w = min(scr_w - 40, base_w)
        init_h = min(max(720, self.winfo_screenheight() - 80), base_h)
        self.geometry(f"{init_w}x{init_h}")
        try: self.state("zoomed")
        except: pass
        
        self.create_project_win = None
        self.captcha_win = None
        self.archive_job_running = False

        # Styles
        style = ttk.Style()
        style.theme_use('clam')
        row_h = int(round(30 * self.ui_scale))
        tender_row_h = int(round(44 * self.ui_scale))
        style.configure("Treeview", rowheight=row_h, font=("Segoe UI", 10))
        style.configure("Wrapped.Treeview", rowheight=tender_row_h, font=("Segoe UI", 10))
        style.configure("Org.Treeview", rowheight=row_h, font=("Segoe UI", 10))
        style.configure("Dashboard.Treeview", rowheight=row_h, font=("Segoe UI", 10))
        style.configure("Tender.Treeview", rowheight=tender_row_h, font=("Segoe UI", 10))
        style.configure("ActiveTender.Treeview", rowheight=tender_row_h, font=("Segoe UI", 10))
        style.configure("ArchivedTender.Treeview", rowheight=tender_row_h, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#cfd8dc")
        style.configure("TLabelFrame", font=("Segoe UI", 11, "bold"))
        
        # Main Layout
        self.container = tk.Frame(self)
        self.container.pack(fill="both", expand=True)
        self.container.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(1, weight=1)

        # 1. SIDEBAR
        self.sidebar = tk.Frame(self.container, bg="#263238", width=220)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.pack_propagate(False)

        tk.Label(self.sidebar, text="BID MANAGER", bg="#263238", fg="white", font=("Segoe UI", 16, "bold"), pady=30).pack(fill="x")

        self.btn_dash = tk.Button(self.sidebar, text="Projects", bg="#37474F", fg="white", relief="flat", font=("Segoe UI", 11), padx=20, pady=10, anchor="w", command=lambda: self.show_frame("Projects"))
        self.btn_dash.pack(fill="x", pady=2)
        
        # NEW BUTTON
        self.btn_online = tk.Button(self.sidebar, text="Online Tenders", bg="#37474F", fg="white", relief="flat", font=("Segoe UI", 11), padx=20, pady=10, anchor="w", command=lambda: self.show_frame("ViewTenders"))
        self.btn_online.pack(fill="x", pady=2)

        # Spacer keeps Settings pinned at the bottom of the sidebar.
        tk.Frame(self.sidebar, bg="#263238").pack(fill="both", expand=True)
        self.btn_settings = tk.Button(self.sidebar, text="⚙ Settings", bg="#37474F", fg="white", relief="flat", font=("Segoe UI", 11), padx=20, pady=10, anchor="w", command=self.open_settings_popup)
        self.btn_settings.pack(fill="x", pady=2, side="bottom")

        # 2. CONTENT AREA
        self.frames = {}
        for F in (Dashboard, ProjectDetails, ViewTenders):
            page_name = "Projects" if F.__name__ == "Dashboard" else F.__name__
            frame = F(parent=self.container, controller=self)
            self.frames[page_name] = frame
            frame.grid(row=0, column=1, sticky="nsew")

        start_page = str(get_user_setting("last_main_view", "Projects") or "Projects")
        if start_page == "Dashboard":
            start_page = "Projects"
        if start_page not in ("Projects", "ViewTenders"):
            start_page = "Projects"
        self.show_frame(start_page)
        
        # Start Polling for Captcha/Logs
        self.check_queues()
        self.start_daily_archive_scheduler()

    def check_queues(self):
        # 1. Logs
        try:
            while True:
                msg = log_queue.get_nowait()
                if "ViewTenders" in self.frames:
                    self.frames["ViewTenders"].append_log(msg)
        except queue.Empty: pass
        
        # 2. Captcha Requests
        try:
            img_data = captcha_req_queue.get_nowait()
            self.open_captcha_popup(img_data)
        except queue.Empty: pass
        
        self.after(200, self.check_queues)

    def start_daily_archive_scheduler(self):
        # Start checks after app has been running for 1 hour, then repeat hourly.
        self.after(60 * 60 * 1000, self._archive_scheduler_tick)

    def _archive_scheduler_tick(self):
        # Hourly check; actual execution is gated by a 12-hour timestamp.
        self.run_daily_archive_if_due()
        self.after(60 * 60 * 1000, self._archive_scheduler_tick)

    def run_daily_archive_if_due(self):
        if self.archive_job_running:
            return
        last_run_raw = ScraperBackend.get_setting("last_auto_archive_utc")
        due = True
        if last_run_raw:
            try:
                last_run = datetime.datetime.fromisoformat(last_run_raw)
                if last_run.tzinfo is None:
                    # Backward compatibility for older naive timestamps.
                    last_run = last_run.replace(tzinfo=datetime.UTC)
                due = (datetime.datetime.now(datetime.UTC) - last_run) >= datetime.timedelta(hours=12)
            except Exception:
                due = True
        if due:
            self.archive_job_running = True
            threading.Thread(target=self._daily_archive_worker, daemon=True).start()

    def _pick_dir_into_var(self, var, parent=None):
        start = str(var.get() or "").strip() or _safe_getcwd()
        picked = filedialog.askdirectory(parent=parent, initialdir=start)
        if picked:
            var.set(picked)

    def open_settings_popup(self):
        win = tk.Toplevel(self)
        win.title("Settings")
        win.geometry("780x320")
        win.transient(self)
        win.grab_set()

        body = tk.Frame(win, padx=14, pady=12)
        body.pack(fill="both", expand=True)
        body.grid_columnconfigure(1, weight=1)

        tk.Label(body, text="Tender Database Folder:", anchor="w").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=(0, 10))
        db_dir_var = tk.StringVar(value=os.path.dirname(_resolve_path(DB_FILE)) or _safe_getcwd())
        tk.Entry(body, textvariable=db_dir_var).grid(row=0, column=1, sticky="ew", pady=(0, 10))
        tk.Button(body, text="Browse", command=lambda: self._pick_dir_into_var(db_dir_var, win)).grid(row=0, column=2, padx=(8, 0), pady=(0, 10))

        tk.Label(body, text="My Tender Projects Folder:", anchor="w").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(0, 10))
        proj_var = tk.StringVar(value=_resolve_path(ROOT_FOLDER))
        tk.Entry(body, textvariable=proj_var).grid(row=1, column=1, sticky="ew", pady=(0, 10))
        tk.Button(body, text="Browse", command=lambda: self._pick_dir_into_var(proj_var, win)).grid(row=1, column=2, padx=(8, 0), pady=(0, 10))

        tk.Label(body, text="Tender Downloads Folder:", anchor="w").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=(0, 10))
        down_var = tk.StringVar(value=_resolve_path(BASE_DOWNLOAD_DIRECTORY))
        tk.Entry(body, textvariable=down_var).grid(row=2, column=1, sticky="ew", pady=(0, 10))
        tk.Button(body, text="Browse", command=lambda: self._pick_dir_into_var(down_var, win)).grid(row=2, column=2, padx=(8, 0), pady=(0, 10))

        tk.Label(
            body,
            text="Database path uses 'tender_manager.db' inside the selected DB folder.",
            fg="#555",
            anchor="w",
            justify="left"
        ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(2, 12))

        btns = tk.Frame(body)
        btns.grid(row=4, column=0, columnspan=3, sticky="ew")

        def save_settings():
            db_dir = str(db_dir_var.get() or "").strip()
            proj_dir = str(proj_var.get() or "").strip()
            down_dir = str(down_var.get() or "").strip()
            if not db_dir or not proj_dir or not down_dir:
                return messagebox.showerror("Settings", "All three paths are required.", parent=win)
            new_db_file = _normalize_db_file(db_dir)
            new_root = _resolve_path(proj_dir)
            new_down = _resolve_path(down_dir)
            try:
                os.makedirs(os.path.dirname(new_db_file) or ".", exist_ok=True)
                os.makedirs(new_root, exist_ok=True)
                os.makedirs(new_down, exist_ok=True)
                save_app_paths_config(new_db_file, new_root, new_down)
                global DB_FILE, ROOT_FOLDER, BASE_DOWNLOAD_DIRECTORY
                old_db = DB_FILE
                DB_FILE = new_db_file
                ROOT_FOLDER = new_root
                BASE_DOWNLOAD_DIRECTORY = new_down
                if _resolve_path(old_db) != _resolve_path(DB_FILE):
                    init_db()
                win.destroy()
                messagebox.showinfo("Settings", "Paths updated successfully.")
            except Exception as e:
                messagebox.showerror("Settings", f"Failed to save settings:\n{e}", parent=win)

        tk.Button(btns, text="Save", bg="#2e7d32", fg="white", command=save_settings).pack(side="left")
        tk.Button(btns, text="Cancel", command=win.destroy).pack(side="right")

    def _daily_archive_worker(self):
        total = 0
        websites_count = 0
        try:
            websites = ScraperBackend.get_websites()
            websites_count = len(websites)
            for sid in websites.keys():
                total += int(ScraperBackend.archive_completed_tenders_logic(sid) or 0)
            ScraperBackend.set_setting("last_auto_archive_utc", datetime.datetime.now(datetime.UTC).isoformat())
            ScraperBackend.log_auto_archive_run(
                status="success",
                archived_count=total,
                archived_status_updated=0,
                websites_count=websites_count,
                notes="12-hour scheduled run"
            )
            log_to_gui(
                f"Auto-archive complete (12-hour schedule). Archived {total} tenders."
            )
            if "ViewTenders" in self.frames:
                self.after(0, self.frames["ViewTenders"].on_site_changed)
        except Exception as e:
            ScraperBackend.log_auto_archive_run(
                status="failed",
                archived_count=total,
                archived_status_updated=0,
                websites_count=websites_count,
                notes=f"12-hour scheduled run failed: {e}"
            )
            log_to_gui(f"Auto-archive failed: {e}")
        finally:
            self.archive_job_running = False

    def open_captcha_popup(self, img_data):
        if self.captcha_win and self.captcha_win.winfo_exists():
            self.captcha_win.destroy()
        
        win = tk.Toplevel(self)
        win.title("Solve Captcha")
        win.geometry("300x250")
        win.attributes('-topmost', True)
        self.captcha_win = win
        
        # Display Image
        try:
            image = Image.open(io.BytesIO(img_data))
            photo = ImageTk.PhotoImage(image)
            lbl_img = tk.Label(win, image=photo)
            lbl_img.image = photo
            lbl_img.pack(pady=10)
        except:
            tk.Label(win, text="[Image Error]").pack(pady=10)
            
        tk.Label(win, text="Enter Captcha:").pack()
        e_cap = tk.Entry(win, font=("Arial", 12))
        e_cap.pack(pady=5)
        e_cap.focus_set()
        
        def submit():
            val = e_cap.get()
            win.destroy()
            captcha_res_queue.put(val) # Send back to thread
            
        def cancel():
            win.destroy()
            captcha_res_queue.put(None)

        tk.Button(win, text="Submit", command=submit, bg="#4CAF50", fg="white").pack(pady=10, fill="x", padx=20)
        win.protocol("WM_DELETE_WINDOW", cancel)

    def show_frame(self, page_name, data=None):
        if page_name == "Dashboard":
            page_name = "Projects"
        frame = self.frames[page_name]
        if page_name == "Projects": frame.load_projects()
        elif page_name == "ProjectDetails" and data: frame.load_project(data)
        frame.tkraise()
        if page_name in ("Projects", "ViewTenders"):
            set_user_setting("last_main_view", page_name)

    def open_create_project_window(self, prefill=None):
        if self.create_project_win is not None and self.create_project_win.winfo_exists():
            self.create_project_win.lift()
            return
        prefill = prefill or {}
        win = tk.Toplevel(self)
        win.title("Create New Tender Project")
        win.geometry("620x580")
        self.create_project_win = win
        
        tk.Label(win, text="Project Title:", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20, pady=(20,5))
        e_title = tk.Entry(win)
        e_title.pack(fill="x", padx=20)
        if prefill.get("title"):
            e_title.insert(0, str(prefill.get("title")))

        tk.Label(win, text="Client / Authority:", font=("Segoe UI", 10)).pack(anchor="w", padx=20, pady=(10,5))
        client_row = tk.Frame(win)
        client_row.pack(fill="x", padx=20)
        client_var = tk.StringVar()
        cb_client = ttk.Combobox(client_row, textvariable=client_var, state="readonly")
        cb_client.pack(side="left", fill="x", expand=True)
        cb_client.bind("<Key>", on_combo_key_cycling)

        def load_client_options():
            opts = []
            raw = ScraperBackend.get_setting("project_client_options", "[]")
            try:
                opts = [str(x).strip() for x in json.loads(raw) if str(x).strip()]
            except Exception:
                opts = []
            conn = sqlite3.connect(DB_FILE)
            try:
                rows = conn.execute(
                    "SELECT DISTINCT TRIM(client_name) FROM projects WHERE TRIM(COALESCE(client_name,''))!='' ORDER BY client_name"
                ).fetchall()
                opts.extend([r[0] for r in rows if r and r[0]])
            finally:
                conn.close()
            opts = sorted(set(opts), key=lambda x: x.lower())
            cb_client["values"] = opts
            preferred_client = str(prefill.get("client_name", "")).strip()
            if preferred_client and preferred_client not in opts:
                opts.append(preferred_client)
                opts = sorted(set(opts), key=lambda x: x.lower())
                cb_client["values"] = opts
                save_client_options(opts)
            if preferred_client:
                client_var.set(preferred_client)
            elif opts and not client_var.get():
                client_var.set(opts[0])
            return opts

        def save_client_options(options):
            clean = sorted(set([str(x).strip() for x in options if str(x).strip()]), key=lambda x: x.lower())
            ScraperBackend.set_setting("project_client_options", json.dumps(clean))

        def manage_clients_popup():
            cwin = tk.Toplevel(win)
            cwin.title("Manage Client / Authority List")
            cwin.geometry("520x460")
            cwin.minsize(460, 360)

            header = tk.Frame(cwin, bg="#ECEFF1", height=44)
            header.pack(fill="x")
            header.pack_propagate(False)
            tk.Label(
                header,
                text="Manage Client / Authority List",
                bg="#ECEFF1",
                fg="#263238",
                font=("Segoe UI", 11, "bold")
            ).pack(side="left", padx=12, pady=10)

            body = tk.Frame(cwin, padx=12, pady=10)
            body.pack(fill="both", expand=True)
            tk.Label(body, text="Clients / Authorities", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))

            frame = tk.Frame(body, bd=1, relief="solid")
            frame.pack(fill="both", expand=True)
            lst = tk.Listbox(frame, selectmode=tk.SINGLE, font=("Segoe UI", 10), activestyle="none")
            ysb = ttk.Scrollbar(frame, orient="vertical", command=lst.yview)
            lst.configure(yscrollcommand=ysb.set)
            lst.pack(side="left", fill="both", expand=True)
            ysb.pack(side="right", fill="y")

            for opt in cb_client["values"]:
                lst.insert(tk.END, opt)

            def add_client():
                name = simpledialog.askstring("Add Client / Authority", "Enter client / authority name:", parent=cwin)
                if name is None:
                    return
                name = name.strip()
                if not name:
                    messagebox.showwarning("Add Client / Authority", "Name cannot be blank.", parent=cwin)
                    return
                existing = list(lst.get(0, tk.END))
                if name not in existing:
                    lst.insert(tk.END, name)
                    updated = sorted(set(list(lst.get(0, tk.END))), key=lambda x: x.lower())
                    lst.delete(0, tk.END)
                    for item in updated:
                        lst.insert(tk.END, item)
                    save_client_options(updated)
                    load_client_options()
                else:
                    messagebox.showinfo("Add Client / Authority", "This name already exists.", parent=cwin)

            def delete_client():
                sel = lst.curselection()
                if not sel:
                    messagebox.showwarning("Delete Client / Authority", "Select a name first.", parent=cwin)
                    return
                name = lst.get(sel[0])
                if not messagebox.askyesno("Delete Client / Authority", f"Delete '{name}' from the list?", parent=cwin):
                    return
                lst.delete(sel[0])
                updated = list(lst.get(0, tk.END))
                save_client_options(updated)
                load_client_options()

            action_row = tk.Frame(body)
            action_row.pack(fill="x", pady=(10, 0))
            tk.Button(action_row, text="Add", bg="#2e7d32", fg="white", width=14, command=add_client).pack(side="left")
            tk.Button(action_row, text="Delete Selected", bg="#d32f2f", fg="white", width=14, command=delete_client).pack(side="left", padx=8)
            tk.Button(action_row, text="Close", width=14, command=cwin.destroy).pack(side="right")

        tk.Button(client_row, text="Manage", bg="#607D8B", fg="white", command=manage_clients_popup).pack(side="left", padx=(8, 0))
        load_client_options()

        tk.Label(win, text="Submission Deadline:", font=("Segoe UI", 10)).pack(anchor="w", padx=20, pady=(10,5))
        deadline_row = tk.Frame(win)
        deadline_row.pack(fill="x", padx=20)

        today = datetime.datetime.now()
        if CALENDAR_SUPPORT:
            date_widget = DateEntry(
                deadline_row,
                date_pattern="dd-mm-yyyy",
                year=today.year,
                month=today.month,
                day=today.day,
                state="readonly",
                width=14
            )
            date_widget.pack(side="left")
        else:
            tk.Label(deadline_row, text="DD").pack(side="left")
            day_var = tk.StringVar(value=today.strftime("%d"))
            cb_day = ttk.Combobox(deadline_row, width=4, state="readonly", textvariable=day_var, values=[f"{d:02d}" for d in range(1, 32)])
            cb_day.pack(side="left", padx=(4, 8))
            tk.Label(deadline_row, text="MM").pack(side="left")
            month_var = tk.StringVar(value=today.strftime("%m"))
            cb_month = ttk.Combobox(deadline_row, width=4, state="readonly", textvariable=month_var, values=[f"{m:02d}" for m in range(1, 13)])
            cb_month.pack(side="left", padx=(4, 8))
            tk.Label(deadline_row, text="YYYY").pack(side="left")
            year_var = tk.StringVar(value=today.strftime("%Y"))
            cb_year = ttk.Combobox(deadline_row, width=6, state="readonly", textvariable=year_var, values=[str(y) for y in range(today.year - 2, today.year + 8)])
            cb_year.pack(side="left", padx=(4, 12))

        tk.Label(deadline_row, text="Time:").pack(side="left", padx=(12, 4))
        hour_var = tk.StringVar(value=today.strftime("%I"))
        cb_hour = ttk.Combobox(deadline_row, width=4, state="readonly", textvariable=hour_var, values=[f"{h:02d}" for h in range(1, 13)])
        cb_hour.pack(side="left")
        tk.Label(deadline_row, text=":").pack(side="left")
        minute_var = tk.StringVar(value=today.strftime("%M"))
        cb_min = ttk.Combobox(deadline_row, width=4, state="readonly", textvariable=minute_var, values=[f"{m:02d}" for m in range(0, 60)])
        cb_min.pack(side="left")
        ampm_var = tk.StringVar(value=today.strftime("%p"))
        cb_ampm = ttk.Combobox(deadline_row, width=5, state="readonly", textvariable=ampm_var, values=["AM", "PM"])
        cb_ampm.pack(side="left", padx=(8, 0))

        def apply_prefill_deadline(deadline_text):
            if not deadline_text:
                return
            dt = None
            txt = str(deadline_text).strip()
            formats = [
                "%d-%b-%Y %I:%M %p",
                "%d-%b-%Y",
                "%d-%m-%Y %I:%M %p",
                "%d-%m-%Y",
                "%d/%m/%Y %I:%M %p",
                "%d/%m/%Y",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d",
            ]
            for fmt in formats:
                try:
                    dt = datetime.datetime.strptime(txt, fmt)
                    break
                except Exception:
                    continue
            if dt is None:
                m = re.search(
                    r"(\d{1,2})[-/ ]([A-Za-z]{3}|\d{1,2})[-/ ](\d{2,4})(?:\s+(\d{1,2}):(\d{2})\s*(AM|PM)?)?",
                    txt,
                    flags=re.IGNORECASE
                )
                if m:
                    dd = int(m.group(1))
                    mm_raw = m.group(2)
                    yy = int(m.group(3))
                    if yy < 100:
                        yy += 2000
                    if mm_raw.isdigit():
                        mm = int(mm_raw)
                    else:
                        try:
                            mm = datetime.datetime.strptime(mm_raw[:3].title(), "%b").month
                        except Exception:
                            mm = today.month
                    hh = int(m.group(4)) if m.group(4) else int(today.strftime("%I"))
                    mi = int(m.group(5)) if m.group(5) else int(today.strftime("%M"))
                    ap = (m.group(6) or today.strftime("%p")).upper()
                    try:
                        dt = datetime.datetime.strptime(f"{dd:02d}-{mm:02d}-{yy} {hh:02d}:{mi:02d} {ap}", "%d-%m-%Y %I:%M %p")
                    except Exception:
                        dt = None
            if dt is None:
                return
            if CALENDAR_SUPPORT:
                try:
                    date_widget.set_date(dt.date())
                except Exception:
                    pass
            else:
                day_var.set(f"{dt.day:02d}")
                month_var.set(f"{dt.month:02d}")
                year_var.set(f"{dt.year:04d}")
            hour_var.set(dt.strftime("%I"))
            minute_var.set(dt.strftime("%M"))
            ampm_var.set(dt.strftime("%p"))

        apply_prefill_deadline(prefill.get("deadline"))

        tk.Label(win, text="Description:", font=("Segoe UI", 10)).pack(anchor="w", padx=20, pady=(10,5))
        e_desc = tk.Text(win, height=4)
        e_desc.pack(fill="x", padx=20)
        if prefill.get("description"):
            e_desc.insert("1.0", str(prefill.get("description")))

        def save():
            title = e_title.get()
            if not title: return messagebox.showerror("Error", "Title required")
            client_name = client_var.get().strip()
            if not client_name:
                return messagebox.showerror("Error", "Client / Authority is required")

            if CALENDAR_SUPPORT:
                date_part = date_widget.get()
            else:
                date_part = f"{day_var.get()}-{month_var.get()}-{year_var.get()}"
            deadline = f"{date_part} {hour_var.get()}:{minute_var.get()} {ampm_var.get()}"
            
            safe_title = "".join(c for c in title if c.isalnum() or c in (' ', '_')).strip()
            folder_path = os.path.join(ROOT_FOLDER, safe_title)
            std_folders = ensure_project_standard_folders(folder_path)

            source_tender_id = str(prefill.get("tender_id", "") or "").strip()
            project_value = str(prefill.get("project_value", "") or "").strip()
            prebid = str(prefill.get("prebid", "") or "").strip()
            tender_folder_path = str(prefill.get("tender_folder_path", "") or "").strip()
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            if source_tender_id:
                exists = c.execute(
                    "SELECT id FROM projects WHERE source_tender_id=? LIMIT 1",
                    (source_tender_id,)
                ).fetchone()
                if exists:
                    conn.close()
                    return messagebox.showinfo(
                        "Duplicate Tender",
                        f"A project for Tender ID '{source_tender_id}' already exists."
                    )
            c.execute(
                "INSERT INTO projects (title, client_name, deadline, description, folder_path, source_tender_id, project_value, prebid) VALUES (?,?,?,?,?,?,?,?)",
                (title, client_name, deadline, e_desc.get("1.0", tk.END).strip(), folder_path, source_tender_id or None, project_value, prebid)
            )
            pid = c.lastrowid
            
            conn.commit()
            conn.close()
            if tender_folder_path and os.path.isdir(tender_folder_path):
                copied = copy_tree_contents(tender_folder_path, std_folders["tender_docs"])
                if copied:
                    log_to_gui(f"Copied {copied} tender item(s) to project Tender Docs: {title}")
            # Keep custom list in sync with new entries.
            current = list(cb_client["values"])
            if client_name and client_name not in current:
                current.append(client_name)
                save_client_options(current)
            win.destroy()
            self.show_frame("Projects")

        tk.Button(win, text="Create Project", bg="#4CAF50", fg="white", command=save).pack(fill="x", padx=20, pady=20)

if __name__ == "__main__":
    init_db()
    app = TenderApp()
    app.mainloop()
